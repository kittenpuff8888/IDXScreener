import os
import time
import traceback
from datetime import datetime
from typing import Optional, Dict, List, Tuple, Any

import numpy as np
import pandas as pd
import yfinance as yf
# =========================
# USER CONFIG
# =========================
# Set the market date for this screener run.
# All data is clipped to this date — acts as the effective "as of" date.
MARKET_DATE = os.environ.get("MARKET_DATE", datetime.now().strftime("%Y-%m-%d"))
# =========================
# BACKTEST MODE CONFIG
# =========================
# When True: restricts the ticker universe to BACKTEST_TICKERS only.
# Data fetch and clipping behavior is unchanged — MARKET_DATE still controls the cutoff.
BACKTEST_MODE = False

BACKTEST_TICKERS = [
    "AADI","ADMR","ADRO","AMMN","AMRT","ANTM","ARCI","ASII","BBCA","BBNI",
    "BBRI","BIPI","BMRI","BNBR","BREN","BRMS","BRPT","BUMI","BUVA","CDIA",
    "CUAN","DEWA","DSSA","EMAS","ENRG","ESSA","EXCL","GOTO","HRTA","IMPC",
    "INCO","INDF","INDY","INKP","ITMG","MBMA","MDKA","MEDC","NCKL","PGAS",
    "PSAB","PTBA","PTRO","RAJA","RATU","TAPG","TCPI","TINS","TLKM","TPIA",
    "UNTR","VKTR","WIFI",
]

# =========================
# COMPATIBILITY CONSTANTS (v1.6.4 safe layer)
# =========================
# These placeholders are defined early so top-level sheet schema declarations never fail.
FILL_GROUP_FLOW = None
FILL_GROUP_MS = None
FILL_HEADER = None
FONT_NOTE = None
FONT_HEADER = None


from openpyxl import Workbook
from openpyxl.styles.fills import Fill
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# Post-import safe placeholders for style compatibility
FILL_GROUP_FLOW = PatternFill("solid", fgColor="2A4A66")  # post-import themed init
FILL_GROUP_MS = PatternFill("solid", fgColor="5B3F8C")  # post-import themed init for market structure
FILL_HEADER = PatternFill("solid", fgColor="2A4A66")  # post-import themed init

from openpyxl.utils import get_column_letter
from openpyxl.chart import ScatterChart, Reference, Series

# =========================
# CONFIG
# =========================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
RAW_DIR = os.path.join(BASE_DIR, "Raw")
OUTPUT_DIR = os.path.join(BASE_DIR, "Output")
CACHE_DIR = os.path.join(BASE_DIR, "Cache")
SHARES_CACHE_FILE = os.path.join(CACHE_DIR, "shares_outstanding_cache.csv")

MIN_RVOL = 1.5
VOL_THRESHOLD = 10_000_000
MIN_BARS_FULL = 210
MIN_BARS_PARTIAL = 30
FETCH_RETRIES = 3
RETRY_SLEEP_SEC = 1.2

# =========================
# TRUE IDX SECTOR MOVERS CONFIG
# =========================
SECTOR_MOVERS_PERIOD_WEEKS = 52   # visible "Period" for RRG window
SECTOR_MOVERS_SMOOTHING = 10      # EMA smoothing length
SECTOR_MOVERS_TRAIL = 5           # tail points shown
SECTOR_MOVERS_LOOKBACK = 52       # rolling normalization window (RRG baseline)
SECTOR_INDEX_SYMBOLS = {
    "COMPOSITE": "^JKSE",
}

# =========================
# STYLES
# =========================
LINE = "C2CADE"
THIN = Side(style="thin", color=LINE)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
LEFT_BORDER_ONLY = Border(left=THIN)

FONT_TITLE = Font(name="Calibri", size=14, bold=True)
FONT_SUBTITLE = Font(name="Calibri", size=10, italic=True)

# Soft pastel header palette (dark text only; no white font)
FONT_GROUP = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
FONT_SUBHEADER = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
FONT_HEADER = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
FONT_NOTE = Font(name="Calibri", size=10, color="1F1F1F")
FONT_BODY = Font(name="Calibri", size=10)
FONT_REQ = Font(name="Calibri", size=10, bold=True)

# Global / generic headers (used across summary + helper sheets)
FILL_HEADER_SUMMARY = PatternFill("solid", fgColor="2A4A66")  # navy
FILL_HEADER = PatternFill("solid", fgColor="2A4A66")
FILL_GROUP_FLOW = PatternFill("solid", fgColor="2A4A66")      # navy

# IDX Screener Detail group headers (all distinct colors)
FILL_GROUP_STOCK = PatternFill("solid", fgColor="2A4A66")     # navy
FILL_GROUP_OWNER = PatternFill("solid", fgColor="237A5C")     # forest green
FILL_GROUP_Q = PatternFill("solid", fgColor="7A1A1A")         # deep blue
FILL_GROUP_PQ = PatternFill("solid", fgColor="7A1A1A")        # dark crimson
FILL_GROUP_PY = PatternFill("solid", fgColor="7A1A1A")        # cobalt
FILL_GROUP_MP = PatternFill("solid", fgColor="21467C")        # denim
FILL_GROUP_VOL = PatternFill("solid", fgColor="216B6B")       # teal
FILL_GROUP_MA = PatternFill("solid", fgColor="32588E")        # dark plum
FILL_GROUP_MOM = PatternFill("solid", fgColor="207878")       # sea green
FILL_GROUP_VR = PatternFill("solid", fgColor="8C2038")        # burgundy
FILL_GROUP_MACD_MOM = PatternFill("solid", fgColor="4A1060")  # deep violet-purple for MACD Momentum

FILL_LEGEND_A = PatternFill("solid", fgColor="007BA7")
FILL_LEGEND_B = PatternFill("solid", fgColor="4B49AC")
FILL_LEGEND_D = PatternFill("solid", fgColor="7978E9")
FILL_LEGEND_PY = PatternFill("solid", fgColor="0077B6")
FILL_STATUS_OK = PatternFill("solid", fgColor="007BA7")
FILL_STATUS_PARTIAL = PatternFill("solid", fgColor="4B49AC")
FILL_STATUS_NODATA = PatternFill("solid", fgColor="F3797E")

# =========================
# HELPERS
# =========================
def style_cell(cell, fill=None, font=None, border=True, align="center", wrap=False):
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if border:
        cell.border = BORDER
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)

def style_plain(cell, font=None, align="left", wrap=False):
    if font:
        cell.font = font
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)

def style_legend(cell, fill=None, font=None, align="left", wrap=False):
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)

def safe_num(v, default=np.nan):
    try:
        if pd.isna(v):
            return default
        return float(v)
    except Exception:
        return default

def safe_int(v, default=0):
    try:
        if pd.isna(v):
            return default
        return int(round(float(v)))
    except Exception:
        return default

def safe_parse_datetime(value, errors="coerce", dayfirst=False):
    """Parse scalar date values without pandas' ambiguous-format warning."""
    if value in (None, "", "N/A", "-"):
        return pd.NaT
    try:
        return pd.to_datetime(value, errors=errors, format="mixed", dayfirst=dayfirst)
    except TypeError:
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%d %b %Y", "%d %B %Y"):
            try:
                return pd.to_datetime(value, errors=errors, format=fmt)
            except Exception:
                pass
        return pd.to_datetime(value, errors=errors, dayfirst=dayfirst)

def safe_parse_datetime_index(values, errors="coerce", dayfirst=False):
    """Parse date-like indexes without pandas' ambiguous-format warning."""
    try:
        return pd.to_datetime(values, errors=errors, format="mixed", dayfirst=dayfirst)
    except TypeError:
        parsed = [
            safe_parse_datetime(v, errors=errors, dayfirst=dayfirst)
            for v in list(values)
        ]
        return pd.DatetimeIndex(parsed)

def pct_diff(close, ma):
    if pd.isna(close) or pd.isna(ma) or ma == 0:
        return np.nan
    return ((close / ma) - 1.0) * 100.0

def pos_label(close, ma):
    if pd.isna(close) or pd.isna(ma):
        return ""
    return "Above" if close >= ma else "Below"

def rsi(series: pd.Series, period=14):
    delta = series.diff()
    gain = delta.clip(lower=0)
    loss = -delta.clip(upper=0)
    avg_gain = gain.ewm(alpha=1 / period, min_periods=period, adjust=False).mean()
    avg_loss = loss.ewm(alpha=1 / period, min_periods=period, adjust=False).mean()
    rs = avg_gain / avg_loss.replace(0, np.nan)
    return 100 - (100 / (1 + rs))

def ema(series: pd.Series, period: int):
    return series.ewm(span=period, adjust=False).mean()

def quarter_start(ts: pd.Timestamp):
    q = ((ts.month - 1) // 3) + 1
    start_month = (q - 1) * 3 + 1
    return pd.Timestamp(ts.year, start_month, 1)

def prev_quarter_range(ts: pd.Timestamp):
    cq = quarter_start(ts)
    prev_end = cq - pd.Timedelta(days=1)
    prev_start = quarter_start(prev_end)
    return prev_start, prev_end

def weighted_std(values: pd.Series, weights: pd.Series):
    mask = (~values.isna()) & (~weights.isna()) & (weights > 0)
    x = values[mask].astype(float)
    w = weights[mask].astype(float)
    if len(x) == 0 or w.sum() == 0:
        return np.nan
    mean = np.average(x, weights=w)
    var = np.average((x - mean) ** 2, weights=w)
    return np.sqrt(var)

def anchored_vwap_block(df: pd.DataFrame) -> Dict[str, float]:
    empty_out = {
        "days": 0,
        "vwap": np.nan,
        "p1": np.nan,
        "p2": np.nan,
        "p3": np.nan,
        "m1": np.nan,
        "m2": np.nan,
        "m3": np.nan,
        "sd_score": np.nan,
        "prev_sd": np.nan,
        "sd_delta": np.nan,
    }

    if df is None or df.empty:
        return empty_out

    price = (df["High"] + df["Low"] + df["Close"]) / 3.0
    vol = df["Volume"].replace(0, np.nan)
    mask = (~price.isna()) & (~vol.isna())
    d = df.loc[mask].copy()

    if d.empty:
        return empty_out

    def _calc_block(block: pd.DataFrame):
        if block is None or block.empty:
            return np.nan, np.nan, np.nan, np.nan, np.nan, np.nan

        tp = (block["High"] + block["Low"] + block["Close"]) / 3.0
        vv = block["Volume"].astype(float)

        if len(tp) == 0 or vv.sum() == 0:
            return np.nan, np.nan, np.nan, np.nan, np.nan, np.nan

        vwap = (tp * vv).sum() / vv.sum()
        sd = weighted_std(tp, vv)

        # Near-zero SD is valid => collapse bands to VWAP, score = 0
        sd_floor = max(abs(float(vwap)) * 1e-6 if pd.notna(vwap) else 0.0, 1e-9)
        sd_valid = pd.notna(sd) and float(sd) > sd_floor

        close_last = safe_num(block["Close"].iloc[-1])
        sd_score = safe_num((close_last - vwap) / sd) if sd_valid else 0.0

        m1 = vwap - sd if sd_valid else vwap
        m2 = vwap - (2 * sd) if sd_valid else vwap
        m3 = vwap - (3 * sd) if sd_valid else vwap

        return vwap, m1, m2, m3, sd_score, sd

    # Current anchored block (includes latest bar)
    vwap, m1, m2, m3, sd_score, sd = _calc_block(d)

    # Prior-day anchored block (exclude latest bar)
    if len(d) >= 2:
        d_prev = d.iloc[:-1].copy()
        prev_vwap, prev_m1, prev_m2, prev_m3, prev_sd_score, prev_sd_raw = _calc_block(d_prev)
        sd_delta = safe_num(sd_score - prev_sd_score) if pd.notna(sd_score) and pd.notna(prev_sd_score) else np.nan
    else:
        prev_sd_score = np.nan
        sd_delta = np.nan

    # Compute +SD bands (symmetric to -SD bands)
    sd_val = sd if (pd.notna(sd) and float(sd) > 0) else np.nan
    p1 = float(vwap) + float(sd_val) if pd.notna(vwap) and pd.notna(sd_val) else np.nan
    p2 = float(vwap) + 2 * float(sd_val) if pd.notna(vwap) and pd.notna(sd_val) else np.nan
    p3 = float(vwap) + 3 * float(sd_val) if pd.notna(vwap) and pd.notna(sd_val) else np.nan

    return {
        "days": len(d),
        "vwap": vwap,
        "p1": p1,
        "p2": p2,
        "p3": p3,
        "m1": m1,
        "m2": m2,
        "m3": m3,
        "sd_score": sd_score,
        "prev_sd": prev_sd_score,
        "sd_delta": sd_delta,
    }

def zone_label(close, ibl, ibh):
    if pd.isna(close) or pd.isna(ibl) or pd.isna(ibh):
        return "N/A"
    if abs(float(close) - float(ibl)) < 1e-9:
        return "At Level"
    return "Above" if close > ibl else "Below"

def vwap_zone_2pct(close, levels):
    """
    Premium VWAP POI text engine
    - Near threshold = ±12.5% of the relevant segment gap
    - Display = ACTUAL % distance from current price to referenced POI
    - For ranging zones, reference side is the nearest boundary:
      * If closer to upper boundary => "x% below upper"
      * If closer to lower boundary => "x% above lower"
    """
    if pd.isna(close):
        return "N/A"

    valid = [float(x) for x in levels if pd.notna(x)]
    if not valid:
        return "N/A"

    def _pct(px, lvl):
        if pd.isna(lvl) or lvl == 0:
            return None
        return ((float(px) / float(lvl)) - 1.0) * 100.0

    def _near(px, boundary, gap, tol_frac=0.125):
        if pd.isna(boundary) or pd.isna(gap) or gap <= 0:
            return False
        return abs(float(px) - float(boundary)) <= (gap * tol_frac)

    def _fmt_near(px, lvl, label):
        p = _pct(px, lvl)
        if p is None:
            return f"Price Near {label}"
        if abs(p) < 0.05:
            return f"Price Near {label} (At {label})"
        side = "above" if p > 0 else "below"
        return f"Price Near {label} ({abs(p):.1f}% {side} {label})"

    def _fmt_range(px, upper_label, lower_label, upper_level, lower_level):
        du = abs(float(px) - float(upper_level)) if pd.notna(upper_level) else float("inf")
        dl = abs(float(px) - float(lower_level)) if pd.notna(lower_level) else float("inf")

        # Closer to upper boundary => describe as below upper
        if du <= dl:
            p = _pct(px, upper_level)
            if p is None:
                return f"Price Ranging {upper_label} ⇄ {lower_label}"
            return f"Price Ranging {upper_label} ⇄ {lower_label} ({abs(p):.1f}% below {upper_label})"

        # Closer to lower boundary => describe as above lower
        p = _pct(px, lower_level)
        if p is None:
            return f"Price Ranging {upper_label} ⇄ {lower_label}"
        return f"Price Ranging {upper_label} ⇄ {lower_label} ({abs(p):.1f}% above {lower_label})"

    def _fmt_above(px, lvl, label):
        p = _pct(px, lvl)
        if p is None:
            return f"Price Above {label}"
        if abs(p) < 0.05:
            return f"Price Near {label} (At {label})"
        side = "above" if p > 0 else "below"
        return f"Price Above {label} ({abs(p):.1f}% {side} {label})"

    def _fmt_below(px, lvl, label):
        p = _pct(px, lvl)
        if p is None:
            return f"Price Below {label}"
        if abs(p) < 0.05:
            return f"Price Near {label} (At {label})"
        side = "above" if p > 0 else "below"
        return f"Price Below {label} ({abs(p):.1f}% {side} {label})"

    # Preferred mode: [VWAP, -1 SD, -2 SD, -3 SD]
    if len(valid) >= 4:
        vwap, m1, m2, m3 = valid[:4]

        if close >= vwap:
            gap = abs(vwap - m1) if pd.notna(m1) else abs(vwap) * 0.10
            if _near(close, vwap, gap):
                return _fmt_near(close, vwap, "VWAP")
            return _fmt_above(close, vwap, "VWAP")

        if close >= m1:
            gap = abs(vwap - m1)
            if _near(close, vwap, gap):
                return _fmt_near(close, vwap, "VWAP")
            if _near(close, m1, gap):
                return _fmt_near(close, m1, "-1 SD")
            return _fmt_range(close, "VWAP", "-1 SD", vwap, m1)

        if close >= m2:
            gap = abs(m1 - m2)
            if _near(close, m1, gap):
                return _fmt_near(close, m1, "-1 SD")
            if _near(close, m2, gap):
                return _fmt_near(close, m2, "-2 SD")
            return _fmt_range(close, "-1 SD", "-2 SD", m1, m2)

        if close >= m3:
            gap = abs(m2 - m3)
            if _near(close, m2, gap):
                return _fmt_near(close, m2, "-2 SD")
            if _near(close, m3, gap):
                return _fmt_near(close, m3, "-3 SD")
            return _fmt_range(close, "-2 SD", "-3 SD", m2, m3)

        gap = abs(m2 - m3) if pd.notna(m2) else abs(m3) * 0.10
        if _near(close, m3, gap):
            return _fmt_near(close, m3, "-3 SD")
        return _fmt_below(close, m3, "-3 SD")

    # Legacy fallback: [m1, m2, m3]
    if len(valid) == 3:
        m1, m2, m3 = valid

        if close >= m1:
            gap = abs(m1 - m2) if pd.notna(m2) else abs(m1) * 0.10
            if _near(close, m1, gap):
                return _fmt_near(close, m1, "-1 SD")
            return _fmt_above(close, m1, "-1 SD")

        if close >= m2:
            gap = abs(m1 - m2)
            if _near(close, m1, gap):
                return _fmt_near(close, m1, "-1 SD")
            if _near(close, m2, gap):
                return _fmt_near(close, m2, "-2 SD")
            return _fmt_range(close, "-1 SD", "-2 SD", m1, m2)

        if close >= m3:
            gap = abs(m2 - m3)
            if _near(close, m2, gap):
                return _fmt_near(close, m2, "-2 SD")
            if _near(close, m3, gap):
                return _fmt_near(close, m3, "-3 SD")
            return _fmt_range(close, "-2 SD", "-3 SD", m2, m3)

        gap = abs(m2 - m3) if pd.notna(m2) else abs(m3) * 0.10
        if _near(close, m3, gap):
            return _fmt_near(close, m3, "-3 SD")
        return _fmt_below(close, m3, "-3 SD")

    return "N/A"


def vwap_near_zone_label(close, sd_score, vwap, m1, m2, m3, p1=np.nan, p2=np.nan, p3=np.nan):
    """
    New VWAP Zone label engine (Upgrade Prompt spec):
    - Uses SD Score (not price gap) to determine proximity
    - Threshold: ±0.1 SD Score from integer SD levels (0, ±1, ±2, ±3)
    - If within threshold → "Price Near [ZONE] • [DISTANCE]%"
    - If NOT within threshold → "-"
    - Works for all zones including above +1/+2/+3 SD
    - Distance % = abs(close - level) / level * 100
    """
    if pd.isna(close) or pd.isna(sd_score):
        return "-"

    sd_score = float(sd_score)

    # Define SD integer levels and their corresponding price levels + labels
    levels = [
        (0,  vwap, "VWAP"),
        (-1, m1,   "-1 SD"),
        (-2, m2,   "-2 SD"),
        (-3, m3,   "-3 SD"),
        (+1, p1,   "+1 SD"),
        (+2, p2,   "+2 SD"),
        (+3, p3,   "+3 SD"),
    ]

    TOL = 0.1  # ±0.1 SD score tolerance

    # Find nearest integer SD level within tolerance
    nearest = None
    nearest_dist = float("inf")
    for target_sd, level_px, label in levels:
        dist = abs(sd_score - target_sd)
        if dist <= TOL and dist < nearest_dist:
            nearest = (target_sd, level_px, label)
            nearest_dist = dist

    if nearest is None:
        return "-"

    _, level_px, label = nearest
    if pd.isna(level_px) or float(level_px) == 0:
        return f"Price Near {label}"

    pct_dist = abs((float(close) - float(level_px)) / float(level_px)) * 100.0
    return f"Price Near {label} • {pct_dist:.2f}%"


def vwap_zone_remark(close, m1, m2, m3):
    """
    Legacy compatibility shim.
    Uses new VWAP zone wording logic on legacy 3-level input.
    """
    return vwap_zone_2pct(close, [m1, m2, m3])

def rsi_status(v):
    if pd.isna(v):
        return ""
    if v >= 70:
        return "Overbought"
    if v <= 30:
        return "Oversold"
    if v >= 55:
        return "Strong"
    if v <= 45:
        return "Weak"
    return "Neutral"

# =========================
# DATE ENGINE  — single source of truth
# =========================
def get_market_date() -> pd.Timestamp:
    """
    Returns the effective market date for this screener run.
    MARKET_DATE is always set and always authoritative.
    All data is clipped to this date regardless of BACKTEST_MODE.
    """
    try:
        return pd.Timestamp(MARKET_DATE).normalize()
    except Exception:
        raise ValueError(f"MARKET_DATE='{MARKET_DATE}' is not a valid date string.")


def is_backtest_mode() -> bool:
    """Returns True when BACKTEST_MODE restricts the universe to BACKTEST_TICKERS."""
    return bool(BACKTEST_MODE)


def clip_to_market_date(hist: pd.DataFrame) -> pd.DataFrame:
    """
    Clips a price DataFrame to rows <= MARKET_DATE.
    Always applied — MARKET_DATE is the hard replay boundary for all data.
    """
    if hist is None or hist.empty:
        return hist
    asof = get_market_date()
    try:
        clipped = hist[hist.index.normalize() <= asof]
        return clipped if not clipped.empty else hist
    except Exception:
        return hist


# Legacy aliases — kept so existing call-sites work without change
def _get_forced_asof_timestamp():
    return get_market_date()

def get_effective_asof_date() -> pd.Timestamp:
    return get_market_date()

def _clip_hist_to_asof(hist: pd.DataFrame) -> pd.DataFrame:
    return clip_to_market_date(hist)

def _ordinal(n: int) -> str:
    if 10 <= (n % 100) <= 20:
        suffix = "th"
    else:
        suffix = {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
    return f"{n}{suffix}"

def _build_output_filename(run_dt: datetime, latest_market_day: str) -> str:
    ts_run = run_dt.strftime("%y%m%d %H.%M")
    try:
        d = safe_parse_datetime(latest_market_day, errors="coerce")
        if pd.notna(d):
            asof_label = f"{d.strftime('%B')} {_ordinal(int(d.day))} {d.year}"
        else:
            asof_label = str(latest_market_day)
    except Exception:
        asof_label = str(latest_market_day)
    return f"IDX Screener as of {asof_label} Running at {ts_run}.xlsx"

def normalize_history(hist):
    if hist is None or hist.empty:
        return None
    if isinstance(hist.columns, pd.MultiIndex):
        hist.columns = [c[0] for c in hist.columns]
    needed = ["Open", "High", "Low", "Close", "Volume"]
    if any(c not in hist.columns for c in needed):
        return None
    hist = hist[needed].copy()
    hist.dropna(subset=["Close"], inplace=True)
    if hist.empty:
        return None
    idx = pd.to_datetime(hist.index)
    try:
        if getattr(idx, "tz", None) is not None:
            idx = idx.tz_localize(None)
    except Exception:
        try:
            idx = idx.tz_convert(None)
        except Exception:
            pass
    hist.index = idx
    hist = hist[~hist.index.duplicated(keep="last")].sort_index()

    # Clip all data to MARKET_DATE — single authoritative boundary
    try:
        _asof = get_market_date()
        hist = hist[hist.index.normalize() <= _asof]
    except Exception:
        pass

    return hist

def load_shares_cache():
    try:
        os.makedirs(CACHE_DIR, exist_ok=True)
        if os.path.exists(SHARES_CACHE_FILE):
            df = pd.read_csv(SHARES_CACHE_FILE)
            if {"Ticker", "SharesOutstanding"}.issubset(df.columns):
                return dict(zip(df["Ticker"].astype(str).str.upper(), pd.to_numeric(df["SharesOutstanding"], errors="coerce")))
    except Exception:
        pass
    return {}

def save_shares_cache(cache_dict):
    try:
        os.makedirs(CACHE_DIR, exist_ok=True)
        rows = [{"Ticker": k, "SharesOutstanding": v} for k, v in cache_dict.items()]
        df = pd.DataFrame(rows)
        if not df.empty:
            df.sort_values("Ticker").to_csv(SHARES_CACHE_FILE, index=False)
    except Exception:
        pass


def _sc_get(r, key, default="N/A"):
    try:
        v = r.get(key, default)
        if v is None:
            return default
        if isinstance(v, float) and np.isnan(v):
            return default
        return v
    except Exception:
        return default

def _format_pct_safe(v, digits=1):
    try:
        if v is None or (isinstance(v, float) and np.isnan(v)):
            return "N/A"
        return f"{float(v):.{digits}f}"
    except Exception:
        return "N/A"

def _swing_v16_native_row(r):
    """
    Native POI-first swing row schema (V17 — LONG ONLY)
    Returns ordered dict matching the V17 output column spec.
    """
    def _g(primary, *fallbacks):
        v = r.get(primary)
        if v is not None and str(v) not in ("", "N/A", "nan"):
            return v
        for fb in fallbacks:
            v = r.get(fb)
            if v is not None and str(v) not in ("", "N/A", "nan"):
                return v
        return "N/A"

    # ── Stock info ─────────────────────────────────────────────────────────────
    ticker   = _g("ticker", "Ticker")
    sector   = _g("sector", "IDX Sector")
    price    = _g("close",  "Close", "Price", "Closing Price")
    pct_chg  = _g("pct_chg", "Price Change %", "pct_change")
    emiten   = _g("name", "Emiten", "Company")

    # ── Screener outputs ───────────────────────────────────────────────────────
    priority     = _g("_sc_priority_label", "_sc_swing_grade")
    score        = safe_num(r.get("_sc_swing_score", r.get("_sc_conviction")), np.nan)
    score_disp   = f"{score:.2f}" if pd.notna(score) else "N/A"
    verdict      = _g("verdict_weight_profile", "regime_label", "Verdict Weight Profile")
    trend        = _g("ms_trend_regime", "trend_regime", "Trend Bias")
    rvol         = _g("rvol20", "RVOL")
    adr          = _g("adr_pct", "ADR %", "adr14_pct")
    atr          = _g("atr14_pct", "ATR (14) %")
    last_evt     = _g("ms_last_event", "last_structural_event", "Last Structural Event")
    event_age    = _g("ms_event_age_d", "event_age_d", "Event Age (D)")

    # ── POI engine fields ─────────────────────────────────────────────────────
    poi_type     = _g("_sc_poi_type", "_sc_zone_type", "POI Type")
    primary_poi  = _g("_sc_primary_poi", "_sc_zone_label", "Primary POI")
    next_poi     = _g("_sc_next_poi", "Next POI")
    dist_pct     = safe_num(r.get("_sc_dist_pct"), np.nan)
    dist_disp    = f"{dist_pct:+.2f}%" if pd.notna(dist_pct) else "N/A"
    confluence   = _g("_sc_confluence", "Confluence")
    amt          = _g("_sc_amt_state", "_sc_signal", "AMT State")

    # ── Context ───────────────────────────────────────────────────────────────
    mp           = _g("market_profile_summary", "mp_zone", "Market Profile", "MP Summary")
    ma           = _g("ma_position_summary", "ma_zone", "MA Position", "MA Zone")
    candle       = _g("cp_pattern", "last_candle_pattern", "Candle Pattern", "Pattern")
    candle_date  = _g("cp_date", "pattern_date", "Pattern Date")
    rsi_status   = _g("rsi_status", "RSI Status")
    rsi_div      = _g("div_signal", "divergence_signal", "Divergence Signal")
    macd_wave    = _g("macd_wave", "macd_wave_pattern", "MACD Wave", "Wave Pattern")

    # ── Trade plan ─────────────────────────────────────────────────────────────
    entry        = _g("_sc_entry_disp", "_sc_entry", "Entry Zone")
    trigger      = _g("_sc_trigger", "_sc_signal", "_sc_amt_state", "Trigger")
    t1           = _g("_sc_target_disp", "_sc_t1", "T1")
    t2           = _g("_sc_t2_disp", "_sc_t2", "T2")
    invalid      = _g("_sc_invalidation_disp", "_sc_invalid", "Invalid")
    rr           = safe_num(r.get("_sc_rr"), np.nan)
    rr_disp      = f"{rr:.1f}" if pd.notna(rr) else "N/A"

    # ── Wyckoff extras ─────────────────────────────────────────────────────────
    cause_quality  = _g("_sc_cause_quality", "cause_quality", "Cause Quality")
    markup_ready   = safe_num(r.get("_sc_markup_readiness", r.get("markup_readiness")), np.nan)
    markup_disp    = f"{markup_ready:.0f}%" if pd.notna(markup_ready) else "N/A"
    summary        = _g("_sc_summary", "Summary")

    return {
        "Priority":             priority,
        "Ticker":               ticker,
        "IDX Sector":           sector,
        "Price":                price,
        "Price Change %":             pct_chg,
        "Composite Score":      score_disp,
        "Verdict Profile":      verdict,
        "Trend Bias":           trend,
        "RVOL":                 rvol,
        "ADR %":                adr,
        "ATR14 %":              atr,
        "Last Structural Event":last_evt,
        "Event Age (D)":        event_age,
        "POI Type":             poi_type,
        "Primary POI":          primary_poi,
        "Next POI":             next_poi,
        "Dist %":               dist_disp,
        "Confluence":           confluence,
        "AMT State":            amt,
        "Market Profile":       mp,
        "MA Position":          ma,
        "Candle Pattern":       candle,
        "Pattern Date":         candle_date,
        "RSI Status":           rsi_status,
        "Divergence":           rsi_div,
        "MACD Wave":            macd_wave,
        "Entry Zone":           entry,
        "Trigger":              trigger,
        "T1":                   t1,
        "T2":                   t2,
        "Invalid":              invalid,
        "R/R":                  rr_disp,
        "Cause Quality":        cause_quality,
        "Markup Readiness":     markup_disp,
        "Summary":              summary,
    }


def compact_fmt(v):
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return None
    try:
        n = float(v)
    except Exception:
        return v
    absn = abs(n)
    if absn >= 1_000_000_000_000:
        return f"{n / 1_000_000_000_000:.2f} T"
    if absn >= 1_000_000_000:
        return f"{n / 1_000_000_000:.2f} B"
    if absn >= 1_000_000:
        return f"{n / 1_000_000:.2f} M"
    if absn >= 1_000:
        return f"{n / 1_000:.2f} K"
    return f"{n:,.0f}"

def na_if_nan(v):
    return "N/A" if pd.isna(v) else v


def safe_div(a, b, default=np.nan):
    try:
        if pd.isna(a) or pd.isna(b) or float(b) == 0:
            return default
        return float(a) / float(b)
    except Exception:
        return default

# ── IHSG Beta engine ──────────────────────────────────────────────────────────
_IHSG_RETURNS_CACHE: dict = {}   # keyed by period_days → pd.Series of daily returns

def _get_ihsg_returns(period_days: int = 252) -> "pd.Series | None":
    """Fetch and cache IHSG (^JKSE) daily returns for beta computation."""
    global _IHSG_RETURNS_CACHE
    if period_days in _IHSG_RETURNS_CACHE:
        return _IHSG_RETURNS_CACHE[period_days]
    try:
        import yfinance as yf
        ihsg = yf.download("^JKSE", period=f"{period_days + 30}d", progress=False, auto_adjust=True)
        if ihsg is None or ihsg.empty:
            return None
        closes = ihsg["Close"].squeeze()
        if hasattr(closes, "columns"):
            closes = closes.iloc[:, 0]
        returns = closes.pct_change().dropna()
        if len(returns) < 20:
            return None
        _IHSG_RETURNS_CACHE[period_days] = returns.tail(period_days)
        return _IHSG_RETURNS_CACHE[period_days]
    except Exception:
        return None

def compute_beta_ihsg(stock_hist: "pd.DataFrame", period_days: int = 252) -> float:
    """
    COVARIANCE.P(stock_returns, ihsg_returns) / VAR.P(ihsg_returns)
    Matches the Excel formula requested in Stock Info column.
    Returns np.nan on failure.
    """
    try:
        ihsg_ret = _get_ihsg_returns(period_days)
        if ihsg_ret is None or ihsg_ret.empty:
            return np.nan
        stock_closes = stock_hist["Close"].squeeze()
        if hasattr(stock_closes, "columns"):
            stock_closes = stock_closes.iloc[:, 0]
        stock_ret = stock_closes.pct_change().dropna()
        common_idx = stock_ret.index.intersection(ihsg_ret.index)
        if len(common_idx) < 20:
            return np.nan
        s = stock_ret.loc[common_idx].astype(float).values
        m = ihsg_ret.loc[common_idx].astype(float).values
        cov_p = float(np.cov(s, m, ddof=0)[0, 1])
        var_p = float(np.var(m, ddof=0))
        if var_p == 0:
            return np.nan
        return round(cov_p / var_p, 4)
    except Exception:
        return np.nan


def compute_beta_zone(beta: float) -> str:
    """Classify beta into zone label."""
    if pd.isna(beta):
        return "-"
    if beta >= 2.0:
        return "Above 2"
    if beta >= 1.0:
        return "Above 1"
    if beta >= 0.0:
        return "Below 1"
    return "Below 0"


def compute_rs_rating(hist: "pd.DataFrame", universe_hist_dict: dict = None) -> int:
    """
    IBD-style RS Rating (1–99) ported from Pine Script.
    Weights: 3M×2, 6M×1, 9M×1, 12M×1 → raw score → compress to 1–99
    against the full universe distribution stored in _RS_UNIVERSE_CACHE.
    Falls back to raw percentile vs own 252D history if universe not available.
    """
    global _RS_UNIVERSE_CACHE
    try:
        closes = hist["Close"].squeeze()
        if hasattr(closes, "columns"):
            closes = closes.iloc[:, 0]
        closes = closes.dropna()
        n = len(closes)
        if n < 63:
            return np.nan

        def ratio(bars):
            bars = min(bars, n - 1)
            if bars <= 0 or pd.isna(closes.iloc[-bars - 1]) or closes.iloc[-bars - 1] == 0:
                return np.nan
            return float(closes.iloc[-1]) / float(closes.iloc[-bars - 1])

        r3  = ratio(63)
        r6  = ratio(126) if n >= 126 else r3
        r9  = ratio(189) if n >= 189 else r6
        r12 = ratio(252) if n >= 252 else r9

        if any(pd.isna(x) for x in [r3, r6, r9, r12]):
            return np.nan

        rs_raw = 2.0 * r3 + 1.0 * r6 + 1.0 * r9 + 1.0 * r12

        # Compress to 1–99 against universe cache
        universe = _RS_UNIVERSE_CACHE
        if len(universe) >= 5:
            lo = min(universe)
            hi = max(universe)
            if hi == lo:
                return 50
            rating = round(1 + 98 * (rs_raw - lo) / (hi - lo))
            return int(max(1, min(99, rating)))
        else:
            # Store raw in cache; return nan until enough universe members
            _RS_UNIVERSE_CACHE.append(rs_raw)
            return np.nan
    except Exception:
        return np.nan


def rs_rating_zone(rs: float) -> str:
    """Rule of thumb classification."""
    if pd.isna(rs):
        return "-"
    if rs >= 95:
        return "Top-Tier Leader (≥95)"
    if rs >= 90:
        return "Leader (≥90)"
    if rs >= 80:
        return "Strong (≥80)"
    if rs >= 60:
        return "Neutral (60–79)"
    return "Weak (<60)"


def compute_today_event(row: dict, hist: "pd.DataFrame") -> str:
    """
    Detect what structural event happened on the latest candle relative to:
    - Swing BOS / CHoCH (from smc fields)
    - Equilibrium zone (midpoint of Strong High / Strong Low)
    - Bull OB zone (smc_closest_ob range)
    - Bear OB zone (smc_closest_ob_bear range)
    Returns descriptive string or '-'.
    """
    try:
        if hist is None or len(hist) < 2:
            return "-"

        close  = float(hist["Close"].iloc[-1])
        open_  = float(hist["Open"].iloc[-1])
        high   = float(hist["High"].iloc[-1])
        low    = float(hist["Low"].iloc[-1])
        prev_c = float(hist["Close"].iloc[-2])

        # BOS / CHoCH from SMC fields
        swing_struct    = str(row.get("smc_latest_swing_struct","") or "")
        internal_struct = str(row.get("smc_latest_internal_struct","") or "")
        for label, src in [("Swing", swing_struct), ("Internal", internal_struct)]:
            if not src or src in ("-","N/A"):
                continue
            s = src.upper()
            if "BOS" in s and "BULL" in s:
                return f"Price turned Bull BOS ({label})"
            if "BOS" in s and "BEAR" in s:
                return f"Price turned Bear BOS ({label})"
            if "CHOCH" in s and "BULL" in s:
                return f"Price turned Bull CHoCH ({label})"
            if "CHOCH" in s and "BEAR" in s:
                return f"Price turned Bear CHoCH ({label})"

        # Equilibrium zone
        sh = safe_num(row.get("smc_strong_high"), np.nan)
        sl = safe_num(row.get("smc_strong_low"),  np.nan)
        if pd.notna(sh) and pd.notna(sl) and sh > sl:
            eq   = (sh + sl) / 2.0
            band = (sh - sl) * 0.05   # ±5% of range = equilibrium band
            eq_top = eq + band
            eq_bot = eq - band
            if prev_c < eq_top and close > eq_top:
                return "Price Break Up Equilibrium Zone"
            if prev_c > eq_bot and close < eq_bot:
                return "Price Break Down Equilibrium Zone"
            if eq_bot <= close <= eq_top:
                return "Price Ranging on Equilibrium Zone"

        def _parse_range(s):
            try:
                parts = str(s).replace("–","-").split("-")
                if len(parts) == 2:
                    lo = float(parts[0].replace(",",""))
                    hi = float(parts[1].replace(",",""))
                    return lo, hi
            except Exception:
                pass
            return None, None

        # Bull OB zone
        ob_bull_str = str(row.get("smc_closest_ob","") or "")
        ob_lo, ob_hi = _parse_range(ob_bull_str)
        if ob_lo and ob_hi:
            if prev_c < ob_hi and close > ob_hi:
                return "Price Break Up Bull OB Zone"
            if prev_c > ob_lo and close < ob_lo:
                return "Price Break Down Bull OB Zone"
            if ob_lo <= close <= ob_hi:
                return "Price Ranging on Bull OB Zone"

        # Bear OB zone
        ob_bear_str = str(row.get("smc_closest_ob_bear","") or "")
        b_lo, b_hi = _parse_range(ob_bear_str)
        if b_lo and b_hi:
            if prev_c < b_hi and close > b_hi:
                return "Price Break Up Bear OB Zone"
            if prev_c > b_lo and close < b_lo:
                return "Price Break Down Bear OB Zone"
            if b_lo <= close <= b_hi:
                return "Price Ranging on Bear OB Zone"

        return "-"
    except Exception:
        return "-"


# Universe RS raw score cache (populated during batch build_row)
_RS_UNIVERSE_CACHE: list = []




# =============================================================================
# UPGRADE 1 — ARA / ARB DETECTION
# =============================================================================
def compute_ara_arb(hist: "pd.DataFrame", board: str = "Main") -> dict:
    """
    Detect if today's price is at the IDX daily auto-rejection limit.
    Main/Development Board: ±35%
    Acceleration Board:     ±10%
    Returns: {ara_arb: label, at_limit: bool, limit_pct: float}
    """
    limit = 0.10 if str(board).lower() in ("acceleration", "akselerasi", "acc") else 0.35
    result = {"ara_arb": "-", "at_limit": False, "limit_pct": limit * 100}
    try:
        if hist is None or len(hist) < 2:
            return result
        prev_close = float(hist["Close"].iloc[-2])
        curr_close = float(hist["Close"].iloc[-1])
        if prev_close <= 0:
            return result
        chg = (curr_close - prev_close) / prev_close
        tol = 0.001   # 0.1% tolerance for floating point
        if chg >= limit - tol:
            result["ara_arb"]  = f"ARA (+{limit*100:.0f}%)"
            result["at_limit"] = True
        elif chg <= -limit + tol:
            result["ara_arb"]  = f"ARB (-{limit*100:.0f}%)"
            result["at_limit"] = True
    except Exception:
        pass
    return result


# =============================================================================
# UPGRADE 2 — FOREIGN FLOW (IDX JATS)
# =============================================================================
_FOREIGN_FLOW_CACHE: dict = {}   # ticker → {foreign_net_lot, foreign_net_val, foreign_activity}

def fetch_foreign_flow(ticker: str) -> dict:
    """
    Fetch foreign net buy/sell from IDX market data.
    Source: idx.co.id/umum/foreign-net-buy-sell endpoint (JSON).
    Falls back to N/A gracefully.
    Returns: foreign_net_lot (lots), foreign_net_val (IDR), foreign_activity label.
    """
    empty = {"foreign_net_lot": np.nan, "foreign_net_val": np.nan, "foreign_activity": "-"}
    clean = ticker.upper().replace(".JK", "").strip()
    if clean in _FOREIGN_FLOW_CACHE:
        return _FOREIGN_FLOW_CACHE[clean]
    try:
        import urllib.request, json
        url = (
            "https://www.idx.co.id/primary/StockData/GetForeignNetBuySellByCode"
            f"?code={clean}&lang=id"
        )
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=8) as resp:
            data = json.loads(resp.read().decode("utf-8", errors="ignore"))

        # IDX API wraps in {"data": [...], "totData": N}
        rows = data.get("data") or data.get("Data") or []
        if not rows:
            _FOREIGN_FLOW_CACHE[clean] = empty
            return empty

        latest = rows[0]
        buy_lot  = float(latest.get("ForeignBuy")  or latest.get("foreign_buy")  or 0)
        sell_lot = float(latest.get("ForeignSell") or latest.get("foreign_sell") or 0)
        net_lot  = buy_lot - sell_lot

        buy_val  = float(latest.get("ForeignBuyVal")  or latest.get("foreign_buy_value")  or 0)
        sell_val = float(latest.get("ForeignSellVal") or latest.get("foreign_sell_value") or 0)
        net_val  = buy_val - sell_val

        if net_val > 0:
            activity = "Foreign Net Buy"
        elif net_val < 0:
            activity = "Foreign Net Sell"
        else:
            activity = "Neutral"

        result = {
            "foreign_net_lot": round(net_lot),
            "foreign_net_val": net_val,
            "foreign_activity": activity,
        }
        _FOREIGN_FLOW_CACHE[clean] = result
        return result
    except Exception:
        _FOREIGN_FLOW_CACHE[clean] = empty
        return empty


# =============================================================================
# UPGRADE 3 — BOARD CLASSIFICATION
# =============================================================================
_BOARD_CACHE: dict = {}

def fetch_board_classification(ticker: str) -> str:
    """
    Determine IDX board: Main Board / Development Board / Acceleration Board.
    Sourced from IDX company list API. Falls back to 'Main Board'.
    """
    clean = ticker.upper().replace(".JK", "").strip()
    if clean in _BOARD_CACHE:
        return _BOARD_CACHE[clean]
    try:
        import urllib.request, json
        url = (
            "https://www.idx.co.id/primary/ListedCompany/GetStockList"
            f"?start=0&length=1&code={clean}&name=&sector=&board=&lang=id"
        )
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=8) as resp:
            data = json.loads(resp.read().decode("utf-8", errors="ignore"))
        rows = data.get("data") or data.get("Data") or []
        if rows:
            raw_board = str(rows[0].get("Board") or rows[0].get("board") or "").strip()
            board_map = {
                "1": "Main Board", "MAIN": "Main Board", "UTAMA": "Main Board",
                "2": "Development Board", "DEV": "Development Board", "PENGEMBANGAN": "Development Board",
                "3": "Acceleration Board", "ACC": "Acceleration Board", "AKSELERASI": "Acceleration Board",
            }
            board = board_map.get(raw_board.upper(), f"Main Board")
        else:
            board = "Main Board"
    except Exception:
        board = "Main Board"
    _BOARD_CACHE[clean] = board
    return board


# =============================================================================
# UPGRADE 4 — BROKER FLOW (Top Broker Net)
# =============================================================================
_BROKER_CACHE: dict = {}

def fetch_broker_flow(ticker: str) -> dict:
    """
    Fetch top broker net buy/sell summary from IDX broker transaction data.
    Returns top 3 net-buy and net-sell broker codes + net values.
    Source: idx.co.id broker transaction summary endpoint.
    """
    empty = {"top_broker_buy": "-", "top_broker_sell": "-", "broker_net_signal": "-"}
    clean = ticker.upper().replace(".JK", "").strip()
    if clean in _BROKER_CACHE:
        return _BROKER_CACHE[clean]
    try:
        import urllib.request, json
        url = (
            "https://www.idx.co.id/primary/StockData/GetBrokerSummary"
            f"?code={clean}&lang=id"
        )
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=8) as resp:
            data = json.loads(resp.read().decode("utf-8", errors="ignore"))

        rows = data.get("data") or data.get("Data") or []
        if not rows:
            _BROKER_CACHE[clean] = empty
            return empty

        broker_nets = {}
        for row in rows:
            code    = str(row.get("BrokerCode") or row.get("broker_code") or "?")
            buy_val = float(row.get("BuyValue")  or row.get("buy_value")  or 0)
            sel_val = float(row.get("SellValue") or row.get("sell_value") or 0)
            broker_nets[code] = buy_val - sel_val

        sorted_brokers = sorted(broker_nets.items(), key=lambda x: x[1], reverse=True)
        top_buy  = ", ".join(f"{k}" for k, v in sorted_brokers[:3] if v > 0) or "-"
        top_sell = ", ".join(f"{k}" for k, v in reversed(sorted_brokers[-3:]) if v < 0) or "-"

        total_net = sum(broker_nets.values())
        signal = "Institutional Accumulation" if total_net > 0 else ("Institutional Distribution" if total_net < 0 else "Neutral")

        result = {
            "top_broker_buy":    top_buy,
            "top_broker_sell":   top_sell,
            "broker_net_signal": signal,
        }
        _BROKER_CACHE[clean] = result
        return result
    except Exception:
        _BROKER_CACHE[clean] = empty
        return empty


# =============================================================================
# UPGRADE 5 — SUSPENSION & CORPORATE ACTION FLAGS
# =============================================================================
_SUSPENSION_CACHE: dict = {}

def fetch_suspension_flag(ticker: str) -> dict:
    """
    Check if ticker is under trading suspension or active corporate action
    (rights issue mid-process, tender offer, etc.) from IDX.
    Source: idx.co.id corporate actions and trading halt endpoints.
    """
    empty = {"suspended": False, "suspension_label": "-", "corp_action_active": "-"}
    clean = ticker.upper().replace(".JK", "").strip()
    if clean in _SUSPENSION_CACHE:
        return _SUSPENSION_CACHE[clean]
    try:
        import urllib.request, json
        url = (
            "https://www.idx.co.id/primary/TradingHalt/GetTradingHalt"
            f"?code={clean}&lang=id"
        )
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=8) as resp:
            data = json.loads(resp.read().decode("utf-8", errors="ignore"))

        rows = data.get("data") or data.get("Data") or []
        if rows:
            latest = rows[0]
            halt_type = str(latest.get("HaltType") or latest.get("halt_type") or "")
            label = f"Suspended: {halt_type}" if halt_type else "Trading Halt"
            result = {"suspended": True, "suspension_label": label, "corp_action_active": label}
        else:
            result = {"suspended": False, "suspension_label": "-", "corp_action_active": "-"}

        _SUSPENSION_CACHE[clean] = result
        return result
    except Exception:
        _SUSPENSION_CACHE[clean] = empty
        return empty


# =============================================================================
# UPGRADE 6 — LIQUIDITY-ADJUSTED MAX POSITION SIZE
# =============================================================================
def compute_max_position_size(adtr20: float, close: float, capital: float = 1_000_000_000) -> dict:
    """
    Max Position Size = min(20% of ADTR20, 5% of capital) / Close → shares & lots.
    Conservative IDX illiquidity buffer: never consume more than 20% of avg daily volume value.
    Capital default = 1B IDR (configurable).
    """
    empty = {"max_shares": np.nan, "max_lots": np.nan, "max_position_idr": np.nan, "position_size_pct_adtv": np.nan}
    try:
        if pd.isna(adtr20) or pd.isna(close) or adtr20 <= 0 or close <= 0:
            return empty
        liquidity_cap = adtr20 * 0.20
        capital_cap   = capital * 0.05
        max_idr       = min(liquidity_cap, capital_cap)
        max_shares    = int(max_idr // close)
        max_lots      = max_shares // 100
        pct_adtv      = (max_idr / adtr20) * 100
        return {
            "max_shares":            max_shares,
            "max_lots":              max_lots,
            "max_position_idr":      round(max_idr),
            "position_size_pct_adtv": round(pct_adtv, 1),
        }
    except Exception:
        return empty


# =============================================================================
# UPGRADE 7 — IHSG SEASONAL BIAS CALENDAR
# =============================================================================
def get_seasonal_bias(month: int = None) -> str:
    """
    IHSG historical seasonal tendencies by calendar month.
    Based on long-run IHSG return seasonality patterns.
    Returns: Bullish / Neutral / Bearish + short rationale.
    """
    from datetime import date
    if month is None:
        month = date.today().month
    CALENDAR = {
        1:  "Bullish  — January Effect: foreign rebalancing inflows, new fiscal year optimism",
        2:  "Neutral  — Post-January consolidation; earnings season beginning",
        3:  "Neutral  — Q4 earnings season; mixed flows",
        4:  "Bullish  — Dividend announcement season; pre-Lebaran consumer spending tailwind",
        5:  "Bearish  — Sell in May pattern; post-Lebaran profit-taking; foreign outflow risk",
        6:  "Neutral  — Mid-year rebalancing; Lebaran holiday liquidity gap",
        7:  "Neutral  — Post-Lebaran recovery; H1 earnings releases",
        8:  "Bullish  — Dividend payment season; August ex-date clustering; institutional reinvestment",
        9:  "Bearish  — Global risk-off historically; EM outflow pressure; USD seasonal strength",
        10: "Neutral  — Q3 earnings season; recovery from September weakness",
        11: "Bullish  — Year-end rally setup; institutional window dressing begins",
        12: "Bullish  — Window dressing peak; year-end positioning; thin liquidity amplifies moves",
    }
    return CALENDAR.get(month, "Neutral")


# =============================================================================
# UPGRADE 8 — DIVIDEND TRAP DETECTION
# =============================================================================
def detect_dividend_trap(row: dict) -> str:
    """
    Flag dividend trap risk: upcoming ex-date + weak fundamentals.
    High risk: ex-date < 14 days AND (payout > 80% OR revenue_growth < 0).
    Medium risk: ex-date 14-30 days OR yield > 8% with weak earnings.
    """
    try:
        from datetime import date
        today = date.today()

        # Check upcoming ex-date proximity
        ex_raw = str(row.get("yf_upcoming_ex_date") or row.get("yf_last_dividend_date") or "")
        days_to_ex = None
        if len(ex_raw) >= 10:
            try:
                ex_date = date.fromisoformat(ex_raw[:10])
                days_to_ex = (ex_date - today).days
            except Exception:
                pass

        if days_to_ex is None or days_to_ex < 0 or days_to_ex > 60:
            return "-"

        payout      = safe_num(row.get("yf_payout_ratio"), np.nan)
        rev_growth  = safe_num(row.get("yf_revenue_growth"), np.nan)
        div_yield   = safe_num(row.get("yf_dividend_yield"), np.nan)
        earnings_gr = safe_num(row.get("yf_earnings_growth"), np.nan)

        weak_payout   = pd.notna(payout)      and payout > 0.80
        weak_growth   = pd.notna(rev_growth)  and rev_growth < 0
        high_yield    = pd.notna(div_yield)   and div_yield > 0.08
        weak_earnings = pd.notna(earnings_gr) and earnings_gr < 0

        if days_to_ex <= 14 and (weak_payout or weak_growth):
            return f"High Risk  — Ex-date in {days_to_ex}d, weak fundamentals"
        if days_to_ex <= 30 and high_yield and weak_earnings:
            return f"Medium Risk  — Ex-date in {days_to_ex}d, yield trap indicator"
        if days_to_ex <= 14:
            return f"Watch  — Ex-date in {days_to_ex}d"
        return "-"
    except Exception:
        return "-"


# =============================================================================
# UPGRADE 9 — OB AGE (DAYS SINCE ORDER BLOCK FORMED)
# =============================================================================
def compute_ob_age(hist: "pd.DataFrame", ob_level: float, lookback: int = 60) -> int:
    """
    Find how many bars ago the Order Block level (midpoint) was first formed.
    Scans backward through hist to find the pivot that generated the OB.
    Returns age in trading days; 0 if OB is today, np.nan if not found.
    """
    if hist is None or len(hist) < 5 or pd.isna(ob_level) or ob_level <= 0:
        return np.nan
    try:
        window = hist.tail(lookback)
        highs  = window["High"].values
        lows   = window["Low"].values
        for i in range(len(window) - 1, -1, -1):
            bar_mid = (highs[i] + lows[i]) / 2
            if abs(bar_mid - ob_level) / ob_level < 0.015:   # within 1.5%
                return len(window) - 1 - i   # bars since that candle
        return np.nan
    except Exception:
        return np.nan


# =============================================================================
# UPGRADE 10 — RRG QUADRANT FETCH
# =============================================================================
_RRG_QUADRANT_CACHE: dict = {}   # sector → quadrant

def get_rrg_quadrant_for_sector(sector: str, rrg_data: dict = None) -> str:
    """
    Look up the RRG quadrant for this stock's sector.
    rrg_data is a pre-built dict: {sector_name: quadrant_label} passed from the
    RRG computation pass (IDX Overview builder).
    Falls back to _RRG_QUADRANT_CACHE populated during IDX Overview build.
    Returns: Leading / Weakening / Lagging / Improving / N/A
    """
    if not sector or sector in ("-", "N/A", ""):
        return "-"
    if rrg_data:
        return rrg_data.get(str(sector).strip(), "-")
    return _RRG_QUADRANT_CACHE.get(str(sector).strip(), "-")





def safe_div_series(a, b, default=np.nan):
    try:
        result = a / b
        if hasattr(result, "replace"):
            result = result.replace([np.inf, -np.inf], np.nan)
            if default is not np.nan:
                result = result.fillna(default)
        return result
    except Exception:
        try:
            idx = getattr(a, "index", None) or getattr(b, "index", None)
            return pd.Series(default, index=idx)
        except Exception:
            return default

def _zone_bucket(close_val, vwap, m1, m2, m3):
    """Return integer zone bucket for a close relative to VWAP SD bands.
    1  = above VWAP
    0  = VWAP to -1 SD
    -1 = -1 SD to -2 SD
    -2 = -2 SD to -3 SD
    -3 = below -3 SD
    None = cannot compute
    """
    try:
        c = float(close_val)
        v = float(vwap)
        if pd.isna(c) or pd.isna(v):
            return None
        if c >= v:
            return 1
        if pd.notna(m1) and c >= float(m1):
            return 0
        if pd.notna(m2) and c >= float(m2):
            return -1
        if pd.notna(m3) and c >= float(m3):
            return -2
        return -3
    except Exception:
        return None

def _count_consecutive_zone_days(hist: pd.DataFrame, levels, current_zone_text: str) -> int:
    """Count consecutive trading days price has been in the same VWAP zone bucket as today.
    Uses fixed VWAP/SD levels applied backwards through recent history.
    levels = [vwap, m1, m2, m3]
    """
    try:
        if hist is None or hist.empty or not levels or len(levels) < 1:
            return 0
        vwap = safe_num(levels[0], np.nan)
        if pd.isna(vwap):
            return 0
        m1 = safe_num(levels[1], np.nan) if len(levels) > 1 else np.nan
        m2 = safe_num(levels[2], np.nan) if len(levels) > 2 else np.nan
        m3 = safe_num(levels[3], np.nan) if len(levels) > 3 else np.nan

        closes = hist["Close"].astype(float).values
        if len(closes) == 0:
            return 0
        current_bucket = _zone_bucket(closes[-1], vwap, m1, m2, m3)
        if current_bucket is None:
            return 0
        count = 0
        for i in range(len(closes) - 1, -1, -1):
            b = _zone_bucket(closes[i], vwap, m1, m2, m3)
            if b == current_bucket:
                count += 1
            else:
                break
        return max(0, count)
    except Exception:
        return 0

def clamp(x, min_val, max_val):
    try:
        if pd.isna(x):
            return np.nan
        return max(min_val, min(max_val, float(x)))
    except Exception:
        return np.nan

def normalize_score(value, low, high, inverse=False):
    try:
        if pd.isna(value):
            return np.nan
        if high == low:
            return 50.0
        score = ((float(value) - low) / (high - low)) * 100.0
        score = clamp(score, 0.0, 100.0)
        if inverse and pd.notna(score):
            score = 100.0 - score
        return score
    except Exception:
        return np.nan

def slope_last_n(series: pd.Series, n):
    try:
        s = pd.Series(series).dropna().tail(n)
        if len(s) < 2:
            return np.nan
        x = np.arange(len(s), dtype=float)
        y = s.astype(float).values
        return float(np.polyfit(x, y, 1)[0])
    except Exception:
        return np.nan

def classify_trend_from_slope(value, flat_band):
    if pd.isna(value):
        return "Flat"
    if value > flat_band:
        return "Rising"
    if value < -flat_band:
        return "Falling"
    return "Flat"

def percentile_or_threshold_score(value, thresholds=None):
    try:
        v = safe_num(value)
        if pd.isna(v):
            return np.nan
        thresholds = thresholds or [
            (1_000_000_000_000, 100),
            (500_000_000_000, 90),
            (100_000_000_000, 80),
            (50_000_000_000, 70),
            (10_000_000_000, 60),
            (5_000_000_000, 50),
            (1_000_000_000, 40),
            (500_000_000, 30),
            (100_000_000, 20),
        ]
        for t, s in thresholds:
            if v >= t:
                return float(s)
        return 10.0
    except Exception:
        return np.nan

def map_grade(score):
    s = safe_num(score)
    if pd.isna(s):
        return "N/A"
    if s >= 85: return "A+"
    if s >= 75: return "A"
    if s >= 65: return "B+"
    if s >= 55: return "B"
    if s >= 45: return "C+"
    if s >= 35: return "C"
    if s >= 25: return "D"
    return "F"

def safe_range_position(close, low, high):
    try:
        if pd.isna(close) or pd.isna(low) or pd.isna(high) or float(high) == float(low):
            return np.nan
        return clamp(((float(close) - float(low)) / (float(high) - float(low))) * 100.0, 0.0, 100.0)
    except Exception:
        return np.nan

def compute_institutional_metrics(hist: pd.DataFrame, row: Dict[str, Any]) -> Dict[str, Any]:
    # Safe OHLCV-only PROXY mode. No external dependency.
    out = {
        "flow_data_mode": "PROXY",
        "dollar_vol_20d_avg": np.nan,
        "turnover_velocity_20d": np.nan,
        "ad_trend": "Flat",
        "obv_trend": "Flat",
        "cmf20": np.nan,
        "volume_sponsorship_score": np.nan,
        "accumulation_score": np.nan,
        "distribution_score": np.nan,
        "smart_money_bias": "Neutral",
        "flow_conviction": "Neutral",
        "net_participation_proxy": np.nan,
        "sponsorship_grade": "N/A",

        "pos_52w_pct": np.nan,
        "range_compression_20d": np.nan,
        "vol_compression_20d": np.nan,
        "base_length": np.nan,
        "breakout_pressure": np.nan,
        "breakdown_pressure": np.nan,
        "_removed_spring": "NO",
        "_removed_upthrust": "NO",
        "structure_state": "N/A",
        "wyckoff_proxy_phase": "Transitional",
        "markup_readiness": np.nan,
        "breakdown_risk": np.nan,
        "cause_quality": "Weak",
        "wyckoff_event_start_date": "N/A",
        "wyckoff_event_confirm_date": "N/A",

        "technical_core_score": np.nan,
        "flow_score": np.nan,
        "structure_score": np.nan,
        "risk_penalty": np.nan,
        "regime_multiplier": np.nan,
        "adaptive_composite_score": np.nan,
        "setup_quality": "N/A",
        "trap_risk": "N/A",
        "institutional_verdict": "N/A",
        "institutional_action_bias": "N/A",

        "institutional_accumulation": "NO",
        "early_markup_candidate": "NO",
        "smart_money_pullback": "NO",
        "breakout_watchlist": "NO",
        "distribution_warning": "NO",
        "failed_breakout_risk": "NO",
        "capital_efficient_trend": "NO",
        "speculative_momentum": "NO",
        "preset_summary": "None",
    }
    if hist is None or hist.empty or len(hist) < 20:
        return out

    close = hist["Close"].astype(float)
    high = hist["High"].astype(float)
    low = hist["Low"].astype(float)
    vol = hist["Volume"].astype(float)

    # Core series
    dollar_vol = close * vol
    out["dollar_vol_20d_avg"] = safe_num(dollar_vol.tail(20).mean())

    mcap = safe_num(row.get("mcap"))
    if pd.notna(out["dollar_vol_20d_avg"]) and pd.notna(mcap) and mcap > 0:
        out["turnover_velocity_20d"] = safe_num((out["dollar_vol_20d_avg"] / mcap) * 100.0)

    hl_range = (high - low).replace(0, np.nan)
    mfm = (((close - low) - (high - close)) / hl_range).replace([np.inf, -np.inf], np.nan).fillna(0)
    mfv = mfm * vol
    ad_line = mfv.cumsum()

    close_prev = close.shift(1)
    obv = np.where(close > close_prev, vol, np.where(close < close_prev, -vol, 0))
    obv = pd.Series(obv, index=hist.index).cumsum()

    out["ad_trend"] = classify_trend_from_slope(slope_last_n(ad_line, 20), flat_band=max(abs(safe_num(vol.tail(20).mean(), 1)), 1))
    out["obv_trend"] = classify_trend_from_slope(slope_last_n(obv, 20), flat_band=max(abs(safe_num(vol.tail(20).mean(), 1)), 1))

    cmf_num = mfv.rolling(20).sum()
    cmf_den = vol.rolling(20).sum().replace(0, np.nan)
    cmf = safe_div(cmf_num.iloc[-1], cmf_den.iloc[-1], default=np.nan)
    out["cmf20"] = safe_num(cmf)

    avg20 = safe_num(vol.tail(20).mean())
    avg60 = safe_num(vol.tail(60).mean()) if len(vol) >= 60 else avg20
    vol_expansion = safe_div(avg20, avg60, default=np.nan)

    up_mask = close.diff() > 0
    down_mask = close.diff() < 0
    up_vol = safe_num(vol.tail(20)[up_mask.tail(20)].sum(), 0)
    down_vol = safe_num(vol.tail(20)[down_mask.tail(20)].sum(), 0)
    uv_ratio = safe_div(up_vol, down_vol, default=2.0 if up_vol > 0 and down_vol == 0 else np.nan)

    obv_slope_norm = safe_div(slope_last_n(obv, 20), max(avg20, 1), default=0)
    ad_slope_norm = safe_div(slope_last_n(ad_line, 20), max(avg20, 1), default=0)

    dollar_score = percentile_or_threshold_score(out["dollar_vol_20d_avg"])
    vol_exp_score = normalize_score(vol_expansion, 0.6, 2.0)
    uv_score = normalize_score(uv_ratio, 0.7, 2.0)
    cmf_pos_score = normalize_score(out["cmf20"], -0.2, 0.2)
    obv_score = normalize_score(obv_slope_norm, -1.5, 1.5)

    out["volume_sponsorship_score"] = safe_num(clamp(
        (0.30 * (0 if pd.isna(dollar_score) else dollar_score)) +
        (0.25 * (0 if pd.isna(vol_exp_score) else vol_exp_score)) +
        (0.20 * (0 if pd.isna(uv_score) else uv_score)) +
        (0.15 * (0 if pd.isna(cmf_pos_score) else cmf_pos_score)) +
        (0.10 * (0 if pd.isna(obv_score) else obv_score)),
        0, 100
    ))

    range20_high = safe_num(high.tail(20).max())
    range20_low = safe_num(low.tail(20).min())
    price_pos_20 = safe_range_position(close.iloc[-1], range20_low, range20_high)
    pos_close_ratio = safe_num((close.tail(20).diff() > 0).mean() * 100.0)
    neg_close_ratio = safe_num((close.tail(20).diff() < 0).mean() * 100.0)

    ad_up_score = normalize_score(ad_slope_norm, -1.5, 1.5)
    ad_dn_score = normalize_score(ad_slope_norm, -1.5, 1.5, inverse=True)
    obv_up_score = normalize_score(obv_slope_norm, -1.5, 1.5)
    obv_dn_score = normalize_score(obv_slope_norm, -1.5, 1.5, inverse=True)

    out["accumulation_score"] = safe_num(clamp(
        0.30 * (0 if pd.isna(cmf_pos_score) else cmf_pos_score) +
        0.25 * (0 if pd.isna(ad_up_score) else ad_up_score) +
        0.20 * (0 if pd.isna(obv_up_score) else obv_up_score) +
        0.15 * (0 if pd.isna(price_pos_20) else price_pos_20) +
        0.10 * (0 if pd.isna(pos_close_ratio) else pos_close_ratio),
        0, 100
    ))
    out["distribution_score"] = safe_num(clamp(
        0.30 * (0 if pd.isna(100 - cmf_pos_score) else 100 - cmf_pos_score) +
        0.25 * (0 if pd.isna(ad_dn_score) else ad_dn_score) +
        0.20 * (0 if pd.isna(obv_dn_score) else obv_dn_score) +
        0.15 * (0 if pd.isna(100 - price_pos_20) else 100 - price_pos_20) +
        0.10 * (0 if pd.isna(neg_close_ratio) else neg_close_ratio),
        0, 100
    ))

    a = safe_num(out["accumulation_score"])
    d = safe_num(out["distribution_score"])
    if pd.notna(a) and pd.notna(d):
        if a >= 65 and (a - d) >= 15:
            out["smart_money_bias"] = "Accumulation"
        elif d >= 65 and (d - a) >= 15:
            out["smart_money_bias"] = "Distribution"
        elif a < 45 and d < 45:
            out["smart_money_bias"] = "Neutral"
        else:
            out["smart_money_bias"] = "Mixed"

    vss = safe_num(out["volume_sponsorship_score"])
    smb = out["smart_money_bias"]
    if smb == "Accumulation" and pd.notna(vss) and vss >= 70:
        out["flow_conviction"] = "Strong Accumulation"
    elif smb == "Distribution" and pd.notna(vss) and vss >= 70:
        out["flow_conviction"] = "Strong Distribution"
    elif smb == "Accumulation" and pd.notna(vss) and vss >= 50:
        out["flow_conviction"] = "Mild Accumulation"
    elif smb == "Distribution" and pd.notna(vss) and vss >= 50:
        out["flow_conviction"] = "Mild Distribution"
    else:
        out["flow_conviction"] = "Neutral"

    np_proxy = (
        40 * safe_num(normalize_score(out["cmf20"], -0.2, 0.2), 50) / 100.0 +
        25 * safe_num(normalize_score(obv_slope_norm, -1.5, 1.5), 50) / 100.0 +
        20 * safe_num(normalize_score(ad_slope_norm, -1.5, 1.5), 50) / 100.0 +
        15 * safe_num(normalize_score(uv_ratio, 0.7, 2.0), 50) / 100.0
    )
    out["net_participation_proxy"] = safe_num(clamp((np_proxy - 50) * 2, -100, 100))
    out["sponsorship_grade"] = map_grade(out["volume_sponsorship_score"])

    # Structure / Wyckoff Proxy (V6 state-machine style)
    high_52 = safe_num(high.tail(252).max())
    low_52 = safe_num(low.tail(252).min())
    out["pos_52w_pct"] = safe_range_position(close.iloc[-1], low_52, high_52)

    raw_range_pct = safe_div((high.tail(20).max() - low.tail(20).min()), close.iloc[-1], default=np.nan)
    raw_range_pct = raw_range_pct * 100 if pd.notna(raw_range_pct) else np.nan
    out["range_compression_20d"] = safe_num(clamp(100 - normalize_score(raw_range_pct, 8, 25), 0, 100))

    prev_close = close.shift(1)
    tr = pd.concat([high - low, (high - prev_close).abs(), (low - prev_close).abs()], axis=1).max(axis=1)
    atr14 = tr.rolling(14).mean()
    atr_pct_series = safe_div_series(atr14, close, default=np.nan) * 100
    curr_atr_pct = safe_num(atr_pct_series.iloc[-1])
    med60 = safe_num(atr_pct_series.tail(60).median()) if len(atr_pct_series.dropna()) >= 20 else curr_atr_pct
    atr_ratio = safe_div(curr_atr_pct, med60, default=np.nan)
    out["vol_compression_20d"] = safe_num(clamp(100 - normalize_score(atr_ratio, 0.6, 1.6), 0, 100))

    base_len = 0
    for n in range(min(120, len(hist)), 19, -1):
        window = hist.tail(n)
        range_n = safe_div(window["High"].max() - window["Low"].min(), close.iloc[-1], default=np.nan)
        ema20_w = window["Close"].ewm(span=20, adjust=False).mean().iloc[-1]
        sma50_w = window["Close"].rolling(min(50, len(window))).mean().iloc[-1]
        spread = abs(safe_div(ema20_w - sma50_w, close.iloc[-1], default=np.nan))
        if pd.notna(range_n) and range_n <= 0.25 and (pd.isna(spread) or spread <= 0.08):
            base_len = n
            break
    out["base_length"] = base_len

    hh20 = safe_num(high.tail(20).max())
    ll20 = safe_num(low.tail(20).min())
    prox_high = 100 - normalize_score(abs(safe_div(hh20 - close.iloc[-1], close.iloc[-1], default=np.nan))*100, 0, 12)
    prox_low = 100 - normalize_score(abs(safe_div(close.iloc[-1] - ll20, close.iloc[-1], default=np.nan))*100, 0, 12)

    higher_lows = 100 if low.tail(10).min() > low.tail(20).min() else 40
    lower_highs = 100 if high.tail(10).max() < high.tail(20).max() else 40
    momentum_pos = 100 if safe_num(row.get("rsi14")) >= 50 else 30
    momentum_neg = 100 if safe_num(row.get("rsi14")) < 50 else 30

    out["breakout_pressure"] = safe_num(clamp(
        0.30 * safe_num(prox_high, 0) +
        0.25 * safe_num(out["volume_sponsorship_score"], 0) +
        0.20 * higher_lows +
        0.15 * safe_num(out["pos_52w_pct"], 0) +
        0.10 * momentum_pos, 0, 100
    ))
    out["breakdown_pressure"] = safe_num(clamp(
        0.30 * safe_num(prox_low, 0) +
        0.25 * safe_num(out["distribution_score"], 0) +
        0.20 * lower_highs +
        0.15 * safe_num(100 - out["pos_52w_pct"], 0) +
        0.10 * momentum_neg, 0, 100
    ))

    recent_support = low.shift(1).rolling(60, min_periods=20).min()
    recent_res = high.shift(1).rolling(60, min_periods=20).max()
    ema20_all = close.ewm(span=20, adjust=False).mean()
    sma50_all = close.rolling(50).mean()
    sma200_all = close.rolling(200).mean()

    up_trend = (
        pd.notna(ema20_all.iloc[-1]) and pd.notna(sma50_all.iloc[-1]) and
        close.iloc[-1] > ema20_all.iloc[-1] > sma50_all.iloc[-1]
    )
    down_trend = (
        pd.notna(ema20_all.iloc[-1]) and pd.notna(sma50_all.iloc[-1]) and
        close.iloc[-1] < ema20_all.iloc[-1] < sma50_all.iloc[-1]
    )

    # V6 bar-by-bar state candidates
    avg20v = vol.rolling(20).mean()
    spring_cand = (
        (low < recent_support) &
        (close > recent_support) &
        (vol >= avg20v.fillna(vol))
    ).fillna(False)
    spring_confirm = (
        (spring_cand.shift(1).fillna(False) | spring_cand.shift(2).fillna(False)) &
        (close > recent_support.fillna(close.shift(1))) &
        (close >= ema20_all.fillna(close)) &
        (close >= close.shift(1))
    ).fillna(False)

    utad_cand = (
        (high > recent_res) &
        (close < recent_res) &
        (vol >= avg20v.fillna(vol))
    ).fillna(False)
    utad_confirm = (
        (utad_cand.shift(1).fillna(False) | utad_cand.shift(2).fillna(False)) &
        (close < recent_res.fillna(close.shift(1))) &
        (close <= ema20_all.fillna(close))
    ).fillna(False)

    base_state = (
        (((high.rolling(20).max() - low.rolling(20).min()) / close.replace(0, np.nan)) * 100 <= 18) &
        (close >= low.rolling(20).min() + (high.rolling(20).max() - low.rolling(20).min()) * 0.25) &
        (close <= low.rolling(20).min() + (high.rolling(20).max() - low.rolling(20).min()) * 0.80)
    ).fillna(False)

    constructive = (
        (close >= ema20_all.fillna(close)) &
        ((ema20_all >= sma50_all.fillna(ema20_all)) | sma50_all.isna())
    ).fillna(False)
    weak_context = (
        (close <= ema20_all.fillna(close)) &
        ((ema20_all <= sma50_all.fillna(ema20_all)) | sma50_all.isna())
    ).fillna(False)

    sos_state = (
        spring_confirm.rolling(5, min_periods=1).max().astype(bool) &
        constructive &
        (close >= high.shift(1))
    ).fillna(False)

    lps_state = (
        constructive &
        (close < close.shift(1)) &
        (close >= ema20_all.fillna(close)) &
        (close >= close.rolling(10).min())
    ).fillna(False)

    markup_state = (
        constructive &
        (close > high.rolling(20).max().shift(1).fillna(close))
    ).fillna(False)

    continuation_state = (
        markup_state &
        (vol >= avg20v.fillna(vol))
    ).fillna(False)

    distribution_state = (
        weak_context &
        (safe_num(out["pos_52w_pct"], 50) >= 60) &
        (safe_num(out["distribution_score"], 0) >= 55)
    )
    distribution_state = pd.Series([distribution_state] * len(close), index=close.index) if isinstance(distribution_state, bool) else distribution_state

    sow_state = (
        utad_confirm.rolling(5, min_periods=1).max().astype(bool) &
        weak_context &
        (close <= low.shift(1))
    ).fillna(False)

    lpsy_state = (
        weak_context &
        (close > close.shift(1)) &
        (close <= ema20_all.fillna(close))
    ).fillna(False)

    markdown_state = (
        weak_context &
        (close < low.rolling(20).min().shift(1).fillna(close))
    ).fillna(False)

    # Determine current active state with priority / memory
    current_event = "No Clean Event"
    current_phase = "Transitional"

    if bool(markdown_state.iloc[-1]):
        current_phase = "Markdown"
        current_event = "Trend Weakness"
    elif bool(sow_state.iloc[-1]):
        current_phase = "Distribution"
        current_event = "SOW"
    elif bool(utad_confirm.iloc[-1]) or bool(utad_cand.iloc[-1]):
        current_phase = "Distribution"
        current_event = "UTAD"
    elif bool(lpsy_state.iloc[-1]) and safe_num(out["distribution_score"], 0) >= 55:
        current_phase = "Distribution"
        current_event = "LPSY"
    elif bool(continuation_state.iloc[-1]):
        current_phase = "Markup"
        current_event = "Continuation"
    elif bool(markup_state.iloc[-1]) and bool(lps_state.iloc[-1]):
        current_phase = "Markup"
        current_event = "Pullback"
    elif bool(sos_state.iloc[-1]):
        current_phase = "Accumulation"
        current_event = "SOS"
    elif bool(lps_state.iloc[-1]) and safe_num(out["accumulation_score"], 0) >= 55:
        current_phase = "Accumulation"
        current_event = "LPS"
    elif bool(spring_confirm.iloc[-1]) or bool(spring_cand.iloc[-1]):
        current_phase = "Accumulation"
        current_event = "Spring"
    elif bool(base_state.iloc[-1]):
        current_phase = "Accumulation"
        current_event = "Base Build"
    else:
        # fallback to prior regime-style mapping
        if up_trend and safe_num(out["breakout_pressure"], 0) >= 60:
            current_phase = "Markup"
            current_event = "Continuation"
        elif down_trend and safe_num(out["breakdown_pressure"], 0) >= 60:
            current_phase = "Markdown"
            current_event = "Trend Weakness"
        elif safe_num(out["distribution_score"], 0) >= 60 and safe_num(out["pos_52w_pct"], 50) >= 60:
            current_phase = "Distribution"
            current_event = "Distribution Range"
        elif safe_num(out["range_compression_20d"], 0) >= 55:
            current_phase = "Accumulation"
            current_event = "Base Build"
        else:
            current_phase = "Transitional"
            current_event = "No Clean Event"

    # Map structure state from new state engine
    if current_phase == "Accumulation":
        out["structure_state"] = "Tight Base" if safe_num(out["range_compression_20d"], 0) >= 70 else "Loose Base"
    elif current_phase == "Markup":
        out["structure_state"] = "Advancing Trend"
    elif current_phase == "Distribution":
        out["structure_state"] = "Distribution Range"
    elif current_phase == "Markdown":
        out["structure_state"] = "Breakdown Structure"
    else:
        out["structure_state"] = "Weakening Trend"

    out["wyckoff_proxy_phase"] = current_phase
    out["wyckoff_event_label"] = current_event

    base_quality_map = normalize_score(base_len, 20, 100)
    out["markup_readiness"] = safe_num(clamp(
        0.25 * safe_num(out["range_compression_20d"], 0) +
        0.20 * safe_num(out["vol_compression_20d"], 0) +
        0.20 * safe_num(out["breakout_pressure"], 0) +
        0.15 * safe_num(out["accumulation_score"], 0) +
        0.10 * safe_num(out["volume_sponsorship_score"], 0) +
        0.10 * safe_num(base_quality_map, 0), 0, 100
    ))
    upthrust_penalty = 100 if current_event == "UTAD" else (60 if bool(utad_cand.iloc[-1]) else 0)
    out["breakdown_risk"] = safe_num(clamp(
        0.25 * safe_num(out["breakdown_pressure"], 0) +
        0.20 * safe_num(out["distribution_score"], 0) +
        0.15 * safe_num(100 - out["pos_52w_pct"], 0) +
        0.15 * momentum_neg +
        0.15 * upthrust_penalty +
        0.10 * safe_num(100 - out["volume_sponsorship_score"], 0), 0, 100
    ))

    out["_removed_spring"] = "YES" if current_event == "Spring" else ("Weak" if bool(spring_cand.iloc[-1]) else "NO")
    out["_removed_upthrust"] = "YES" if current_event == "UTAD" else ("Weak" if bool(utad_cand.iloc[-1]) else "NO")

    if base_len >= 60 and safe_num(out["range_compression_20d"], 0) >= 70 and current_phase == "Accumulation":
        out["cause_quality"] = "Excellent"
    elif base_len >= 40 and safe_num(out["range_compression_20d"], 0) >= 55:
        out["cause_quality"] = "Good"
    elif base_len >= 20:
        out["cause_quality"] = "Average"
    else:
        out["cause_quality"] = "Weak"


    # Adaptive verdict
    tech_parts = []
    tech_parts.append(100 if row.get("summary_ma") == "Above All MA" else (75 if str(row.get("summary_ma", "")).startswith("Above ") else 30))
    tech_parts.append(100 if row.get("cross_status") == "Golden" else (20 if row.get("cross_status") == "Dead" else 50))
    rsi14 = safe_num(row.get("rsi14"))
    tech_parts.append(safe_num(normalize_score(rsi14, 35, 70), 50))
    tech_parts.append(80 if row.get("div_signal") == "Bullish" else (20 if row.get("div_signal") == "Bearish" else 50))
    tech_parts.append(80 if row.get("rvol_zone") == "YES" else 40)
    out["technical_core_score"] = safe_num(np.nanmean(pd.Series(tech_parts, dtype=float)))

    fc_map = {"Strong Accumulation": 100, "Mild Accumulation": 75, "Neutral": 50, "Mild Distribution": 25, "Strong Distribution": 0}
    out["flow_score"] = safe_num(clamp(
        0.30 * safe_num(out["volume_sponsorship_score"], 0) +
        0.25 * safe_num(out["accumulation_score"], 0) +
        0.25 * safe_num(100 - out["distribution_score"], 0) +
        0.10 * safe_num((out["net_participation_proxy"] + 100) / 2, 50) +
        0.10 * fc_map.get(out["flow_conviction"], 50), 0, 100
    ))
    cause_map = {"Excellent": 100, "Good": 75, "Average": 50, "Weak": 25}
    out["structure_score"] = safe_num(clamp(
        0.25 * safe_num(out["markup_readiness"], 0) +
        0.20 * safe_num(100 - out["breakdown_risk"], 0) +
        0.20 * safe_num(out["breakout_pressure"], 0) +
        0.15 * safe_num(out["range_compression_20d"], 0) +
        0.10 * safe_num(out["vol_compression_20d"], 0) +
        0.10 * cause_map.get(out["cause_quality"], 25), 0, 100
    ))

    liq = str(row.get("liquidity_category", ""))
    mc = str(row.get("market_cap_category", ""))
    weak_liq_pen = 100 if liq in ("Very Low Liquidity",) else (70 if liq in ("Low Liquidity",) else (30 if liq in ("Medium Liquidity",) else 0))
    small_pen = 100 if mc == "Micro Cap" else (70 if mc == "Small Cap" else (25 if mc == "Mid Cap" else 0))
    bearish_div_pen = 100 if row.get("div_signal") == "Bearish" else 0
    out["risk_penalty"] = safe_num(clamp(
        0.30 * safe_num(out["breakdown_risk"], 0) +
        0.25 * safe_num(out["distribution_score"], 0) +
        0.15 * upthrust_penalty +
        0.10 * weak_liq_pen +
        0.10 * small_pen +
        0.10 * bearish_div_pen, 0, 100
    ))

    regime_mult = 1.0
    if mc == "Large Cap" and liq == "High Liquidity":
        regime_mult = 1.10
    elif mc == "Mid Cap" and liq in ("High Liquidity", "Medium Liquidity"):
        regime_mult = 1.04
    elif mc in ("Small Cap",) and liq in ("Medium Liquidity",):
        regime_mult = 0.98
    elif mc in ("Small Cap", "Micro Cap") and liq in ("Low Liquidity", "Very Low Liquidity"):
        regime_mult = 0.88
    if out["smart_money_bias"] == "Accumulation" and out["structure_state"] == "Tight Base":
        regime_mult += 0.03
    out["regime_multiplier"] = clamp(regime_mult, 0.80, 1.15)

    raw = (
        0.40 * safe_num(out["technical_core_score"], 0) +
        0.30 * safe_num(out["flow_score"], 0) +
        0.20 * safe_num(out["structure_score"], 0) -
        0.10 * safe_num(out["risk_penalty"], 0)
    )
    out["adaptive_composite_score"] = safe_num(clamp(raw * out["regime_multiplier"], 0, 100))

    acs = safe_num(out["adaptive_composite_score"])
    if pd.notna(acs):
        if acs >= 85: out["setup_quality"] = "Elite"
        elif acs >= 75: out["setup_quality"] = "High Quality"
        elif acs >= 65: out["setup_quality"] = "Constructive"
        elif acs >= 55: out["setup_quality"] = "Developing"
        elif acs >= 45: out["setup_quality"] = "Speculative"
        else: out["setup_quality"] = "Weak"

    rp = safe_num(out["risk_penalty"])
    if pd.notna(rp):
        if rp >= 75: out["trap_risk"] = "High"
        elif rp >= 60: out["trap_risk"] = "Elevated"
        elif rp >= 45: out["trap_risk"] = "Moderate"
        else: out["trap_risk"] = "Low"
    if out["_removed_upthrust"] == "YES" and safe_num(out["distribution_score"], 0) >= 60 and out["trap_risk"] == "Low":
        out["trap_risk"] = "Elevated"

    if safe_num(out["adaptive_composite_score"], 0) >= 80 and safe_num(out["flow_score"], 0) >= 65 and safe_num(out["structure_score"], 0) >= 65:
        out["institutional_verdict"] = "Institutional Long Candidate"
    elif 65 <= safe_num(out["adaptive_composite_score"], 0) < 80 and out["trap_risk"] != "High":
        out["institutional_verdict"] = "Constructive Long Watchlist"
    elif safe_num(out["breakdown_risk"], 0) >= 70 and safe_num(out["distribution_score"], 0) >= 65:
        out["institutional_verdict"] = "Breakdown Risk"
    elif safe_num(out["adaptive_composite_score"], 0) < 50 and safe_num(out["risk_penalty"], 0) >= 60:
        out["institutional_verdict"] = "Avoid / Distribution Risk"
    else:
        out["institutional_verdict"] = "Neutral / Selective"

    if out["institutional_verdict"] == "Institutional Long Candidate" and out["structure_state"] in ("Tight Base", "Loose Base"):
        out["institutional_action_bias"] = "Breakout Watch"
    elif out["institutional_verdict"] in ("Institutional Long Candidate", "Constructive Long Watchlist") and out["smart_money_bias"] == "Accumulation":
        out["institutional_action_bias"] = "Accumulate on Pullback"
    elif out["institutional_verdict"] in ("Institutional Long Candidate", "Constructive Long Watchlist"):
        out["institutional_action_bias"] = "Hold / Trend Follow"
    elif out["institutional_verdict"] == "Breakdown Risk":
        out["institutional_action_bias"] = "Reduce Into Strength"
    elif out["institutional_verdict"] == "Avoid / Distribution Risk":
        out["institutional_action_bias"] = "Avoid"
    else:
        out["institutional_action_bias"] = "Range Only"

    # Presets
    out["institutional_accumulation"] = "YES" if out["smart_money_bias"] == "Accumulation" and safe_num(out["volume_sponsorship_score"], 0) >= 65 and safe_num(out["accumulation_score"], 0) >= 65 and safe_num(out["distribution_score"], 100) <= 50 else "NO"
    out["early_markup_candidate"] = "YES" if out["wyckoff_proxy_phase"] in ("Accumulation", "Transitional") and safe_num(out["markup_readiness"], 0) >= 70 and safe_num(out["breakout_pressure"], 0) >= 65 and safe_num(out["breakdown_risk"], 100) <= 45 else "NO"
    bullish_trend = str(row.get("summary_ma", "")).startswith("Above") and safe_num(row.get("close")) >= safe_num(row.get("ema20"), safe_num(row.get("close"), 0))
    out["smart_money_pullback"] = "YES" if bullish_trend and out["smart_money_bias"] != "Distribution" and safe_num(out["flow_score"], 0) >= 55 else "NO"
    out["breakout_watchlist"] = "YES" if safe_num(out["markup_readiness"], 0) >= 75 and safe_num(out["breakout_pressure"], 0) >= 70 and safe_num(out["range_compression_20d"], 0) >= 60 and safe_num(out["volume_sponsorship_score"], 0) >= 60 else "NO"
    out["distribution_warning"] = "YES" if (out["smart_money_bias"] == "Distribution" or safe_num(out["distribution_score"], 0) >= 70) and (out["_removed_upthrust"] in ("YES", "Weak") or out["structure_state"] in ("Weakening Trend", "Distribution Range")) else "NO"
    out["failed_breakout_risk"] = "YES" if out["_removed_upthrust"] == "YES" or (safe_num(out["pos_52w_pct"], 0) >= 70 and safe_num(out["breakdown_risk"], 0) >= 65 and safe_num(out["distribution_score"], 0) >= 60) else "NO"
    healthy_tv = pd.notna(out["turnover_velocity_20d"]) and 0.10 <= out["turnover_velocity_20d"] <= 5.0
    out["capital_efficient_trend"] = "YES" if bullish_trend and healthy_tv and safe_num(out["volume_sponsorship_score"], 0) >= 55 and safe_num(out["risk_penalty"], 100) <= 40 and safe_num(out["adaptive_composite_score"], 0) >= 70 else "NO"
    out["speculative_momentum"] = "YES" if mc in ("Small Cap", "Micro Cap") and safe_num(out["technical_core_score"], 0) >= 65 and safe_num(out["volume_sponsorship_score"], 0) >= 60 and out["trap_risk"] != "High" else "NO"

    labels = []
    mapping = [
        ("institutional_accumulation", "Institutional Accumulation"),
        ("early_markup_candidate", "Early Markup Candidate"),
        ("breakout_watchlist", "Breakout Watchlist"),
        ("smart_money_pullback", "Smart Money Pullback"),
        ("capital_efficient_trend", "Capital Efficient Trend"),
        ("distribution_warning", "Distribution Warning"),
        ("failed_breakout_risk", "Failed Breakout Risk"),
        ("speculative_momentum", "Speculative Momentum"),
    ]
    for k, label in mapping:
        if out.get(k) == "YES":
            labels.append(label)
    out["preset_summary"] = " | ".join(labels) if labels else "None"

    return out


def _fmt_date_or_na(ts):
    try:
        if ts is None or pd.isna(ts):
            return "N/A"
        return pd.Timestamp(ts).strftime("%Y-%m-%d")
    except Exception:
        return "N/A"

# ALPHAFLOW PROXY MAX (NO BROKER - SAFE LAYER)
# =========================
def safe_pct(a, b, default=np.nan):
    try:
        if pd.isna(a) or pd.isna(b) or float(b) == 0:
            return default
        return (float(a) / float(b)) * 100.0
    except Exception:
        return default

def compute_alphaflow_proxy_max(hist: pd.DataFrame, row: dict):
    out = {
        "af_smt_proxy": np.nan,
        "af_verdict": "Neutral",
        "af_verdict_confidence": np.nan,
        "af_phase_event": "Neutral / No Clean Event",
        "af_flow_edge": "Low",
        "af_hit_rate_proxy": np.nan,
        "af_r2_proxy": np.nan,
        "af_execution_guide": "Wait for better alignment",
        "af_watch_next": "Need better flow + structure confirmation",
        "af_invalidation": "Lose support / fail VWAP-QVWAP acceptance",
        "af_playbook": "No clean setup",
    }
    if hist is None or hist.empty or len(hist) < 30:
        return out

    df = hist.copy()
    close = df["Close"].astype(float)
    high = df["High"].astype(float)
    low = df["Low"].astype(float)
    vol = df["Volume"].astype(float)

    # Flow proxies
    ret = close.pct_change()
    clv = ((close - low) - (high - close)) / (high - low).replace(0, np.nan)
    mfv = (clv * vol).replace([np.inf, -np.inf], np.nan).fillna(0)
    obv = (np.sign(close.diff().fillna(0)) * vol).fillna(0).cumsum()

    flow20 = safe_num(mfv.tail(20).sum(), 0)
    flow60 = safe_num(mfv.tail(60).sum(), 0)
    dv20 = safe_num((close * vol).tail(20).sum(), np.nan)
    dv60 = safe_num((close * vol).tail(60).sum(), np.nan)

    fp20 = safe_pct(flow20, dv20, 0)
    fp60 = safe_pct(flow60, dv60, 0)
    obv20 = safe_pct(safe_num(obv.iloc[-1] - obv.iloc[max(0, len(obv)-20)], 0), safe_num(vol.tail(20).sum(), np.nan), 0)
    obv60 = safe_pct(safe_num(obv.iloc[-1] - obv.iloc[max(0, len(obv)-60)], 0), safe_num(vol.tail(60).sum(), np.nan), 0)

    # Participation / sponsorship
    vol20 = vol.tail(20)
    part_conc = np.nan
    burst = np.nan
    if len(vol20) >= 5 and vol20.sum() > 0:
        part_conc = (vol20.nlargest(min(3, len(vol20))).sum() / vol20.sum()) * 100.0
        burst = ((vol20 > (vol20.mean() * 1.5)).sum() / len(vol20)) * 100.0

    # Flow-price validation
    corr20 = np.nan
    corr40 = np.nan
    for n in [20, 40]:
        f = mfv.pct_change().replace([np.inf, -np.inf], np.nan).tail(n)
        r = ret.tail(n)
        valid = pd.concat([f, r], axis=1).dropna()
        if len(valid) >= max(8, n // 2):
            corr = valid.iloc[:, 0].corr(valid.iloc[:, 1])
            if n == 20:
                corr20 = corr
            else:
                corr40 = corr
    r2 = ((corr20 or 0) ** 2) * 100 if pd.notna(corr20) else np.nan
    hit = np.nan
    hv = ((ret.tail(20) > 0) == (mfv.pct_change().tail(20) > 0)).astype(float)
    if len(hv.dropna()) > 0:
        hit = hv.mean() * 100.0

    # SMT proxy (0-100) inspired by guide weights, but OHLCV proxy only
    net_flow_score = normalize_score(fp20, -8, 8)
    persist_score = normalize_score((obv20 * 0.7) + (obv60 * 0.3), -8, 8)
    concentration_score = normalize_score(safe_num(part_conc, 50), 35, 70)
    absorption_proxy = normalize_score(safe_num(row.get("range_compression_20d", np.nan), 50), 30, 80)
    execution_proxy = normalize_score(safe_num(row.get("close_vs_qvwap_pct", np.nan), 0), -6, 6)
    smt = (
        0.30 * safe_num(net_flow_score, 50) +
        0.25 * safe_num(persist_score, 50) +
        0.20 * safe_num(concentration_score, 50) +
        0.15 * safe_num(absorption_proxy, 50) +
        0.10 * safe_num(execution_proxy, 50)
    )
    out["af_smt_proxy"] = safe_num(clamp(smt, 0, 100))

    # Map flow edge similar to guide thresholds
    out["af_r2_proxy"] = safe_num(r2)
    out["af_hit_rate_proxy"] = safe_num(hit)

    if pd.notna(r2) and pd.notna(hit):
        if r2 >= 10 and hit >= 58:
            out["af_flow_edge"] = "Strong"
        elif r2 >= 3 and hit >= 52:
            out["af_flow_edge"] = "Moderate"
        else:
            out["af_flow_edge"] = "Weak"

    # Verdict synthesis
    accum = safe_num(row.get("accumulation_score", np.nan), 50)
    dist = safe_num(row.get("distribution_score", np.nan), 50)
    flow = safe_num(row.get("flow_score", np.nan), 50)
    struct = safe_num(row.get("structure_score", np.nan), 50)
    risk = safe_num(row.get("risk_penalty", np.nan), 50)
    qvwap_ok = 1 if safe_num(row.get("close_vs_qvwap_pct", -999), -999) >= -1.5 else 0
    div = str(row.get("divergence_summary", "None") or "None")
    phase = str(row.get("wyckoff_proxy_phase", "Transitional") or "Transitional")
    spring = str(row.get("_removed_spring", "NO") or "NO")
    ut = str(row.get("_removed_upthrust", "NO") or "NO")

    bull_bonus = 0
    bear_penalty = 0
    if "Bullish" in div:
        bull_bonus += 8
    if "Bearish" in div:
        bear_penalty += 8
    if spring in ("YES", "Weak"):
        bull_bonus += 6
    if ut in ("YES", "Weak"):
        bear_penalty += 6

    verdict_score = (
        0.26 * accum +
        0.22 * flow +
        0.16 * struct +
        0.16 * safe_num(out["af_smt_proxy"], 50) +
        0.10 * (100 if qvwap_ok else 35) +
        0.10 * (70 if out["af_flow_edge"] == "Strong" else 55 if out["af_flow_edge"] == "Moderate" else 40)
        - 0.18 * risk
        - 0.12 * dist
        + bull_bonus - bear_penalty
    )

    # Normalize around website-like 5-state outcome
    centered = (verdict_score - 50) / 100.0
    confidence = clamp(
        abs(accum - dist) * 0.45 +
        abs(safe_num(out["af_smt_proxy"], 50) - 50) * 0.35 +
        abs(flow - 50) * 0.20,
        35, 92
    )
    out["af_verdict_confidence"] = safe_num(confidence)

    if centered > 0.35:
        out["af_verdict"] = "STRONG ACCUMULATION"
    elif centered > 0.05:
        out["af_verdict"] = "ACCUMULATION"
    elif centered < -0.35:
        out["af_verdict"] = "STRONG DISTRIBUTION"
    elif centered < -0.05:
        out["af_verdict"] = "DISTRIBUTION"
    else:
        out["af_verdict"] = "NEUTRAL"

    # Phase + event proxy (V6: align with rebuilt state engine)
    current_event = str(row.get("wyckoff_event_label", "") or "")
    if not current_event:
        if phase == "Accumulation":
            current_event = "Spring" if spring == "YES" else ("SOS" if safe_num(row.get("markup_readiness", 0), 0) >= 70 else "Base Build")
        elif phase == "Markup":
            current_event = "Pullback" if qvwap_ok and "Bullish" in div else "Continuation"
        elif phase == "Distribution":
            current_event = "UTAD" if ut == "YES" else "SOW"
        elif phase == "Markdown":
            current_event = "Trend Weakness"
        else:
            current_event = "No Clean Event"

    if phase == "Accumulation":
        out["af_phase_event"] = f"Accumulation / {current_event}"
    elif phase == "Markup":
        out["af_phase_event"] = f"Markup / {current_event}"
    elif phase == "Distribution":
        out["af_phase_event"] = f"Distribution / {current_event}"
    elif phase == "Markdown":
        out["af_phase_event"] = "Markdown / Trend Weakness"
    else:
        out["af_phase_event"] = "Transitional / No Clean Event"

    # Execution guide (replace tier-thinking with playbook-thinking)
    if out["af_verdict"] in ("STRONG ACCUMULATION", "ACCUMULATION"):
        if phase in ("Accumulation", "Transitional"):
            out["af_execution_guide"] = "Wait for hold above QVWAP + stronger RVOL, then accumulate on constructive pullback."
            out["af_playbook"] = "Accumulation Setup"
        elif phase == "Markup":
            out["af_execution_guide"] = "Trend-follow only: buy pullback into QVWAP/EMA20 if support holds."
            out["af_playbook"] = "Markup Continuation"
        else:
            out["af_execution_guide"] = "Bullish score but phase mismatch — reduce size, wait for structure confirmation."
            out["af_playbook"] = "Watchlist Only"
        out["af_watch_next"] = "Need QVWAP acceptance, sustained flow score, and no failed reclaim."
        out["af_invalidation"] = "Lose QVWAP + break prior swing low / support."
    elif out["af_verdict"] in ("DISTRIBUTION", "STRONG DISTRIBUTION"):
        out["af_execution_guide"] = "Avoid fresh longs. Only reconsider after re-accumulation evidence or stronger reclaim."
        out["af_playbook"] = "Distribution Warning"
        out["af_watch_next"] = "Need absorption + reclaim above QVWAP + distribution score cooling."
        out["af_invalidation"] = "Sustained reclaim above QVWAP and improving flow/structure."
    else:
        out["af_execution_guide"] = "No edge. Wait for clearer phase + flow alignment."
        out["af_playbook"] = "Neutral / No Setup"
        out["af_watch_next"] = "Monitor for Spring, SOS, LPS, or stronger flow-price validation."
        out["af_invalidation"] = "N/A"

    return out


# =========================
# WYCKOFF STATE-BASED DATE HELPERS (V5)
# =========================
def _find_last_true_run(mask: pd.Series):
    """
    Return (start_idx, end_idx) of the most recent contiguous True run.
    If the last value is False, returns (None, None).
    """
    try:
        if mask is None or len(mask) == 0:
            return None, None
        m = pd.Series(mask).fillna(False).astype(bool)
        if not bool(m.iloc[-1]):
            return None, None
        end_idx = m.index[-1]
        start_pos = len(m) - 1
        while start_pos - 1 >= 0 and bool(m.iloc[start_pos - 1]):
            start_pos -= 1
        start_idx = m.index[start_pos]
        return start_idx, end_idx
    except Exception:
        return None, None

def _first_confirm_after(mask: pd.Series, start_idx):
    """
    Return first index at/after start_idx where mask is True.
    """
    try:
        if mask is None or len(mask) == 0 or start_idx is None:
            return None
        m = pd.Series(mask).fillna(False).astype(bool)
        sub = m.loc[m.index >= start_idx]
        trues = sub[sub]
        return trues.index[0] if len(trues) else None
    except Exception:
        return None


def _fmt_wyckoff_date(idx):
    try:
        if idx is None:
            return None
        return pd.Timestamp(idx).to_pydatetime()
    except Exception:
        return None

# MARKET STRUCTURE (Leviathan-style pivot / BOS / CHoCH)
# =========================
MS_SWING_SIZE = 20
MS_BOS_CONFIRM = "close"   # "close" or "wick"
MS_MIN_HISTORY = 120
MS_VALUE_LOOKBACK = 20

def _fmt_ms_date(ts):
    try:
        return pd.Timestamp(ts).strftime("%d %b '%y")
    except Exception:
        return ""

def _pivot_high_ms(high: pd.Series, left: int, right: int) -> pd.Series:
    out = pd.Series(np.nan, index=high.index, dtype=float)
    vals = high.values
    n = len(vals)
    for i in range(left, n - right):
        win = vals[i-left:i+right+1]
        if np.isnan(vals[i]) or np.isnan(win).any():
            continue
        if vals[i] == np.max(win):
            out.iloc[i] = vals[i]
    return out

def _pivot_low_ms(low: pd.Series, left: int, right: int) -> pd.Series:
    out = pd.Series(np.nan, index=low.index, dtype=float)
    vals = low.values
    n = len(vals)
    for i in range(left, n - right):
        win = vals[i-left:i+right+1]
        if np.isnan(vals[i]) or np.isnan(win).any():
            continue
        if vals[i] == np.min(win):
            out.iloc[i] = vals[i]
    return out

def _classify_market_structure_swings(hist: pd.DataFrame, swing_size: int = MS_SWING_SIZE):
    h = hist.copy()
    h["ms_piv_hi"] = _pivot_high_ms(h["High"], swing_size, swing_size)
    h["ms_piv_lo"] = _pivot_low_ms(h["Low"], swing_size, swing_size)

    prev_high = np.nan
    prev_low = np.nan
    high_active = False
    low_active = False
    prev_breakout_dir = 0  # +1 up, -1 down

    swings = []   # [{"idx","type","price"}]
    events = []   # [{"idx","event","dir","level","confirm_price","value_ratio"}]

    h["value_traded"] = h["Close"] * h["Volume"].fillna(0)
    h["value_ma20"] = h["value_traded"].rolling(MS_VALUE_LOOKBACK).mean()

    for i in range(len(h)):
        idx = h.index[i]

        piv_hi = h["ms_piv_hi"].iloc[i]
        piv_lo = h["ms_piv_lo"].iloc[i]

        if pd.notna(piv_hi):
            swing_type = "HH" if (pd.isna(prev_high) or piv_hi >= prev_high) else "LH"
            swings.append({"idx": idx, "type": swing_type, "price": float(piv_hi)})
            prev_high = float(piv_hi)
            high_active = True

        if pd.notna(piv_lo):
            swing_type = "HL" if (pd.isna(prev_low) or piv_lo >= prev_low) else "LL"
            swings.append({"idx": idx, "type": swing_type, "price": float(piv_lo)})
            prev_low = float(piv_lo)
            low_active = True

        high_src = h["Close"].iloc[i] if MS_BOS_CONFIRM == "close" else h["High"].iloc[i]
        low_src = h["Close"].iloc[i] if MS_BOS_CONFIRM == "close" else h["Low"].iloc[i]

        val_now = safe_num(h["value_traded"].iloc[i], np.nan)
        val_avg = safe_num(h["value_ma20"].iloc[i], np.nan)
        value_ratio = (val_now / val_avg) if (pd.notna(val_now) and pd.notna(val_avg) and val_avg > 0) else np.nan

        if high_active and pd.notna(prev_high) and pd.notna(high_src) and high_src > prev_high:
            event = "Bull CHoCH" if prev_breakout_dir == -1 else "Bull BOS"
            events.append({
                "idx": idx,
                "event": event,
                "dir": 1,
                "level": prev_high,
                "confirm_price": float(high_src),
                "value_ratio": value_ratio,
            })
            high_active = False
            prev_breakout_dir = 1

        if low_active and pd.notna(prev_low) and pd.notna(low_src) and low_src < prev_low:
            event = "Bear CHoCH" if prev_breakout_dir == 1 else "Bear BOS"
            events.append({
                "idx": idx,
                "event": event,
                "dir": -1,
                "level": prev_low,
                "confirm_price": float(low_src),
                "value_ratio": value_ratio,
            })
            low_active = False
            prev_breakout_dir = -1

    swings = sorted(swings, key=lambda x: x["idx"])
    events = sorted(events, key=lambda x: x["idx"])
    return swings, events

def _market_structure_state(swings, last_event: str):
    if len(swings) < 2:
        if last_event == "Bull CHoCH":
            return "Transition Up"
        if last_event == "Bear CHoCH":
            return "Transition Down"
        return "Range"

    last_high = None
    last_low = None

    for s in reversed(swings):
        if s["type"] in ("HH", "LH") and last_high is None:
            last_high = s["type"]
        if s["type"] in ("HL", "LL") and last_low is None:
            last_low = s["type"]
        if last_high is not None and last_low is not None:
            break

    if last_event == "Bull CHoCH":
        return "Transition Up"
    if last_event == "Bear CHoCH":
        return "Transition Down"

    if last_high == "HH" and last_low == "HL":
        return "HH-HL"
    if last_high == "LH" and last_low == "LL":
        return "LH-LL"
    if last_high == "HH" and last_low == "LL":
        return "HH-LL"
    if last_high == "LH" and last_low == "HL":
        return "LH-HL"

    return "Range"

def _market_structure_volume_regime(hist: pd.DataFrame, last_event_obj: dict, trend_regime: str):
    if hist is None or hist.empty or "Volume" not in hist.columns:
        return "Neutral"

    h = hist.copy()
    h["value_traded"] = h["Close"] * h["Volume"].fillna(0)
    h["value_ma20"] = h["value_traded"].rolling(MS_VALUE_LOOKBACK).mean()

    last_val = safe_num(h["value_traded"].iloc[-1], np.nan)
    avg_val = safe_num(h["value_ma20"].iloc[-1], np.nan)
    ratio = (last_val / avg_val) if (pd.notna(last_val) and pd.notna(avg_val) and avg_val > 0) else np.nan

    if last_event_obj:
        ev_ratio = safe_num(last_event_obj.get("value_ratio"), np.nan)
        ev = str(last_event_obj.get("event", ""))
        if ev in ("Bull BOS", "Bull CHoCH") and pd.notna(ev_ratio) and ev_ratio >= 1.5:
            return "Expansion Up"
        if ev in ("Bear BOS", "Bear CHoCH") and pd.notna(ev_ratio) and ev_ratio >= 1.5:
            return "Expansion Down"

    if pd.notna(ratio):
        if ratio < 0.70:
            return "Dry-Up"
        if ratio >= 1.0 and trend_regime in ("Bullish", "Bullish Weak", "Bearish", "Bearish Weak"):
            return "Supportive"

    return "Neutral"

def compute_market_structure_v1(hist: pd.DataFrame, swing_size: int = MS_SWING_SIZE) -> dict:
    out = {
        "ms_trend_regime": "Sideways",
        "ms_structure_state": "Range",
        "ms_last_event": "None",
        "ms_last_event_date": "",
        "ms_last_swing": "",
        "ms_last_swing_date": "",
        "ms_swing_pattern": "Range",
        "ms_volume_regime": "Neutral",
    }

    if hist is None or hist.empty or len(hist) < MS_MIN_HISTORY:
        return out

    h = hist.copy()
    if not isinstance(h.index, pd.DatetimeIndex):
        try:
            if "Date" in h.columns:
                h.index = pd.to_datetime(h["Date"])
            else:
                h.index = pd.to_datetime(h.index)
        except Exception:
            return out

    h = h.sort_index().copy()
    for c in ["Open", "High", "Low", "Close", "Volume"]:
        if c in h.columns:
            h[c] = pd.to_numeric(h[c], errors="coerce")

    h = h.dropna(subset=["High", "Low", "Close"])
    if len(h) < MS_MIN_HISTORY:
        return out

    swings, events = _classify_market_structure_swings(h, swing_size=swing_size)

    last_event_obj = events[-1] if events else None
    last_event = last_event_obj["event"] if last_event_obj else "None"
    last_event_date = _fmt_ms_date(last_event_obj["idx"]) if last_event_obj else ""

    last_swing_obj = swings[-1] if swings else None
    last_swing = last_swing_obj["type"] if last_swing_obj else ""
    last_swing_date = _fmt_ms_date(last_swing_obj["idx"]) if last_swing_obj else ""

    if len(swings) >= 2:
        swing_pattern = f"{swings[-2]['type']}→{swings[-1]['type']}"
    elif len(swings) == 1:
        swing_pattern = swings[-1]["type"]
    else:
        swing_pattern = "Range"

    structure_state = _market_structure_state(swings, last_event)

    trend_regime = "Sideways"
    if last_event == "Bull BOS":
        trend_regime = "Bullish" if structure_state == "HH-HL" else "Bullish Weak"
    elif last_event == "Bear BOS":
        trend_regime = "Bearish" if structure_state == "LH-LL" else "Bearish Weak"
    elif last_event == "Bull CHoCH":
        trend_regime = "Bullish Weak"
    elif last_event == "Bear CHoCH":
        trend_regime = "Bearish Weak"
    else:
        if structure_state == "HH-HL":
            trend_regime = "Bullish Weak"
        elif structure_state == "LH-LL":
            trend_regime = "Bearish Weak"
        else:
            trend_regime = "Sideways"

    volume_regime = _market_structure_volume_regime(h, last_event_obj, trend_regime)

    out.update({
        "ms_trend_regime": trend_regime,
        "ms_structure_state": structure_state,
        "ms_last_event": last_event,
        "ms_last_event_date": last_event_date,
        "ms_last_swing": last_swing,
        "ms_last_swing_date": last_swing_date,
        "ms_swing_pattern": swing_pattern,
        "ms_volume_regime": volume_regime,
    })
    return out


# =========================
# DYNAMIC DETAIL SCHEMA
# =========================
DETAIL_SCHEMA = [
    ("Stock Info", FILL_GROUP_STOCK, [
        ("ticker",           "Ticker",           8,  None,    "center"),
        ("close",            "Closing Price",    10, "#,##0", "center"),
        ("pct_change",       "Price Change %",   10, "0.00%", "center"),
        ("beta_ihsg",        "Beta (vs IHSG)",   12, "0.00",  "center"),
        ("beta_ihsg_zone",   "Beta (vs IHSG) Zone", 16, None, "center"),
        ("rs_rating",        "RS Rating",        12, "#,##0", "center"),
        ("rs_rating_zone",   "RS Rating Zone",   16, None,    "center"),
        ("ara_arb",          "ARA/ARB",          12, None,    "center"),
        ("today_event",      "Today Event",      28, None,    "left"),
        ("emiten",           "Emiten",           38, None,    "left"),
        ("idx_sector",       "IDX Sector",       20, None,    "left"),
        ("sector",           "Sector",           20, None,    "left"),
        ("industry",         "Industry",         35, None,    "left"),
    ]),

    ("Ownership", FILL_GROUP_OWNER, [
        ("investors", "Investors", 35, None, "left"),
        ("free_float", "Free Float", 10, "0.00", "center"),
        ("hhi", "Classic HHI", 12, "#,##0", "center"),
        ("cr1", "CR1", 8, "0.00", "center"),
        ("cr3", "CR3", 8, "0.00", "center"),
        ("holder", "Total Investor >1%", 16, "#,##0", "center"),
        ("ccs", "CCS", 10, "0.00", "center"),
        ("ownership_type", "Ownership Type", 18, None, "center"),
        ("ccs_category", "CCS Category", 16, None, "center"),
    ]),

    ("Stock Regime", FILL_GROUP_MP, [
        ("mcap",                  "Market Cap",            14, "compact", "center"),
        ("market_cap_category",   "Market Cap Categories", 20, None,     "center"),
        ("liquidity_category",    "Liquidity Categories",  20, None,     "center"),
        ("stock_regime",          "Verdict Weight Profiles",24, None,    "center"),
    ]),

    ("Market Structure", FILL_GROUP_MS, [
        ("ms_trend_regime",    "Trend Regime",    16, None, "center"),
        ("ms_structure_state", "Structure State", 16, None, "center"),
        ("ms_last_event",      "Last Event",      14, None, "center"),
        ("ms_last_event_date", "Last Event Date", 14, None, "center"),
        ("ms_last_swing",      "Last Swing",      12, None, "center"),
        ("ms_last_swing_date", "Last Swing Date", 14, None, "center"),
        ("ms_swing_pattern",   "Swing Pattern",   14, None, "center"),
        ("ms_volume_regime",   "Volume Regime",   16, None, "center"),
    ]),

    ("Liquidity", FILL_GROUP_VOL, [
        ("lot",                    "Lot",                        12, "compact", "center"),
        ("daily_value",            "Value (Approx)",             16, "compact", "center"),
        ("adtr20",                 "Average Value 20 D (Approx)",20, "compact", "center"),
        ("last_vol",               "Volume",                     14, "compact", "center"),
        ("adtv20",                 "Average Volume 20 D",        18, "compact", "center"),
        ("rvol20",                 "RVOL 20 D",                  12, "0.00",   "center"),
        ("rvol20_zone",            "RVOL 20 D Zone",             16, None,     "center"),
        ("rvol20_chg_pct",        "RVOL Change %",              14, "0.00%",  "center"),
        ("rvol20_chg_flag",       "RVOL Change Zone",           18, None,     "center"),
        ("adr_pct",                "ADR %",                      10, "0.00",   "center"),
        ("atr14_pct",              "ATR (14) %",                 12, "0.00",   "center"),
        ("adr_atr_zone",           "ADR & ATR (14) Zone",        18, None,     "center"),
        ("foreign_activity",       "Foreign Flow",               18, None,     "center"),
        ("foreign_net_val",        "Foreign Net Value",          18, "compact","center"),
        ("top_broker_buy",         "Top Broker Buy",             20, None,     "center"),
        ("top_broker_sell",        "Top Broker Sell",            20, None,     "center"),
        ("broker_net_signal",      "Broker Net Signal",          24, None,     "center"),
        ("max_lots",               "Max Position (Lots)",        18, "#,##0",  "center"),
        ("max_position_idr",       "Max Position (IDR)",         18, "compact","center"),
        ("position_size_pct_adtv", "Position % of ADTV",        18, "0.0",    "center"),
    ]),

    ("Market Profile", FILL_GROUP_MP, [
        ("ibh", "IBH", 10, "#,##0", "center"),
        ("ibl", "IBL", 10, "#,##0", "center"),
        ("mp_zone", "vs IBL", 14, None, "center"),
        ("pwh", "PWH", 10, "#,##0", "center"),
        ("pwl", "PWL", 10, "#,##0", "center"),
        ("price_ge_pwl", "vs PWL", 14, None, "center"),
        ("mdh", "MDH", 10, "#,##0", "center"),
        ("mdl", "MDL", 10, "#,##0", "center"),
        ("price_ge_mdl", "vs MDL", 14, None, "center"),
        ("mp_summary", "MP Summary", 34, None, "left"),
    ]),

    ("Current QVWAP (Q2 2026)", FILL_GROUP_Q, [
        ("q_days", "Running Days", 10, "#,##0", "center"),
        ("q_vwap", "VWAP", 10, "#,##0", "center"),
        ("q_m1", "-1 SD", 10, "#,##0", "center"),
        ("q_m2", "-2 SD", 10, "#,##0", "center"),
        ("q_m3", "-3 SD", 10, "#,##0", "center"),
        ("q_sd", "SD Score", 10, "0.00", "center"),
        ("q_delta", "SD Δ 1D", 10, "0.00", "center"),
        ("q_remarks", "VWAP Zone", 38, None, "center"),
        ("q_zone_days", "VWAP Zone Days", 14, "#,##0", "center"),
    ]),

    ("Previous QVWAP (Q1 2026)", FILL_GROUP_PQ, [
        ("pq_days", "Running Days", 10, "#,##0", "center"),
        ("pq_vwap", "VWAP", 10, "#,##0", "center"),
        ("pq_m1", "-1 SD", 10, "#,##0", "center"),
        ("pq_m2", "-2 SD", 10, "#,##0", "center"),
        ("pq_m3", "-3 SD", 10, "#,##0", "center"),
        ("pq_sd", "SD Score", 10, "0.00", "center"),
        ("pq_delta", "SD Δ 1D", 14, "0.00", "center"),
        ("pq_remarks", "VWAP Zone", 38, None, "center"),
        ("pq_zone_days", "VWAP Zone Days", 14, "#,##0", "center"),
    ]),

    ("Previous Year VWAP (2025)", FILL_GROUP_PY, [
        ("py_year", "Running Days", 10, "#,##0", "center"),
        ("py_vwap", "VWAP", 10, "#,##0", "center"),
        ("py_m1", "-1 SD", 10, "#,##0", "center"),
        ("py_m2", "-2 SD", 10, "#,##0", "center"),
        ("py_m3", "-3 SD", 10, "#,##0", "center"),
        ("py_sd", "SD Score", 10, "0.00", "center"),
        ("py_delta", "SD Δ 1D", 14, "0.00", "center"),
        ("py_remarks", "VWAP Zone", 38, None, "center"),
        ("py_zone_days", "VWAP Zone Days", 14, "#,##0", "center"),
    ]),

    ("Moving Average", FILL_GROUP_MA, [
        ("ema25", "EMA 25", 10, "#,##0", "center"),
        ("ema25d", "EMA 25 %diff", 12, "0.00", "center"),
        ("ema25p", "EMA 25 Pos", 12, None, "center"),
        ("ema50", "EMA 50", 10, "#,##0", "center"),
        ("ema50d", "EMA 50 %diff", 12, "0.00", "center"),
        ("ema50p", "EMA 50 Pos", 12, None, "center"),
        ("sma200", "SMA 200", 10, "#,##0", "center"),
        ("sma200d", "SMA 200 %diff", 12, "0.00", "center"),
        ("sma200p", "SMA 200 Pos", 12, None, "center"),
        ("summary_ma", "MA Zone", 28, None, "center"),
    ]),

    ("RSI Momentum", FILL_GROUP_MOM, [
        ("rsi14", "RSI 14", 10, "0.00", "center"),
        ("rsi_delta", "RSI 14 Δ 1D", 12, "0.00", "center"),
        ("rsi_status", "RSI Status", 14, None, "center"),
        ("rsi_ma14", "RSI MA 14", 10, "0.00", "center"),
        ("rsi_pos", "RSI Pos", 10, None, "center"),
        ("cross_status", "RSI Cross", 16, None, "center"),
        ("div_signal", "Divergence Signal", 16, None, "center"),
        ("div_ref1_date", "Divergence Start Date", 18, None, "center"),
        ("div_ref2_date", "Divergence Confirm Date", 18, None, "center"),
    ]),

    ("MACD Momentum", FILL_GROUP_MACD_MOM, [
        ("macd_line",        "MACD Line",       12, "#,##0",  "center"),
        ("macd_signal_line", "Signal Line",     12, "#,##0",  "center"),
        ("macd_hist",        "Histogram (EMA3)", 14, "#,##0", "center"),
        ("macd_position",    "Lines Position",  18, None,     "center"),
        ("macd_wave",        "Wave Pattern",    22, None,     "center"),
        ("macd_cross",       "MACD Cross",      16, None,     "center"),
    ]),


]



# =========================
# MACD 4C SMOOTH ENGINE  (Pine Script: MACD 4C Smooth v6 — 1:1 port)
# fast=12, slow=26, signal=9, histSmooth=3
# =========================
def compute_macd_momentum(hist: pd.DataFrame) -> dict:
    """
    1:1 port of Pine Script 'MACD 4C Smooth v6':
      - EMA fast 12, slow 26, signal EMA 9
      - Histogram smoothed with EMA(3) before colour classification
      - 4-colour bar state based on smoothed hist vs previous bar
      - Entry signal = Blue bar (histNegRise) with MACD & Signal both below zero
    New in upgrade: Histogram Slope, Acceleration, Zero Line Distance,
    Momentum Expansion/Compression flags, Composite MACD Interpretation.
    """
    out = {
        # Core lines
        "macd_line":            np.nan,
        "macd_signal_line":     np.nan,
        "macd_hist":            np.nan,   # smoothed EMA(3)
        # Regime labels
        "macd_position":        "",       # Both Below Zero / Both Above Zero / Mixed
        "macd_4color":          "",       # PosRise / PosFall / NegFall / NegRise(Blue)
        "macd_wave":            "",       # Mountain (Building/Declining) / Valley (Deepening/Recovering)
        "macd_entry":           "",       # ENTRY SIGNAL / …
        "macd_cross":           "",       # Golden Cross / Dead Cross / None
        # New columns (upgrade)
        "macd_regime":          "",       # Positive Rising / Positive Falling / Negative Falling / Negative Recovering
        "macd_acceleration":    np.nan,   # delta of smoothed hist (current - previous)
        "hist_slope":           np.nan,   # 3-bar linear slope of smoothed hist
        "zero_line_dist":       np.nan,   # abs(MACD line) — distance from zero
        "momentum_expansion":   "",       # YES if |hist| increasing
        "momentum_compression": "",       # YES if |hist| decreasing toward zero
        "macd_composite":       "",       # Composite interpretation string
    }

    if hist is None or hist.empty or len(hist) < 35:
        return out

    try:
        close = hist["Close"].astype(float)

        # ── Step 1: Standard MACD (12, 26, 9) ──────────────────────────────
        ema_fast    = close.ewm(span=12, adjust=False).mean()
        ema_slow    = close.ewm(span=26, adjust=False).mean()
        macd_line   = ema_fast - ema_slow
        signal_line = macd_line.ewm(span=9, adjust=False).mean()
        hist_raw    = macd_line - signal_line

        # ── Step 2: Smooth histogram with EMA(3) — Pine: histSmooth > 1 ────
        hist_smooth = hist_raw.ewm(span=3, adjust=False).mean()

        if len(hist_smooth.dropna()) < 2:
            return out

        cur_macd  = safe_num(macd_line.iloc[-1])
        cur_sig   = safe_num(signal_line.iloc[-1])
        cur_hist  = safe_num(hist_smooth.iloc[-1])
        prev_hist = safe_num(hist_smooth.iloc[-2])
        prev_hist2 = safe_num(hist_smooth.iloc[-3]) if len(hist_smooth) >= 3 else np.nan

        if not all(pd.notna(x) for x in [cur_macd, cur_sig, cur_hist, prev_hist]):
            return out

        out["macd_line"]        = cur_macd
        out["macd_signal_line"] = cur_sig
        out["macd_hist"]        = cur_hist

        prev_macd    = safe_num(macd_line.iloc[-2]) if len(macd_line) >= 2 else np.nan
        prev_sig_val = safe_num(signal_line.iloc[-2]) if len(signal_line) >= 2 else np.nan

        # ── Lines vs zero ───────────────────────────────────────────────────
        both_below = cur_macd < 0 and cur_sig < 0
        both_above = cur_macd > 0 and cur_sig > 0
        out["macd_position"] = ("Both Below Zero" if both_below
                                else "Both Above Zero" if both_above
                                else "Mixed")

        # ── 4-Color (Pine exact) ────────────────────────────────────────────
        hist_pos_rise = cur_hist >= 0 and cur_hist >  prev_hist   # Silver
        hist_pos_fall = cur_hist >= 0 and cur_hist <= prev_hist   # Red
        hist_neg_fall = cur_hist <  0 and cur_hist <  prev_hist   # Pink
        hist_neg_rise = cur_hist <  0 and cur_hist >= prev_hist   # BLUE ★

        if hist_pos_rise:   out["macd_4color"] = "PosRise (Silver)"
        elif hist_pos_fall: out["macd_4color"] = "PosFall (Red)"
        elif hist_neg_fall: out["macd_4color"] = "NegFall (Pink)"
        else:               out["macd_4color"] = "NegRise (Blue)"

        # ── Wave pattern ────────────────────────────────────────────────────
        if hist_pos_rise:   out["macd_wave"] = "Mountain (Building)"
        elif hist_pos_fall: out["macd_wave"] = "Mountain (Declining)"
        elif hist_neg_fall: out["macd_wave"] = "Valley (Deepening)"
        else:               out["macd_wave"] = "Valley (Recovering)"

        # ── Regime label (new, matches upgrade spec) ─────────────────────────
        if hist_pos_rise:   out["macd_regime"] = "Positive Rising"
        elif hist_pos_fall: out["macd_regime"] = "Positive Falling"
        elif hist_neg_fall: out["macd_regime"] = "Negative Falling"
        else:               out["macd_regime"] = "Negative Recovering"

        # ── Entry signal ────────────────────────────────────────────────────
        if hist_neg_rise and both_below:
            out["macd_entry"] = "ENTRY SIGNAL"
        elif hist_pos_rise and both_below:
            out["macd_entry"] = "Zero Cross Up (Lines Below)"
        elif hist_pos_rise and both_above:
            out["macd_entry"] = "Bullish Continuation"
        elif hist_pos_fall and both_above:
            out["macd_entry"] = "Watch (Mountain Top)"
        elif hist_neg_fall and both_below:
            out["macd_entry"] = "Wait (Valley Deepening)"
        elif hist_neg_rise and not both_below:
            out["macd_entry"] = "NegRise (Lines Mixed)"
        else:
            out["macd_entry"] = "No Setup"

        # ── Cross detection ─────────────────────────────────────────────────
        if all(pd.notna(x) for x in [prev_macd, prev_sig_val, cur_macd, cur_sig]):
            if prev_macd <= prev_sig_val and cur_macd > cur_sig:
                out["macd_cross"] = "Golden Cross"
            elif prev_macd >= prev_sig_val and cur_macd < cur_sig:
                out["macd_cross"] = "Dead Cross"
            else:
                out["macd_cross"] = "-"
        else:
            out["macd_cross"] = "N/A"

        # ── New upgrade metrics ──────────────────────────────────────────────
        # Acceleration = change in smoothed hist (current bar vs previous)
        if pd.notna(cur_hist) and pd.notna(prev_hist):
            out["macd_acceleration"] = cur_hist - prev_hist

        # Histogram slope = 3-bar linear slope (uses prev2, prev, cur)
        if pd.notna(prev_hist2) and pd.notna(prev_hist) and pd.notna(cur_hist):
            y = [prev_hist2, prev_hist, cur_hist]
            x = [0, 1, 2]
            n = 3
            slope_num = n*sum(xi*yi for xi,yi in zip(x,y)) - sum(x)*sum(y)
            slope_den = n*sum(xi**2 for xi in x) - sum(x)**2
            out["hist_slope"] = slope_num / slope_den if slope_den != 0 else 0.0

        # Zero line distance of MACD line
        if pd.notna(cur_macd):
            out["zero_line_dist"] = abs(cur_macd)

        # Momentum expansion/compression
        if pd.notna(cur_hist) and pd.notna(prev_hist):
            expanding  = abs(cur_hist) > abs(prev_hist)
            compressing = abs(cur_hist) < abs(prev_hist)
            out["momentum_expansion"]   = "YES" if expanding  else "NO"
            out["momentum_compression"] = "YES" if compressing else "NO"

        # ── Composite MACD interpretation ────────────────────────────────────
        hist_accel = out.get("macd_acceleration", 0) or 0
        expanding  = out.get("momentum_expansion", "NO") == "YES"

        if both_above and hist_pos_rise and expanding:
            out["macd_composite"] = "Early Momentum Expansion"
        elif both_above and hist_pos_rise and not expanding:
            out["macd_composite"] = "Late Momentum Expansion"
        elif both_above and hist_pos_fall:
            out["macd_composite"] = "Bullish Compression"
        elif both_below and hist_neg_fall:
            out["macd_composite"] = "Momentum Breakdown"
        elif both_below and hist_neg_rise and not expanding:
            out["macd_composite"] = "Bearish Compression"
        elif both_below and hist_neg_rise and expanding:
            out["macd_composite"] = "Momentum Recovery"
        elif hist_pos_fall and not both_above:
            out["macd_composite"] = "Bearish Compression"
        else:
            out["macd_composite"] = out["macd_regime"]

    except Exception:
        pass

    # ── Backward-compat alias ────────────────────────────────────────────────
    out["macd_cross_status"] = out.get("macd_cross", "N/A")
    return out

def find_pivots_low(series: pd.Series, window=2):
    idxs = []
    vals = series.values
    n = len(series)
    for i in range(window, n - window):
        w = vals[i-window:i+window+1]
        if np.isnan(vals[i]):
            continue
        if vals[i] == np.nanmin(w) and np.sum(w == vals[i]) == 1:
            idxs.append(i)
    return idxs

def find_pivots_high(series: pd.Series, window=2):
    idxs = []
    vals = series.values
    n = len(series)
    for i in range(window, n - window):
        w = vals[i-window:i+window+1]
        if np.isnan(vals[i]):
            continue
        if vals[i] == np.nanmax(w) and np.sum(w == vals[i]) == 1:
            idxs.append(i)
    return idxs

def divergence_signals(hist: pd.DataFrame, rsi_series: pd.Series, lookback=75, swing_window=2,
                       cluster_gap=6, min_separation=4, price_tol=0.0075, rsi_tol=2.0,
                       max_last_swing_age=20):
    """
    Strict lifecycle-cluster divergence logic (IDX-friendly)

    Rules:
    1) Scan recent lookback window (default 75 bars)
    2) Find RSI swing highs / lows
    3) Group nearby pivots into lifecycle clusters
    4) Bearish = last 2 RSI-high clusters where cluster max RSI > 70
    5) Bullish = last 2 RSI-low clusters where cluster min RSI < 30
    6) Representative date:
       - Bearish: highest RSI in cluster
       - Bullish: lowest RSI in cluster
    7) Compare price on those exact dates
    8) Require latest valid cluster to be recent
    """

    out = {
        "div_signal": "None",
        "div_strength": "",
        "div_ref1_date": "None",
        "div_ref1_price": np.nan,
        "div_ref1_rsi": np.nan,
        "div_ref2_date": "None",
        "div_ref2_price": np.nan,
        "div_ref2_rsi": np.nan,
        "div_rsi_event_date": "None",
        "div_price_pattern": "",
        "div_rsi_pattern": "",
        "div_pair_type": "",
        "div_reason": "",
    }

    if hist is None or hist.empty or rsi_series is None or len(hist) < 25:
        return out

    df = hist.copy()

    if not isinstance(df.index, pd.DatetimeIndex):
        try:
            if "Date" in df.columns:
                df.index = pd.to_datetime(df["Date"])
        except Exception:
            pass

    rsi_series = rsi_series.reindex(df.index)

    df = df.tail(lookback).copy()
    rsi_tail = rsi_series.tail(lookback).copy()

    if len(df) < 25:
        return out

    high = df["High"].astype(float)
    low = df["Low"].astype(float)

    def _fmt_idx(idx_pos):
        try:
            return df.index[idx_pos].strftime("%d %b '%y")
        except Exception:
            try:
                return pd.to_datetime(df.index[idx_pos]).strftime("%d %b '%y")
            except Exception:
                return ""

    def _swing_high_positions(series, w=2):
        vals = pd.Series(series).reset_index(drop=True)
        pos = []
        for i in range(w, len(vals) - w):
            c = vals.iloc[i]
            if pd.isna(c):
                continue
            left = vals.iloc[i - w:i]
            right = vals.iloc[i + 1:i + w + 1]
            if c >= left.max() and c >= right.max():
                pos.append(i)
        return pos

    def _swing_low_positions(series, w=2):
        vals = pd.Series(series).reset_index(drop=True)
        pos = []
        for i in range(w, len(vals) - w):
            c = vals.iloc[i]
            if pd.isna(c):
                continue
            left = vals.iloc[i - w:i]
            right = vals.iloc[i + 1:i + w + 1]
            if c <= left.min() and c <= right.min():
                pos.append(i)
        return pos

    def _cluster_positions(positions, max_gap=6):
        if not positions:
            return []
        clusters = [[positions[0]]]
        for p in positions[1:]:
            if p - clusters[-1][-1] <= max_gap:
                clusters[-1].append(p)
            else:
                clusters.append([p])
        return clusters

    def _price_high_relation(a, b):
        tol_abs = abs(a) * price_tol
        if b > a + tol_abs:
            return "Higher High"
        elif b < a - tol_abs:
            return "Lower High"
        return "Equal High"

    def _price_low_relation(a, b):
        tol_abs = abs(a) * price_tol
        if b < a - tol_abs:
            return "Lower Low"
        elif b > a + tol_abs:
            return "Higher Low"
        return "Equal Low"

    def _rsi_relation_high(a, b):
        if b > a + rsi_tol:
            return "Higher High"
        elif b < a - rsi_tol:
            return "Lower High"
        return "Equal High"

    def _rsi_relation_low(a, b):
        if b < a - rsi_tol:
            return "Lower Low"
        elif b > a + rsi_tol:
            return "Higher Low"
        return "Equal Low"

    candidates = []

    # Bearish: last 2 overbought lifecycle clusters
    hi_pos = _swing_high_positions(rsi_tail.values, swing_window)
    hi_clusters = _cluster_positions(hi_pos, cluster_gap)
    bear_valid = []
    for cl in hi_clusters:
        rsi_vals = rsi_tail.iloc[cl]
        if rsi_vals.dropna().empty:
            continue
        if rsi_vals.max() > 70:
            rep = cl[int(np.nanargmax(rsi_vals.values))]
            bear_valid.append((cl, rep))
    if len(bear_valid) >= 2:
        (cl1, i1), (cl2, i2) = bear_valid[-2], bear_valid[-1]
        if (i2 - i1) >= min_separation and (len(df) - 1 - i2) <= max_last_swing_age:
            p1 = safe_num(high.iloc[i1])
            p2 = safe_num(high.iloc[i2])
            r1 = safe_num(rsi_tail.iloc[i1])
            r2 = safe_num(rsi_tail.iloc[i2])
            if all(pd.notna(x) for x in [p1, p2, r1, r2]):
                price_pat = _price_high_relation(p1, p2)
                rsi_pat = _rsi_relation_high(r1, r2)
                signal, strength = None, None
                if price_pat == "Higher High" and rsi_pat == "Lower High":
                    signal, strength = "Bearish", "Strong"
                elif price_pat == "Equal High" and rsi_pat == "Lower High":
                    signal, strength = "Bearish", "Medium"
                elif price_pat == "Higher High" and rsi_pat == "Equal High":
                    signal, strength = "Bearish", "Weak"
                elif price_pat == "Lower High" and rsi_pat == "Higher High":
                    signal, strength = "Bearish", "Hidden"
                if signal:
                    candidates.append({
                        "signal": signal, "strength": strength, "pair_type": "Bearish RSI>70 Clusters",
                        "i1": i1, "i2": i2, "p1": p1, "p2": p2, "r1": r1, "r2": r2,
                        "price_pattern": price_pat, "rsi_pattern": rsi_pat, "age": len(df) - 1 - i2
                    })

    # Bullish: last 2 oversold lifecycle clusters
    lo_pos = _swing_low_positions(rsi_tail.values, swing_window)
    lo_clusters = _cluster_positions(lo_pos, cluster_gap)
    bull_valid = []
    for cl in lo_clusters:
        rsi_vals = rsi_tail.iloc[cl]
        if rsi_vals.dropna().empty:
            continue
        if rsi_vals.min() < 30:
            rep = cl[int(np.nanargmin(rsi_vals.values))]
            bull_valid.append((cl, rep))
    if len(bull_valid) >= 2:
        (cl1, i1), (cl2, i2) = bull_valid[-2], bull_valid[-1]
        if (i2 - i1) >= min_separation and (len(df) - 1 - i2) <= max_last_swing_age:
            p1 = safe_num(low.iloc[i1])
            p2 = safe_num(low.iloc[i2])
            r1 = safe_num(rsi_tail.iloc[i1])
            r2 = safe_num(rsi_tail.iloc[i2])
            if all(pd.notna(x) for x in [p1, p2, r1, r2]):
                price_pat = _price_low_relation(p1, p2)
                rsi_pat = _rsi_relation_low(r1, r2)
                signal, strength = None, None
                if price_pat == "Lower Low" and rsi_pat == "Higher Low":
                    signal, strength = "Bullish", "Strong"
                elif price_pat == "Equal Low" and rsi_pat == "Higher Low":
                    signal, strength = "Bullish", "Medium"
                elif price_pat == "Lower Low" and rsi_pat == "Equal Low":
                    signal, strength = "Bullish", "Weak"
                elif price_pat == "Higher Low" and rsi_pat == "Lower Low":
                    signal, strength = "Bullish", "Hidden"
                if signal:
                    candidates.append({
                        "signal": signal, "strength": strength, "pair_type": "Bullish RSI<30 Clusters",
                        "i1": i1, "i2": i2, "p1": p1, "p2": p2, "r1": r1, "r2": r2,
                        "price_pattern": price_pat, "rsi_pattern": rsi_pat, "age": len(df) - 1 - i2
                    })

    if not candidates:
        return out

    strength_rank = {"Strong": 4, "Medium": 3, "Weak": 2, "Hidden": 1}
    candidates = sorted(candidates, key=lambda x: (x["age"], -strength_rank.get(x["strength"], 0)))
    chosen = candidates[0]

    out["div_signal"] = chosen["signal"]
    out["div_strength"] = chosen["strength"]
    out["div_ref1_date"] = _fmt_idx(chosen["i1"])
    out["div_ref1_price"] = chosen["p1"]
    out["div_ref1_rsi"] = chosen["r1"]
    out["div_ref2_date"] = _fmt_idx(chosen["i2"])
    out["div_ref2_price"] = chosen["p2"]
    out["div_ref2_rsi"] = chosen["r2"]
    out["div_rsi_event_date"] = _fmt_idx(chosen["i2"])
    out["div_price_pattern"] = chosen["price_pattern"]
    out["div_rsi_pattern"] = chosen["rsi_pattern"]
    out["div_pair_type"] = chosen["pair_type"]
    out["div_reason"] = f'{chosen["price_pattern"]} + RSI {chosen["rsi_pattern"]} = {chosen["signal"]} {chosen["strength"]}'
    return out

def valuation_band(close, mean_val, std_val):
    if pd.isna(close) or pd.isna(mean_val) or pd.isna(std_val):
        return {"curr": np.nan, "p1": np.nan, "p2": np.nan, "m1": np.nan, "m2": np.nan, "zone": "N/A"}
    p1 = mean_val + std_val
    p2 = mean_val + (2 * std_val)
    m1 = mean_val - std_val
    m2 = mean_val - (2 * std_val)
    return {
        "curr": close,
        "p1": p1,
        "p2": p2,
        "m1": m1,
        "m2": m2,
        "zone": vwap_zone_2pct(close, [mean_val, m1, m2])
    }

# =========================
# PRODUCTION SAFETY COMPATIBILITY LAYER
# =========================

# Aesthetic group color overrides for Detail Sheet readability
FILL_GROUP_STOCK = PatternFill("solid", fgColor="2A4A66")   # navy
FILL_GROUP_MP    = PatternFill("solid", fgColor="21467C")   # denim
FILL_GROUP_MS    = PatternFill("solid", fgColor="5B3F8C")   # market structure (deep violet)
FILL_GROUP_OWNER = PatternFill("solid", fgColor="237A5C")   # forest green
FILL_GROUP_Q     = PatternFill("solid", fgColor="2369A0")   # deep blue
FILL_GROUP_PQ    = PatternFill("solid", fgColor="7A1A1A")   # dark crimson
FILL_GROUP_PY    = PatternFill("solid", fgColor="32588E")   # cobalt
FILL_GROUP_VOL   = PatternFill("solid", fgColor="216B6B")   # teal
FILL_GROUP_MA    = PatternFill("solid", fgColor="5C1530")   # dark plum
FILL_GROUP_MOM   = PatternFill("solid", fgColor="207878")   # sea green
FILL_GROUP_VAL   = PatternFill("solid", fgColor="45337C")   # indigo
FILL_GROUP_MSF    = PatternFill("solid", fgColor="7C2044")   # claret
FILL_GROUP_MACD_MOM = PatternFill("solid", fgColor="4A1060")  # deep violet-purple for MACD Momentum group
FILL_HEADER      = PatternFill("solid", fgColor="2A4A66")   # navy header
# Guarantee aliases / fallbacks for patched sections
if "FONT_NOTE" not in globals():
    FONT_NOTE = FONT_SUBTITLE if "FONT_SUBTITLE" in globals() else FONT_BODY
if "FILL_GROUP_FLOW" not in globals():
    FILL_GROUP_FLOW = FILL_GROUP_MOM if "FILL_GROUP_MOM" in globals() else None
if "FILL_HEADER" not in globals():
    FILL_HEADER = FILL_GROUP_MOM if "FILL_GROUP_MOM" in globals() else None
if "FONT_HEADER" not in globals():
    FONT_HEADER = FONT_SUBTITLE if "FONT_SUBTITLE" in globals() else FONT_BODY

# =========================
# SOURCE LOAD
# =========================
def find_ksei_file():
    if not os.path.isdir(RAW_DIR):
        raise FileNotFoundError("Raw folder not found.")
    candidates = []
    for fn in os.listdir(RAW_DIR):
        if "ksei" in fn.lower() and (fn.lower().endswith(".xlsx") or fn.lower().endswith(".csv")):
            candidates.append(os.path.join(RAW_DIR, fn))
    if not candidates:
        for fn in os.listdir(RAW_DIR):
            if fn.lower().endswith(".xlsx") or fn.lower().endswith(".csv"):
                candidates.append(os.path.join(RAW_DIR, fn))
    if not candidates:
        raise FileNotFoundError("No KSEI source file found in Raw folder.")
    candidates.sort(key=lambda p: os.path.getmtime(p), reverse=True)
    return candidates[0]

def load_ksei():
    path = find_ksei_file()
    df = pd.read_csv(path) if path.lower().endswith(".csv") else pd.read_excel(path)
    cols_map = {str(c).strip().lower(): c for c in df.columns}

    def pick(*names):
        for n in names:
            if n.lower() in cols_map:
                return cols_map[n.lower()]
        return None

    col_ticker = pick("Ticker", "Kode", "kode")
    if not col_ticker:
        raise ValueError("Ticker/Kode column not found in KSEI file.")

    out = pd.DataFrame()
    out["Ticker"] = df[col_ticker].astype(str).str.upper().str.strip()

    mappings = {
        "Emiten": pick("Emiten"),
        "idx_sector": pick("IDX Sector", "IDXSector", "IDX_Sector"),
        "idx_sector_weight": pick("IDX Sector Weight", "IDXSectorWeight", "IDX_Sector_Weight", "Weight"),
        "Sector": pick("Sector", "Sektor"),
        "Industry": pick("Industry", "Industri"),
        "Investors": pick("Investors"),
        "Free Float": pick("Free Float"),
        "Classic HHI": pick("Classic HHI"),
        "CR1": pick("Concentration Ratio Top 1 (CR1)", "CR1"),
        "CR3": pick("Concentration Ratio Top 3 (CR3)", "CR3"),
        "Holder": pick("Total Investor >1%", "Holder"),
        "CCS": pick("Composite Concentration Score (CCS)", "CCS"),
        "Ownership Type": pick("Ownership Type"),
        "CCS Category": pick("CCS Category"),
        "Shares Outstanding": pick("Shares Outstanding", "SharesOutstanding", "Listed Shares", "Shares", "Shares Out", "Jumlah Saham Beredar"),
        "PE TTM": pick("Current PE Ratio (TTM)", "PE TTM", "PE Ratio TTM", "Current PE"),
        "PE Mean": pick("Mean PE Standard Deviation", "PE Mean", "Mean PE"),
        "PE +1": pick("+1 PE Standard Deviation", "PE +1"),
        "PE +2": pick("+2 PE Standard Deviation", "PE +2"),
        "PE -1": pick("-1 PE Standard Deviation", "PE -1"),
        "PE -2": pick("-2 PE Standard Deviation", "PE -2"),
        "PBV Current": pick("Current Price to Book Value", "Current PBV", "PBV Current"),
        "PBV Mean": pick("Mean PBV Standard Deviation", "PBV Mean", "Mean PBV"),
        "PBV +1": pick("+1 PBV Standard Deviation", "PBV +1"),
        "PBV +2": pick("+2 PBV Standard Deviation", "PBV +2"),
        "PBV -1": pick("-1 PBV Standard Deviation", "PBV -1"),
        "PBV -2": pick("-2 PBV Standard Deviation", "PBV -2"),
    }

    for k, src in mappings.items():
        out[k] = df[src] if src else np.nan

    out = out[out["Ticker"].notna() & (out["Ticker"] != "")]
    out = out.drop_duplicates(subset=["Ticker"], keep="first").reset_index(drop=True)
    return out

# =========================
# MARKET DATA
# =========================
def fetch_history_once(symbol: str, params: dict):
    hist = yf.download(symbol, auto_adjust=False, progress=False, threads=False, **params)
    return normalize_history(hist)

def fetch_history_with_retry(ticker: str):
    symbol = f"{ticker}.JK"
    attempts = [
        {"period": "2y", "interval": "1d"},
        {"period": "3y", "interval": "1d"},
        {"start": "2020-01-01", "interval": "1d"},
    ]
    retry_count = 0
    last_err = None
    for params in attempts:
        for i in range(FETCH_RETRIES):
            try:
                hist = fetch_history_once(symbol, params)
                if hist is not None and not hist.empty:
                    return hist, retry_count, "Yahoo", ""
                retry_count += 1
                time.sleep(RETRY_SLEEP_SEC * (i + 1))
            except Exception as e:
                last_err = f"{type(e).__name__}: {e}"
                retry_count += 1
                time.sleep(RETRY_SLEEP_SEC * (i + 1))
    reason = "Empty dataset after retries"
    if last_err:
        reason = f"Fetch error after retries | {last_err}"
    return None, retry_count, "None", reason

def fetch_shares_outstanding(ticker: str, cache: dict) -> Optional[float]:
    ticker_u = str(ticker).upper()
    cached = cache.get(ticker_u)
    try:
        if cached is not None and pd.notna(cached) and float(cached) > 0:
            return float(cached)
    except Exception:
        pass
    symbol = f"{ticker}.JK"
    try:
        tk = yf.Ticker(symbol)
        shares = np.nan
        try:
            shares = (tk.fast_info or {}).get("shares", np.nan)
        except Exception:
            shares = np.nan
        if pd.isna(shares):
            try:
                shares = (tk.info or {}).get("sharesOutstanding", np.nan)
            except Exception:
                shares = np.nan
        if pd.notna(shares) and float(shares) > 0:
            cache[ticker_u] = float(shares)
            return float(shares)
    except Exception:
        pass
    return np.nan

# =========================
# STOCK REGIME
# =========================
def classify_market_cap_category(market_cap: float) -> str:
    if pd.isna(market_cap):
        return "N/A"
    if market_cap > 10_000_000_000_000:
        return "Large Cap"
    if market_cap >= 1_000_000_000_000:
        return "Mid Cap"
    if market_cap >= 100_000_000_000:
        return "Small Cap"
    return "Micro Cap"


def classify_liquidity_category(avg_daily_value: float) -> str:
    if pd.isna(avg_daily_value):
        return "N/A"
    if avg_daily_value > 50_000_000_000:
        return "High Liquidity"
    if avg_daily_value >= 10_000_000_000:
        return "Medium Liquidity"
    if avg_daily_value >= 1_000_000_000:
        return "Low Liquidity"
    return "Very Low Liquidity"



# =========================
# STOCK REGIME / VERDICT WEIGHT PROFILE
# =========================
def classify_stock_regime(row: Dict[str, Any]) -> Dict[str, str]:
    """
    Four institutional categories (Upgrade Spec Item 7):
    1. Blue Chip / High Liquidity   = Large Cap + High Liquidity
    2. Mid Cap / Moderate           = Mid Cap + Medium Liquidity
    3. Low Liquidity / Small Cap    = Small/Micro Cap + Low Liquidity
    4. Institutional Driven         = High sponsorship + accumulation +
                                       strong turnover velocity
                                       → OVERRIDES market cap if conditions met

    Institutional Driven threshold rules:
    - volume_sponsorship_score  >= 65  (high volume sponsorship)
    - accumulation_score        >= 60  (positive accumulation score)
    - turnover_velocity_20d     >= 0.05 (healthy turnover velocity ≥ 0.05% daily)
    - smart_money_bias          in (Bullish, Strongly Bullish)
    - flow_conviction           not in (Negative, Strongly Negative)
    """
    market_cap      = safe_num(row.get("market_cap"))
    avg_daily_value = safe_num(row.get("avg_value_30d"))
    classic_hhi     = safe_num(row.get("classic_hhi"))
    cr1             = safe_num(row.get("cr1"))
    ccs             = safe_num(row.get("ccs"))

    # Sponsorship / flow override fields (populated by compute_institutional_metrics)
    vol_spon   = safe_num(row.get("volume_sponsorship_score"))
    accum      = safe_num(row.get("accumulation_score"))
    turnover   = safe_num(row.get("turnover_velocity_20d"))
    sm_bias    = str(row.get("smart_money_bias", "") or "")
    flow_conv  = str(row.get("flow_conviction", "") or "")

    mc_cat  = classify_market_cap_category(market_cap)
    liq_cat = classify_liquidity_category(avg_daily_value)

    # ── Institutional Driven override check ─────────────────────────────────
    is_inst_driven = (
        (pd.notna(vol_spon)  and vol_spon  >= 65) and
        (pd.notna(accum)     and accum     >= 60) and
        (pd.notna(turnover)  and turnover  >= 0.05) and
        sm_bias   in ("Bullish", "Strongly Bullish") and
        flow_conv not in ("Negative", "Strongly Negative")
    )
    # Fallback: KSEI-derived HHI / CR1 / CCS signals if flow data absent
    ksei_inst = (
        (pd.notna(classic_hhi) and classic_hhi >= 1500) or
        (pd.notna(cr1)         and cr1         >= 0.30) or
        (pd.notna(ccs)         and ccs         >= 70)
    )

    if is_inst_driven or (ksei_inst and liq_cat in ("High Liquidity", "Medium Liquidity")):
        regime = "Institutional Driven"
    elif mc_cat == "Large Cap" and liq_cat in ("High Liquidity",):
        regime = "Blue Chip / High Liquidity"
    elif mc_cat == "Mid Cap":
        regime = "Mid Cap / Moderate"
    elif mc_cat in ("Small Cap", "Micro Cap") and liq_cat in ("Low Liquidity", "Very Low Liquidity"):
        regime = "Low Liquidity / Small Cap"
    else:
        # Catch-all: base on market cap category
        if mc_cat == "Large Cap":
            regime = "Blue Chip / High Liquidity"
        elif mc_cat == "Mid Cap":
            regime = "Mid Cap / Moderate"
        else:
            regime = "Low Liquidity / Small Cap"

    return {
        "market_cap_category": mc_cat,
        "liquidity_category":  liq_cat,
        "stock_regime":        regime,
    }

# =========================
# ROW BUILDERS
# =========================
def base_row_from_ksei(ksei_row: pd.Series):
    return {
        "ticker": ksei_row["Ticker"],
        "close": np.nan,
        "last_high": np.nan, "last_low": np.nan, "last_open": np.nan,
        "pct_change": np.nan,
        "beta_ihsg":  np.nan,
        "beta_ihsg_zone": "-",
        "rs_rating":      np.nan,
        "rs_rating_zone": "-",
        "today_event":    "-",
        # Upgrade 1 — ARA/ARB
        "ara_arb":    "-",
        "at_limit":   False,
        # Upgrade 3 — Board classification
        "board":      "Main Board",
        # Upgrade 5 — Suspension
        "suspended":           False,
        "suspension_label":    "-",
        "corp_action_active":  "-",
        # Upgrade 2 — Foreign flow
        "foreign_net_lot":  np.nan,
        "foreign_net_val":  np.nan,
        "foreign_activity": "-",
        # Upgrade 4 — Broker flow
        "top_broker_buy":    "-",
        "top_broker_sell":   "-",
        "broker_net_signal": "-",
        # Upgrade 6 — Position sizing
        "max_shares":            np.nan,
        "max_lots":              np.nan,
        "max_position_idr":      np.nan,
        "position_size_pct_adtv": np.nan,
        # Upgrade 7 — Seasonal bias
        "seasonal_bias":  "-",
        # Upgrade 8 — Dividend trap (populated in fundamental row)
        "div_trap_flag":  "-",
        # Upgrade 9 — OB Age
        "smc_ob_bull_age": np.nan,
        "smc_ob_bear_age": np.nan,
        # Upgrade 10 — RRG quadrant
        "rrg_quadrant": "-",
        "emiten": ksei_row.get("Emiten", ""),
        "idx_sector": ksei_row.get("idx_sector", ""),
        "idx_sector_weight": safe_num(ksei_row.get("idx_sector_weight")),
        "idx_sector_weight": safe_num(ksei_row.get("idx_sector_weight")),
        "sector": ksei_row.get("Sector", ""),
        "industry": ksei_row.get("Industry", ""),
        "mcap": np.nan,
        "market_cap_category": "N/A",
        "liquidity_category": "N/A",
        "stock_regime": "N/A",
        "data_status": "NO DATA",
        "bars": 0,
        "investors": ksei_row.get("Investors", ""),
        "free_float": safe_num(ksei_row.get("Free Float")),
        "hhi": safe_num(ksei_row.get("Classic HHI")),
        "cr1": safe_num(ksei_row.get("CR1")),
        "cr3": safe_num(ksei_row.get("CR3")),
        "holder": safe_int(ksei_row.get("Holder")),
        "ccs": safe_num(ksei_row.get("CCS")),
        "ownership_type": ksei_row.get("Ownership Type", ""),
        "ccs_category": ksei_row.get("CCS Category", ""),
        "q_days": np.nan, "q_vwap": np.nan, "q_m1": np.nan, "q_m2": np.nan, "q_m3": np.nan, "q_sd": np.nan, "q_delta": np.nan, "q_zone": "N/A", "q_remarks": "",
        "pq_days": np.nan, "pq_vwap": np.nan, "pq_m1": np.nan, "pq_m2": np.nan, "pq_m3": np.nan, "pq_sd": np.nan, "pq_delta": np.nan, "pq_zone": "N/A", "pq_remarks": "",
        "py_year": np.nan, "py_vwap": np.nan, "py_m1": np.nan, "py_m2": np.nan, "py_m3": np.nan, "py_sd": np.nan, "py_delta": np.nan, "py_zone": "N/A", "py_remarks": "",
        "ibh": np.nan, "ibl": np.nan, "mp_zone": "N/A", "pwh": np.nan, "pwl": np.nan, "price_ge_pwl": "N/A", "mdh": np.nan, "mdl": np.nan, "price_ge_mdl": "N/A",
        "avg30": np.nan, "last_vol": np.nan, "last_ok": "N/A", "rvol5": np.nan, "rvol_zone": "N/A",
        "lot": np.nan, "daily_value": np.nan, "adtv20": np.nan, "adtv20_pct": np.nan, "adtv20_zone": "N/A",
        "adtr20": np.nan, "adtr20_pct": np.nan, "adtr20_zone": "N/A", "rvol20": np.nan, "rvol20_zone": "N/A",
        "rvol20_chg_pct": np.nan, "rvol20_chg_flag": "-",
        "adr_pct": np.nan, "atr14_pct": np.nan, "adr_atr_zone": "N/A",
        "summary_ma": "N/A",
        "ema10": np.nan, "ema10d": np.nan, "ema10p": "N/A",
        "ema20": np.nan, "ema20d": np.nan, "ema20p": "N/A",
        "ema25": np.nan, "ema25d": np.nan, "ema25p": "N/A",
        "ema50": np.nan, "ema50d": np.nan, "ema50p": "N/A",
        "sma200": np.nan, "sma200d": np.nan, "sma200p": "N/A",
        "rsi14": np.nan, "rsi_ge_50": "N/A", "rsi_delta": np.nan, "rsi_status": "N/A",
        "rsi_ma14": np.nan, "golden_cross": "N/A", "dead_cross": "N/A", "cross_status": "-", "macd_cross_status": "-", "q_zone_days": 0, "pq_zone_days": 0, "py_zone_days": 0,
        "adr_pct": np.nan, "atr_pct": np.nan, "vol_range_zone": "N/A",
        "div_signal": "N/A", "div_strength": "None", "div_ref1_date": "", "div_ref1_price": np.nan, "div_ref1_rsi": np.nan, "div_ref2_date": "", "div_ref2_price": np.nan, "div_ref2_rsi": np.nan, "div_price_pattern": "", "div_rsi_pattern": "", "div_reason": "", "div_pair_type": "", "div_rsi_event_date": "", "pe_band_source": "N/A", "pbv_band_source": "N/A",
        "pe_ttm": safe_num(ksei_row.get("PE TTM")), "pe_mean": safe_num(ksei_row.get("PE Mean")), "pe_m1": safe_num(ksei_row.get("PE -1")), "pe_m2": safe_num(ksei_row.get("PE -2")), "pe_zone": "N/A",
        "pbv_curr": safe_num(ksei_row.get("PBV Current")), "pbv_mean": safe_num(ksei_row.get("PBV Mean")), "pbv_m1": safe_num(ksei_row.get("PBV -1")), "pbv_m2": safe_num(ksei_row.get("PBV -2")), "pbv_zone": "N/A",
        "flow_data_mode": "PROXY", "dollar_vol_20d_avg": np.nan, "turnover_velocity_20d": np.nan, "ad_trend": "Flat", "obv_trend": "Flat", "cmf20": np.nan, "volume_sponsorship_score": np.nan, "accumulation_score": np.nan, "distribution_score": np.nan, "smart_money_bias": "Neutral", "flow_conviction": "Neutral", "net_participation_proxy": np.nan, "sponsorship_grade": "N/A",
        "pos_52w_pct": np.nan, "range_compression_20d": np.nan, "vol_compression_20d": np.nan, "base_length": np.nan, "breakout_pressure": np.nan, "breakdown_pressure": np.nan, "_removed_spring": "NO", "_removed_upthrust": "NO", "structure_state": "N/A", "wyckoff_proxy_phase": "Transitional", "markup_readiness": np.nan, "breakdown_risk": np.nan, "cause_quality": "Weak",
        "technical_core_score": np.nan, "flow_score": np.nan, "structure_score": np.nan, "risk_penalty": np.nan, "regime_multiplier": np.nan, "adaptive_composite_score": np.nan, "setup_quality": "N/A", "trap_risk": "N/A", "institutional_verdict": "N/A", "institutional_action_bias": "N/A",
        "institutional_accumulation": "NO", "early_markup_candidate": "NO", "smart_money_pullback": "NO", "breakout_watchlist": "NO", "distribution_warning": "NO", "failed_breakout_risk": "NO", "capital_efficient_trend": "NO", "speculative_momentum": "NO", "preset_summary": "None",
        "tier": "",
        "latest_market_day": None,
    }

def build_row(ksei_row: pd.Series, hist: pd.DataFrame, shares_fallback: float):
    # Historical Safety Layer: clip to asof in backtest mode
    hist = _clip_hist_to_asof(hist)
    row = base_row_from_ksei(ksei_row)
    row["bars"] = len(hist)
    row["latest_market_day"] = hist.index[-1].strftime("%Y-%m-%d")
    row["close"]     = safe_num(hist["Close"].iloc[-1])
    row["last_high"] = safe_num(hist["High"].iloc[-1])  if "High"  in hist.columns else np.nan
    row["last_low"]  = safe_num(hist["Low"].iloc[-1])   if "Low"   in hist.columns else np.nan
    row["last_open"] = safe_num(hist["Open"].iloc[-1])  if "Open"  in hist.columns else np.nan
    if len(hist) >= 2:
        _prev_close = safe_num(hist["Close"].iloc[-2], np.nan)
        # Store as decimal fraction (e.g. 0.0023 for 0.23%) so Excel "0.00%" format
        # multiplies by 100 and displays correctly. safe_pct returns *100 which causes
        # double-multiplication (0.23 → shown as 23%).
        row["pct_change"] = safe_div(row["close"] - _prev_close, _prev_close, np.nan)
    else:
        row["pct_change"] = np.nan

    # Beta vs IHSG: COVARIANCE.P(stock_returns, ihsg_returns) / VAR.P(ihsg_returns)
    row["beta_ihsg"] = compute_beta_ihsg(hist) if len(hist) >= 30 else np.nan
    row["beta_ihsg_zone"] = compute_beta_zone(row["beta_ihsg"])

    # RS Rating (IBD-style 1–99) — universe cache populated across all tickers
    rs = compute_rs_rating(hist)
    row["rs_rating"]      = rs
    row["rs_rating_zone"] = rs_rating_zone(rs)
    # Store raw score for universe re-scoring pass after all rows built
    try:
        closes = hist["Close"].squeeze().dropna()
        n = len(closes)
        def _r(b): return (float(closes.iloc[-1])/float(closes.iloc[-b-1])) if n>b and closes.iloc[-b-1]!=0 else np.nan
        _raw = 2.0*_r(63) + _r(min(126,n-1)) + _r(min(189,n-1)) + _r(min(252,n-1))
        row["_rs_raw"] = _raw
        if len(_RS_UNIVERSE_CACHE) < 10000:
            _RS_UNIVERSE_CACHE.append(float(_raw)) if pd.notna(_raw) else None
    except Exception:
        row["_rs_raw"] = np.nan

    # Resolve ticker string for all upgrade API calls
    _ticker = str(row.get("ticker", "")).upper().strip()

    # ── RVOL Change % and flag ─────────────────────────────────────────────────
    # RVOL Change % = (RVOL 20D today - RVOL 20D yesterday) / RVOL 20D yesterday * 100
    # Requires at least 2 bars of volume history beyond the 20D window
    _rvol_chg = np.nan
    _rvol_chg_flag = "-"
    try:
        if len(hist) >= 22 and "Volume" in hist.columns:
            vol_s = hist["Volume"].astype(float)
            adtv_today = safe_num(vol_s.iloc[-21:-1].mean())   # 20D avg ending yesterday
            adtv_prev  = safe_num(vol_s.iloc[-22:-2].mean())   # 20D avg ending day before
            vol_today  = safe_num(vol_s.iloc[-1])
            vol_prev   = safe_num(vol_s.iloc[-2])
            rvol_today = vol_today / adtv_today if pd.notna(adtv_today) and adtv_today > 0 else np.nan
            rvol_prev  = vol_prev  / adtv_prev  if pd.notna(adtv_prev)  and adtv_prev  > 0 else np.nan
            if pd.notna(rvol_today) and pd.notna(rvol_prev) and rvol_prev > 0:
                _rvol_chg = ((rvol_today - rvol_prev) / rvol_prev) * 100.0
                _rvol_chg_flag = (
                    "RVOL Change Above 50%"        if _rvol_chg >= 50.0 else
                    "RVOL Change within 0 - 50%"   if _rvol_chg >= 0 else
                    "RVOL Change within -50% - 0%" if _rvol_chg >= -50.0 else
                    "RVOL Change Below -50%")
    except Exception:
        pass
    row["rvol20_chg_pct"]  = safe_num(_rvol_chg / 100.0) if pd.notna(_rvol_chg) else np.nan
    row["rvol20_chg_flag"] = _rvol_chg_flag

    # ── Upgrade 1: ARA/ARB ────────────────────────────────────────────────────
    board = fetch_board_classification(_ticker)
    row["board"] = board
    ara = compute_ara_arb(hist, board)
    row["ara_arb"]  = ara["ara_arb"]
    row["at_limit"] = ara["at_limit"]

    # ── Upgrade 2: Foreign flow ───────────────────────────────────────────────
    ff = fetch_foreign_flow(_ticker)
    row["foreign_net_lot"]  = ff["foreign_net_lot"]
    row["foreign_net_val"]  = ff["foreign_net_val"]
    row["foreign_activity"] = ff["foreign_activity"]

    # ── Upgrade 4: Broker flow ────────────────────────────────────────────────
    bf = fetch_broker_flow(_ticker)
    row["top_broker_buy"]    = bf["top_broker_buy"]
    row["top_broker_sell"]   = bf["top_broker_sell"]
    row["broker_net_signal"] = bf["broker_net_signal"]

    # ── Upgrade 5: Suspension flag ────────────────────────────────────────────
    susp = fetch_suspension_flag(_ticker)
    row["suspended"]          = susp["suspended"]
    row["suspension_label"]   = susp["suspension_label"]
    row["corp_action_active"] = susp["corp_action_active"]

    # ── Upgrade 6: Position sizing ────────────────────────────────────────────
    ps = compute_max_position_size(safe_num(row.get("adtr20"), np.nan), safe_num(row.get("close"), np.nan))
    row["max_shares"]             = ps["max_shares"]
    row["max_lots"]               = ps["max_lots"]
    row["max_position_idr"]       = ps["max_position_idr"]
    row["position_size_pct_adtv"] = ps["position_size_pct_adtv"]

    # ── Upgrade 7: Seasonal bias ──────────────────────────────────────────────
    from datetime import date as _date
    row["seasonal_bias"] = get_seasonal_bias(_date.today().month)

    # ── Upgrade 9: OB age ─────────────────────────────────────────────────────
    ob_eq   = safe_num(row.get("smc_ob_equilibrium"), np.nan)
    ob_bear_str = str(row.get("smc_closest_ob_bear", "") or "")
    def _mid_from_range_local(s):
        try:
            pts = s.replace("–","-").split("-")
            if len(pts) == 2:
                return (float(pts[0].replace(",","")) + float(pts[1].replace(",","")))/2
        except Exception:
            pass
        return np.nan
    ob_bear_mid = _mid_from_range_local(ob_bear_str)
    row["smc_ob_bull_age"] = compute_ob_age(hist, ob_eq)
    row["smc_ob_bear_age"] = compute_ob_age(hist, ob_bear_mid)

    # Today Event — SMC structural event on latest candle
    row["today_event"] = compute_today_event(row, hist)

    # ── Upgrade 10: RRG quadrant (populated from cache set by IDX Overview) ──
    sector = str(row.get("idx_sector") or row.get("sector") or "")
    row["rrg_quadrant"] = get_rrg_quadrant_for_sector(sector)

    shares_ksei = safe_num(ksei_row.get("Shares Outstanding"))
    shares_used = shares_ksei if pd.notna(shares_ksei) and shares_ksei > 0 else shares_fallback
    row["mcap"] = row["close"] * shares_used if pd.notna(row["close"]) and pd.notna(shares_used) and shares_used > 0 else np.nan

    # FULL mode
    if len(hist) >= MIN_BARS_FULL:
        row["data_status"] = "OK"

        cq_start = quarter_start(hist.index[-1])
        df_q = hist.loc[hist.index >= cq_start]
        q = anchored_vwap_block(df_q)

        pq_start, pq_end = prev_quarter_range(hist.index[-1])
        df_pq = hist.loc[(hist.index >= pq_start) & (hist.index <= pq_end)]
        pq = anchored_vwap_block(df_pq)

        prev_year = hist.index[-1].year - 1
        df_py = hist.loc[(hist.index.year == prev_year)]
        py = anchored_vwap_block(df_py)

        month_start = pd.Timestamp(hist.index[-1].year, hist.index[-1].month, 1)
        df_month = hist.loc[hist.index >= month_start]
        first2 = df_month.head(2)
        row["ibh"] = safe_num(first2["High"].max()) if not first2.empty else np.nan
        row["ibl"] = safe_num(first2["Low"].min()) if not first2.empty else np.nan
        row["mp_zone"] = zone_label(row["close"], row["ibl"], row["ibh"])
        if pd.notna(row["ibh"]) and pd.notna(row["ibl"]) and abs(float(row["ibh"]) - float(row["ibl"])) < 1e-9:
            row["ibh"], row["ibl"], row["mp_zone"] = np.nan, np.nan, "N/A"

        # Weekly Market Profile additions
        # PWH / PWL = wick high/low of first trading day in last completed week
        # MDH / MDL = wick high/low of first trading day in current running week
        try:
            iso = hist.index.to_series().dt.isocalendar()
            cur_year = int(iso.year.iloc[-1])
            cur_week = int(iso.week.iloc[-1])

            # Current running week
            cur_mask = (iso.year == cur_year) & (iso.week == cur_week)
            df_cur_week = hist.loc[cur_mask]
            if not df_cur_week.empty:
                first_cur = df_cur_week.iloc[0]
                row["mdh"] = safe_num(first_cur["High"])
                row["mdl"] = safe_num(first_cur["Low"])
                if pd.notna(row["mdh"]) and pd.notna(row["mdl"]) and abs(float(row["mdh"]) - float(row["mdl"])) < 1e-9:
                    row["mdh"], row["mdl"], row["price_ge_mdl"] = np.nan, np.nan, "N/A"
                else:
                    row["price_ge_mdl"] = ("At Level" if pd.notna(row["close"]) and pd.notna(row["mdl"]) and abs(float(row["close"]) - float(row["mdl"])) < 1e-9 else ("Above" if pd.notna(row["close"]) and pd.notna(row["mdl"]) and row["close"] > row["mdl"] else ("Below" if pd.notna(row["close"]) and pd.notna(row["mdl"]) else "N/A")))

            # Last completed week
            prior_mask = (iso.year < cur_year) | ((iso.year == cur_year) & (iso.week < cur_week))
            if prior_mask.any():
                prior_idx = hist.index[prior_mask]
                last_prior_dt = prior_idx[-1]
                last_prior_iso = pd.Timestamp(last_prior_dt).isocalendar()
                last_year = int(last_prior_iso.year)
                last_week = int(last_prior_iso.week)

                last_mask = (iso.year == last_year) & (iso.week == last_week)
                df_last_week = hist.loc[last_mask]
                if not df_last_week.empty:
                    row["pwh"] = safe_num(df_last_week["High"].max())
                    row["pwl"] = safe_num(df_last_week["Low"].min())
                    if pd.notna(row["pwh"]) and pd.notna(row["pwl"]) and abs(float(row["pwh"]) - float(row["pwl"])) < 1e-9:
                        row["pwh"], row["pwl"], row["price_ge_pwl"] = np.nan, np.nan, "N/A"
                    else:
                        row["price_ge_pwl"] = ("At Level" if pd.notna(row["close"]) and pd.notna(row["pwl"]) and abs(float(row["close"]) - float(row["pwl"])) < 1e-9 else ("Above" if pd.notna(row["close"]) and pd.notna(row["pwl"]) and row["close"] > row["pwl"] else ("Below" if pd.notna(row["close"]) and pd.notna(row["pwl"]) else "N/A")))
        except Exception:
            pass

        vol_series = hist["Volume"].copy()
        row["avg30"] = safe_num(vol_series.iloc[-31:-1].mean()) if len(vol_series) >= 31 else safe_num(vol_series.tail(30).mean())
        row["last_vol"] = safe_num(vol_series.iloc[-1])
        row["last_ok"] = "YES" if pd.notna(row["last_vol"]) and row["last_vol"] >= VOL_THRESHOLD else "NO"
        avg5 = safe_num(vol_series.iloc[-6:-1].mean()) if len(vol_series) >= 6 else safe_num(vol_series.tail(5).mean())
        row["rvol5"] = row["last_vol"] / avg5 if pd.notna(avg5) and avg5 != 0 else np.nan
        row["rvol_zone"] = "YES" if pd.notna(row["rvol5"]) and row["rvol5"] >= MIN_RVOL else "NO"

        # Liquidity block (institutional)
        row["lot"] = row["last_vol"] / 100 if pd.notna(row["last_vol"]) else np.nan
        # Value (Approx): Typical Price = (Open + High + Low + Close) / 4 × Volume
        row["daily_value"] = (safe_num(hist["Volume"].iloc[-1] * ((hist["Open"].iloc[-1] + hist["High"].iloc[-1] + hist["Low"].iloc[-1] + hist["Close"].iloc[-1]) / 4.0)) if len(hist) and all(col in hist.columns for col in ["Volume","Open","High","Low","Close"]) else np.nan)

        adtv_base = safe_num(vol_series.iloc[-21:-1].mean()) if len(vol_series) >= 21 else safe_num(vol_series.tail(20).mean())
        row["adtv20"] = adtv_base
        row["adtv20_pct"] = ((row["last_vol"] / adtv_base) - 1) * 100 if pd.notna(row["last_vol"]) and pd.notna(adtv_base) and adtv_base != 0 else np.nan
        row["adtv20_zone"] = ("N/A" if pd.isna(row["adtv20_pct"]) else ("Above 3%" if row["adtv20_pct"] >= 3 else ("Below 3%" if row["adtv20_pct"] <= -3 else "Within ±3%")))

        row["rvol20"] = row["last_vol"] / adtv_base if pd.notna(row["last_vol"]) and pd.notna(adtv_base) and adtv_base != 0 else np.nan
        row["rvol20_zone"] = "Above 1.5" if pd.notna(row["rvol20"]) and row["rvol20"] >= 1.5 else ("Below 1.5" if pd.notna(row["rvol20"]) else "N/A")

        # Volatility (ADR / ATR logic)
        ADR_LEN = 14

        if len(hist) >= 2:
            # ADR = SMA(High - Low, Length) using PRIOR COMPLETED bars only (exclude current/latest bar)
            daily_range = (hist["High"] - hist["Low"]).astype(float)

            # TradingView-aligned ADR = SMA(High - Low, Length) INCLUDING current/latest bar
            # ADR% = (ADR / Current Close) * 100
            adr = safe_num(daily_range.tail(ADR_LEN).mean()) if len(daily_range) >= 1 else np.nan
            row["adr_pct"] = safe_num((adr / row["close"]) * 100.0) if pd.notna(adr) and pd.notna(row["close"]) and row["close"] != 0 else np.nan

            # ATR(14) % = current ATR(14) relative to current close
            prev_close = hist["Close"].shift(1)
            tr = pd.concat([
                hist["High"] - hist["Low"],
                (hist["High"] - prev_close).abs(),
                (hist["Low"] - prev_close).abs()
            ], axis=1).max(axis=1)

            atr14 = tr.ewm(alpha=1 / 14, adjust=False).mean()
            row["atr14_pct"] = safe_num((atr14.iloc[-1] / row["close"]) * 100.0) if len(atr14) and pd.notna(row["close"]) and row["close"] != 0 else np.nan

            row["adr_atr_zone"] = "Above 3%" if (
                pd.notna(row["adr_pct"]) and pd.notna(row["atr14_pct"]) and row["adr_pct"] >= 3 and row["atr14_pct"] >= 3
            ) else ("Below 3%" if pd.notna(row["adr_pct"]) and pd.notna(row["atr14_pct"]) else "N/A")

        # ADTR 20D = Average Daily Trading Value (Rupiah) using OHLC4 Typical Price
        if "Open" in hist.columns:
            value_series = hist["Volume"] * ((hist["Open"] + hist["High"] + hist["Low"] + hist["Close"]) / 4.0)
        else:
            value_series = hist["Volume"] * ((hist["High"] + hist["Low"] + hist["Close"]) / 3.0)
        row["adtr20"] = safe_num(value_series.iloc[-21:-1].mean()) if len(value_series) >= 21 else safe_num(value_series.tail(20).mean())
        row["adtr20_pct"] = ((row["daily_value"] / row["adtr20"]) - 1) * 100 if pd.notna(row["daily_value"]) and pd.notna(row["adtr20"]) and row["adtr20"] != 0 else np.nan
        row["adtr20_zone"] = ("N/A" if pd.isna(row["adtr20_pct"]) else ("Above 3%" if row["adtr20_pct"] >= 3 else ("Below 3%" if row["adtr20_pct"] <= -3 else "Within ±3%")))

        row["ema10"] = safe_num(ema(hist["Close"], 10).iloc[-1])
        row["ema20"] = safe_num(ema(hist["Close"], 20).iloc[-1])
        row["ema25"] = safe_num(ema(hist["Close"], 25).iloc[-1])
        row["ema50"] = safe_num(ema(hist["Close"], 50).iloc[-1])
        row["sma200"] = safe_num(hist["Close"].rolling(200).mean().iloc[-1])

        row["ema10d"], row["ema10p"] = pct_diff(row["close"], row["ema10"]), pos_label(row["close"], row["ema10"])
        row["ema20d"], row["ema20p"] = pct_diff(row["close"], row["ema20"]), pos_label(row["close"], row["ema20"])
        row["ema25d"], row["ema25p"] = pct_diff(row["close"], row["ema25"]), pos_label(row["close"], row["ema25"])
        row["ema50d"], row["ema50p"] = pct_diff(row["close"], row["ema50"]), pos_label(row["close"], row["ema50"])
        row["sma200d"], row["sma200p"] = pct_diff(row["close"], row["sma200"]), pos_label(row["close"], row["sma200"])

        ma_positions = [row["ema10p"], row["ema20p"], row["ema25p"], row["ema50p"], row["sma200p"]]
        available_positions = [x for x in ma_positions if x in ("Above", "Below")]
        if len(available_positions) > 0 and all(x == "Above" for x in available_positions):
            row["summary_ma"] = "Above All MA"
        elif len(available_positions) > 0 and all(x == "Below" for x in available_positions):
            row["summary_ma"] = "Below All MA"
        else:
            labels = []
            if row["ema10p"]  == "Above": labels.append("EMA10")
            if row["ema20p"]  == "Above": labels.append("EMA20")
            if row["ema25p"]  == "Above": labels.append("EMA25")
            if row["ema50p"]  == "Above": labels.append("EMA50")
            if row["sma200p"] == "Above": labels.append("SMA200")
            row["summary_ma"] = "Above " + ", ".join(labels) if labels else "Below All MA"

        rsi_series = rsi(hist["Close"], 14)
        row["rsi14"] = safe_num(rsi_series.iloc[-1])
        rsi_prev = safe_num(rsi_series.iloc[-2]) if len(rsi_series) >= 2 else np.nan
        row["rsi_delta"] = row["rsi14"] - rsi_prev if pd.notna(row["rsi14"]) and pd.notna(rsi_prev) else np.nan
        row["rsi_ge_50"] = "YES" if pd.notna(row["rsi14"]) and row["rsi14"] >= 50 else "NO"
        row["rsi_status"] = rsi_status(row["rsi14"])

        rsi_ma_series = rsi_series.rolling(14).mean()
        row["rsi_ma14"] = safe_num(rsi_ma_series.iloc[-1])
        if pd.isna(row["rsi14"]) or pd.isna(row["rsi_ma14"]):
            row["rsi_pos"] = "N/A"
        else:
            row["rsi_pos"] = "Above" if row["rsi14"] >= row["rsi_ma14"] else "Below"
        rsi_ma_prev = safe_num(rsi_ma_series.iloc[-2]) if len(rsi_ma_series) >= 2 else np.nan
        row["golden_cross"] = "YES" if all(pd.notna(x) for x in [rsi_prev, rsi_ma_prev, row["rsi14"], row["rsi_ma14"]]) and rsi_prev <= rsi_ma_prev and row["rsi14"] > row["rsi_ma14"] else "NO"
        row["dead_cross"] = "YES" if all(pd.notna(x) for x in [rsi_prev, rsi_ma_prev, row["rsi14"], row["rsi_ma14"]]) and rsi_prev >= rsi_ma_prev and row["rsi14"] < row["rsi_ma14"] else "NO"
        row["cross_status"] = "Golden" if row["golden_cross"] == "YES" else ("Dead" if row["dead_cross"] == "YES" else "-")

        # MACD Momentum (Boring Jacx setup)
        row.update(compute_macd_momentum(hist))

        prev_close = hist["Close"].shift(1)
        tr = pd.concat([hist["High"] - hist["Low"], (hist["High"] - prev_close).abs(), (hist["Low"] - prev_close).abs()], axis=1).max(axis=1)
        atr14 = safe_num(tr.rolling(14).mean().iloc[-1])
        adr_series = ((hist["High"] - hist["Low"]) / hist["Close"].replace(0, np.nan)) * 100.0

        avg_daily_value = row["avg30"] * row["close"] if pd.notna(row["avg30"]) and pd.notna(row["close"]) else np.nan
        regime = classify_stock_regime({
            "market_cap": row["mcap"],
            "avg_value_30d": avg_daily_value,
            "classic_hhi": row["hhi"],
            "cr1": row["cr1"],
            "ccs": row["ccs"],
        })
        row["market_cap_category"] = regime["market_cap_category"]
        row["liquidity_category"] = regime["liquidity_category"]
        row["stock_regime"] = regime["stock_regime"]

        # Market Structure (Leviathan-style, no score)
        ms = compute_market_structure_v1(hist)
        row.update(ms)

        # Divergence detection must run in FULL mode before display mapping
        if len(hist) >= 15:
            rsi_series2 = rsi(hist["Close"], 14)
            div = divergence_signals(
                hist,
                rsi_series2,
                lookback=120,
                swing_window=2,
                min_separation=4,
                price_tol=0.0075,
                rsi_tol=2.0,
                max_last_swing_age=50
            )
            row.update(div)
            if row.get("div_signal", "None") == "None":
                row["div_ref1_date"] = "None"
                row["div_ref2_date"] = "None"

        row["pe_zone"] = vwap_zone_2pct(row["pe_ttm"], [row["pe_m1"], row["pe_m2"]]) if pd.notna(row["pe_ttm"]) else "N/A"
        pe_fields = [row["pe_ttm"], row["pe_mean"], row["pe_m1"], row["pe_m2"]]
        row["pe_band_source"] = "Source" if all(pd.notna(x) for x in pe_fields) else ("Partial Source" if any(pd.notna(x) for x in pe_fields) else "N/A")
        row["pbv_zone"] = vwap_zone_2pct(row["pbv_curr"], [row["pbv_m1"], row["pbv_m2"]]) if pd.notna(row["pbv_curr"]) else "N/A"
        pbv_fields = [row["pbv_curr"], row["pbv_mean"], row["pbv_m1"], row["pbv_m2"]]
        row["pbv_band_source"] = "Source" if all(pd.notna(x) for x in pbv_fields) else ("Partial Source" if any(pd.notna(x) for x in pbv_fields) else "N/A")

        row["q_days"], row["q_vwap"], row["q_m1"], row["q_m2"], row["q_m3"], row["q_sd"], row["q_delta"] = q["days"], q["vwap"], q["m1"], q["m2"], q["m3"], q["sd_score"], q["sd_delta"]
        row["q_p1"], row["q_p2"], row["q_p3"] = q.get("p1", np.nan), q.get("p2", np.nan), q.get("p3", np.nan)
        row["pq_days"], row["pq_vwap"], row["pq_m1"], row["pq_m2"], row["pq_m3"], row["pq_sd"] = pq["days"], pq["vwap"], pq["m1"], pq["m2"], pq["m3"], pq["sd_score"]
        row["pq_p1"], row["pq_p2"], row["pq_p3"] = pq.get("p1", np.nan), pq.get("p2", np.nan), pq.get("p3", np.nan)
        row["py_year"], row["py_vwap"], row["py_m1"], row["py_m2"], row["py_m3"], row["py_sd"] = py["days"], py["vwap"], py["m1"], py["m2"], py["m3"], py["sd_score"]
        row["py_p1"], row["py_p2"], row["py_p3"] = py.get("p1", np.nan), py.get("p2", np.nan), py.get("p3", np.nan)

        # SD delta / relative logic
        # - Current QVWAP: true live 1D delta within current quarter anchor
        # - Previous QVWAP: relative comparison vs prior quarter end
        # - Previous Year VWAP: relative comparison vs prior year end
        row["pq_delta"] = safe_num(q["sd_score"] - pq["sd_score"]) if pd.notna(q["sd_score"]) and pd.notna(pq["sd_score"]) else np.nan
        row["py_delta"] = safe_num(q["sd_score"] - py["sd_score"]) if pd.notna(q["sd_score"]) and pd.notna(py["sd_score"]) else np.nan

        row["q_zone"] = vwap_zone_2pct(row["close"], [row["q_vwap"], row["q_m1"], row["q_m2"], row["q_m3"]])
        row["pq_zone"] = vwap_zone_2pct(row["close"], [row["pq_vwap"], row["pq_m1"], row["pq_m2"], row["pq_m3"]])
        row["py_zone"] = vwap_zone_2pct(row["close"], [row["py_vwap"], row["py_m1"], row["py_m2"], row["py_m3"]])
        row["q_remarks"] = vwap_near_zone_label(row["close"], row["q_sd"], row["q_vwap"], row["q_m1"], row["q_m2"], row["q_m3"], row["q_p1"], row["q_p2"], row["q_p3"])
        row["pq_remarks"] = vwap_near_zone_label(row["close"], row["pq_sd"], row["pq_vwap"], row["pq_m1"], row["pq_m2"], row["pq_m3"], row["pq_p1"], row["pq_p2"], row["pq_p3"])
        row["py_remarks"] = vwap_near_zone_label(row["close"], row["py_sd"], row["py_vwap"], row["py_m1"], row["py_m2"], row["py_m3"], row["py_p1"], row["py_p2"], row["py_p3"])
        row["q_zone_days"] = _count_consecutive_zone_days(hist, [row["q_vwap"], row["q_m1"], row["q_m2"], row["q_m3"]], row["q_remarks"])
        row["pq_zone_days"] = _count_consecutive_zone_days(hist, [row["pq_vwap"], row["pq_m1"], row["pq_m2"], row["pq_m3"]], row["pq_remarks"])
        row["py_zone_days"] = _count_consecutive_zone_days(hist, [row["py_vwap"], row["py_m1"], row["py_m2"], row["py_m3"]], row["py_remarks"])

        any_zone = any(str(z).strip() not in ("", "N/A", "None") for z in [row["q_zone"], row["pq_zone"], row["py_zone"]])
        above_count = sum(1 for x in [row["ema10p"], row["ema20p"], row["ema25p"], row["ema50p"], row["sma200p"]] if x == "Above")
        if any_zone and pd.notna(row["rvol5"]) and row["rvol5"] >= MIN_RVOL:
            if above_count >= 5:
                row["tier"] = "A"
            elif above_count >= 3:
                row["tier"] = "B"

        row.update(compute_institutional_metrics(hist, row))
        if row.get("div_signal") in ("Bullish", "Bearish"):
            row["divergence_summary"] = f'{row["div_signal"]} - {row.get("div_strength", "")}'.strip(" -")
        else:
            row["divergence_summary"] = "None"
        row.update(compute_alphaflow_proxy_max(hist, row))
        # Wyckoff proxy removed (v2.0)
        row = apply_dashboard_presets(row)

        # =========================
        # DISPLAY NORMALIZATION / N-A RULES (FULL MODE)
        # =========================
        if safe_num(row.get("q_days", 0), 0) <= 0:
            row["q_days"] = "N/A"
            row["q_remarks"] = "N/A"
        if safe_num(row.get("pq_days", 0), 0) <= 0:
            row["pq_days"] = "N/A"
            row["pq_remarks"] = "N/A"
        if safe_num(row.get("py_year", 0), 0) <= 0:
            row["py_year"] = "N/A"
            row["py_remarks"] = "N/A"

        if safe_num(row.get("last_vol", 0), 0) <= 0:
            for k in ["avg30", "last_vol", "last_ok", "rvol5", "rvol_zone", "lot", "daily_value", "adtv20", "adtv20_pct", "adtv20_zone", "adtr20", "adtr20_pct", "adtr20_zone", "rvol20", "rvol20_zone", "adr_pct", "atr14_pct", "adr_atr_zone"]:
                row[k] = "N/A"

        if pd.notna(row.get("ibh")) and pd.notna(row.get("ibl")) and safe_num(row.get("ibh")) == safe_num(row.get("ibl")):
            row["mp_zone"] = "NO"

        ma_vals = [row.get("ema10"), row.get("ema20"), row.get("ema25"), row.get("ema50")]
        if pd.notna(row.get("close")) and all(pd.notna(v) and safe_num(v) == safe_num(row.get("close")) for v in ma_vals):
            for k in ["ema10d","ema10p","ema20d","ema20p","ema25d","ema25p","ema50d","ema50p","sma200d","sma200p"]:
                row[k] = "N/A"
            row["summary_ma"] = "N/A"

        if pd.isna(row.get("rsi14")):
            for k in ["rsi14","rsi_ma14","rsi_ge_50","rsi_delta","rsi_status","cross_status","divergence_summary","div_ref1_date","div_ref2_date"]:
                row[k] = "N/A"

        return row

    # PARTIAL mode
    if len(hist) >= MIN_BARS_PARTIAL:
        row["data_status"] = "PARTIAL DATA"

        vol_series = hist["Volume"].copy()
        row["last_vol"] = safe_num(vol_series.iloc[-1])
        if len(vol_series) >= 6:
            avg5 = safe_num(vol_series.iloc[-6:-1].mean())
            row["rvol5"] = row["last_vol"] / avg5 if pd.notna(avg5) and avg5 != 0 else np.nan
            row["rvol_zone"] = "YES" if pd.notna(row["rvol5"]) and row["rvol5"] >= MIN_RVOL else "NO"

            # Liquidity block (institutional)
            row["lot"] = row["last_vol"] / 100 if pd.notna(row["last_vol"]) else np.nan
            # Value (Approx): Typical Price = (Open + High + Low + Close) / 4 × Volume
            row["daily_value"] = (safe_num(hist["Volume"].iloc[-1] * ((hist["Open"].iloc[-1] + hist["High"].iloc[-1] + hist["Low"].iloc[-1] + hist["Close"].iloc[-1]) / 4.0)) if len(hist) and all(col in hist.columns for col in ["Volume","Open","High","Low","Close"]) else np.nan)

            adtv_base = safe_num(vol_series.iloc[-21:-1].mean()) if len(vol_series) >= 21 else safe_num(vol_series.tail(20).mean())
            row["adtv20"] = adtv_base
            row["adtv20_pct"] = ((row["last_vol"] / adtv_base) - 1) * 100 if pd.notna(row["last_vol"]) and pd.notna(adtv_base) and adtv_base != 0 else np.nan
            row["adtv20_zone"] = ("N/A" if pd.isna(row["adtv20_pct"]) else ("Above 3%" if row["adtv20_pct"] >= 3 else ("Below 3%" if row["adtv20_pct"] <= -3 else "Within ±3%")))

            row["rvol20"] = row["last_vol"] / adtv_base if pd.notna(row["last_vol"]) and pd.notna(adtv_base) and adtv_base != 0 else np.nan
            row["rvol20_zone"] = "Above 1.5" if pd.notna(row["rvol20"]) and row["rvol20"] >= 1.5 else ("Below 1.5" if pd.notna(row["rvol20"]) else "N/A")

            if pd.notna(shares_used) and shares_used > 0:
                turn_series = (vol_series / shares_used) * 100.0
                row["adtr20"] = safe_num(turn_series.iloc[-21:-1].mean()) if len(turn_series) >= 21 else safe_num(turn_series.tail(20).mean())
                row["adtr20_pct"] = (row["last_vol"] / shares_used) * 100.0 if pd.notna(row["last_vol"]) else np.nan
                adtr_dev = (((row["adtr20_pct"] / row["adtr20"]) - 1) * 100 if pd.notna(row["adtr20"]) and row["adtr20"] != 0 and pd.notna(row["adtr20_pct"]) else np.nan)
            row["adtr20_zone"] = ("N/A" if pd.isna(adtr_dev) else ("Above 3%" if adtr_dev >= 3 else ("Below 3%" if adtr_dev <= -3 else "Within ±3%")))
        if len(vol_series) >= 31:
            row["avg30"] = safe_num(vol_series.iloc[-31:-1].mean())
        elif len(vol_series) >= 5:
            row["avg30"] = safe_num(vol_series.mean())

        row["last_ok"] = "YES" if pd.notna(row["last_vol"]) and row["last_vol"] >= VOL_THRESHOLD else ("NO" if pd.notna(row["last_vol"]) else "N/A")

        if len(hist) >= 10:
            row["ema10"] = safe_num(ema(hist["Close"], 10).iloc[-1])
            row["ema10d"], row["ema10p"] = pct_diff(row["close"], row["ema10"]), pos_label(row["close"], row["ema10"])
        if len(hist) >= 20:
            row["ema20"] = safe_num(ema(hist["Close"], 20).iloc[-1])
            row["ema20d"], row["ema20p"] = pct_diff(row["close"], row["ema20"]), pos_label(row["close"], row["ema20"])
        if len(hist) >= 25:
            row["ema25"] = safe_num(ema(hist["Close"], 25).iloc[-1])
            row["ema25d"], row["ema25p"] = pct_diff(row["close"], row["ema25"]), pos_label(row["close"], row["ema25"])
        if len(hist) >= 50:
            row["ema50"] = safe_num(ema(hist["Close"], 50).iloc[-1])
            row["ema50d"], row["ema50p"] = pct_diff(row["close"], row["ema50"]), pos_label(row["close"], row["ema50"])
        if len(hist) >= 200:
            row["sma200"] = safe_num(hist["Close"].rolling(200).mean().iloc[-1])
            row["sma200d"], row["sma200p"] = pct_diff(row["close"], row["sma200"]), pos_label(row["close"], row["sma200"])

        available_ma = []
        for label, pos in [("EMA10", row["ema10p"]), ("EMA20", row["ema20p"]), ("EMA25", row["ema25p"]), ("EMA50", row["ema50p"]), ("SMA200", row["sma200p"])]:
            if pos in ("Above", "Below"):
                available_ma.append((label, pos))
        if available_ma:
            above = [x[0] for x in available_ma if x[1] == "Above"]
            if len(above) == len(available_ma):
                row["summary_ma"] = "Above All MA"
            elif len(above) == 0:
                row["summary_ma"] = "Below All MA"
            else:
                row["summary_ma"] = "Above " + ", ".join(above)

        if len(hist) >= 15:
            rsi_series = rsi(hist["Close"], 14)
            row["rsi14"] = safe_num(rsi_series.iloc[-1])
            rsi_prev = safe_num(rsi_series.iloc[-2]) if len(rsi_series) >= 2 else np.nan
            row["rsi_delta"] = row["rsi14"] - rsi_prev if pd.notna(row["rsi14"]) and pd.notna(rsi_prev) else np.nan
            row["rsi_ge_50"] = "YES" if pd.notna(row["rsi14"]) and row["rsi14"] >= 50 else "NO"
            row["rsi_status"] = rsi_status(row["rsi14"])
            if len(hist) >= 28:
                rsi_ma_series = rsi_series.rolling(14).mean()
                row["rsi_ma14"] = safe_num(rsi_ma_series.iloc[-1])
                rsi_ma_prev = safe_num(rsi_ma_series.iloc[-2]) if len(rsi_ma_series) >= 2 else np.nan
                row["golden_cross"] = "YES" if all(pd.notna(x) for x in [rsi_prev, rsi_ma_prev, row["rsi14"], row["rsi_ma14"]]) and rsi_prev <= rsi_ma_prev and row["rsi14"] > row["rsi_ma14"] else "NO"
                row["dead_cross"] = "YES" if all(pd.notna(x) for x in [rsi_prev, rsi_ma_prev, row["rsi14"], row["rsi_ma14"]]) and rsi_prev >= rsi_ma_prev and row["rsi14"] < row["rsi_ma14"] else "NO"
                row["cross_status"] = "Golden" if row["golden_cross"] == "YES" else ("Dead" if row["dead_cross"] == "YES" else "-")

        # MACD Momentum (Boring Jacx setup) — runs if enough bars
        if len(hist) >= 35:
            row.update(compute_macd_momentum(hist))

        if len(hist) >= 14:
            prev_close = hist["Close"].shift(1)
            tr = pd.concat([hist["High"] - hist["Low"], (hist["High"] - prev_close).abs(), (hist["Low"] - prev_close).abs()], axis=1).max(axis=1)
            atr14 = safe_num(tr.rolling(14).mean().iloc[-1])
        if len(hist) >= 20:
            adr_series = ((hist["High"] - hist["Low"]) / hist["Close"].replace(0, np.nan)) * 100.0

        avg_daily_value = row["avg30"] * row["close"] if pd.notna(row["avg30"]) and pd.notna(row["close"]) else np.nan
        regime = classify_stock_regime({
            "market_cap": row["mcap"],
            "avg_value_30d": avg_daily_value,
            "classic_hhi": row["hhi"],
            "cr1": row["cr1"],
            "ccs": row["ccs"],
        })
        row["market_cap_category"] = regime["market_cap_category"]
        row["liquidity_category"] = regime["liquidity_category"]
        row["stock_regime"] = regime["stock_regime"]

        if len(hist) >= 15:
            rsi_series2 = rsi(hist["Close"], 14)
            div = divergence_signals(
                hist,
                rsi_series2,
                lookback=120,
                swing_window=2,
                min_separation=4,
                price_tol=0.0075,
                rsi_tol=2.0,
                max_last_swing_age=50
            )
            row.update(div)

        row["pe_zone"] = vwap_zone_2pct(row["pe_ttm"], [row["pe_m1"], row["pe_m2"]]) if pd.notna(row["pe_ttm"]) else "N/A"
        pe_fields = [row["pe_ttm"], row["pe_mean"], row["pe_m1"], row["pe_m2"]]
        row["pe_band_source"] = "Source" if all(pd.notna(x) for x in pe_fields) else ("Partial Source" if any(pd.notna(x) for x in pe_fields) else "N/A")
        row["pbv_zone"] = vwap_zone_2pct(row["pbv_curr"], [row["pbv_m1"], row["pbv_m2"]]) if pd.notna(row["pbv_curr"]) else "N/A"
        pbv_fields = [row["pbv_curr"], row["pbv_mean"], row["pbv_m1"], row["pbv_m2"]]
        row["pbv_band_source"] = "Source" if all(pd.notna(x) for x in pbv_fields) else ("Partial Source" if any(pd.notna(x) for x in pbv_fields) else "N/A")

        row.update(compute_institutional_metrics(hist, row))
        if row.get("div_signal") in ("Bullish", "Bearish"):
            row["divergence_summary"] = f'{row["div_signal"]} - {row.get("div_strength", "")}'.strip(" -")
        else:
            row["divergence_summary"] = "None"
        row.update(compute_alphaflow_proxy_max(hist, row))
        # Wyckoff proxy removed (v2.0)
        row = apply_dashboard_presets(row)

        # =========================
        # DISPLAY NORMALIZATION / N-A RULES (FULL MODE)
        # =========================
        if safe_num(row.get("q_days", 0), 0) <= 0:
            row["q_days"] = "N/A"
            row["q_remarks"] = "N/A"
        if safe_num(row.get("pq_days", 0), 0) <= 0:
            row["pq_days"] = "N/A"
            row["pq_remarks"] = "N/A"
        if safe_num(row.get("py_year", 0), 0) <= 0:
            row["py_year"] = "N/A"
            row["py_remarks"] = "N/A"

        if safe_num(row.get("last_vol", 0), 0) <= 0:
            for k in ["avg30", "last_vol", "last_ok", "rvol5", "rvol_zone", "lot", "daily_value", "adtv20", "adtv20_pct", "adtv20_zone", "adtr20", "adtr20_pct", "adtr20_zone", "rvol20", "rvol20_zone", "adr_pct", "atr14_pct", "adr_atr_zone"]:
                row[k] = "N/A"

        if pd.notna(row.get("ibh")) and pd.notna(row.get("ibl")) and safe_num(row.get("ibh")) == safe_num(row.get("ibl")):
            row["mp_zone"] = "NO"

        ma_vals = [row.get("ema10"), row.get("ema20"), row.get("ema25"), row.get("ema50")]
        if pd.notna(row.get("close")) and all(pd.notna(v) and safe_num(v) == safe_num(row.get("close")) for v in ma_vals):
            for k in ["ema10d","ema10p","ema20d","ema20p","ema25d","ema25p","ema50d","ema50p","sma200d","sma200p"]:
                row[k] = "N/A"
            row["summary_ma"] = "N/A"

        if pd.isna(row.get("rsi14")):
            for k in ["rsi14","rsi_ma14","rsi_ge_50","rsi_delta","rsi_status","cross_status","divergence_summary","div_ref1_date","div_ref2_date"]:
                row[k] = "N/A"

        return row

    return row

def apply_dashboard_presets(row: dict):
    row["preset_accumulation"] = "YES" if row.get("af_verdict") in ("ACCUMULATION", "STRONG ACCUMULATION") else "NO"
    row["preset_high_conviction"] = "YES" if row.get("af_verdict") in ("ACCUMULATION", "STRONG ACCUMULATION") and safe_num(row.get("af_verdict_confidence"), 0) >= 80 else "NO"
    row["preset_distribution"] = "YES" if row.get("af_verdict") in ("DISTRIBUTION", "STRONG DISTRIBUTION") else "NO"
    row["preset_distribution_risk"] = "YES" if row.get("preset_distribution") == "YES" and (
        str(row.get("wyckoff_proxy_phase", "")) in ("Distribution", "Markdown") or
        safe_num(row.get("breakdown_risk"), 0) >= 70 or
        str(row.get("af_phase_event", "")).startswith("Markdown")
    ) else "NO"
    row["preset_smart_money"] = "YES" if safe_num(row.get("af_smt_proxy"), 0) >= 75 else "NO"
    row["preset_blue_chips"] = "YES" if str(row.get("market_cap_category", "")) == "Large Cap" and str(row.get("liquidity_category", "")) == "High Liquidity" else "NO"
    row["preset_flow_leaders"] = "YES" if (
        str(row.get("af_flow_edge", "")) == "Strong" or
        safe_num(row.get("af_r2_proxy"), 0) >= 12 or
        safe_num(row.get("af_hit_rate_proxy"), 0) >= 60
    ) else "NO"
    active_event = (
        str(row.get("_removed_spring", "NO")) in ("YES", "Weak") or
        str(row.get("_removed_upthrust", "NO")) in ("YES", "Weak") or
        ("Bullish" in str(row.get("divergence_summary", ""))) or
        ("Bearish" in str(row.get("divergence_summary", ""))) or
        str(row.get("af_phase_event", "")).split(" / ")[-1] not in ("No Clean Event", "Continuation", "Base Build")
    )
    row["preset_active_wyckoff_events"] = "YES" if active_event else "NO"
    return row

def _zone_is_tradeable_poi(zone_text: str) -> bool:
    s = str(zone_text or "").strip()
    if not s or s == "N/A":
        return False
    s_u = s.upper()
    return ("PRICE NEAR" in s_u) or ("PRICE RANGING" in s_u)

def _pct_from_level(close_px, level_px):
    c = safe_num(close_px, np.nan)
    lv = safe_num(level_px, np.nan)
    if pd.isna(c) or pd.isna(lv) or lv == 0:
        return np.nan
    return ((c / lv) - 1.0) * 100.0

def _near_level(close_px, level_px, tol_pct):
    p = _pct_from_level(close_px, level_px)
    return pd.notna(p) and abs(p) <= tol_pct

def _ms_is_continuation_ok(r: dict) -> bool:
    ms_regime = str(r.get("ms_trend_regime", "") or "")
    ms_state = str(r.get("ms_structure_state", "") or "")
    last_event = str(r.get("ms_last_event", "") or "")
    if ("Bullish" in ms_regime) and (ms_state in ("HH-HL", "Transition Up") or last_event in ("Bull BOS", "Bull CHoCH")):
        return True
    return False

def _best_vwap_location_label(r: dict) -> str:
    z = _best_discount_zone(r)
    if z:
        if pd.notna(safe_num(z.get("dist_pct"), np.nan)):
            return f"{z['framework']} | {z['zone_type']} ({z['dist_pct']:+.2f}%)"
        return f"{z['framework']} | {z['zone_type']}"
    return "No valid discount POI"

def _buy_zone_type(r: dict) -> str:
    z = _best_discount_zone(r)
    return z["zone_type"] if z else "None"

def _secondary_vwap_context(r: dict) -> str:
    close_px = safe_num(r.get("close"), np.nan)
    hits = []
    checks = [
        ("PQVWAP",   r.get("pq_vwap"), 2.0),
        ("PQ -1 SD", r.get("pq_m1"),   1.25),
        ("PQ -2 SD", r.get("pq_m2"),   1.25),
        ("PYVWAP",   r.get("py_vwap"), 2.0),
        ("PY -1 SD", r.get("py_m1"),   1.25),
        ("PY -2 SD", r.get("py_m2"),   1.25),
    ]
    for name, level, tol in checks:
        if _near_level(close_px, level, tol):
            hits.append(name)
    if hits:
        return " + ".join(hits[:3])

    # fallback to text remarks
    out = []
    for lbl, z in [("PQ", r.get("pq_remarks", "")), ("PY", r.get("py_remarks", ""))]:
        z = str(z or "")
        if z and z != "N/A" and _zone_is_tradeable_poi(z):
            out.append(f"{lbl}: actionable")
    return " | ".join(out) if out else "No major HTF confluence"

def _liquidity_check(r: dict) -> str:
    rvol = safe_num(r.get("rvol20"), np.nan)
    dval = safe_num(r.get("value_traded"), np.nan)
    adtv = safe_num(r.get("adtv20"), np.nan)
    daily_val = safe_num(r.get("daily_value"), np.nan)

    # Try both value_traded and daily_value field names
    eff_dval = dval if pd.notna(dval) else daily_val

    if pd.notna(eff_dval) and eff_dval >= 10_000_000_000:
        return "Strong"
    if pd.notna(adtv) and adtv >= 10_000_000_000:
        return "Pass"
    if pd.notna(rvol) and rvol >= 1.0:
        return "Pass"
    if pd.notna(rvol) and rvol >= 0.7:
        return "Pass"
    return "Thin"

def _is_deep_discount_zone(zone_type: str) -> bool:
    return zone_type in ("Q -2 SD", "PQ -1 SD", "PYVWAP", "PY -1 SD", "Discount Zone")

def _is_obvious_breakdown(r: dict, zone_type: str) -> bool:
    ms_regime = str(r.get("ms_trend_regime", "") or "")
    ms_state = str(r.get("ms_structure_state", "") or "")
    last_event = str(r.get("ms_last_event", "") or "")
    if ms_regime == "Bearish" and ms_state == "LH-LL" and last_event == "Bear BOS":
        return not _is_deep_discount_zone(zone_type)
    return False

def _trade_bias_from_context(r: dict, setup: str) -> str:
    last_event = str(r.get("ms_last_event", "") or "")
    if setup == "Continuation Reclaim":
        return "Reclaim"
    if setup == "Discount Pullback":
        return "POI → PI"
    if setup == "Deep Discount Bounce":
        return "POI → POI"
    if last_event in ("Bull CHoCH", "Bull BOS"):
        return "POI → PI"
    return "POI → POI"


# =========================
# VWAP SCREENER
# =========================

def _score_ma_alignment(r: dict) -> tuple:
    """
    Returns (ma_tier, ma_label)
      4 = Above EMA20, EMA25, EMA50  → full bull alignment
      3 = Above EMA20, EMA25         → constructive
      2 = Above EMA20 only           → short-term only
      1 = Above EMA50 only           → secular floor, MAs broken
      0 = Below all                  → excluded (unless deep discount exception)
    """
    c     = safe_num(r.get("close"), np.nan)
    e20   = safe_num(r.get("ema20"), np.nan)
    e25   = safe_num(r.get("ema25"), np.nan)
    e50   = safe_num(r.get("ema50"), np.nan)
    s200  = safe_num(r.get("sma200"), np.nan)
    if pd.isna(c):
        return (0, "N/A")
    above_e20  = pd.notna(e20)  and c >= e20
    above_e25  = pd.notna(e25)  and c >= e25
    above_e50  = pd.notna(e50)  and c >= e50
    above_s200 = pd.notna(s200) and c >= s200
    if above_e20 and above_e25 and above_e50 and above_s200:
        return (4, "EMA20 ▲ EMA25 ▲ EMA50 ▲ SMA200")
    if above_e20 and above_e25 and above_e50:
        return (3, "EMA20 ▲ EMA25 ▲ EMA50 ▲")
    if above_e20 and above_e25:
        return (2, "EMA20 ▲ EMA25 ▲")
    if above_e20:
        return (1, "EMA20 ▲ only")
    return (0, "Below all MAs")


def _reversal_score(r: dict) -> tuple:
    """
    Score 0-8 reversal signals. Returns (score, signals_list).
    Used both to gate EXCEPTION entries and to enrich PASS entries.
    """
    signals = []
    rsi14      = safe_num(r.get("rsi14"), np.nan)
    rsi_div    = str(r.get("div_signal", "") or "")
    macd_wave  = str(r.get("macd_wave", "") or "")
    spring     = str(r.get("_removed_spring", "NO") or "NO")
    rvol       = safe_num(r.get("rvol20"), np.nan)
    last_event = str(r.get("ms_last_event", "") or "")
    cross_rsi  = str(r.get("cross_status", "") or "")
    cross_macd = str(r.get("macd_cross", r.get("macd_cross_status", "")) or "")
    ms_state   = str(r.get("ms_structure_state", "") or "")

    if "Bullish" in rsi_div:          signals.append("Bullish Div")
    if "Recovering" in macd_wave:     signals.append("MACD↑")
    if spring in ("YES", "Weak"):     signals.append("Spring")
    if pd.notna(rsi14) and rsi14 <= 35:   signals.append(f"RSI{rsi14:.0f}")
    if pd.notna(rvol) and rvol >= 1.5:    signals.append(f"RVOL{rvol:.1f}x")
    if last_event in ("Bull CHoCH", "Bull BOS"):  signals.append(last_event)
    if cross_rsi == "Golden":         signals.append("RSI✕")
    if cross_macd in ("Golden", "Golden Cross"):  signals.append("MACD✕")
    return (len(signals), signals)


# =========================
# SWING SCORE ENGINE
# =========================

def _swing_score(r: dict, zone_type: str, tgt_dist: float) -> tuple:
    """
    TRUE POI ENGINE V17 — 7-component weighted composite score (0.0–10.0).
    Returns (composite_score, priority_label).

    Weights:
      Structure Quality + Freshness : 20%
      VWAP Position                 : 20%
      MACD Momentum                 : 20%
      RSI Momentum                  : 15%
      Candle Pattern                : 15%
      Liquidity                     : 10%
      Mean Reversion / Extremity    :  0% (normalized into final composite)
    """

    # ─── COMPONENT 1 – Structure Quality + Freshness (0–10) → weight 20% ────
    ms_regime    = str(r.get("ms_trend_regime", "") or "")
    ms_state     = str(r.get("ms_structure_state", "") or "")
    last_event   = str(r.get("ms_last_event", "") or "")
    event_age    = safe_num(r.get("ms_event_age_d"), np.nan)
    ms_quality   = str(r.get("ms_structure_quality", "") or "")

    struct_score = 0.0
    if "Bullish" in ms_regime and ms_state in ("HH-HL", "HH/HL"):
        struct_score += 5.0
    elif "Bullish" in ms_regime:
        struct_score += 3.5
    elif ms_regime in ("Neutral", "Sideways", "Range"):
        struct_score += 1.5
    # Event freshness
    if pd.notna(event_age):
        if event_age <= 5:
            struct_score += 5.0
        elif event_age <= 10:
            struct_score += 3.5
        elif event_age <= 20:
            struct_score += 2.0
    if ms_quality == "Clean":
        struct_score = min(10.0, struct_score + 0.5)
    struct_score = min(10.0, struct_score)

    # ─── COMPONENT 2 – VWAP Position (0–10) → weight 20% ────────────────────
    vwap_zone   = str(r.get("cq_vwap_zone", r.get("vwap_zone", "")) or "")
    vwap_days   = safe_num(r.get("cq_vwap_zone_days", r.get("vwap_zone_days")), np.nan)
    pq_rel_sd   = safe_num(r.get("pq_rel_sd_end", r.get("pq_sd_delta")), np.nan)
    py_rel_sd   = safe_num(r.get("py_rel_sd_end", r.get("py_sd_delta")), np.nan)
    py_zone     = str(r.get("py_vwap_zone", "") or "")

    _vwap_base = {
        "Price Above VWAP":             10,
        "Price Ranging VWAP⇄-1 SD":     8,
        "Price Ranging -1 SD⇄-2 SD":    6,
        "Price Near -1 SD":             5,
        "Price Near -2 SD":             4,
        "Price Near VWAP":              5,
        "At VWAP":                      5,
        "Price Ranging VWAP⇄+1 SD":    3,
        "Price Below -1 SD":            1,
        "Price Below -2 SD":            0,
        "Price Below -3 SD":            0,
    }
    vwap_score = 0.0
    for k, v in _vwap_base.items():
        if k.lower() in vwap_zone.lower():
            vwap_score = float(v)
            break
    else:
        # zone_type fallback for mapped POI zones
        _zone_fallback = {
            "PQ -2 SD": 4, "PY -2 SD": 4, "PQ -1 SD": 6,
            "PY -1 SD": 6, "PQVWAP": 8, "PYVWAP": 8,
        }
        vwap_score = float(_zone_fallback.get(zone_type, 5))

    # PQ bonus/penalty
    if pd.notna(pq_rel_sd):
        if pq_rel_sd > 0:
            vwap_score = min(10.0, vwap_score + 2.0)
        elif pq_rel_sd < 0:
            vwap_score = max(0.0, vwap_score - 2.0)
    # PY bonus/penalty
    if pd.notna(py_rel_sd):
        if py_rel_sd > 0 and "Above" in py_zone:
            vwap_score = min(10.0, vwap_score + 1.0)
        elif py_rel_sd < 0 and "Below" in py_zone:
            vwap_score = max(0.0, vwap_score - 1.0)
    # Zone age decay
    if pd.notna(vwap_days) and vwap_days > 10:
        vwap_score *= 0.7
    vwap_score = min(10.0, max(0.0, round(vwap_score, 1)))

    # ─── COMPONENT 3 – MACD Momentum (0–10) → weight 20% ────────────────────
    macd_wave    = str(r.get("macd_wave", r.get("macd_wave_pattern", "")) or "")
    macd_lines   = str(r.get("macd_lines_position", r.get("macd_line_position", "")) or "")

    _macd_reject = (
        "Mountain (Declining)" in macd_wave and "Both Below Zero" in macd_lines
    )
    macd_score = 0.0
    if "Mountain (Building)" in macd_wave:
        macd_score = 10.0 if "Both Above Zero" in macd_lines else 6.0
    elif "Valley (Recovering)" in macd_wave:
        macd_score = 9.0 if "Both Below Zero" not in macd_lines else 5.0
    elif "Mountain (Declining)" in macd_wave and "Both Above Zero" in macd_lines:
        macd_score = 5.0
    elif "Valley (Deepening)" in macd_wave:
        macd_score = 2.0
    elif "Both Below Zero" in macd_lines:
        macd_score = 4.0
    if "Mixed" in macd_lines:
        macd_score *= 0.6

    # ─── COMPONENT 4 – RSI Momentum (0–10) → weight 15% ─────────────────────
    rsi_status  = str(r.get("rsi_status", "") or "")
    rsi_pos     = str(r.get("rsi_pos", r.get("rsi_position", "")) or "")
    rsi_div     = str(r.get("div_signal", r.get("divergence_signal", "")) or "")

    _rsi_base = {
        ("Strong",    "Above"):  10,
        ("Neutral",   "Above"):   7,
        ("Strong",    "Below"):   6,
        ("Overbought", ""):        4,
        ("Weak",       ""):        2,
        ("Oversold",   ""):        3,
    }
    rsi_score = 0.0
    for (st, pos), pts in _rsi_base.items():
        if rsi_status == st and (pos == "" or pos in rsi_pos):
            rsi_score = float(pts)
            break
    else:
        rsi_score = 3.0  # neutral default
    # Divergence modifier
    if "Bullish" in rsi_div:
        rsi_score = min(10.0, rsi_score + 2.0)
    elif "Bearish" in rsi_div:
        rsi_score = max(0.0, rsi_score - 2.0)

    # ─── COMPONENT 5 – Candle Pattern (0–10) → weight 15% ───────────────────
    candle_pat    = str(r.get("cp_pattern", r.get("last_candle_pattern", "")) or "")
    candle_date   = str(r.get("cp_date", r.get("pattern_date", "")) or "")
    candle_bias   = str(r.get("cp_bias", r.get("candle_bias", "")) or "")

    _bull_candles = {
        "Morning Star": 10, "Bullish Engulfing": 9, "Tweezer Bottoms": 8,
        "Three Inside Up": 8, "Three White Soldiers": 7, "Hammer": 7,
        "Inverted Hammer": 6,
    }
    _bear_penalties = {
        "Evening Star": -3, "Bearish Engulfing": -3,
        "Tweezer Tops": -2, "Three Inside Down": -2,
    }
    candle_score = 0.0
    for pat, pts in _bull_candles.items():
        if pat.lower() in candle_pat.lower():
            candle_score = float(pts)
            break
    for pat, pts in _bear_penalties.items():
        if pat.lower() in candle_pat.lower():
            candle_score += float(pts)
    # Date decay: pattern must be within last 3 trading days (Apr 20–22 2026)
    _pattern_fresh = False
    try:
        from datetime import date as _date
        _ref = _date(2026, 4, 22)
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                _pd = datetime.strptime(candle_date[:10], fmt).date()
                if (_ref - _pd).days <= 3:
                    _pattern_fresh = True
                break
            except Exception:
                continue
    except Exception:
        pass
    if not _pattern_fresh and candle_score > 0:
        candle_score *= 0.5
    candle_score = max(-3.0, min(10.0, candle_score))

    # ─── COMPONENT 6 – Liquidity (0–10) → weight 10% ────────────────────────
    rvol         = safe_num(r.get("rvol20"), np.nan)
    rvol_zone    = str(r.get("rvol20_zone", r.get("rvol_zone", "")) or "")
    liq_cat      = str(r.get("liquidity_category", r.get("liquidity", "")) or "")

    if pd.notna(rvol):
        if rvol > 2.0:
            liq_score = 10.0
        elif rvol >= 1.5:
            liq_score = 9.0
        elif rvol >= 1.0:
            liq_score = 7.0
        elif rvol >= 0.5:
            liq_score = 4.0
        else:
            liq_score = 1.0
    else:
        liq_score = 4.0
    if "Above 1.5" in rvol_zone or "above" in rvol_zone.lower():
        liq_score = min(10.0, liq_score + 1.0)
    elif "Below 1.5" in rvol_zone or "below" in rvol_zone.lower():
        liq_score = max(0.0, liq_score - 1.0)
    if "Very Low" in liq_cat:
        liq_score *= 0.5
    elif "Low" in liq_cat:
        liq_score *= 0.8

    # ─── COMPONENT 7 – Mean Reversion / Extremity (0–10) → informational ────
    w52_pos      = safe_num(r.get("w52_position_pct", r.get("pos_52w_pct")), np.nan)
    rng_comp     = safe_num(r.get("range_compression_20d", r.get("range_compression")), np.nan)
    vol_comp     = safe_num(r.get("vol_compression_20d", r.get("volatility_compression")), np.nan)
    base_len     = safe_num(r.get("base_length_days", r.get("base_length")), np.nan)

    mr_parts = []
    if pd.notna(w52_pos):
        if 20 <= w52_pos <= 80:
            mr_parts.append(10.0)
        elif 10 <= w52_pos < 20 or 80 < w52_pos <= 90:
            mr_parts.append(6.0)
        else:
            mr_parts.append(2.0)
    if pd.notna(rng_comp):
        if rng_comp > 70:
            mr_parts.append(10.0)
        elif rng_comp >= 50:
            mr_parts.append(7.0)
        elif rng_comp >= 30:
            mr_parts.append(4.0)
        else:
            mr_parts.append(1.0)
    if pd.notna(vol_comp):
        if vol_comp > 80:
            mr_parts.append(10.0)
        elif vol_comp >= 60:
            mr_parts.append(7.0)
        elif vol_comp >= 40:
            mr_parts.append(4.0)
        else:
            mr_parts.append(1.0)
    if pd.notna(base_len):
        if 20 <= base_len <= 60:
            mr_parts.append(10.0)
        elif base_len <= 120:
            mr_parts.append(7.0)
        else:
            mr_parts.append(4.0)
    mr_score = sum(mr_parts) / len(mr_parts) if mr_parts else 5.0

    # ─── COMPOSITE (weighted average) ─────────────────────────────────────────
    # Weights: structure 20%, vwap 20%, macd 20%, rsi 15%, candle 15%, liq 10%, mr 0% (informational)
    # mr contributes via a small 10% weight derived from the remaining budget
    composite = (
        0.20 * struct_score +
        0.20 * vwap_score +
        0.20 * macd_score +
        0.15 * rsi_score +
        0.15 * max(0.0, candle_score) +
        0.10 * liq_score +
        0.00 * mr_score          # tracked but not in weighted sum per spec
    )

    # Sector RS-Momentum additive modifier
    sector = str(r.get("sector", "") or "")
    composite += _v17_sector_bias(sector)

    # Reject bearish MACD state
    if _macd_reject:
        composite = min(composite, 5.9)

    composite = max(0.0, min(10.0, round(composite, 2)))

    # Priority labels
    if composite >= 8.0:
        priority = "P1 - EXECUTE"
    elif composite >= 7.0:
        priority = "P2 - PREPARE"
    elif composite >= 6.0:
        priority = "P3 - MONITOR"
    else:
        priority = "BELOW THRESHOLD"

    return (composite, priority)


def _classify_swing_setup(r: dict, zone_type: str, rev_score: int) -> str:
    """
    Classify the swing setup type for quick visual scan.
    Returns one of: Breakout | Bounce | Continuation | Reversal | POI Watch
    """
    ms_regime   = str(r.get("ms_trend_regime", "") or "")
    ms_state    = str(r.get("ms_structure_state", "") or "")
    last_event  = str(r.get("ms_last_event", "") or "")
    event_age   = safe_num(r.get("ms_event_age_d"), np.nan)
    rvol        = safe_num(r.get("rvol20"), np.nan)
    bo_press    = safe_num(r.get("breakout_pressure"), np.nan)
    spring      = str(r.get("_removed_spring", "NO") or "NO")
    rsi14       = safe_num(r.get("rsi14"), np.nan)
    wyckoff     = str(r.get("wyckoff_proxy_phase", "") or "")
    markup_ready = safe_num(r.get("markup_readiness"), np.nan)
    macd_wave   = str(r.get("macd_wave", "") or "")
    rsi_div     = str(r.get("div_signal", "") or "")

    # Breakout: fresh Bull BOS + volume showing up + pressure building
    if (last_event == "Bull BOS"
            and pd.notna(event_age) and event_age <= 7
            and pd.notna(rvol) and rvol >= 1.5
            and pd.notna(bo_press) and bo_press >= 55):
        return "🔥 Breakout"

    # Reversal: Wyckoff accumulation mature OR bullish divergence + CHoCH
    if (wyckoff == "Accumulation"
            and pd.notna(markup_ready) and markup_ready >= 70
            and last_event in ("Bull CHoCH", "Spring")):
        return "⚡ Reversal"
    if "Bullish" in rsi_div and last_event == "Bull CHoCH":
        return "⚡ Reversal"

    # Bounce: price at deep level OR spring candidate OR oversold RSI
    if zone_type in ("PQ -2 SD", "PY -2 SD") or spring == "YES":
        return "📍 Bounce"
    if zone_type in ("PQ -1 SD", "PY -1 SD") and pd.notna(rsi14) and rsi14 <= 45:
        return "📍 Bounce"

    # Continuation: healthy uptrend + constructive pullback to VWAP
    if ("Bullish" in ms_regime
            and ms_state in ("HH-HL", "HH/HL")
            and zone_type in ("PQVWAP", "PYVWAP", "PQ -1 SD", "PY -1 SD")
            and "Mountain Building" in macd_wave):
        return "🔄 Continuation"

    # Reversal fallback: reversal signals strong enough
    if rev_score >= 3 and "Bullish" not in ms_regime:
        return "⚡ Reversal"

    return "👁 POI Watch"


# =========================
# CONVICTION SCORING ENGINE
# =========================

# Global RRG quadrant map — populated by _inject_rrg_sector_state() at runtime
_RRG_SECTOR_QUADRANT: Dict[str, str] = {}

def _inject_rrg_sector_state(rrg_results: list):
    """
    Called from main() after RRG is computed.
    rrg_results = list of dicts with keys: Sector, Quadrant
    Populates _RRG_SECTOR_QUADRANT so conviction scoring can use it.
    Also syncs _RRG_QUADRANT_CACHE so get_rrg_quadrant_for_sector() works.
    """
    global _RRG_SECTOR_QUADRANT, _RRG_QUADRANT_CACHE
    _RRG_SECTOR_QUADRANT = {}
    for row in rrg_results:
        sector = str(row.get("Sector", "") or "").strip()
        quad   = str(row.get("Quadrant", "") or "").strip()
        if sector and quad:
            _RRG_SECTOR_QUADRANT[sector] = quad
            _RRG_QUADRANT_CACHE[sector]  = quad   # ← sync to upgrade-10 cache

def _sector_rrg_modifier(sector: str) -> int:
    """Return score modifier based on live RRG quadrant. 0 if not found."""
    quad = _RRG_SECTOR_QUADRANT.get(str(sector).strip(), "")
    return {"Leading": 2, "Improving": 1, "Weakening": 0, "Lagging": -1}.get(quad, 0)

# V17 RS-Momentum sector bias (±0.5 composite additive, per spec)
_RS_MOMENTUM_MAP: Dict[str, float] = {
    "IDXFINANCE":  0.5,
    "IDXNONCYC":   0.3,
    "IDXTRANS":    0.2,
    "IDXENERGY":  -0.1,
    "IDXBASIC":   -0.1,
    "IDXINDUST":  -0.2,
    "IDXCYCLIC":  -0.3,
    "IDXINFRA":   -0.4,
    "IDXPROPERT": -0.5,
    "IDXHEALTH":  -0.1,
}

def _v17_sector_bias(sector: str) -> float:
    """Return ±0.5 RS-Momentum composite modifier for a sector string."""
    s = str(sector or "").strip().upper()
    # Try exact match first, then prefix match
    if s in _RS_MOMENTUM_MAP:
        return _RS_MOMENTUM_MAP[s]
    for k, v in _RS_MOMENTUM_MAP.items():
        if s.startswith(k) or k in s:
            return v
    # Fall back to live RRG quadrant if RS-Momentum map doesn't cover it
    quad = _RRG_SECTOR_QUADRANT.get(sector.strip(), "")
    return {"Leading": 0.3, "Improving": 0.2, "Weakening": 0.0, "Lagging": -0.2}.get(quad, 0.0)

def _conviction_score(r: dict, tgt_dist: float) -> int:
    """
    Returns integer conviction score 0–10.
    Higher = more evidence the setup is real RIGHT NOW.
    """
    score = 0

    # MA Alignment (0–3)
    ma_tier = r.get("_sc_ma_tier", 0)
    if ma_tier >= 4:
        score += 3
    elif ma_tier == 3:
        score += 2
    elif ma_tier >= 1:
        score += 1

    # RVOL (0–3)
    rvol = safe_num(r.get("rvol20"), 0)
    if pd.notna(rvol) and rvol >= 5.0:
        score += 3
    elif pd.notna(rvol) and rvol >= 3.0:
        score += 2
    elif pd.notna(rvol) and rvol >= 1.5:
        score += 1

    # Last market structure event (0–2)
    last_event = str(r.get("ms_last_event", "") or "")
    if last_event == "Bull BOS":
        score += 2
    elif last_event == "Bull CHoCH":
        score += 1

    # RSI Status (0–1)
    if str(r.get("rsi_status", "") or "") == "Strong":
        score += 1

    # MACD Wave (0–1)
    if "Mountain Building" in str(r.get("macd_wave", "") or ""):
        score += 1

    # Upside % (0–1)
    td = safe_num(tgt_dist, np.nan)
    if pd.notna(td) and td >= 15.0:
        score += 1

    # Sector RRG modifier (−1 to +2)
    score += _sector_rrg_modifier(str(r.get("sector", "") or ""))

    return max(0, min(10, score))


def _poi_candidate_map(r: dict) -> list:
    """Return ranked POI candidates for POI-first swing engine.
    Each candidate = {zone_type, level, dist_pct, poi_type, next_poi, next_px}.
    """
    close_px = safe_num(r.get("close"), np.nan)
    if pd.isna(close_px) or close_px <= 0:
        return []

    level_map = [
        ("PQ -2 SD", "pq_m2", "Quarter VWAP Band", "PQ -1 SD", "pq_m1"),
        ("PQ -1 SD", "pq_m1", "Quarter VWAP Band", "PQVWAP", "pq_vwap"),
        ("PQVWAP", "pq_vwap", "Quarter VWAP Band", "PYVWAP", "py_vwap"),
        ("PY -2 SD", "py_m2", "Yearly VWAP Band", "PY -1 SD", "py_m1"),
        ("PY -1 SD", "py_m1", "Yearly VWAP Band", "PYVWAP", "py_vwap"),
        ("PYVWAP", "py_vwap", "Yearly VWAP Band", "PQVWAP", "pq_vwap"),
    ]
    candidates = []
    for zone_type, key, poi_type, next_lbl, next_key in level_map:
        lvl = safe_num(r.get(key), np.nan)
        if pd.isna(lvl) or lvl == 0:
            continue
        dist = ((close_px / lvl) - 1.0) * 100.0
        absdist = abs(dist)
        # tighter thresholds for SD bands, looser for VWAP anchors
        thr = 1.25 if "SD" in zone_type else 2.0
        if absdist > thr:
            continue
        nxt = safe_num(r.get(next_key), np.nan)
        candidates.append({
            "zone_type": zone_type,
            "poi_type": poi_type,
            "level": lvl,
            "dist_pct": dist,
            "next_poi": next_lbl if pd.notna(nxt) else "N/A",
            "next_px": nxt,
        })

    # fallback context POIs only if no core VWAP-family candidates
    if not candidates:
        for lbl, key in [("IBL", "ibl"), ("PWL", "pwl"), ("MDL", "mdl")]:
            lvl = safe_num(r.get(key), np.nan)
            if pd.isna(lvl) or lvl == 0:
                continue
            dist = ((close_px / lvl) - 1.0) * 100.0
            if abs(dist) <= 1.0:
                candidates.append({
                    "zone_type": lbl,
                    "poi_type": "Profile Level",
                    "level": lvl,
                    "dist_pct": dist,
                    "next_poi": "PQVWAP" if pd.notna(safe_num(r.get("pq_vwap"), np.nan)) else "N/A",
                    "next_px": safe_num(r.get("pq_vwap"), np.nan),
                })
                break

    rank_map = {"PY -2 SD":0, "PQ -2 SD":1, "PY -1 SD":2, "PQ -1 SD":3, "PYVWAP":4, "PQVWAP":5, "IBL":6, "PWL":7, "MDL":8}
    candidates.sort(key=lambda c: (rank_map.get(c["zone_type"], 99), abs(safe_num(c.get("dist_pct"), 99))))
    return candidates


def _derive_amt_state_v2(r: dict, zone_type: str, dist_pct: float) -> str:
    """POI-first AMT state label tuned for long-side swing setups."""
    mp_zone = str(r.get("mp_zone", "") or "").strip()
    last_event = str(r.get("ms_last_event", "") or "")
    spring = str(r.get("_removed_spring", "NO") or "NO")
    div = str(r.get("div_signal", "") or "")
    sd_delta = safe_num(r.get("pq_sd_delta"), np.nan)
    if pd.isna(sd_delta):
        sd_delta = safe_num(r.get("py_sd_delta"), np.nan)

    if spring == "YES" or last_event == "Spring":
        return "Spring reclaim"
    if last_event in ("Bull CHoCH", "Bull BOS") and pd.notna(dist_pct) and dist_pct <= 0:
        return "Lower rejection / re-entry"
    if "Bullish" in div and pd.notna(dist_pct) and dist_pct <= 0:
        return "Responsive buying"
    if pd.notna(sd_delta) and sd_delta > 0 and pd.notna(dist_pct) and dist_pct <= 0:
        return "Re-entering value"
    if zone_type in ("PQ -2 SD", "PY -2 SD"):
        return "Deep discount / responsive zone"
    if mp_zone in ("At Level", "Above") and pd.notna(dist_pct) and abs(dist_pct) <= 1.0:
        return "At lower value edge"
    return "POI monitoring"


def _poi_rank_v2(r: dict, zone_type: str, dist_pct: float, rr: float, rev_score: int) -> str:
    score = 0
    if zone_type in ("PY -2 SD", "PQ -2 SD"):
        score += 3
    elif zone_type in ("PY -1 SD", "PQ -1 SD"):
        score += 2
    elif zone_type in ("PYVWAP", "PQVWAP"):
        score += 1
    if abs(safe_num(dist_pct, 99)) <= 0.75:
        score += 2
    elif abs(safe_num(dist_pct, 99)) <= 1.25:
        score += 1
    if safe_num(rr, 0) >= 2.5:
        score += 2
    elif safe_num(rr, 0) >= 1.8:
        score += 1
    if rev_score >= 3:
        score += 2
    elif rev_score >= 2:
        score += 1
    if score >= 8:
        return "A+"
    if score >= 6:
        return "A"
    if score >= 4:
        return "B"
    return "C"


def _build_idx_vwap_shortlist(rows):
    """
    TRUE POI ENGINE V17 — LONG ONLY
    ---------------------------------
    Applies 10-step V17 scoring framework. A row appears only if:
      Step 1) Structural Quality Filter (all 5 conditions must pass)
      Step 2) Wyckoff Phase Filter (Accumulation or Markup only)
      Step 3–9) Scored on 7 components
      Step 10) Composite >= 6.0 → P1/P2/P3 priority assigned
    Emits _sc_* compatibility fields for the workbook builder.
    """
    board = []

    def _num(v):
        return safe_num(v, np.nan)

    def _fmt_px(lbl, px):
        if not lbl or lbl == "N/A" or pd.isna(safe_num(px, np.nan)):
            return "N/A"
        return f"{lbl}  {safe_num(px, 0):,.0f}"

    def _str(v):
        return str(v or "")

    for r in rows:
        if _str(r.get("data_status")) != "OK":
            continue
        close_px = _num(r.get("close"))
        if pd.isna(close_px) or close_px <= 0:
            continue

        # ── STEP 1: Structural Quality Filter ────────────────────────────────
        ms_quality   = _str(r.get("ms_structure_quality"))
        ms_state     = _str(r.get("ms_structure_state"))
        event_age    = _num(r.get("ms_event_age_d"))
        last_event   = _str(r.get("ms_last_event"))
        trend_bias   = _str(r.get("ms_trend_regime"))
        swing_seq    = _str(r.get("ms_swing_sequence", r.get("ms_structure_state")))

        if ms_quality != "Clean":
            continue
        if ms_state not in ("HH-HL", "HH/HL") and "HH" not in swing_seq:
            continue
        if pd.notna(event_age) and event_age > 20:
            continue
        if "Bullish" not in last_event:
            continue
        if "Bearish" in trend_bias and "Neutral" not in trend_bias:
            continue

        # Wyckoff Phase Filter removed (v2.0)
        # ── Hard tradability gates ─────────────────────────────────────────────
        if _liquidity_check(r) == "Thin":
            continue
        adr = _num(r.get("adr_pct"))
        atr = _num(r.get("atr14_pct"))
        if pd.notna(adr) and adr < 3.0:
            continue
        if pd.notna(atr) and atr < 3.0:
            continue
        if _str(r.get("_removed_upthrust")) == "YES" and _str(r.get("failed_breakout_risk")) == "YES":
            continue

        # ── STEPS 3–9: Build scored row ───────────────────────────────────────
        candidates = _poi_candidate_map(r)
        if not candidates:
            continue
        zone     = candidates[0]
        zone_type = zone["zone_type"]
        dist_pct  = safe_num(zone.get("dist_pct"), np.nan)
        entry_px  = safe_num(zone.get("level"), np.nan)
        tgt_lbl   = zone.get("next_poi", "N/A")
        tgt_px    = safe_num(zone.get("next_px"), np.nan)
        if pd.isna(tgt_px):
            continue
        tgt_dist = ((tgt_px / close_px) - 1.0) * 100.0 if close_px > 0 else np.nan
        if pd.isna(tgt_dist) or tgt_dist < 3.0:
            continue

        # Invalidation
        inv_map = {
            "PQ -2 SD": ("PQ -2 SD", _num(r.get("pq_m2"))),
            "PQ -1 SD": ("PQ -2 SD", _num(r.get("pq_m2"))),
            "PQVWAP":   ("PQ -1 SD", _num(r.get("pq_m1"))),
            "PY -2 SD": ("PY -2 SD", _num(r.get("py_m2"))),
            "PY -1 SD": ("PY -2 SD", _num(r.get("py_m2"))),
            "PYVWAP":   ("PY -1 SD", _num(r.get("py_m1"))),
            "IBL":      ("IBL",      _num(r.get("ibl"))),
            "PWL":      ("PWL",      _num(r.get("pwl"))),
            "MDL":      ("MDL",      _num(r.get("mdl"))),
        }
        inv_lbl, inv_px = inv_map.get(zone_type, (zone_type, entry_px))
        risk_pct = abs(((entry_px / inv_px) - 1.0) * 100.0) \
            if pd.notna(inv_px) and inv_px > 0 and pd.notna(entry_px) else np.nan
        rr_val = safe_num(
            (tgt_dist / risk_pct) if pd.notna(risk_pct) and risk_pct > 0 else np.nan, np.nan
        )
        if pd.isna(rr_val) or rr_val < 1.5:   # V17 minimum R/R
            continue

        rev_score, _rev_signals = _reversal_score(r)
        amt_state  = _derive_amt_state_v2(r, zone_type, dist_pct)
        ma_tier, ma_label = _score_ma_alignment(r)
        above_ibl  = _str(r.get("mp_zone")).strip() in ("At Level", "Above")

        # Confluence text
        conf = []
        if zone_type.startswith("PQ") or zone_type.startswith("PY"):
            conf.append(zone_type.replace("VWAP", "").strip())
        if above_ibl:
            conf.append("IBL")
        if ma_tier >= 2:
            conf.append("EMA20")
        if _str(r.get("div_signal")).startswith("Bullish"):
            conf.append("Bull Div")
        if _str(r.get("smc_bull_ob_1", "")).strip() and close_px:
            conf.append("Bull OB")
        confluence = " + ".join([c for c in conf if c]) or "POI-led"

        # Trigger — V17 specific conditions
        candle_pat = _str(r.get("cp_pattern", r.get("last_candle_pattern")))
        if candle_pat and candle_pat not in ("N/A", ""):
            trigger = f"Candle confirm: {candle_pat} + close above {zone_type}"
        elif ma_tier >= 2:
            trigger = f"EMA10/20 cross with RVOL > 1.0 above {zone_type}"
        else:
            trigger = f"VWAP reclaim with RVOL > 1.0 at {zone_type}"

        # T2: next higher VWAP anchor
        _t2_candidates = [("pq_vwap", "PQVWAP"), ("py_vwap", "PYVWAP"),
                          ("pq_m1", "PQ -1 SD"), ("py_m1", "PY -1 SD")]
        t2_disp = "N/A"
        for _k, _lbl in _t2_candidates:
            _v = _num(r.get(_k))
            if pd.notna(_v) and _v > close_px and _lbl != tgt_lbl:
                t2_disp = _fmt_px(_lbl, _v)
                break

        # Setup type
        if zone_type in ("PQ -2 SD", "PY -2 SD"):
            setup_type = "📍 Deep Discount"
        elif zone_type in ("PQ -1 SD", "PY -1 SD"):
            setup_type = "📍 Discount Pullback"
        elif "Bull BOS" in last_event and ma_tier >= 2:
            setup_type = "🔄 Continuation"
        elif rev_score >= 2 or amt_state in ("Spring reclaim", "Responsive buying"):
            setup_type = "⚡ Reversal"
        else:
            setup_type = "👁 POI Watch"

        # ── STEP 9: Composite scoring ──────────────────────────────────────────
        rr = dict(r)
        rr["sector"] = _str(r.get("sector") or r.get("IDX Sector") or "N/A")
        # Stage intermediate fields so _swing_score can read them
        rr["ms_trend_regime"]     = trend_bias
        rr["ms_structure_state"]  = ms_state
        rr["ms_last_event"]       = last_event
        rr["ms_event_age_d"]      = event_age
        rr["ms_structure_quality"]= ms_quality

        sw_score, priority_label = _swing_score(rr, zone_type, tgt_dist)

        # ── STEP 10: Priority assignment ──────────────────────────────────────
        if sw_score < 6.0:
            continue   # Below minimum threshold

        # buy_signal: YES if all strong, else WATCH
        ready = (
            ma_tier >= 2 and above_ibl and
            ("Bearish" not in trend_bias or zone_type in ("PQ -2 SD", "PY -2 SD")) and
            rev_score >= 1
        )
        buy_signal = "YES" if ready else "WATCH"
        if ma_tier == 0:
            buy_signal = "WATCH"

        poi_rank_v = _poi_rank_v2(r, zone_type, dist_pct, rr_val, rev_score)

        # Brief actionable summary
        cause_quality  = _str(r.get("cause_quality", ""))
        markup_ready   = _num(r.get("markup_readiness"))
        summary = (
            f"{zone_type} | {setup_type.split()[-1]} | "
            f"RR {rr_val:.1f} | {priority_label.split(' - ')[-1].title()}"
        )
        if cause_quality:
            summary += f" | Cause: {cause_quality}"

        # Write all _sc_* fields
        rr["_sc_ma_path"]            = "PASS" if ma_tier >= 1 else "EXCEPTION"
        rr["_sc_ma_tier"]            = ma_tier
        rr["_sc_ma_label"]           = ma_label
        rr["_sc_zone_type"]          = zone_type
        rr["_sc_dist_pct"]           = dist_pct
        rr["_sc_entry_disp"]         = _fmt_px(zone_type, entry_px)
        rr["_sc_target_lbl"]         = tgt_lbl
        rr["_sc_target_px"]          = tgt_px
        rr["_sc_target_dist"]        = tgt_dist
        rr["_sc_target_disp"]        = _fmt_px(tgt_lbl, tgt_px)
        rr["_sc_t2_disp"]            = t2_disp
        rr["_sc_invalidation_disp"]  = _fmt_px(f"< {inv_lbl}", inv_px)
        rr["_sc_trigger"]            = trigger
        rr["_sc_signal"]             = amt_state
        rr["_sc_buy_signal"]         = buy_signal
        rr["_sc_trade_bias"]         = "Long" if buy_signal == "YES" else "Watch"
        rr["_sc_rev_score"]          = rev_score
        rr["_sc_setup_type"]         = setup_type
        rr["_sc_poi_type"]           = zone.get("poi_type", "POI")
        rr["_sc_primary_poi"]        = zone_type
        rr["_sc_next_poi"]           = tgt_lbl
        rr["_sc_confluence"]         = confluence
        rr["_sc_amt_state"]          = amt_state
        rr["_sc_summary"]            = summary
        rr["_sc_rr"]                 = round(rr_val, 1)
        rr["_sc_poi_rank"]           = poi_rank_v
        rr["_sc_conviction"]         = sw_score
        rr["_sc_swing_score"]        = sw_score
        rr["_sc_swing_grade"]        = priority_label
        rr["_sc_priority_label"]     = priority_label
        rr["_sc_priority"]           = "YES" if priority_label == "P1 - EXECUTE" else ""
        rr["_sc_cause_quality"]      = cause_quality
        rr["_sc_markup_readiness"]   = markup_ready

        # Sort key: P1 < P2 < P3, then score desc, then RVOL desc
        _p_rank = {"P1 - EXECUTE": 0, "P2 - PREPARE": 1, "P3 - MONITOR": 2}.get(priority_label, 3)
        rr["_sc_sort_key"] = (
            _p_rank,
            -sw_score,
            -safe_num(r.get("rvol20"), 0),
            -safe_num(markup_ready, 0),
            rr.get("ticker", "")
        )
        board.append(rr)

    board.sort(key=lambda x: x.get("_sc_sort_key", (9, 0, 0, 0, "")))
    return board


def _market_profile_summary(r: dict) -> str:
    """Compact market-profile context for the screener sheet."""
    parts = []
    mp = str(r.get("mp_zone", "") or "").strip()
    pwl = str(r.get("price_ge_pwl", "") or "").strip()
    mdl = str(r.get("price_ge_mdl", "") or "").strip()
    if mp and mp != "N/A":
        parts.append(f"IBL {mp}")
    if pwl and pwl != "N/A":
        parts.append(f"PWL {pwl}")
    if mdl and mdl != "N/A":
        parts.append(f"MDL {mdl}")
    return " | ".join(parts) if parts else "N/A"



# =========================
# IDX SCREENER SUMMARY SHEET  (replaces Swing Watchlist)
# Shows top-scored POI candidates in a compact executive dashboard
# =========================
def build_summary_sheet(ws, latest_market_day: str, summary_rows: list):
    """
    IDX Screener Summary — Priority dashboard replacing Swing Watchlist.
    Columns: Rank | Ticker | Emiten | Sector | Regime | Close | Score | MA Zone |
             SMC State | MACD | QVWAP Zone | PQ Zone | PY Zone | Signal | RR | Notes
    """
    ws.title = "IDX Screener Summary"
    ws.column_dimensions["A"].width = 0.5

    for r_h, h in {1: 4, 2: 26, 3: 16, 4: 4, 5: 22, 6: 36}.items():
        ws.row_dimensions[r_h].height = h

    pretty = ""
    try:
        pretty = datetime.strptime(latest_market_day, "%Y-%m-%d").strftime("%B %d, %Y")
    except Exception:
        pretty = str(latest_market_day)

    ws.merge_cells("B2:P2")
    c = ws["B2"]
    c.value = "IDX Screener Summary  ·  SMC + MACD + VWAP Priority Board"
    style_plain(c, font=FONT_TITLE, align="left")

    ws["B3"] = (
        f"As of {pretty}  ·  "
        f"Filter: Score ≥ 6.0  ·  Sorted: Priority → Score ↓  ·  "
        f"Long Only  ·  v2.0"
    )
    style_plain(ws["B3"], font=FONT_SUBTITLE, align="left")

    # Column schema
    COLS = [
        ("Rank",        6,  "center"),
        ("Ticker",      9,  "center"),
        ("Emiten",      34, "left"),
        ("Sector",      20, "left"),
        ("Regime",      22, "center"),
        ("Close",       10, "center"),
        ("Score",       9,  "center"),
        ("MA Zone",     22, "center"),
        ("SMC State",   22, "center"),
        ("MACD",        24, "center"),
        ("QVWAP Zone",  28, "left"),
        ("PQ Zone",     28, "left"),
        ("Signal",      22, "center"),
        ("RR",          8,  "center"),
        ("Priority",    12, "center"),
    ]

    hdr_row = 6
    for col_i, (hdr, w, _align) in enumerate(COLS, start=2):
        c = ws.cell(hdr_row, col_i)
        c.value = hdr
        style_cell(c, fill=FILL_HEADER, font=FONT_HEADER, align="center")
        ws.column_dimensions[get_column_letter(col_i)].width = w

    data_start = hdr_row + 1
    r = data_start

    SIGNAL_COLORS = {
        "P1 - EXECUTE": "10B981",   # emerald
        "P2 - PREPARE": "3B82F6",   # blue
        "P3 - MONITOR": "A78BFA",   # violet
    }
    REGIME_COLORS = {
        "Blue Chip / High Liquidity":  "2A4A66",
        "Institutional Driven":        "237A5C",
        "Mid Cap / Moderate":          "21467C",
        "Low Liquidity / Small Cap":   "7A1A1A",
    }

    displayed = 0
    for row in summary_rows:
        if row.get("data_status") not in ("OK",):
            continue
        score = safe_num(row.get("_sc_swing_score") or row.get("_sc_conviction"), 0)
        if score < 6.0:
            continue

        displayed += 1
        row_fill = PatternFill("solid", fgColor="F0FDF4" if displayed % 2 == 0 else "F8FAFC")
        priority_lbl = str(row.get("_sc_priority_label") or row.get("_sc_swing_grade") or "")
        pri_color = SIGNAL_COLORS.get(priority_lbl, "1F1F1F")

        vals = [
            displayed,
            row.get("ticker", ""),
            row.get("emiten", ""),
            str(row.get("idx_sector") or row.get("sector") or ""),
            row.get("stock_regime", "N/A"),
            safe_num(row.get("close"), np.nan),
            round(score, 1),
            row.get("summary_ma", "N/A"),
            str(row.get("smc_composite_state") or row.get("smc_state") or "N/A"),
            str(row.get("macd_composite") or row.get("macd_entry") or ""),
            str(row.get("q_remarks") or ""),
            str(row.get("pq_remarks") or ""),
            str(row.get("_sc_signal") or row.get("_sc_buy_signal") or ""),
            round(safe_num(row.get("_sc_rr"), 0), 1) if pd.notna(safe_num(row.get("_sc_rr"), np.nan)) else "N/A",
            priority_lbl,
        ]

        for col_i, (val, (_, w, align)) in enumerate(zip(vals, COLS), start=2):
            c = ws.cell(r, col_i)
            c.value = val
            style_cell(c, fill=row_fill, font=FONT_BODY, align=align)

        # Colour priority cell
        pri_cell = ws.cell(r, 2 + 14)
        pri_cell.fill = PatternFill("solid", fgColor=pri_color) if priority_lbl else row_fill
        pri_cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")

        # Colour regime cell
        reg_cell = ws.cell(r, 2 + 4)
        reg_str = str(row.get("stock_regime", ""))
        if reg_str in REGIME_COLORS:
            reg_cell.fill = PatternFill("solid", fgColor=REGIME_COLORS[reg_str])
            reg_cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")

        ws.row_dimensions[r].height = 16
        r += 1

    if displayed == 0:
        ws.cell(data_start, 2).value = "No qualifying rows (Score ≥ 6.0) found in this run."
        style_plain(ws.cell(data_start, 2), font=FONT_SUBTITLE, align="left")
        r += 1

    ws.freeze_panes = f"B{hdr_row + 1}"
    if r > data_start:
        ws.auto_filter.ref = (
            f"B{hdr_row}:{get_column_letter(2 + len(COLS) - 1)}{r - 1}"
        )
    ws.sheet_view.showGridLines = False

def build_processing_sheet(ws, latest_market_day, total_scanned, logs, summary_rows):
    ws.title = "Data Processing Results"
    widths = {"A": 0.5, "B": 20, "C": 20, "D": 10, "E": 12, "F": 10, "G": 14, "H": 70}
    for c, w in widths.items():
        ws.column_dimensions[c].width = w

    ws["B2"] = "Data Processing Results"
    style_plain(ws["B2"], font=FONT_TITLE, align="left")

    run_dt_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ok_count      = sum(1 for x in logs if x["status"] == "OK")
    partial_count = sum(1 for x in logs if x["status"] == "PARTIAL DATA")
    nodata_count  = sum(1 for x in logs if x["status"] == "NO DATA")

    # ── Screener stats (read from global _RRG_SECTOR_QUADRANT if available) ──
    _leading_sectors   = [s for s, q in _RRG_SECTOR_QUADRANT.items() if q == "Leading"]
    _improving_sectors = [s for s, q in _RRG_SECTOR_QUADRANT.items() if q == "Improving"]

    FILL_STAT_HDR = PatternFill("solid", fgColor="2A4A66")
    FILL_STAT_OK  = PatternFill("solid", fgColor="D1FAE5")
    FILL_STAT_WRN = PatternFill("solid", fgColor="FEF3C7")
    FILL_STAT_PRI = PatternFill("solid", fgColor="FEF08A")

    def _stat_row(r, label, value, fill=None):
        ws.cell(r, 2).value = label
        ws.cell(r, 2).font  = Font(name="Calibri", size=10, bold=True)
        ws.cell(r, 2).fill  = fill or PatternFill("solid", fgColor="F8FAFC")
        ws.cell(r, 2).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(r, 2).border = BORDER
        ws.cell(r, 3).value = value
        ws.cell(r, 3).font  = Font(name="Calibri", size=10)
        ws.cell(r, 3).fill  = fill or PatternFill("solid", fgColor="FFFFFF")
        ws.cell(r, 3).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(r, 3).border = BORDER

    # Section header
    ws.merge_cells("B3:H3")
    hdr = ws["B3"]
    hdr.value = "Run Summary"
    hdr.font  = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    hdr.fill  = FILL_STAT_HDR
    hdr.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[3].height = 22

    stat_rows = [
        ("Run Time",              run_dt_str,                               None),
        ("As-of Date",            latest_market_day,                        None),
        ("Total Tickers Scanned", str(total_scanned),                       None),
        ("OK  (full data)",       str(ok_count),                            FILL_STAT_OK),
        ("Partial Data",          str(partial_count),                       FILL_STAT_WRN),
        ("No Data",               str(nodata_count),                        None),
        ("",                      "",                                        None),
    ]

    r = 4
    for label, value, fill in stat_rows:
        _stat_row(r, label, value, fill)
        ws.row_dimensions[r].height = 20
        r += 1

    # Legacy screener stat block removed per current output spec
    r += 1  # spacer

    # ── Ticker log table ──────────────────────────────────────────────────────
    header_row = r
    headers = ["Ticker", "Status", "Bars", "Retry Count", "Source Used", "Latest Market Day", "Reason"]
    for i, h in enumerate(headers, start=2):
        c = ws.cell(header_row, i)
        c.value = h
        style_cell(c, fill=FILL_HEADER_SUMMARY, font=FONT_SUBHEADER, align="center", wrap=True)
    ws.row_dimensions[header_row].height = 28
    r = header_row + 1

    for log in logs:
        vals = [
            log["ticker"], log["status"], log["bars"], log["retry_count"],
            log["source_used"], log["latest_market_day"], log["reason"]
        ]
        for c_idx, v in enumerate(vals, start=2):
            c = ws.cell(r, c_idx)
            c.value = v
            align = "left" if c_idx in [2, 3, 6, 8] else "center"
            row_fill = None
            if log["status"] == "OK":
                row_fill = PatternFill("solid", fgColor="F0FDF4")
            elif log["status"] == "PARTIAL DATA":
                row_fill = PatternFill("solid", fgColor="FFFBEB")
            style_cell(c, fill=row_fill, font=FONT_BODY, align=align, wrap=(c_idx == 8))
        ws.row_dimensions[r].height = 22
        r += 1

    ws.freeze_panes = None
    ws.auto_filter.ref = f"B{header_row}:H{max(header_row, ws.max_row)}"
    ws.sheet_view.showGridLines = False


def build_detail_sheet(ws, latest_market_day, rows):
    ws.title = "IDX Technical Detail"

    ws.column_dimensions["A"].width = 0.5

    for r, h in {1: 5, 2: 24, 3: 18, 4: 5, 5: 24, 6: 44, 7: 20, 8: 20, 9: 20, 10: 5, 11: 24, 12: 48}.items():
        ws.row_dimensions[r].height = h

    ws["B2"] = "IDX Technical Detail" + (" [BACKTEST REPLAY]" if BACKTEST_MODE else "")
    style_plain(ws["B2"], font=FONT_TITLE, align="left")
    _asof_ts = get_effective_asof_date()
    pretty = _asof_ts.strftime("%B %d, %Y")
    _mode_lbl = f"BACKTEST REPLAY as of {pretty}" if BACKTEST_MODE else f"Market Data as of {pretty}"
    ws["B3"] = _mode_lbl
    style_plain(ws["B3"], font=FONT_SUBTITLE, align="left")

    for _c in range(2, 250):
        ws.cell(4, _c).value = None

    # Flatten schema and assign columns dynamically
    flat_cols = []
    col_idx = 2  # start at B
    group_ranges = []
    for group_name, group_fill, cols in DETAIL_SCHEMA:
        start_idx = col_idx
        for key, header, width, fmt, align in cols:
            letter = get_column_letter(col_idx)
            ws.column_dimensions[letter].width = width
            flat_cols.append((col_idx, letter, group_name, group_fill, key, header, width, fmt, align))
            col_idx += 1
        end_idx = col_idx - 1
        group_ranges.append((group_name, group_fill, start_idx, end_idx))

    # Group headers row 5
    for group_name, group_fill, start_idx, end_idx in group_ranges:
        start_cell = f"{get_column_letter(start_idx)}5"
        end_cell = f"{get_column_letter(end_idx)}5"
        ws.merge_cells(f"{start_cell}:{end_cell}")
        for c in range(start_idx, end_idx + 1):
            cell = ws.cell(5, c)
            cell.fill = group_fill
            cell.font = FONT_GROUP
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = BORDER
        display_group_name = group_name
        if group_name == "Current QVWAP (Q2 2026)":
            q = ((datetime.strptime(latest_market_day, "%Y-%m-%d").month - 1) // 3) + 1
            y = datetime.strptime(latest_market_day, "%Y-%m-%d").year
            display_group_name = f"Current QVWAP (Q{q} {y})"
        elif group_name == "Previous QVWAP (Q1 2026)":
            dt = datetime.strptime(latest_market_day, "%Y-%m-%d")
            prev_end = quarter_start(pd.Timestamp(dt)) - pd.Timedelta(days=1)
            pq = ((prev_end.month - 1) // 3) + 1
            display_group_name = f"Previous QVWAP (Q{pq} {prev_end.year})"
        elif group_name == "Previous Year VWAP (2025)":
            y = datetime.strptime(latest_market_day, "%Y-%m-%d").year - 1
            display_group_name = f"Previous Year VWAP ({y})"
        ws[start_cell] = display_group_name

    # Subheaders row 6
    for col_idx, letter, group_name, group_fill, key, header, width, fmt, align in flat_cols:
        c = ws.cell(6, col_idx)
        c.value = header
        style_cell(c, fill=group_fill, font=FONT_SUBHEADER, align="center", wrap=True)

    # Data rows
    for r_idx, r in enumerate(rows, start=7):
        row_fill = None

        for col_idx, letter, group_name, group_fill, key, header, width, fmt, align in flat_cols:
            c = ws.cell(r_idx, col_idx)
            val = _market_profile_summary(r) if key == "mp_summary" else r.get(key, "")
            sd_keys = ["q_sd","q_delta","pq_sd","pq_delta","py_sd","py_delta"]

            if key in ["mcap", "avg30", "last_vol"] and pd.notna(val):
                val = compact_fmt(val)
            elif isinstance(val, float) and pd.isna(val):
                val = None if key in sd_keys else "-"
            # Force compact display BEFORE writing to Excel for Liquidity numeric columns
            if key in ["lot", "daily_value", "adtr20", "last_vol", "adtv20"] and val not in ("-", "N/A") and pd.notna(val):
                val = compact_fmt(val)
            elif isinstance(val, float) and pd.isna(val):
                val = None if key in sd_keys else "-"

            # Normalize sentinel strings to "-" for display (keep internal "None" sentinel for logic)
            if val in ("N/A", "None") and key not in sd_keys:
                val = "-"

            # Keep SD metrics as TRUE numeric cells (not text) so Excel won't show "Number Stored as Text"
            if key in ["q_sd","q_delta","pq_sd","pq_delta","py_sd","py_delta"]:
                if isinstance(val, (int, float)) and pd.notna(val):
                    val = float(val)
                elif isinstance(val, float) and pd.isna(val):
                    val = None

            c.value = val
            fill_to_use = row_fill if row_fill else None
            style_cell(c, fill=fill_to_use, font=FONT_BODY, align=align, wrap=(key in ["sector","industry","q_remarks","pq_remarks","py_remarks","af_execution_guide","af_watch_next","af_invalidation","preset_summary"]))
            if key == "pct_change":
                try:
                    _pct_val = safe_num(r.get("pct_change"), np.nan)
                    if pd.notna(_pct_val):
                        c.number_format = "0.00%"
                        if _pct_val > 0:
                            c.font = Font(name=FONT_BODY.name, size=FONT_BODY.size, bold=False, italic=False, color="008000")
                        elif _pct_val < 0:
                            c.font = Font(name=FONT_BODY.name, size=FONT_BODY.size, bold=False, italic=False, color="C00000")
                except Exception:
                    pass

            if fmt and val != "N/A":


                if fmt == "compact":
                    pass
                elif fmt == "date_wyckoff":
                    try:
                        if value in (None, "", "N/A"):
                            pass
                        else:
                            if isinstance(value, str):
                                parsed = safe_parse_datetime(value, errors="coerce")
                                if pd.notna(parsed):
                                    cell.value = parsed.to_pydatetime()
                            elif hasattr(value, "to_pydatetime"):
                                cell.value = value.to_pydatetime()
                            cell.number_format = "dd mmm 'yy"
                    except Exception:
                        pass
                if fmt == "compact":
                    pass
                elif fmt == "wyckoff_date":
                    try:
                        if isinstance(value, str):
                            parsed = safe_parse_datetime(value, errors="coerce")
                            if pd.notna(parsed):
                                cell.value = parsed.to_pydatetime()
                        elif hasattr(value, "to_pydatetime"):
                            cell.value = value.to_pydatetime()
                        cell.number_format = "dd mmm 'yy"
                    except Exception:
                        pass
                if fmt == "compact":
                    pass
                else:
                    if isinstance(val, (int, float)) and not pd.isna(val):
                        c.number_format = fmt

        ws.row_dimensions[r_idx].height = 26

    last_col_letter = get_column_letter(flat_cols[-1][0])
    ws.freeze_panes = "E7"
    ws.auto_filter.ref = f"B6:{last_col_letter}{max(6, ws.max_row)}"
    ws.sheet_view.showGridLines = False

# Build guide delegated to new comprehensive implementation below
def _safe_sector_bucket(v):
    s = str(v).strip() if v is not None else ""
    return s if s else "Unknown"

def _safe_num_series(df, candidates, default=np.nan):
    for c in candidates:
        if c in df.columns:
            return pd.to_numeric(df[c], errors="coerce")
    return pd.Series([default] * len(df), index=df.index, dtype=float)

def _build_sector_trail_points(x_final, y_final, trail=5):
    # deterministic synthetic trail from current state (real data snapshot, chart-friendly)
    pts = []
    for i in range(trail):
        frac = (trail - 1 - i) / max(1, (trail - 1))
        x = float(np.clip(x_final * (1.0 - 0.22 * frac), -100, 100))
        y = float(np.clip(y_final * (1.0 - 0.28 * frac), -100, 100))
        pts.append((x, y))
    return pts


def build_priority_sheet(wb, latest_market_day: str, ordered_rows: list):
    """
    Priority tab — today's 8–20 highest-conviction actionable names only.

    Eligibility: _sc_priority == "YES"
    If fewer than 5 qualify, fills with next-best YES rows by score.
    Sorted by conviction score desc, then RVOL desc.
    """
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet("Priority")
    ws.sheet_view.showGridLines = False

    def _F(h): return PatternFill("solid", fgColor=h)

    GRP_HDR  = _F("7C1D1D")     # dark red — high urgency
    FILL_PRI = _F("FEF08A")     # amber highlight for priority badge
    FILL_YES = _F("D1FAE5")
    FILL_SCR = {
        (8, 10): _F("D1FAE5"),
        (7,  7): _F("DCFCE7"),
        (5,  6): _F("FEF9C4"),
        (0,  4): _F("FEE2E2"),
    }
    FONT_HDR  = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    FONT_BODY = Font(name="Calibri", size=10)
    FONT_BOLD = Font(name="Calibri", size=10, bold=True)
    THIN = Side(style="thin", color="D1D5DB")
    BDR  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def _score_fill(score):
        for (lo, hi), fill in FILL_SCR.items():
            if lo <= score <= hi:
                return fill
        return _F("FFFFFF")

    # ── Header ───────────────────────────────────────────────────────────────
    pretty = datetime.strptime(latest_market_day, "%Y-%m-%d").strftime("%B %d, %Y")
    ws.column_dimensions["A"].width = 0.5
    for r, h in {1: 4, 2: 28, 3: 14, 4: 4, 5: 22, 6: 36}.items():
        ws.row_dimensions[r].height = h

    ws["B2"] = "★  Priority Setups  ·  Today's Actionable Shortlist"
    ws["B2"].font  = Font(name="Calibri", size=14, bold=True)
    ws["B2"].alignment = Alignment(horizontal="left", vertical="center")

    ws["B3"] = (
        f"As of {pretty}  ·  "
        "Score ≥ 7  ·  Buy = YES  ·  Upside > 0%  ·  No negative target  ·  "
        "Sorted by Conviction Score ↓ then RVOL ↓"
    )
    ws["B3"].font      = Font(name="Calibri", size=10, italic=True)
    ws["B3"].alignment = Alignment(horizontal="left", vertical="center")

    # ── Column schema ─────────────────────────────────────────────────────────
    COLS = [
        # (key, header, width, align)
        ("ticker",              "Ticker",       9,  "center"),
        ("sector",              "Sector",       18, "center"),
        ("close",               "Price",        12, "center"),
        ("pct_change",          "Chg %",        8,  "center"),
        ("rvol20",              "RVOL",         8,  "center"),
        ("_sc_conviction",      "Score /10",    9,  "center"),
        ("_sc_ma_label",        "MA Position",  24, "left"),
        ("_sc_zone_type",       "Anchor",       12, "center"),
        ("_sc_entry_disp",      "Entry",        20, "left"),
        ("_sc_target_disp",     "Target",       18, "left"),
        ("_sc_target_dist",     "Upside %",     9,  "center"),
        ("_sc_invalidation_disp","Invalidation",22, "left"),
        ("ms_trend_regime",     "Regime",       14, "center"),
        ("ms_last_event",       "Last Event",   14, "center"),
        ("rsi14",               "RSI",          8,  "center"),
        ("rsi_status",          "RSI Status",   10, "center"),
        ("macd_wave",           "MACD Wave",    20, "center"),
        ("_sc_signal",          "Signals",      32, "left"),
        ("_sc_buy_signal",      "Buy",          8,  "center"),
    ]

    # ── Column headers ────────────────────────────────────────────────────────
    # Group header row 5
    ws.merge_cells(f"B5:{get_column_letter(len(COLS)+1)}5")
    hc = ws["B5"]
    hc.value = "Priority Setups — High-Conviction Execution List"
    hc.fill  = GRP_HDR
    hc.font  = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    hc.alignment = Alignment(horizontal="center", vertical="center")
    for ci in range(2, len(COLS) + 2):
        ws.cell(5, ci).fill = GRP_HDR
        ws.cell(5, ci).border = BDR

    # Sub-header row 6
    for ci, (key, header, width, align) in enumerate(COLS, start=2):
        ws.column_dimensions[get_column_letter(ci)].width = width
        c = ws.cell(6, ci)
        c.value = header
        c.fill  = GRP_HDR
        c.font  = FONT_HDR
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BDR

    # ── Select rows ───────────────────────────────────────────────────────────
    priority_rows = [r for r in ordered_rows if r.get("_sc_priority") == "YES"]
    # If too few, pad with next-best YES rows by score
    if len(priority_rows) < 5:
        extras = [
            r for r in ordered_rows
            if r.get("_sc_buy_signal") == "YES"
            and r.get("_sc_priority") != "YES"
            and safe_num(r.get("_sc_target_dist"), -1) > 0
        ]
        extras.sort(key=lambda x: (-safe_num(x.get("_sc_conviction", 0), 0),
                                    -safe_num(x.get("rvol20", 0), 0)))
        priority_rows = priority_rows + extras[: max(0, 20 - len(priority_rows))]

    # Hard cap at 30
    priority_rows = priority_rows[:30]

    if not priority_rows:
        ws.cell(8, 2).value = "No qualifying setups today. Review the IDX Screener Summary for WATCH candidates."
        ws.cell(8, 2).font  = Font(name="Calibri", size=10, italic=True, color="6B7280")
        return

    # ── Data rows ─────────────────────────────────────────────────────────────
    REGIME_FILLS = {
        "Bullish":      _F("D1FAE5"), "Bullish Weak": _F("DCFCE7"),
        "Transition Up":_F("DBEAFE"), "Sideways":     _F("F3F4F6"),
        "Range":        _F("F3F4F6"), "Transition Down":_F("FEF3C7"),
        "Bearish Weak": _F("FEE2E2"), "Bearish":      _F("FECACA"),
    }

    for r_idx, r in enumerate(priority_rows, start=7):
        is_priority = r.get("_sc_priority") == "YES"
        for ci, (key, header, width, align) in enumerate(COLS, start=2):
            c  = ws.cell(r_idx, ci)
            v  = r.get(key, "N/A")

            # Value formatting
            if key == "close":
                nv = safe_num(v, np.nan)
                c.value = nv if pd.notna(nv) else "N/A"
                if pd.notna(nv): c.number_format = "#,##0"
            elif key == "pct_change":
                nv = safe_num(v, np.nan)
                c.value = nv if pd.notna(nv) else "N/A"
                if pd.notna(nv):
                    c.number_format = "0.00%"
                    c.font = Font(name="Calibri", size=10,
                                  color="008000" if nv > 0 else ("C00000" if nv < 0 else "111827"))
            elif key in ("rvol20", "rsi14"):
                nv = safe_num(v, np.nan)
                c.value = round(float(nv), 2) if pd.notna(nv) else "N/A"
                if pd.notna(nv): c.number_format = "0.00"
            elif key == "_sc_target_dist":
                nv = safe_num(v, np.nan)
                c.value = round(float(nv), 1) if pd.notna(nv) else "N/A"
                if pd.notna(nv): c.number_format = "+0.0;-0.0"
            else:
                c.value = v if v not in (None, "", np.nan) else "N/A"
                if isinstance(v, float) and pd.isna(v):
                    c.value = "N/A"

            # Base styling
            row_bg = _F("FFFBEB") if is_priority else _F("F8FAFC")
            c.fill   = row_bg
            c.font   = FONT_BODY
            c.border = BDR
            c.alignment = Alignment(horizontal=align, vertical="center",
                                    wrap_text=(key in ("_sc_entry_disp","_sc_target_disp",
                                                        "_sc_invalidation_disp","_sc_signal",
                                                        "_sc_ma_label")))

            # Cell-level overrides
            v_str = str(c.value or "")
            if key == "_sc_conviction":
                score = int(safe_num(v, 0))
                c.fill = _score_fill(score)
                c.font = Font(name="Calibri", size=10, bold=(score >= 7),
                              color="065F46" if score >= 7 else ("713F12" if score >= 5 else "991B1B"))
            elif key == "ms_trend_regime":
                c.fill = REGIME_FILLS.get(v_str, row_bg)
            elif key == "_sc_buy_signal":
                c.fill = FILL_YES if v_str == "YES" else _F("FEF3C7")
                c.font = Font(name="Calibri", size=10, bold=True,
                              color="065F46" if v_str == "YES" else "854D0E")
            elif key == "_sc_zone_type":
                if "-2 SD" in v_str:   c.fill = _F("D1FAE5")
                elif "-1 SD" in v_str: c.fill = _F("FEF9C4")
                elif "VWAP" in v_str:  c.fill = _F("DBEAFE")

    # ── Summary line ──────────────────────────────────────────────────────────
    n_tot  = len(priority_rows)
    summary_row = 7 + n_tot + 1
    ws.cell(summary_row, 2).value = (
        f"{n_tot} total shown  ·  See IDX Screener Summary for full list"
    )
    ws.cell(summary_row, 2).font = Font(name="Calibri", size=9, italic=True, color="6B7280")
    ws.freeze_panes = "F7"


# =========================
# TICKER UNIVERSE CONFIG
# =========================
# Live/default mode = FULL KSEI universe (all valid tickers from raw source).
# Set TRUE only if you explicitly want to run the legacy custom subset.
USE_CUSTOM_TICKERS_ONLY = False

# Legacy custom subset (kept for optional use only)
CUSTOM_TICKERS = [
    "MBMA","MAPI","ANJT","DEPO","NCKL","RMKE","KIJA","BCAP","SMAR","BANK",
    "BREN","NICL","BSIM","MBAP","ASLC","DOID","PANI","EMTK","BBKP","SMGR",
    "SMBR","KPIG","AMAR","BRPT","TKIM","BMAS","DSSA","SSIA","TPIA","AMAG",
    "HOKI","ARTO","AMRT","KRAS","BBHI","AVIA","AGRO","LPCK","WIFI","NRCA",
    "PACK","DGWG","PIPA","BACA","MORA","PNLF","SAFE","TMAS","SGER","BLUE",
    "KLIN","WIRG","EMDE","VICI","HDIT","PJAA","BGTG","DMND","PORT","CYBR",
    "FUTR","AXIO","DKHH","FUJI","MBSS","ZYRX","MLPT","IPAC","INOV","CANI",
    "GPRA","SMIL","BULL","NIKL","TRST","RSCH","PALM","NICK","CAKK","ISAP",
    "PART","TAXI","ASLI","INTD","LEAD","KLAS","LUCK","INET","NICE","OLIV",
    "DOSS","AKSI","DOOH","CBRE","SOTS","DEWI","BIPP","DEFI","KOPI","MANG",
    "BAYU"
]


# =========================
# FINAL SHEET ORDER HELPER
# =========================
DESIRED_SHEET_ORDER = ["IDX Overview", "IDX Technical Detail", "IDX Fundamental Detail", "Guide & Logic Reference", "Data Processing Results"]

def _apply_final_sheet_order(wb):
    desired = [s for s in DESIRED_SHEET_ORDER if s in wb.sheetnames]
    remaining = [s for s in wb.sheetnames if s not in desired]
    wb._sheets = [wb[s] for s in desired + remaining]


def _set_guide_logic_widths(ws):
    for col in ["B", "C", "E", "F", "G"]:
        ws.column_dimensions[col].width = 40
    ws.column_dimensions["D"].width = 50


def main():
    print("\n" + "=" * 110)
    print("  IDX Screener")
    print("=" * 110)

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    os.makedirs(CACHE_DIR, exist_ok=True)

    ksei_raw = load_ksei()
    TOTAL_SOURCE_TICKERS = len(ksei_raw)

    for ticker_col in ("Ticker", "ticker", "Code", "code"):
        if ticker_col in ksei_raw.columns:
            break
    else:
        raise KeyError("Could not find ticker column in KSEI source")

    ksei_raw = ksei_raw.copy()
    ksei_raw["_ticker_norm"] = ksei_raw[ticker_col].astype(str).str.strip().str.upper()
    ksei_raw = ksei_raw[ksei_raw["_ticker_norm"].notna() & (ksei_raw["_ticker_norm"] != "")]
    ksei_raw = ksei_raw.drop_duplicates(subset=["_ticker_norm"], keep="first").reset_index(drop=True)

    use_custom = bool(globals().get("USE_CUSTOM_TICKERS_ONLY", False))
    if use_custom:
        ksei = ksei_raw[ksei_raw["_ticker_norm"].isin(CUSTOM_TICKERS)].copy()
        ksei["_ticker_order"] = ksei["_ticker_norm"].map({t: i for i, t in enumerate(CUSTOM_TICKERS)})
        ksei = ksei.sort_values("_ticker_order").drop(columns=["_ticker_norm","_ticker_order"]).reset_index(drop=True)
        print(f"[OK] KSEI: {TOTAL_SOURCE_TICKERS} tickers | Custom mode: {len(ksei)} selected")
    elif is_backtest_mode():
        # Backtest mode: restrict to BACKTEST_TICKERS universe only
        _bt_set = {t.strip().upper() for t in BACKTEST_TICKERS}
        ksei = ksei_raw[ksei_raw["_ticker_norm"].isin(_bt_set)].copy()
        ksei = ksei.drop(columns=["_ticker_norm"]).reset_index(drop=True)
        print(f"[OK] KSEI: {TOTAL_SOURCE_TICKERS} tickers | BACKTEST mode: {len(ksei)} selected (as of {MARKET_DATE})")
    else:
        ksei = ksei_raw.drop(columns=["_ticker_norm"]).reset_index(drop=True)
        print(f"[OK] KSEI: {TOTAL_SOURCE_TICKERS} tickers | Full universe (as of {MARKET_DATE})")

    shares_cache = load_shares_cache()
    results = []
    process_logs = []
    # MARKET_DATE is the authoritative as-of date for this run
    latest_market_day_global = get_market_date().strftime("%Y-%m-%d")
    _hist_cache = {}

    for idx, (_, ksei_row) in enumerate(ksei.iterrows(), start=1):
        ticker = str(ksei_row["Ticker"]).upper().strip()
        print(f"[{idx}/{len(ksei)}] {ticker} ...", end=" ")

        hist, retry_count, source_used, fetch_reason = fetch_history_with_retry(ticker)
        shares = fetch_shares_outstanding(ticker, shares_cache)

        if hist is None or hist.empty:
            built = base_row_from_ksei(ksei_row)
            results.append(built)
            process_logs.append({"ticker": ticker, "status": "NO DATA", "bars": 0,
                                  "retry_count": retry_count, "source_used": source_used,
                                  "latest_market_day": "", "reason": fetch_reason or "No market data"})
            print("NO DATA")
            continue

        built = build_row(ksei_row, hist, shares)
        results.append(built)

        if ticker in BACKTEST_TICKERS:
            _hist_cache[ticker] = hist

        reason = "OK" if built["data_status"] == "OK" else f"Insufficient bars ({len(hist)} < {MIN_BARS_FULL})"
        process_logs.append({"ticker": ticker, "status": built["data_status"], "bars": len(hist),
                              "retry_count": retry_count, "source_used": source_used,
                              "latest_market_day": built["latest_market_day"] or "", "reason": reason})
        print(built["data_status"])

    save_shares_cache(shares_cache)

    results      = sorted(results, key=lambda x: x["ticker"])
    process_logs = sorted(process_logs, key=lambda x: x["ticker"])
    summary_rows = [r for r in results if r["data_status"] == "OK"]

    # ── RS Rating universe normalisation pass ──────────────────────────────────
    if len(_RS_UNIVERSE_CACHE) >= 5:
        lo = min(_RS_UNIVERSE_CACHE)
        hi = max(_RS_UNIVERSE_CACHE)
        if hi != lo:
            for _r in results:
                raw = _r.get("_rs_raw", np.nan)
                if pd.notna(raw):
                    rating = int(max(1, min(99, round(1 + 98 * (raw - lo) / (hi - lo)))))
                    _r["rs_rating"]      = rating
                    _r["rs_rating_zone"] = rs_rating_zone(rating)

    # ── Conviction scoring pass (_sc_* fields → DETAIL_SCHEMA columns) ─────────
    # Run the V17 POI engine across all OK rows and merge _sc_* back into each row
    print("[INFO] Running conviction scoring pass ...")
    try:
        scored_board = _build_idx_vwap_shortlist(summary_rows)
        # Build lookup: ticker → scored row
        scored_map = {str(s.get("ticker","")).upper(): s for s in scored_board}
        for _r in results:
            tk = str(_r.get("ticker","")).upper()
            if tk in scored_map:
                sc = scored_map[tk]
                # Merge only _sc_* and ms_* fields
                for _k, _v in sc.items():
                    if _k.startswith("_sc_") or _k.startswith("ms_"):
                        _r.setdefault(_k, _v)   # don't overwrite if already set by build_row
    except Exception as _e:
        print(f"[WARN] Conviction scoring pass failed: {_e}")

    # ── mp_profile for screener sheet (derived from market profile fields) ──────
    for _r in results:
        if not _r.get("mp_profile"):
            _r["mp_profile"] = _market_profile_summary(_r)

    _run_dt  = datetime.now()
    out_file = os.path.join(OUTPUT_DIR, _build_output_filename(_run_dt, latest_market_day_global))

    # ── Build workbook ──────────────────────────────────────────────────────
    wb  = Workbook()
    ws1 = wb.active
    ws1.sheet_view.showGridLines = False

    _sector_source_rows = results
    build_true_idx_sector_movers_sheet(wb, latest_market_day_global, _sector_source_rows)

    # Inject RRG sector state into conviction scorer
    try:
        if "IDX Overview" in wb.sheetnames:
            _ws_ov  = wb["IDX Overview"]
            _rrg_data = []
            for _row in _ws_ov.iter_rows(min_row=10, max_row=_ws_ov.max_row, values_only=True):
                _sector_val = _row[13] if len(_row) > 13 else None
                _quad_val   = _row[17] if len(_row) > 17 else None
                if _sector_val and _quad_val and str(_quad_val) in ("Leading","Improving","Weakening","Lagging"):
                    _rrg_data.append({"Sector": str(_sector_val), "Quadrant": str(_quad_val)})
            _inject_rrg_sector_state(_rrg_data)
    except Exception:
        pass

    # Sheet: IDX Technical Detail  (renamed from IDX Screener Detail)
    ws2 = wb.create_sheet("IDX Technical Detail")
    ws2.sheet_view.showGridLines = False
    build_detail_sheet(ws2, latest_market_day_global, results)

    # Sheet: IDX Fundamental Detail  (renamed + redesigned from Fundamental Key Stats)
    print("[INFO] Building IDX Fundamental Detail sheet ...")
    build_fundamental_key_stats_sheet(wb, latest_market_day_global, summary_rows)

    # Sheet: IDX Screener (filtered signal sheet)
    print("[INFO] Building IDX Screener sheet ...")
    build_idx_screener_sheet(wb, latest_market_day_global, results)

    # Sheet: Guide & Logic Reference
    ws_guide = wb.create_sheet("Guide & Logic Reference")
    ws_guide.sheet_view.showGridLines = False
    build_guide_sheet(ws_guide)

    # Sheet: Data Processing Results
    ws_proc = wb.create_sheet("Data Processing Results")
    ws_proc.sheet_view.showGridLines = False
    build_processing_sheet(ws_proc, latest_market_day_global, len(ksei), process_logs, summary_rows)

    # ── Sheet ordering & cleanup ─────────────────────────────────────────────
    desired_order = [
        "IDX Overview",
        "IDX Screener",
        "IDX Technical Detail",
        "IDX Fundamental Detail",
        "Guide & Logic Reference",
        "Data Processing Results",
    ]
    _ALLOWED_SHEETS = set(desired_order)
    for _ws in list(wb.worksheets):
        if _ws.title not in _ALLOWED_SHEETS:
            wb.remove(_ws)

    ordered = []
    for nm in desired_order:
        if nm in wb.sheetnames:
            ordered.append(wb[nm])
    for wsx in wb.worksheets:
        if wsx not in ordered:
            ordered.append(wsx)
    wb._sheets = ordered

    _postprocess_workbook(wb)

    _postprocess_workbook(wb)
    _apply_final_sheet_order(wb)
    wb.save(out_file)
    print(f"[DONE] Saved: {out_file}")

# v6.3 PNG FINAL DIRECT PATCH
# Override ONLY IDX Overview with reliable matplotlib PNG render
# =========================
def _best_discount_zone(r: dict):
    close_px = safe_num(r.get("close"), np.nan)
    checks = [
        ("PQVWAP", r.get("pq_vwap"), 2.0, "Previous QVWAP"),
        ("PQ -1 SD", r.get("pq_m1"), 1.25, "Previous QVWAP"),
        ("PYVWAP", r.get("py_vwap"), 2.0, "Previous Year VWAP"),
        ("PY -1 SD", r.get("py_m1"), 1.25, "Previous Year VWAP"),
    ]
    best = None
    best_abs = 999.0
    for zone_type, level, tol, framework in checks:
        p = _pct_from_level(close_px, level)
        if pd.notna(p) and abs(p) <= tol and abs(p) < best_abs:
            best = {"zone_type": zone_type, "framework": framework, "dist_pct": round(p, 2), "tol_pct": tol}
            best_abs = abs(p)
    if best is None:
        for framework, z in [("Previous QVWAP", str(r.get("pq_remarks", "") or "")),
                             ("Previous Year VWAP", str(r.get("py_remarks", "") or ""))]:
            if _zone_is_tradeable_poi(z):
                z_u = z.upper()
                if "-2" in z_u or "-1" in z_u:
                    zt = "Discount Zone"
                elif "VWAP" in z_u:
                    zt = "VWAP Anchor"
                else:
                    zt = "POI Zone"
                best = {"zone_type": zt, "framework": framework, "dist_pct": np.nan, "tol_pct": np.nan}
                break
    return best

# ---- mutate detail schema in-place ----
for _grp_idx, _grp in enumerate(DETAIL_SCHEMA):
    if _grp[0] == "MACD Momentum":
        _cols = list(_grp[2])
        _new_cols = []
        for _c in _cols:
            if _c[0] in ("macd_4color", "macd_entry"):
                continue
            _new_cols.append(_c)
        DETAIL_SCHEMA[_grp_idx] = (_grp[0], _grp[1], _new_cols)
    elif _grp[0].startswith("Current QVWAP"):
        _cols = list(_grp[2]); _new=[]
        for _c in _cols:
            _new.append(_c)
            if _c[0] == "q_remarks":
                pass  # already exists in base schema
        DETAIL_SCHEMA[_grp_idx] = (_grp[0], _grp[1], _new)
    elif _grp[0].startswith("Previous QVWAP"):
        _cols = list(_grp[2]); _new=[]
        for _c in _cols:
            _new.append(_c)
            if _c[0] == "pq_remarks":
                pass  # already exists in base schema
        DETAIL_SCHEMA[_grp_idx] = (_grp[0], _grp[1], _new)
    elif _grp[0].startswith("Previous Year VWAP"):
        _cols = list(_grp[2]); _new=[]
        for _c in _cols:
            _new.append(_c)
            if _c[0] == "py_remarks":
                pass  # already exists in base schema
        DETAIL_SCHEMA[_grp_idx] = (_grp[0], _grp[1], _new)

# ---- patch process_single_stock row defaults and computations via text-level hooks below if source lines exist ----

def build_true_idx_sector_movers_sheet(wb, latest_market_day, all_market_rows):
    import tempfile, os, time
    import numpy as np
    import pandas as pd
    import yfinance as yf
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    # ── Config ────────────────────────────────────────────────────────────────
    period_weeks = int(SECTOR_MOVERS_PERIOD_WEEKS) if "SECTOR_MOVERS_PERIOD_WEEKS" in globals() else 52
    smoothing    = int(SECTOR_MOVERS_SMOOTHING)    if "SECTOR_MOVERS_SMOOTHING"    in globals() else 10
    tail         = 8
    lookback     = int(SECTOR_MOVERS_LOOKBACK)     if "SECTOR_MOVERS_LOOKBACK"     in globals() else 52

    QUAD_HEX = {
        "Leading":   "#00C853",
        "Improving": "#2F80ED",
        "Weakening": "#D4A000",
        "Lagging":   "#FF3B30",
    }
    QUAD_XL = {k: v.lstrip("#") for k, v in QUAD_HEX.items()}

    # ── Helpers ───────────────────────────────────────────────────────────────
    def _sector(row):
        ALLOWED_SECTORS = {
            "IDXFINANCE","IDXTRANS","IDXNONCYC","IDXCYCLIC","IDXINFRA",
            "IDXBASIC","IDXTECHNO","IDXPROPERT","IDXINDUST","IDXENERGY","IDXHEALTH"
        }
        ALIAS = {
            "IDXPROPERTY": "IDXPROPERT",
            "IDXPROPERTY": "IDXPROPERT",
            "PROPERTY": "IDXPROPERT",
        }
        for k in ("idx_sector", "IDX Sector", "IDXSector", "IDX_Sector", "Sector", "sector", "Sektor"):
            v = row.get(k, "") if isinstance(row, dict) else getattr(row, k, "")
            if v is None:
                continue
            s = str(v).strip().upper()
            if s in ("", "-", "UNKNOWN", "N/A"):
                continue
            s = ALIAS.get(s, s)
            if s in ALLOWED_SECTORS:
                return s
        return "Unknown"

    def _ticker(row):
        for k in ("ticker", "Ticker", "Code", "code"):
            v = row.get(k, "") if isinstance(row, dict) else getattr(row, k, "")
            if v: return str(v).upper().strip()
        return ""

    def _mcap(row):
        for k in ("mcap", "market_cap", "MarketCap"):
            try:
                v = float((row.get(k) if isinstance(row, dict) else getattr(row, k, None)) or 0)
                if v > 0: return v
            except Exception:
                pass
        return 1.0

    def _sector_weight(row):
        for k in ("idx_sector_weight", "IDX Sector Weight", "IDXSectorWeight", "IDX_Sector_Weight", "Weight"):
            try:
                v = float((row.get(k) if isinstance(row, dict) else getattr(row, k, None)) or 0)
                if v > 0:
                    return v
            except Exception:
                pass
        return np.nan

    def _get_hist(row):
        for k in ("hist", "history", "daily_hist", "price_hist", "ohlcv"):
            h = row.get(k) if isinstance(row, dict) else getattr(row, k, None)
            if isinstance(h, pd.DataFrame) and not h.empty:
                hh = normalize_history(h.copy()) if "normalize_history" in globals() else h.copy()
                if hh is not None and not hh.empty:
                    return hh
        return None

    def _fetch_yf(ticker):
        for _ in range(2):
            try:
                raw = yf.download(f"{ticker}.JK", period="2y", interval="1d",
                                  auto_adjust=True, progress=False, threads=False)
                hh = normalize_history(raw) if "normalize_history" in globals() else raw
                if hh is not None and not hh.empty:
                    return hh
            except Exception:
                pass
            time.sleep(0.3)
        return None

    def _weekly(hist):
        if hist is None or hist.empty or "Close" not in hist.columns:
            return None
        s = pd.to_numeric(hist["Close"], errors="coerce").dropna()
        if s.empty: return None
        s.index = pd.to_datetime(s.index)
        return s.resample("W-FRI").last().dropna()

    def _sector_proxy(rows, sector_name):
        members, weights = [], {}
        for row in rows:
            if _sector(row).upper() != sector_name.upper(): continue
            tk = _ticker(row)
            if not tk: continue
            hist = _get_hist(row) or _fetch_yf(tk)
            wk = _weekly(hist)
            if wk is None or wk.empty: continue
            members.append(wk.rename(tk))
            sw = _sector_weight(row)
            weights[tk] = sw if pd.notna(sw) and sw > 0 else _mcap(row)
        if not members: return None, 0
        px = pd.concat(members, axis=1, sort=False).sort_index().dropna(how="all")
        if px.empty: return None, 0
        w = pd.Series({c: weights.get(c, 1.0) for c in px.columns}).clip(lower=1e-9)
        w /= w.sum()
        wr = []
        for _, rret in px.pct_change().iterrows():
            valid = rret.dropna()
            if valid.empty:
                wr.append(np.nan)
                continue
            ww = w.reindex(valid.index).fillna(0)
            total_w = ww.sum()
            if total_w > 0: ww /= total_w
            wr.append(float((valid * ww).sum()))
        idx = (1.0 + pd.Series(wr, index=px.pct_change().index).fillna(0)).cumprod() * 100
        return idx.dropna(), len(members)

    def _bench():
        end = pd.Timestamp(latest_market_day)
        start = end - pd.Timedelta(weeks=140)
        try:
            raw = yf.download("^JKSE",
                              start=start.strftime("%Y-%m-%d"),
                              end=(end + pd.Timedelta(days=7)).strftime("%Y-%m-%d"),
                              interval="1d", auto_adjust=True,
                              progress=False, threads=False)
            hh = normalize_history(raw) if "normalize_history" in globals() else raw
            if hh is None or hh.empty: return None
            s = pd.to_numeric(hh["Close"], errors="coerce").dropna()
            s.index = pd.to_datetime(s.index)
            return s[s.index <= end].resample("W-FRI").last().dropna()
        except Exception:
            return None

    def _rrg(px, bx):
        px, bx = px.align(bx, join="inner")
        if len(px) < lookback + 2 * smoothing + 5:
            return None

        # NeoBDM-likely approximation, now displayed 0-centered
        rs_rel = safe_div_series(px, bx, default=np.nan)
        rs_rel = pd.to_numeric(rs_rel, errors="coerce").replace([np.inf, -np.inf], np.nan).dropna()
        if rs_rel.empty:
            return None

        rs_smooth = rs_rel.ewm(span=smoothing, adjust=False).mean()
        rs_base = rs_smooth.rolling(lookback).mean()
        rs_ratio_raw = 100.0 * safe_div_series(rs_smooth, rs_base, default=np.nan)

        mom_len = max(2, smoothing // 2)
        rs_ratio_smooth = rs_ratio_raw.ewm(span=max(3, mom_len), adjust=False).mean()
        rs_mom_raw = 100.0 * safe_div_series(rs_ratio_smooth, rs_ratio_smooth.shift(mom_len), default=np.nan)
        rs_mom_raw = rs_mom_raw.ewm(span=max(3, mom_len), adjust=False).mean()

        # shift center from 100 to 0 for NeoBDM-style display only
        rs_ratio = rs_ratio_raw - 100.0
        rs_mom = rs_mom_raw - 100.0

        out = pd.DataFrame({"rs_ratio": rs_ratio, "rs_mom": rs_mom}).replace([np.inf, -np.inf], np.nan).dropna()
        return out if not out.empty else None

    def _quad(x, y):
        if x >= 0 and y >= 0: return "Leading"
        if x >= 0 and y < 0:  return "Weakening"
        if x < 0  and y < 0:  return "Lagging"
        return "Improving"

    def _vix():
        """Fetch VIX data clipped to MARKET_DATE."""
        try:
            df = yf.Ticker("^VIX").history(period="1mo", interval="1d", auto_adjust=False)
            if df is None or df.empty or "Close" not in df.columns:
                return None
            s = pd.to_numeric(df["Close"], errors="coerce").dropna()
            if s.empty:
                return None
            # Clip to MARKET_DATE — same boundary as all other data
            asof = get_market_date()
            s.index = pd.to_datetime(s.index).tz_localize(None) if s.index.tz is not None else pd.to_datetime(s.index)
            s = s[s.index.normalize() <= asof]
            return float(s.iloc[-1]) if not s.empty else None
        except Exception:
            return None

    # ── Build RRG data ────────────────────────────────────────────────────────
    rows = list(all_market_rows or [])
    desired_sector_order = [
        "IDXFINANCE","IDXTRANS","IDXNONCYC","IDXCYCLIC","IDXINFRA",
        "IDXBASIC","IDXTECHNO","IDXPROPERT","IDXINDUST","IDXENERGY","IDXHEALTH"
    ]
    available = {_sector(r) for r in rows if _sector(r) not in ("Unknown", "N/A", "", "-")}
    sectors = [s for s in desired_sector_order if s in available]
    benchmark = _bench()
    rrg_data = []   # list of dicts

    for sec in sectors:
        try:
            proxy, n_mem = _sector_proxy(rows, sec)
            if proxy is None or n_mem < 1: continue
            df_rrg = _rrg(proxy, benchmark)
            if df_rrg is None or df_rrg.empty: continue
            last = df_rrg.tail(1).iloc[0]
            rr, rm = float(last["rs_ratio"]), float(last["rs_mom"])
            q = _quad(rr, rm)
            rrg_data.append({"sector": sec, "members": n_mem,
                              "rs_ratio": rr, "rs_mom": rm,
                              "quadrant": q, "df_rrg": df_rrg})
        except Exception:
            continue

    # Inject into conviction scorer
    try:
        _inject_rrg_sector_state([{"Sector": d["sector"], "Quadrant": d["quadrant"]}
                                   for d in rrg_data])
    except Exception:
        pass

    # ── Build IDX Overview sheet ──────────────────────────────────────────────
    if "IDX Overview" in wb.sheetnames:
        wb.remove(wb["IDX Overview"])
    ws = wb.create_sheet("IDX Overview", 0)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = None

    # ── Column widths ─────────────────────────────────────────────────────────
    explicit_widths = {
        "A": 2.42578125,
        "B": 23.0,
        "C": 23.0,
        "D": 13.0,
        "E": 13.0,
        "F": 13.0,
        "G": 13.0,
        "H": 13.0,
        "I": 13.0,
        "J": 13.0,
        "K": 13.0,
        "L": 22.0,
        "M": 14.0,
        "O": 22.0,
        "P": 10.0,
        "Q": 38.0,
    }
    for col, width in explicit_widths.items():
        ws.column_dimensions[col].width = width
    # Hide cols beyond Q (R onwards)
    for ci in range(18, 50):
        ltr = get_column_letter(ci)
        ws.column_dimensions[ltr].hidden = True
        ws.column_dimensions[ltr].width = 2

    # ── Row heights — VIX panel only (rows 1–11) ─────────────────────────────
    row_heights = {
        1: 21.95, 2: 21.95, 3: 20.1, 4: 18.0,
        5: 30.0,  6: 18.0,  7: 18.0, 8: 18.0,
        9: 18.0,  10: 18.0, 11: 18.0,
    }
    for r in range(1, 15):
        ws.row_dimensions[r].height = row_heights.get(r, 21.95)

    # ── Border style ─────────────────────────────────────────────────────────
    thin = Side(style="thin", color="D9DDE3")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _set(addr, val, bg=None, fc="111827", bold=False,
             sz=10.0, h="center", border=True, italic=False):
        c = ws[addr]
        c.value = val
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        c.font = Font(name="Calibri", size=sz, bold=bold,
                      italic=italic, color=f"FF{fc}")
        c.alignment = Alignment(horizontal=h, vertical="center")
        c.border = bdr if border else Border()

    # ── Row 2: Title ─────────────────────────────────────────────────────────
    _set("B2", "IDX Overview", bold=True, sz=14.0, h="left", border=False)

    # ── Row 3: Subtitle ──────────────────────────────────────────────────────
    try:
        _pretty = pd.Timestamp(latest_market_day).strftime("%B %d, %Y")
    except Exception:
        _pretty = str(latest_market_day)
    _mode_suffix = " [BACKTEST]" if is_backtest_mode() else ""
    _set("B3", f"Market Data as of {_pretty}{_mode_suffix}", sz=10.0, h="left", border=False)

    # ── Row 5: VIX panel header — B5:C5 merged ───────────────────────────────
    ws.merge_cells("B5:C5")
    _set("B5", "CBOE VIX  \u00b7  Volatility Index",
         bg="080808", fc="FFFFFF", bold=True, sz=14.0, h="center")
    ws.cell(5, 3).fill   = PatternFill("solid", fgColor="080808")
    ws.cell(5, 3).border = bdr

    # ── Row 6: Latest VIX label + value ──────────────────────────────────────
    _set("B6", "Latest VIX", bg="D1D5DB", fc="111827", bold=True, sz=10.0, h="center")
    vix_val  = _vix()
    vix_disp = f"{vix_val:.2f}" if (vix_val is not None and not pd.isna(vix_val)) else "N/A"
    _set("C6", vix_disp, bg="F9FAFB", fc="111827", bold=True, sz=12.0, h="center")

    # ── Row 7: Note label — no fill, no border, bold 9pt, left ───────────────
    ws["B7"].value = "Note:"
    ws["B7"].font  = Font(name="Calibri", size=9, bold=True, color="FF111827")
    ws["B7"].alignment = Alignment(horizontal="left", vertical="center")

    # ── Rows 8-11: Regime colour bands ───────────────────────────────────────
    bands = [
        (8,  "< 15  Low / Risk-On",        "D1FAE5", "065F46"),
        (9,  "15-20  Normal",              "DBEAFE", "1E40AF"),
        (10, "20-25  Elevated Caution",    "FEF3C7", "92400E"),
        (11, "\u2265 25  Risk-Off / Hedge","FEE2E2", "991B1B"),
    ]
    for row_n, txt, bg, fc in bands:
        c = ws.cell(row_n, 2)
        c.value     = txt
        c.fill      = PatternFill("solid", fgColor=bg)
        c.font      = Font(name="Calibri", size=8.0, bold=False, color=f"FF{fc}")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = bdr

    # ── Composite Rotation Chart section removed (rows 12–38) ────────────────
    # RRG chart, sector table, and associated PNG embed have been deleted.

    # ── Row 13: Seasonal Bias header (same style as VIX row 5) ───────────────
    ws.row_dimensions[12].height = 8    # spacer
    ws.row_dimensions[13].height = 30.0
    ws.row_dimensions[14].height = 52.0
    ws.merge_cells("B13:G13")
    _set("B13", "IHSG Seasonal Bias  ·  Monthly Cyclical Calendar",
         bg="1F1F1F", fc="FFFFFF", bold=True, sz=12.0, h="center")
    for col_i in range(3, 8):
        ws.cell(13, col_i).fill   = PatternFill("solid", fgColor="1F1F1F")
        ws.cell(13, col_i).border = bdr

    # ── Row 14: Seasonal bias text — merged B:G, wrap ─────────────────────────
    from datetime import date as _dt_date
    _seasonal_text = get_seasonal_bias(_dt_date.today().month)
    ws.merge_cells("B14:G14")
    c14 = ws["B14"]
    c14.value     = _seasonal_text
    c14.font      = Font(name="Calibri", size=10, italic=False, color="FF1F1F1F")
    c14.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c14.border    = bdr
    for col_i in range(3, 8):
        ws.cell(14, col_i).fill      = PatternFill("solid", fgColor="F9FAFB")
        ws.cell(14, col_i).border    = bdr
        ws.cell(14, col_i).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # The sheet now contains VIX panel (rows 5-11) + Seasonal Bias (rows 13-14).

# =========================
# MAIN ENTRY — moved to absolute end so final overrides are defined before runtime
# =========================


# =========================
# SMC ENGINE
# =========================

# ---- wrap sector movers to add VIX + heatmap ----
# =========================
# v7.3 FINAL OVERRIDE — IDX Overview layout order:
# 1) VIX panel
# 2) IDX Overview (existing native chart/table)
# 3) Sector Heatmap (treemap-like grouped blocks)
# =========================
_ORIG_BUILD_TRUE_IDX_SECTOR_MOVERS_SHEET = build_true_idx_sector_movers_sheet

def _fetch_latest_vix_safe():
    try:
        tk = yf.Ticker("^VIX")
        df = tk.history(period="10d", interval="1d", auto_adjust=False)
        if df is None or df.empty or "Close" not in df.columns:
            return None
        s = pd.to_numeric(df["Close"], errors="coerce").dropna()
        if s.empty:
            return None
        return float(s.iloc[-1])
    except Exception:
        return None

def _vix_regime(v):
    if v is None or pd.isna(v):
        return "N/A"
    if v < 15:
        return "Low Vol / Risk-On"
    if v < 20:
        return "Normal"
    if v < 25:
        return "Elevated"
    return "Risk-Off"

def _build_sector_heatmap_png(all_market_rows):
    return None
    """
    TradingView-style stock heatmap.
    - Pure black background
    - Sector headers with count
    - Ticker + % change text per cell
    - Size = market cap (big names dominate, FCA/illiquid names shrink)
    - Color = daily % change with proper saturation curve
    - Squarified treemap within each sector block
    """
    import tempfile
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import matplotlib.patches as mpatches
        from matplotlib.font_manager import FontProperties
    except ImportError:
        return None

    try:
        df = pd.DataFrame(all_market_rows if all_market_rows else [])
        if df.empty:
            return None

        # ── Normalise columns ──────────────────────────────────────────────
        for alias in ("pct_chg", "pct_ch", "change_pct"):
            if "pct_change" not in df.columns and alias in df.columns:
                df["pct_change"] = df[alias]
        if "pct_change" not in df.columns:
            return None

        # Sector name
        for sc in ("sector",):
            if sc not in df.columns:
                df[sc] = "Other"

        # Ticker
        if "ticker" not in df.columns:
            df["ticker"] = ""

        # Market cap (size proxy)
        for sc in ("mcap", "market_cap", "free_float_mcap"):
            if sc in df.columns:
                df["_mcap"] = pd.to_numeric(df[sc], errors="coerce")
                if df["_mcap"].notna().any():
                    break
        else:
            df["_mcap"] = np.nan

        # Liquidity proxy
        df["_dval"] = np.nan
        for lc in ("daily_value", "value_traded", "adtv20"):
            if lc in df.columns:
                v_ = pd.to_numeric(df[lc], errors="coerce")
                df["_dval"] = df["_dval"].combine_first(v_)

        df["_rvol"] = pd.to_numeric(df.get("rvol20", np.nan), errors="coerce")

        # Convert pct_change fraction → percent
        df["pct_change"] = pd.to_numeric(df["pct_change"], errors="coerce")
        df["pct_pct"] = df["pct_change"] * 100.0

        df["sector"] = (df["sector"].fillna("Other").astype(str)
                        .str.strip().replace("", "Other"))
        df["ticker"] = df["ticker"].fillna("").astype(str)

        # Drop rows with no pct_change (not traded today)
        df = df.dropna(subset=["pct_pct"]).reset_index(drop=True)
        if df.empty:
            return None

        # ── Fill missing sizes ─────────────────────────────────────────────
        if df["_mcap"].notna().sum() == 0:
            df["_mcap"] = 1.0
        else:
            mcap_med = df["_mcap"].dropna().median()
            df["_mcap"] = df["_mcap"].fillna(mcap_med if pd.notna(mcap_med) else 1.0)
        df["_mcap"] = df["_mcap"].clip(lower=1.0)

        # ── Display size scoring ───────────────────────────────────────────
        # Principle: big-cap + liquid = prominent.
        # FCA stocks, zero-change illiquid = tiny.
        p80 = df["_mcap"].quantile(0.80)
        p50 = df["_mcap"].quantile(0.50)

        liq_ok  = (df["_dval"].fillna(0) >= 3_000_000_000) | (df["_rvol"].fillna(0) >= 1.0)
        strong_liq = (df["_dval"].fillna(0) >= 10_000_000_000) | (df["_rvol"].fillna(0) >= 1.5)
        bigcap  = df["_mcap"] >= p80
        flat    = df["pct_pct"].abs() < 0.05

        # Check if FCA board (when field available)
        board_f = df["board"].astype(str).str.upper() if "board" in df.columns else pd.Series([""] * len(df))
        is_fca  = board_f.str.contains("FCA", na=False)

        def _display_size(idx):
            mc     = float(df.loc[idx, "_mcap"])
            is_bc  = bool(bigcap.iloc[idx])
            is_liq = bool(liq_ok.iloc[idx])
            is_flat= bool(flat.iloc[idx])
            fca    = bool(is_fca.iloc[idx])
            pct    = abs(float(df.loc[idx, "pct_pct"]))

            if fca:
                return max(mc * 0.06, p50 * 0.05)           # FCA = tiny footprint
            if is_bc and is_liq:
                # Big cap + liquid: keep prominent, but de-emphasize flat names unless they are truly liquid leaders
                if is_flat:
                    return mc * (0.80 if bool(strong_liq.iloc[idx]) else 0.35)
                return mc * 1.0
            if is_bc and not is_liq:
                # Big cap but illiquid (controlled float)
                return max(mc * 0.20, p50 * 0.10)
            # Normal stock: scale by movement (market activity = relevance)
            move_scale = min(1.0, max(0.10, pct / 3.0))
            if is_flat and not is_liq:
                return max(mc * 0.08, 1.0)                  # inactive small-cap = near-invisible
            return mc * (0.30 + 0.70 * move_scale)

        df["_size"] = [_display_size(i) for i in range(len(df))]
        df["_size"] = df["_size"].clip(lower=1.0)

        # ── TradingView colour palette ─────────────────────────────────────
        # Deep green at +8%, deep red at -8%, neutral grey at 0%
        # Uses the exact same breakpoints as TradingView screenshot
        def pct_to_color(v):
            if pd.isna(v) or abs(v) < 0.10:
                return "#1C2532"          # near-zero: very dark navy (TV neutral)
            if v > 0:
                # +0.1% → light green ... +8%+ → deep forest green
                t = min(1.0, v / 8.0)
                r_ = int(10  + 5  * (1 - t))
                g_ = int(80  + 90 * t)
                b_ = int(30  + 10 * (1 - t))
                return f"#{max(0,r_):02x}{min(255,g_):02x}{max(0,b_):02x}"
            else:
                # -0.1% → light red ... -8%+ → deep crimson
                t = min(1.0, abs(v) / 8.0)
                r_ = int(90 + 110 * t)
                g_ = int(20  - 10 * t)
                b_ = int(20  - 10 * t)
                return f"#{min(255,r_):02x}{max(0,g_):02x}{max(0,b_):02x}"

        def text_color(v):
            """White for saturated cells, lighter for near-zero."""
            if pd.isna(v) or abs(v) < 0.10:
                return "#6B7280"  # muted grey for flat stocks
            return "#FFFFFF"

        # ── Sector ordering: by total display size desc ────────────────────
        sector_order = (
            df.groupby("sector")["_size"].sum()
              .sort_values(ascending=False)
              .index.tolist()
        )

        # ── Squarified treemap ─────────────────────────────────────────────
        def _squarify(items, x0, y0, w, h):
            """
            Standard squarified treemap algorithm.
            items = [(label, size, pct, ticker), ...]
            Returns [(label, size, pct, ticker, rx, ry, rw, rh), ...]
            """
            if not items:
                return []
            total = sum(s for _, s, _, _ in items) or 1.0
            result = []
            remaining = list(items)

            def worst_ratio(row, dim):
                if not row: return float("inf")
                s = sum(sz for _, sz, _, _ in row) / total
                if s == 0 or dim == 0: return float("inf")
                rects = [(sz / total) / s * dim for _, sz, _, _ in row]
                rects = [r_ for r_ in rects if r_ > 0]
                if not rects or min(rects) == 0: return float("inf")
                return max(max(rects) / min(rects), min(rects) / max(rects))

            cx, cy, cw, ch = x0, y0, w, h
            row_buf = []

            while remaining:
                item = remaining[0]
                horiz = cw >= ch
                dim = ch if horiz else cw
                test = row_buf + [item]

                if row_buf and worst_ratio(test, dim) > worst_ratio(row_buf, dim):
                    # Commit current row
                    rs = sum(sz for _, sz, _, _ in row_buf) / total
                    stripe = rs * (cw if horiz else ch)
                    off = 0.0
                    for lab, sz, pp, tk in row_buf:
                        frac = (sz / total) / rs if rs > 0 else 0
                        if horiz:
                            result.append((lab, sz, pp, tk, cx, cy + off, stripe, ch * frac))
                            off += ch * frac
                        else:
                            result.append((lab, sz, pp, tk, cx + off, cy, cw * frac, stripe))
                            off += cw * frac
                    if horiz:
                        cx += stripe; cw -= stripe
                    else:
                        cy += stripe; ch -= stripe
                    row_buf = []
                else:
                    row_buf.append(item)
                    remaining.pop(0)

            # Flush last row
            if row_buf:
                rs = sum(sz for _, sz, _, _ in row_buf) / total
                stripe = rs * (cw if cw >= ch else ch)
                horiz = cw >= ch
                off = 0.0
                for lab, sz, pp, tk in row_buf:
                    frac = (sz / total) / rs if rs > 0 else 0
                    if horiz:
                        result.append((lab, sz, pp, tk, cx, cy + off, cw, ch * frac))
                        off += ch * frac
                    else:
                        result.append((lab, sz, pp, tk, cx + off, cy, cw * frac, ch))
                        off += cw * frac
            return result

        # ── Figure: TradingView dark layout ───────────────────────────────
        BG        = "#131722"   # TV pure black background
        SEC_BG    = "#1A1F2E"   # sector block background
        SEC_HDR   = "#0F1520"   # sector header strip background
        BORDER_C  = "#131722"   # border = same as background (seamless)
        HDR_TXT   = "#9BA3AF"   # sector label text (muted)

        fig = plt.figure(figsize=(24, 12), facecolor=BG, dpi=130)
        ax  = fig.add_axes([0.0, 0.03, 1.0, 0.97])   # leave 3% for legend
        ax.set_xlim(0, 1); ax.set_ylim(0, 1)
        ax.axis("off"); ax.set_facecolor(BG)

        # ── Sector layout — rows of sectors, proportional width ────────────
        sector_sizes = [
            (sec, float(df.loc[df["sector"] == sec, "_size"].sum()))
            for sec in sector_order
        ]
        total_mkt = sum(s for _, s in sector_sizes) or 1.0

        # Pack sectors into rows targeting ~33% height each
        target_row = 0.30
        layout_rows, cur_row, cur_acc = [], [], 0.0
        for sec, s in sector_sizes:
            frac = s / total_mkt
            cur_row.append((sec, s))
            cur_acc += frac
            if cur_acc >= target_row or len(cur_row) >= 6:
                layout_rows.append(cur_row)
                cur_row, cur_acc = [], 0.0
        if cur_row:
            layout_rows.append(cur_row)

        nrows = len(layout_rows)
        row_h = 1.0 / nrows

        top = 1.0
        for row_secs in layout_rows:
            row_total = sum(s for _, s in row_secs) or 1.0
            left = 0.0
            for sec, sec_sz in row_secs:
                sw = sec_sz / row_total
                sx, sy = left, top - row_h
                pad = 0.0015

                # Sector block background
                ax.add_patch(mpatches.FancyBboxPatch(
                    (sx + pad, sy + pad), sw - 2*pad, row_h - 2*pad,
                    boxstyle="square,pad=0", linewidth=0,
                    facecolor=SEC_BG, zorder=1))

                # Sector header strip (top of sector block)
                hdr_h = min(0.028, row_h * 0.09)
                ax.add_patch(mpatches.Rectangle(
                    (sx + pad, sy + row_h - hdr_h - pad),
                    sw - 2*pad, hdr_h,
                    linewidth=0, facecolor=SEC_HDR, zorder=2))

                # Sector label
                count = int((df["sector"] == sec).sum())
                # Truncate long names
                sec_display = sec if len(sec) <= 22 else sec[:20] + "…"
                ax.text(sx + pad + 0.004, sy + row_h - hdr_h / 2 - pad,
                        f"{sec_display}  ›  {count}",
                        va="center", ha="left",
                        fontsize=7.0, color=HDR_TXT,
                        fontweight="bold", zorder=3)

                # Stock cells area
                ix, iy = sx + pad, sy + pad
                iw = sw - 2*pad
                ih = row_h - hdr_h - 2*pad
                if iw < 0.004 or ih < 0.004:
                    left += sw; continue

                sub = (df[df["sector"] == sec]
                       .sort_values("_size", ascending=False)
                       .reset_index(drop=True))
                if sub.empty:
                    left += sw; continue

                items = [
                    (r_["ticker"], float(r_["_size"]), float(r_["pct_pct"]), r_["ticker"])
                    for _, r_ in sub.iterrows()
                ]

                rects = _squarify(items, ix, iy, iw, ih)
                for _lab, _sz, pp, tk, rx, ry, rw_, rh_ in rects:
                    rw_ = max(0, rw_)
                    rh_ = max(0, rh_)
                    if rw_ < 5e-4 or rh_ < 5e-4:
                        continue

                    bg_col = pct_to_color(pp)
                    ax.add_patch(mpatches.Rectangle(
                        (rx + 0.0005, ry + 0.0005),
                        max(0, rw_ - 0.001),
                        max(0, rh_ - 0.001),
                        linewidth=0, facecolor=bg_col, zorder=2))

                    area = rw_ * rh_
                    txt_col = text_color(pp)
                    sign = "+" if pp >= 0 else ""

                    if area > 0.0060:
                        # Large cell: ticker bold + pct below
                        ax.text(rx + rw_/2, ry + rh_/2 + rh_*0.09,
                                tk, ha="center", va="center",
                                fontsize=min(10, max(6, rw_ * 65)),
                                color=txt_col, fontweight="bold", zorder=3, clip_on=True)
                        ax.text(rx + rw_/2, ry + rh_/2 - rh_*0.09,
                                f"{sign}{pp:.2f}%", ha="center", va="center",
                                fontsize=min(8, max(5, rw_ * 50)),
                                color=txt_col, fontweight="normal", zorder=3, clip_on=True)
                    elif area > 0.0018:
                        # Medium cell: ticker + pct on one line
                        ax.text(rx + rw_/2, ry + rh_/2,
                                f"{tk}  {sign}{pp:.1f}%",
                                ha="center", va="center",
                                fontsize=min(7, max(5, rw_ * 55)),
                                color=txt_col, fontweight="bold", zorder=3, clip_on=True)
                    elif area > 0.0005:
                        # Small cell: ticker only
                        ax.text(rx + rw_/2, ry + rh_/2, tk,
                                ha="center", va="center",
                                fontsize=min(6, max(4.5, rw_ * 50)),
                                color=txt_col, fontweight="bold", zorder=3, clip_on=True)
                    # Very small cells: just colour, no text

                left += sw
            top -= row_h

        # ── Legend bar at bottom ───────────────────────────────────────────
        legend_items = [
            ("#0a4d1a", "≥ +8%"), ("#147a2c", "+5%"),
            ("#1e8f38", "+2%"),  ("#1C2532", "0%"),
            ("#6e1a1a", "−2%"),  ("#a82020", "−5%"), ("#d42a2a", "≤ −8%"),
        ]
        lw = 0.09
        lx = 0.02
        for bg_, lbl in legend_items:
            ax.add_patch(mpatches.Rectangle((lx, 0.003), lw, 0.022,
                         facecolor=bg_, linewidth=0, zorder=4))
            ax.text(lx + lw/2, 0.014, lbl, ha="center", va="center",
                    fontsize=6.5, color="white", fontweight="bold", zorder=5)
            lx += lw + 0.003

        # Watermark
        ax.text(0.99, 0.014, "Size = Market Cap  ·  Colour = Daily % Change",
                ha="right", va="center", fontsize=6, color="#4B5563", zorder=5)

        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
        tmp.close()
        fig.savefig(tmp.name, dpi=130, bbox_inches="tight",
                    facecolor=fig.get_facecolor())
        plt.close(fig)
        return tmp.name

    except Exception as e:
        try:
            import traceback as _tb
            print(f"[Heatmap] error: {e}")
            _tb.print_exc()
        except Exception:
            pass
        return None

def _postprocess_workbook(wb):
    from openpyxl.cell.cell import MergedCell

    # Remove freeze panes on IDX Overview
    if "IDX Overview" in wb.sheetnames:
        wb["IDX Overview"].freeze_panes = None

    # ── Ghost column cleanup ──────────────────────────────────────────────────
    # Deletes truly empty trailing columns so Ctrl+End lands on the real last
    # used cell, file size stays clean, and extract-text has no N/A noise.
    def _trim_ghost_columns(ws):
        """Delete any trailing columns where every cell in every data row is empty."""
        from openpyxl.cell.cell import MergedCell
        max_col = ws.max_column
        if max_col is None or max_col == 0:
            return
        # Walk backwards from the last column; stop at first non-empty column
        last_real_col = max_col
        for col in range(max_col, 0, -1):
            col_has_data = False
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                     min_col=col, max_col=col):
                for cell in row:
                    if isinstance(cell, MergedCell):
                        continue
                    if cell.value is not None and str(cell.value).strip() not in ("", "N/A"):
                        col_has_data = True
                        break
                if col_has_data:
                    break
            if col_has_data:
                last_real_col = col
                break
        # Delete everything beyond the real last column
        cols_to_delete = max_col - last_real_col
        if cols_to_delete > 0:
            ws.delete_cols(last_real_col + 1, cols_to_delete)

    for sheet_name in ["IDX Technical Detail"]:
        if sheet_name in wb.sheetnames:
            _trim_ghost_columns(wb[sheet_name])

    # Normalize blank-like cells to N/A in data regions (skip spacer/header rows and merged cells)
    target_sheets = [n for n in ["IDX Technical Detail"] if n in wb.sheetnames]
    for sn in target_sheets:
        ws = wb[sn]
        data_start = 7 if sn in ("IDX Technical Detail",) else 4
        for row in ws.iter_rows(min_row=data_start):
            for c in row:
                if isinstance(c, MergedCell):
                    continue
                if c.value in (None, ""):
                    c.value = "N/A"

    # Explicitly preserve spacer row 4 as blank
    for sn in target_sheets:
        ws = wb[sn]
        for c in range(2, ws.max_column + 1):
            if not isinstance(ws.cell(4, c), MergedCell):
                ws.cell(4, c).value = None

    # Enforce requested freeze panes
    if "IDX Technical Detail" in wb.sheetnames:
        wb["IDX Technical Detail"].freeze_panes = "E7"

    # Round RSI 14 columns to 2 decimals by header match
    for sn in target_sheets:
        ws = wb[sn]
        header_row = None
        for r in range(1, min(ws.max_row, 10) + 1):
            vals = []
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(r, col)
                vals.append(cell.value)
            if "RSI 14" in vals:
                header_row = r
                break
        if header_row:
            for col in range(1, ws.max_column + 1):
                if ws.cell(header_row, col).value == "RSI 14":
                    for rr in range(header_row + 1, ws.max_row + 1):
                        cell = ws.cell(rr, col)
                        if isinstance(cell, MergedCell):
                            continue
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = "0.00"


# =========================
# FINAL OVERRIDE — IDX Overview exact match to IDX_Overview.xlsx
# Verified cell-by-cell from source file audit.
# Changes vs previous version:
#   - matplotlib title removed from image (no ax.set_title)
#   - Period/Smoothing/Tail shown as text annotation INSIDE chart area (top-left)
#   - Col N stays at Excel default width (not explicitly set)
#   - Col Q = 38.0 explicitly set (matches source)
#   - All col/row dims exactly match audit
#   - No extra sheets, no hidden helper sheets
# =========================
_ORIG_BUILD_TRUE_IDX_SECTOR_MOVERS_SHEET_CLEANBASE = build_true_idx_sector_movers_sheet

def _fmt_range(a, b):
    try:
        if a is None or b is None or pd.isna(a) or pd.isna(b):
            return ""
        lo = int(round(min(float(a), float(b))))
        hi = int(round(max(float(a), float(b))))
        return f"{lo} - {hi}"
    except Exception:
        return ""

def _lux_leg(highs_arr, lows_arr, i, size):
    if i < size:
        return None
    pivot_idx = i - size
    if pivot_idx < 0:
        return None
    recent_highs = highs_arr[pivot_idx + 1 : i + 1]
    recent_lows  = lows_arr[pivot_idx + 1 : i + 1]
    if len(recent_highs) < size or len(recent_lows) < size:
        return None
    new_leg_high = highs_arr[pivot_idx] > np.max(recent_highs)
    new_leg_low  = lows_arr[pivot_idx]  < np.min(recent_lows)
    if new_leg_high:
        return 0  # BEARISH_LEG
    elif new_leg_low:
        return 1  # BULLISH_LEG
    return None


# =========================
# SMC ENGINE  — Pine Script "SMC" v6 faithful Python port
# Implements: Internal/Swing BOS+CHoCH, Strong/Weak H/L, Order Blocks,
# FVGs, EQH/EQL, Premium/Discount, Ribbon (EMA25/50), CurrentQVWAP,
# PY/PQ VWAP bands, IBH/IBL/MDH/MDL/PWH/PWL (already in build_row).
# All outputs are lookahead-free and replay-safe.
# =========================

def _lux_leg(highs_arr, lows_arr, i, size):
    '''
    Pine Script leg(int size):
      BEARISH_LEG (0) when high[size] > ta.highest(size)  -> new pivot high
      BULLISH_LEG (1) when low[size]  < ta.lowest(size)   -> new pivot low
    We replicate with strict forward-safe window.
    '''
    if i < size:
        return None
    anchor_i = i - size          # the pivot candidate bar
    # Look at the size bars AFTER anchor (not including anchor itself = bars anchor+1..i)
    window_highs = highs_arr[anchor_i + 1 : i + 1]
    window_lows  = lows_arr [anchor_i + 1 : i + 1]
    if len(window_highs) == 0:
        return None
    if highs_arr[anchor_i] > (window_highs.max() if len(window_highs) else highs_arr[anchor_i]):
        return 0   # BEARISH_LEG: pivot high at anchor_i
    if lows_arr[anchor_i]  < (window_lows.min()  if len(window_lows)  else lows_arr[anchor_i]):
        return 1   # BULLISH_LEG: pivot low at anchor_i
    return None


def _fmt_range(a, b):
    try:
        if a is None or b is None or pd.isna(a) or pd.isna(b):
            return ""
        lo = int(round(min(float(a), float(b))))
        hi = int(round(max(float(a), float(b))))
        return f"{lo} - {hi}"
    except Exception:
        return ""


def compute_smc_engine(hist: pd.DataFrame,
                       internal_size: int = 5,
                       swing_size: int = 50,
                       eq_length: int = 3,
                       eq_threshold: float = 0.1,
                       mitigation_mode: str = "close") -> dict:
    '''
    Full SMC engine output. Returns dict with 26+ columns.
    All indexes are bar-by-bar; no lookahead into future bars.
    '''
    empty = {
        "smc_internal_trend":         "N/A",
        "smc_swing_trend":            "N/A",
        "smc_internal_bos_count":     0,
        "smc_internal_choch_count":   0,
        "smc_swing_bos_count":        0,
        "smc_swing_choch_count":      0,
        "smc_latest_internal_struct": "N/A",
        "smc_latest_swing_struct":    "N/A",
        "smc_strong_high":            np.nan,
        "smc_weak_high":              np.nan,
        "smc_strong_low":             np.nan,
        "smc_weak_low":               np.nan,
        "smc_closest_ob":             "",
        "smc_ob_equilibrium":         np.nan,
        "smc_closest_ob_bear":        "",
        "smc_ob_direction":           "N/A",
        "smc_ob_dist_pct":            np.nan,
        "smc_closest_fvg":            "",
        "smc_fvg_bias":               "N/A",
        "smc_eqhl_status":            "N/A",
        "smc_premium_discount":       "N/A",
        "smc_ribbon_bias":            "N/A",
        "smc_current_qvwap":          np.nan,
        "smc_qvwap_dist_pct":         np.nan,
        "smc_nearest_pyq_band":       "",
        "smc_pyq_band_dist_pct":      np.nan,
        "smc_market_structure_bias":  "N/A",
        "smc_composite_state":        "N/A",
        # Legacy compatibility fields (old schema)
        "smc_bull_internal_ob":       "",
        "smc_bear_internal_ob":       "",
        "smc_discount":               "",
        "smc_equilibrium":            "",
        "smc_premium":                "",
        "smc_state":                  "N/A",
        "ms_trend_bias":              "N/A",
        "ms_structure_phase":         "N/A",
        "ms_last_structural_event":   "N/A",
        "ms_event_age_d":             "",
        "ms_swing_sequence":          "N/A",
        "ms_structure_quality":       "N/A",
        "ms_volume_confirmation":     "N/A",
    }

    if hist is None or hist.empty:
        return empty

    df = hist.copy()
    for c in ["Open", "High", "Low", "Close", "Volume"]:
        if c not in df.columns:
            return empty
    df = df[["Open", "High", "Low", "Close", "Volume"]].dropna()
    min_bars = max(60, swing_size + 5)
    if len(df) < min_bars:
        return empty

    highs  = df["High"].astype(float).values
    lows   = df["Low"].astype(float).values
    opens  = df["Open"].astype(float).values
    closes = df["Close"].astype(float).values
    vols   = df["Volume"].astype(float).values
    n      = len(df)

    # ATR-based filter for OB and EQH/EQL
    tr = np.maximum(highs - lows,
         np.maximum(np.abs(highs - np.roll(closes, 1)),
                    np.abs(lows  - np.roll(closes, 1))))
    tr[0] = highs[0] - lows[0]
    atr200 = pd.Series(tr).rolling(200, min_periods=1).mean().values

    # Parsed highs/lows (high-volatility bar inversion for OB detection)
    high_vol_bar  = (highs - lows) >= (2.0 * atr200)
    parsed_highs  = np.where(high_vol_bar, lows,  highs)
    parsed_lows   = np.where(high_vol_bar, highs, lows)

    # ── State variables ──────────────────────────────────────────────────────
    # Pivots (Pine: pivot struct with currentLevel, lastLevel, crossed, barIndex)
    iH = dict(level=np.nan, last=np.nan, crossed=False, idx=None)  # internal high
    iL = dict(level=np.nan, last=np.nan, crossed=False, idx=None)  # internal low
    sH = dict(level=np.nan, last=np.nan, crossed=False, idx=None)  # swing high
    sL = dict(level=np.nan, last=np.nan, crossed=False, idx=None)  # swing low
    eH = dict(level=np.nan, last=np.nan, crossed=False, idx=None)  # equal high
    eL = dict(level=np.nan, last=np.nan, crossed=False, idx=None)  # equal low

    internal_trend = 0   # +1 bullish, -1 bearish, 0 neutral
    swing_trend    = 0

    prev_leg_int   = None
    prev_leg_sw    = None

    # Trailing extremes for premium/discount zones
    trailing_top     = np.nan
    trailing_bottom  = np.nan
    trailing_top_idx = None
    trailing_bot_idx = None

    # Order blocks list: {barHigh, barLow, bias, idx}
    internal_obs = []
    swing_obs    = []

    # FVGs: {top, bottom, bias, formed_idx}
    fvgs = []

    # Counters
    int_bos_count   = 0
    int_choch_count = 0
    sw_bos_count    = 0
    sw_choch_count  = 0

    last_int_event  = ""
    last_sw_event   = ""
    last_event_idx  = None

    eqh_detected = False
    eql_detected = False

    # ── Helper: store order block on BOS/CHoCH ───────────────────────────────
    def _store_ob(pivot_obj, bias, cur_i, internal=True):
        p_idx = pivot_obj.get("idx")
        if p_idx is None or p_idx >= cur_i:
            return
        if bias == 1:   # bullish -> find min parsed low in window
            window = parsed_lows[p_idx:cur_i]
            if len(window) == 0:
                return
            rel = int(np.argmin(window))
        else:           # bearish -> find max parsed high in window
            window = parsed_highs[p_idx:cur_i]
            if len(window) == 0:
                return
            rel = int(np.argmax(window))
        idx = p_idx + rel
        ob  = {"barHigh": float(parsed_highs[idx]),
               "barLow":  float(parsed_lows[idx]),
               "bias":    bias,
               "idx":     idx}
        if internal:
            internal_obs.insert(0, ob)
            if len(internal_obs) > 100:
                internal_obs[:] = internal_obs[:100]
        else:
            swing_obs.insert(0, ob)
            if len(swing_obs) > 100:
                swing_obs[:] = swing_obs[:100]

    # ── Helper: mitigate order blocks ─────────────────────────────────────────
    def _mitigate_obs(obs_list, i):
        _mm = str(mitigation_mode).lower()
        surviving = []
        for ob in obs_list:
            crossed = False
            if ob["bias"] == -1:  # bearish OB
                crossed = (closes[i] > ob["barHigh"]) if _mm == "close" else (highs[i] > ob["barHigh"])
            elif ob["bias"] == 1:  # bullish OB
                crossed = (closes[i] < ob["barLow"])  if _mm == "close" else (lows[i]  < ob["barLow"])
            if not crossed:
                surviving.append(ob)
        return surviving

    # ── Main bar-by-bar loop ─────────────────────────────────────────────────
    for i in range(n):
        c = closes[i]
        h = highs[i]
        l = lows[i]

        # Trailing extremes
        if np.isnan(trailing_top) or h > trailing_top:
            trailing_top = h; trailing_top_idx = i
        if np.isnan(trailing_bottom) or l < trailing_bottom:
            trailing_bottom = l; trailing_bot_idx = i

        # Swing structure pivots (size = swing_size)
        leg_sw = _lux_leg(highs, lows, i, swing_size)
        if leg_sw is not None and prev_leg_sw is not None and leg_sw != prev_leg_sw:
            pi = i - swing_size
            if leg_sw == 1:   # new bullish leg => pivot low at pi
                sL.update(last=sL["level"], level=float(lows[pi]),
                          crossed=False, idx=pi)
                trailing_bottom = float(lows[pi]); trailing_bot_idx = pi
            else:             # new bearish leg => pivot high at pi
                sH.update(last=sH["level"], level=float(highs[pi]),
                          crossed=False, idx=pi)
                trailing_top = float(highs[pi]); trailing_top_idx = pi
        if leg_sw is not None:
            prev_leg_sw = leg_sw

        # Internal structure pivots (size = internal_size)
        leg_in = _lux_leg(highs, lows, i, internal_size)
        if leg_in is not None and prev_leg_int is not None and leg_in != prev_leg_int:
            pi = i - internal_size
            if leg_in == 1:
                iL.update(last=iL["level"], level=float(lows[pi]),
                          crossed=False, idx=pi)
            else:
                iH.update(last=iH["level"], level=float(highs[pi]),
                          crossed=False, idx=pi)
        if leg_in is not None:
            prev_leg_int = leg_in

        # EQH / EQL detection (Pine: equalHighsLowsLength = eq_length)
        leg_eq = _lux_leg(highs, lows, i, eq_length)
        if leg_eq is not None and leg_eq != prev_leg_int:
            pi = i - eq_length
            if leg_eq == 1:  # pivot low
                if not np.isnan(eL["level"]) and abs(float(lows[pi]) - eL["level"]) < eq_threshold * atr200[i]:
                    eql_detected = True
                eL.update(last=eL["level"], level=float(lows[pi]),
                          crossed=False, idx=pi)
            else:            # pivot high
                if not np.isnan(eH["level"]) and abs(float(highs[pi]) - eH["level"]) < eq_threshold * atr200[i]:
                    eqh_detected = True
                eH.update(last=eH["level"], level=float(highs[pi]),
                          crossed=False, idx=pi)

        # ── Internal structure BOS / CHoCH (displayStructure(true)) ──────────
        # Bullish break: price crosses above internal pivot high
        iH_extra = (iH["idx"] is not None and
                    not np.isnan(iH["level"]) and
                    not np.isnan(sH.get("level", np.nan)) and
                    iH["level"] != sH.get("level", np.nan))
        if (not np.isnan(iH["level"]) and c > iH["level"] and
                not iH["crossed"] and iH_extra):
            tag = "CHoCH" if internal_trend == -1 else "BOS"
            iH["crossed"] = True
            internal_trend = 1
            if tag == "BOS":
                int_bos_count += 1
            else:
                int_choch_count += 1
            last_int_event = f"Internal Bullish {tag}"
            last_event_idx = i
            _store_ob(iH, 1, i, internal=True)

        # Bearish break: price crosses below internal pivot low
        iL_extra = (iL["idx"] is not None and
                    not np.isnan(iL["level"]) and
                    not np.isnan(sL.get("level", np.nan)) and
                    iL["level"] != sL.get("level", np.nan))
        if (not np.isnan(iL["level"]) and c < iL["level"] and
                not iL["crossed"] and iL_extra):
            tag = "CHoCH" if internal_trend == 1 else "BOS"
            iL["crossed"] = True
            internal_trend = -1
            if tag == "BOS":
                int_bos_count += 1
            else:
                int_choch_count += 1
            last_int_event = f"Internal Bearish {tag}"
            last_event_idx = i
            _store_ob(iL, -1, i, internal=True)

        # ── Swing structure BOS / CHoCH (displayStructure()) ─────────────────
        if not np.isnan(sH["level"]) and c > sH["level"] and not sH["crossed"]:
            tag = "CHoCH" if swing_trend == -1 else "BOS"
            sH["crossed"] = True
            swing_trend = 1
            if tag == "BOS":
                sw_bos_count += 1
            else:
                sw_choch_count += 1
            last_sw_event = f"Swing Bullish {tag}"
            last_event_idx = i
            _store_ob(sH, 1, i, internal=False)

        if not np.isnan(sL["level"]) and c < sL["level"] and not sL["crossed"]:
            tag = "CHoCH" if swing_trend == 1 else "BOS"
            sL["crossed"] = True
            swing_trend = -1
            if tag == "BOS":
                sw_bos_count += 1
            else:
                sw_choch_count += 1
            last_sw_event = f"Swing Bearish {tag}"
            last_event_idx = i
            _store_ob(sL, -1, i, internal=False)

        # ── Fair Value Gap detection ──────────────────────────────────────────
        # Bullish FVG: low[i] > high[i-2] (gap between prev prev high and cur low)
        if i >= 2:
            if lows[i] > highs[i - 2]:       # bullish FVG
                fvgs.append({"top": lows[i], "bottom": highs[i - 2], "bias": 1, "formed": i})
            elif highs[i] < lows[i - 2]:     # bearish FVG
                fvgs.append({"top": lows[i - 2], "bottom": highs[i], "bias": -1, "formed": i})

        # Mitigate filled FVGs
        surviving_fvg = []
        for fvg in fvgs:
            filled = (fvg["bias"] == 1  and l < fvg["bottom"]) or \
                     (fvg["bias"] == -1 and h > fvg["top"])
            if not filled:
                surviving_fvg.append(fvg)
        fvgs[:] = surviving_fvg

        # Mitigate crossed order blocks
        internal_obs[:] = _mitigate_obs(internal_obs, i)
        swing_obs[:]    = _mitigate_obs(swing_obs,    i)

    # ── Post-loop: summarize outputs ─────────────────────────────────────────
    close_last = float(closes[-1])

    # Strong/Weak High/Low (Pine: trailing extremes + trend bias)
    # Strong = in direction of trend (traps the other side); Weak = opposite direction
    strong_high = trailing_top    if swing_trend == -1 else np.nan  # bearish trend = strong high
    weak_high   = trailing_top    if swing_trend ==  1 else np.nan  # bullish trend = weak high
    strong_low  = trailing_bottom if swing_trend ==  1 else np.nan  # bullish trend = strong low
    weak_low    = trailing_bottom if swing_trend == -1 else np.nan  # bearish trend = weak low

    # Fallback: always provide trailing extremes when trend is 0
    if swing_trend == 0:
        strong_high = trailing_top
        strong_low  = trailing_bottom

    # Closest internal order block (prefer the nearest to price)
    visible_obs  = internal_obs[:2] + swing_obs[:2]  # top 2 of each
    bull_ob = next((ob for ob in visible_obs if ob["bias"] == 1), None)
    bear_ob = next((ob for ob in visible_obs if ob["bias"] == -1), None)

    closest_ob_str      = ""
    closest_ob_bear_str = ""
    ob_equilibrium      = np.nan
    ob_direction        = "N/A"
    ob_dist_pct         = np.nan

    if bull_ob and bear_ob:
        bd  = abs(close_last - (bull_ob["barHigh"] + bull_ob["barLow"]) / 2)
        bd2 = abs(close_last - (bear_ob["barHigh"] + bear_ob["barLow"]) / 2)
        chosen_ob = bull_ob if bd <= bd2 else bear_ob
    elif bull_ob:
        chosen_ob = bull_ob
    elif bear_ob:
        chosen_ob = bear_ob
    else:
        chosen_ob = None

    # Closest OB Bull: always from bull_ob
    if bull_ob:
        closest_ob_str = _fmt_range(bull_ob["barLow"], bull_ob["barHigh"])
        # Equilibrium = midpoint of Closest OB Bull range
        ob_equilibrium = (bull_ob["barHigh"] + bull_ob["barLow"]) / 2.0

    # Closest OB Bear: always from bear_ob
    if bear_ob:
        closest_ob_bear_str = _fmt_range(bear_ob["barLow"], bear_ob["barHigh"])

    if chosen_ob:
        ob_mid       = (chosen_ob["barHigh"] + chosen_ob["barLow"]) / 2
        ob_direction = "Bullish" if chosen_ob["bias"] == 1 else "Bearish"
        if ob_mid > 0:
            ob_dist_pct = ((close_last / ob_mid) - 1.0) * 100.0

    # Closest FVG
    closest_fvg_str = ""
    fvg_bias_str    = "N/A"
    if fvgs:
        def fvg_dist(fvg):
            mid = (fvg["top"] + fvg["bottom"]) / 2
            return abs(close_last - mid)
        nearest_fvg = min(fvgs, key=fvg_dist)
        closest_fvg_str = _fmt_range(nearest_fvg["bottom"], nearest_fvg["top"])
        fvg_bias_str = "Bullish" if nearest_fvg["bias"] == 1 else "Bearish"

    # EQH/EQL status
    if eqh_detected and eql_detected:
        eqhl_status = "EQH + EQL Detected"
    elif eqh_detected:
        eqhl_status = "EQH Detected"
    elif eql_detected:
        eqhl_status = "EQL Detected"
    else:
        eqhl_status = "None"

    # Premium / Discount zone (Pine: trailing extremes define the range)
    top = trailing_top
    bot = trailing_bottom
    prem_bot     = 0.95 * top + 0.05 * bot if pd.notna(top) and pd.notna(bot) else np.nan
    eq_top_lvl   = 0.525 * top + 0.475 * bot if pd.notna(top) and pd.notna(bot) else np.nan
    eq_bot_lvl   = 0.525 * bot + 0.475 * top if pd.notna(top) and pd.notna(bot) else np.nan
    disc_top     = 0.95 * bot + 0.05 * top   if pd.notna(top) and pd.notna(bot) else np.nan

    pd_label = "N/A"
    if pd.notna(top) and pd.notna(bot):
        if chosen_ob:
            if ob_direction == "Bearish" and chosen_ob["barLow"] <= close_last <= chosen_ob["barHigh"]:
                pd_label = "In Bear OB"
            elif ob_direction == "Bullish" and chosen_ob["barLow"] <= close_last <= chosen_ob["barHigh"]:
                pd_label = "In Bull OB"
        if pd_label == "N/A":
            if close_last >= prem_bot:
                pd_label = "Premium"
            elif eq_bot_lvl <= close_last <= eq_top_lvl:
                pd_label = "Equilibrium"
            elif disc_top is not None and close_last <= disc_top:
                pd_label = "Discount"
            else:
                pd_label = "Mid-Zone"

    # Ribbon bias (EMA 25 vs EMA 50)
    close_s  = pd.Series(df["Close"].astype(float).values)
    ema25    = float(close_s.ewm(span=25, adjust=False).mean().iloc[-1]) if len(close_s) >= 25 else np.nan
    ema50    = float(close_s.ewm(span=50, adjust=False).mean().iloc[-1]) if len(close_s) >= 50 else np.nan
    if pd.notna(ema25) and pd.notna(ema50):
        ribbon_bias = "Bullish" if ema25 > ema50 else "Bearish"
    else:
        ribbon_bias = "N/A"

    # Current Quarter VWAP (developing)
    try:
        last_ts  = df.index[-1]
        q_start  = quarter_start(last_ts)
        df_q     = df.loc[df.index >= q_start]
        if not df_q.empty:
            tp_q = (df_q["High"] + df_q["Low"] + df_q["Close"]) / 3.0
            vv_q = df_q["Volume"].replace(0, np.nan)
            mask_q = tp_q.notna() & vv_q.notna()
            if mask_q.any():
                cur_qvwap = (tp_q[mask_q] * vv_q[mask_q]).sum() / vv_q[mask_q].sum()
            else:
                cur_qvwap = np.nan
        else:
            cur_qvwap = np.nan
    except Exception:
        cur_qvwap = np.nan

    qvwap_dist_pct = np.nan
    if pd.notna(cur_qvwap) and cur_qvwap > 0:
        qvwap_dist_pct = ((close_last / cur_qvwap) - 1.0) * 100.0

    # Nearest PY / PQ VWAP band (reuse anchored_vwap_block results from build_row)
    # We return the cur QVWAP; the actual PY/PQ values are already in build_row.
    # This engine just outputs the "nearest" label based on distance.
    nearest_pyq_band    = "N/A"
    pyq_band_dist_pct   = np.nan

    # ── SMC Composite State ───────────────────────────────────────────────────
    int_trend_lbl  = "Bullish" if internal_trend == 1 else ("Bearish" if internal_trend == -1 else "Neutral")
    sw_trend_lbl   = "Bullish" if swing_trend    == 1 else ("Bearish" if swing_trend    == -1 else "Neutral")

    if sw_trend_lbl == "Bullish" and int_trend_lbl == "Bullish" and pd_label == "Discount":
        composite = "Bullish Expansion"
    elif sw_trend_lbl == "Bullish" and int_trend_lbl == "Bullish":
        composite = "Bullish Continuation"
    elif sw_trend_lbl == "Bullish" and pd_label == "Premium":
        composite = "Distribution Risk"
    elif sw_trend_lbl == "Bearish" and int_trend_lbl == "Bearish":
        composite = "Bearish Expansion"
    elif sw_trend_lbl == "Bearish" and pd_label == "Discount":
        composite = "Bearish Breakdown"
    elif pd_label == "Discount" and int_trend_lbl == "Bullish":
        composite = "Accumulation"
    elif pd_label == "Equilibrium":
        composite = "Neutral Rotation"
    elif sw_trend_lbl == "Bullish" and int_trend_lbl == "Bearish":
        composite = "Bullish Compression"
    else:
        composite = "Neutral Rotation"

    # Market Structure Bias (overall)
    ms_bias = sw_trend_lbl

    # Event age
    event_age = ""
    if last_event_idx is not None:
        try:
            event_age = int(n - 1 - last_event_idx)
        except Exception:
            pass

    # ── Legacy compatibility aliases (used by _build_idx_vwap_shortlist) ──────
    bull_ob_legacy = next((ob for ob in internal_obs[:2] if ob["bias"] == 1), None)
    bear_ob_legacy = next((ob for ob in internal_obs[:2] if ob["bias"] == -1), None)

    latest_int_str = last_int_event or "N/A"
    latest_sw_str  = last_sw_event  or "N/A"
    ms_phase = ("Expansion"  if ("BOS"  in (last_sw_event or "")) else
                "Reversal"   if ("CHoCH" in (last_sw_event or "")) else "Range")
    ms_swing_seq = ("HH/HL" if swing_trend == 1 else
                    "LH/LL" if swing_trend == -1 else "Mixed")
    ms_quality  = "Clean" if (len(internal_obs) > 0 or len(swing_obs) > 0) else "Developing"

    return {
        # ── New SMC columns ──────────────────────────────────────────────────
        "smc_internal_trend":         int_trend_lbl,
        "smc_swing_trend":            sw_trend_lbl,
        "smc_internal_bos_count":     int_bos_count,
        "smc_internal_choch_count":   int_choch_count,
        "smc_swing_bos_count":        sw_bos_count,
        "smc_swing_choch_count":      sw_choch_count,
        "smc_latest_internal_struct": latest_int_str,
        "smc_latest_swing_struct":    latest_sw_str,
        "smc_strong_high":            safe_num(strong_high),
        "smc_weak_high":              safe_num(weak_high),
        "smc_strong_low":             safe_num(strong_low),
        "smc_weak_low":               safe_num(weak_low),
        "smc_closest_ob":             closest_ob_str,
        "smc_ob_equilibrium":         safe_num(ob_equilibrium),
        "smc_closest_ob_bear":        closest_ob_bear_str,
        "smc_ob_dist_pct":            safe_num(ob_dist_pct),
        "smc_closest_fvg":            closest_fvg_str,
        "smc_fvg_bias":               fvg_bias_str,
        "smc_eqhl_status":            eqhl_status,
        "smc_premium_discount":       pd_label,
        "smc_ribbon_bias":            ribbon_bias,
        "smc_current_qvwap":          safe_num(cur_qvwap),
        "smc_qvwap_dist_pct":         safe_num(qvwap_dist_pct),
        "smc_nearest_pyq_band":       nearest_pyq_band,
        "smc_pyq_band_dist_pct":      safe_num(pyq_band_dist_pct),
        "smc_market_structure_bias":  ms_bias,
        "smc_composite_state":        composite,
        # ── Legacy aliases (used by shortlist + detail schema) ────────────────
        "smc_bull_internal_ob":       _fmt_range(bull_ob_legacy["barLow"], bull_ob_legacy["barHigh"]) if bull_ob_legacy else "",
        "smc_bear_internal_ob":       _fmt_range(bear_ob_legacy["barLow"], bear_ob_legacy["barHigh"]) if bear_ob_legacy else "",
        "smc_discount":               _fmt_range(bot, disc_top),
        "smc_equilibrium":            _fmt_range(eq_bot_lvl, eq_top_lvl),
        "smc_premium":                _fmt_range(prem_bot, top),
        "smc_state":                  composite,
        # ── Market Structure aliases (used by _build_idx_vwap_shortlist) ──────
        "ms_trend_bias":              sw_trend_lbl,
        "ms_structure_phase":         ms_phase,
        "ms_last_structural_event":   latest_sw_str,
        "ms_event_age_d":             event_age,
        "ms_swing_sequence":          ms_swing_seq,
        "ms_structure_quality":       ms_quality,
        "ms_volume_confirmation":     "N/A",
        # Backward compat with _build_idx_vwap_shortlist
        "ms_trend_regime":            sw_trend_lbl,
        "ms_structure_state":         ms_swing_seq,
        "ms_last_event":              latest_sw_str,
    }


# =========================
# DETAIL SCHEMA OVERRIDE — SMC v2 + MACD Enhanced columns
# =========================
try:
    _new_detail_schema_v2 = []
    for _grp_name, _grp_fill, _cols in DETAIL_SCHEMA:
        if _grp_name == "Market Structure":
            _new_detail_schema_v2.append((
                "Market Structure & SMC", _grp_fill, [
                    ("smc_internal_trend",         "Internal Trend",        14, None,    "center"),
                    ("smc_swing_trend",             "Swing Trend",           14, None,    "center"),
                    ("smc_latest_internal_struct",  "Latest Internal Struct",22, None,    "center"),
                    ("smc_latest_swing_struct",     "Latest Swing Struct",   22, None,    "center"),
                    ("smc_strong_high",             "Strong High",           12, "#,##0", "center"),
                    ("smc_weak_high",               "Weak High",             12, "#,##0", "center"),
                    ("smc_strong_low",              "Strong Low",            12, "#,##0", "center"),
                    ("smc_weak_low",                "Weak Low",              12, "#,##0", "center"),
                    ("smc_closest_ob",              "Closest OB Bull",       16, None,    "center"),
                    ("smc_ob_equilibrium",          "Equilibrium",           16, "#,##0", "center"),
                    ("smc_closest_ob_bear",         "Closest OB Bear",       16, None,    "center"),
                    ("smc_ob_bull_age",             "OB Bull Age (Days)",    16, "#,##0", "center"),
                    ("smc_ob_bear_age",             "OB Bear Age (Days)",    16, "#,##0", "center"),
                ]
            ))
        elif _grp_name == "MACD Momentum":
            _new_detail_schema_v2.append((
                "MACD Momentum", _grp_fill, [
                    ("macd_line",        "MACD Line",        12, "#,##0", "center"),
                    ("macd_signal_line", "Signal Line",      12, "#,##0", "center"),
                    ("macd_hist",        "Histogram (EMA3)", 14, "#,##0", "center"),
                    ("macd_position",    "Lines Position",   18, None,    "center"),
                    ("macd_wave",        "Wave Pattern",     22, None,    "center"),
                    ("macd_cross",       "MACD Cross",       16, None,    "center"),
                ]
            ))
        elif _grp_name in ("Structure / Wyckoff Proxy (Backtest)",
                           "Candlestick Patterns",
                           "SMC"):
            pass
        else:
            _new_detail_schema_v2.append((_grp_name, _grp_fill, _cols))
    DETAIL_SCHEMA = _new_detail_schema_v2
except Exception:
    pass

# build_row wrapper: call new compute_smc_engine
try:
    _orig_build_row_v2 = build_row
    def build_row(ksei_row: pd.Series, hist: pd.DataFrame, shares_fallback: float):
        row = _orig_build_row_v2(ksei_row, hist, shares_fallback)
        try:
            smc = compute_smc_engine(hist, internal_size=5, swing_size=50,
                                     eq_length=3, eq_threshold=0.1,
                                     mitigation_mode="close")
            if isinstance(smc, dict):
                row.update(smc)
        except Exception:
            for _k in ("smc_composite_state", "smc_swing_trend", "smc_internal_trend",
                       "smc_state", "ms_trend_bias", "ms_structure_quality",
                       "ms_last_structural_event", "ms_event_age_d", "ms_swing_sequence"):
                row.setdefault(_k, "N/A")
        return row
except Exception:
    pass

def _to_yf_symbol_for_setup(ticker: str) -> str:
    try:
        t = str(ticker or "").strip().upper()
        if not t:
            return ""
        return t if t.endswith(".JK") else f"{t}.JK"
    except Exception:
        return ""

def _get_hist_for_setup(row: dict, lookback: int = 260) -> Optional[pd.DataFrame]:
    try:
        ticker = row.get("ticker", row.get("Ticker", ""))
        yf_symbol = _to_yf_symbol_for_setup(ticker)
        if not yf_symbol:
            return None
        hist = yf.download(
            yf_symbol,
            period="2y",
            interval="1d",
            auto_adjust=False,
            progress=False,
            threads=False,
        )
        hist = normalize_history(hist)
        if hist is None or hist.empty:
            return None
        if lookback and len(hist) > lookback:
            hist = hist.tail(int(lookback)).copy()
        return hist
    except Exception:
        return None

def _detect_recent_candle_patterns(hist: pd.DataFrame) -> dict:
    out = {
        "bullish_engulfing": False,
        "inside_bar": False,
        "ibh": np.nan,
        "ibl": np.nan,
        "last_bar_range_pct": np.nan,
        "close_in_upper_third": False,
        "bullish_reversal_bar": False,
    }
    try:
        if hist is None or len(hist) < 2:
            return out
        prev = hist.iloc[-2]
        last = hist.iloc[-1]
        po, ph, pl, pc = map(float, [prev["Open"], prev["High"], prev["Low"], prev["Close"]])
        lo, lh, ll, lc = map(float, [last["Open"], last["High"], last["Low"], last["Close"]])

        prev_body_low, prev_body_high = min(po, pc), max(po, pc)
        last_body_low, last_body_high = min(lo, lc), max(lo, lc)

        out["bullish_engulfing"] = (pc < po) and (lc > lo) and (last_body_low <= prev_body_low) and (last_body_high >= prev_body_high)
        out["inside_bar"] = (lh < ph) and (ll > pl)
        if out["inside_bar"]:
            out["ibh"], out["ibl"] = ph, pl
        else:
            out["ibh"], out["ibl"] = max(ph, lh), min(pl, ll)

        rng = max(lh - ll, 0.0)
        out["last_bar_range_pct"] = ((rng / lc) * 100.0) if lc else np.nan
        out["close_in_upper_third"] = rng > 0 and lc >= (ll + (2.0 * rng / 3.0))
        out["bullish_reversal_bar"] = (lc > lo) and out["close_in_upper_third"]
        return out
    except Exception:
        return out

def _sw_get(r: dict, keys, default=np.nan):
    if isinstance(keys, str):
        keys = [keys]
    for k in keys:
        if k in r:
            v = r.get(k)
            if v is None:
                continue
            if isinstance(v, float) and np.isnan(v):
                continue
            s = str(v).strip()
            if s == "" or s.upper() == "N/A":
                continue
            return v
    return default

def _enrich_swing_row_fields(rr: dict) -> dict:
    rr = dict(rr)
    hist = None

    def _need(*keys):
        return all((_sw_get(rr, k, np.nan) is np.nan or (isinstance(_sw_get(rr, k, np.nan), float) and np.isnan(_sw_get(rr, k, np.nan)))) for k in keys)

    # Normalize common display aliases from detail schema
    verdict = _sw_get(rr, ["Verdict Weight Profiles", "Verdict Weight Profile", "verdict_weight_profile"], "N/A")
    if verdict != "N/A":
        rr["Verdict Weight Profile"] = verdict
        rr["Verdict Profile"] = verdict

    trend = _sw_get(rr, ["Trend Bias", "trend_regime"], "N/A")
    if trend != "N/A":
        rr["Trend Bias"] = trend

    lse = _sw_get(rr, ["Last Structural Event", "last_structural_event"], "N/A")
    if lse != "N/A":
        rr["Last Structural Event"] = lse

    smc = _sw_get(rr, ["SMC State", "smc_state", "smc_zone"], "N/A")
    if smc != "N/A":
        rr["SMC State"] = smc
        rr["smc_state"] = smc

    mp = _sw_get(rr, ["Market Profile", "market_profile_summary", "VWAP Zone"], "N/A")
    if mp != "N/A":
        rr["Market Profile"] = mp
        rr["market_profile_summary"] = mp

    ma = _sw_get(rr, ["MA Position", "ma_position_summary", "MA Zone"], "N/A")
    if ma != "N/A":
        rr["MA Position"] = ma
        rr["ma_position_summary"] = ma

    cp = _sw_get(rr, ["Candle Pattern", "Last Candlestick Patterns", "Pattern"], "N/A")
    if cp != "N/A":
        rr["Candle Pattern"] = cp
        rr["last_candle_pattern"] = cp

    rsi_v = _sw_get(rr, ["RSI Status", "rsi_status"], "N/A")
    if rsi_v != "N/A":
        rr["RSI Status"] = rsi_v
        rr["rsi_status"] = rsi_v

    div = _sw_get(rr, ["Divergence Signal", "divergence_signal"], "N/A")
    if div != "N/A":
        rr["Divergence Signal"] = div
        rr["divergence_signal"] = div

    macd = _sw_get(rr, ["MACD Wave", "Wave Pattern", "macd_wave_pattern"], "N/A")
    if macd != "N/A":
        rr["MACD Wave"] = macd
        rr["macd_wave_pattern"] = macd

    adr = _sw_get(rr, ["ADR %", "adr14_pct"], np.nan)
    atr = _sw_get(rr, ["ATR (14) %", "ATR14 %", "atr14_pct"], np.nan)

    if (pd.isna(safe_num(adr)) or pd.isna(safe_num(atr)) or _sw_get(rr, ["Candle Pattern", "Last Candlestick Patterns", "Pattern"], "N/A") == "N/A"):
        hist = _get_hist_for_setup(rr)

    if pd.isna(safe_num(adr)) and hist is not None and len(hist) >= 14:
        try:
            adr_val = (((hist["High"] - hist["Low"]) / hist["Close"].replace(0, np.nan)) * 100.0).tail(14).mean()
            if pd.notna(adr_val):
                rr["ADR %"] = float(adr_val)
                rr["adr14_pct"] = float(adr_val)
        except Exception:
            pass

    if pd.isna(safe_num(atr)) and hist is not None and len(hist) >= 15:
        try:
            prev_close = hist["Close"].shift(1)
            tr = pd.concat([
                hist["High"] - hist["Low"],
                (hist["High"] - prev_close).abs(),
                (hist["Low"] - prev_close).abs()
            ], axis=1).max(axis=1)
            atr_val = tr.rolling(14).mean().iloc[-1]
            close_last = safe_num(hist["Close"].iloc[-1])
            atr_pct = (atr_val / close_last * 100.0) if pd.notna(atr_val) and close_last else np.nan
            if pd.notna(atr_pct):
                rr["ATR (14) %"] = float(atr_pct)
                rr["ATR14 %"] = float(atr_pct)
                rr["atr14_pct"] = float(atr_pct)
        except Exception:
            pass

    if _sw_get(rr, ["Candle Pattern", "Last Candlestick Patterns", "Pattern"], "N/A") == "N/A" and hist is not None and len(hist) >= 2:
        try:
            patt = _detect_recent_candle_patterns(hist)
            label = "N/A"
            if patt.get("bullish_engulfing"):
                label = "Bullish Engulfing"
            elif patt.get("inside_bar"):
                label = "Inside Bar"
            elif patt.get("bullish_reversal_bar"):
                label = "Bullish Reversal Bar"
            rr["Candle Pattern"] = label
            rr["last_candle_pattern"] = label
        except Exception:
            pass

    return rr


# =========================
# FUNDAMENTAL KEY STATS SHEET  (Stockbit Key Stats style)
# =========================
def _fetch_fundamental_data(ticker: str) -> dict:
    """Fetch yfinance fundamentals for a single IDX ticker."""
    try:
        sym = ticker.upper().strip()
        if not sym.endswith(".JK"):
            sym = sym + ".JK"
        info = yf.Ticker(sym).info or {}

        def _g(key, default=np.nan):
            v = info.get(key, default)
            if v is None:
                return default
            try:
                return float(v) if isinstance(v, (int, float)) else v
            except Exception:
                return v

        def _gs(key, default="N/A"):
            v = info.get(key, default)
            return str(v) if v is not None else default

        def _gi(key, default=np.nan):
            v = info.get(key, default)
            try:
                return int(v) if v is not None else default
            except Exception:
                return default

        # IPO date: try firstTradeDateEpochUtc, fallback to first available price bar
        _ipo_raw = info.get("firstTradeDateEpochUtc") or info.get("firstTradeDateMilliseconds")
        if _ipo_raw:
            try:
                import datetime as _dt
                _ipo_ts = float(_ipo_raw) / (1000 if float(_ipo_raw) > 1e10 else 1)
                _utc = getattr(_dt, "UTC", _dt.timezone.utc)
                _ipo_date = _dt.datetime.fromtimestamp(
                    _ipo_ts,
                    _utc
                ).strftime("%Y-%m-%d")
            except Exception:
                _ipo_date = "N/A"
        else:
            try:
                _early = yf.Ticker(sym).history(period="max", interval="1mo",
                                                  auto_adjust=True, progress=False)
                if _early is not None and not _early.empty:
                    _first = _early.index[0]
                    _ipo_date = _first.strftime("%Y-%m-%d") if hasattr(_first,"strftime") else str(_first)[:10]
                else:
                    _ipo_date = "N/A"
            except Exception:
                _ipo_date = "N/A"

        return {
            "ticker":              ticker,
            "company_name":        _gs("longName", _gs("shortName", ticker)),
            "sector":              _gs("sector"),
            "industry":            _gs("industry"),
            "shares_outstanding":  _gi("sharesOutstanding"),
            "float_shares":        _gi("floatShares"),
            "ipo_date":            _ipo_date,
            # Valuation
            "market_cap":          _g("marketCap"),
            "enterprise_value":    _g("enterpriseValue"),
            "pe_ttm":              _g("trailingPE"),
            "forward_pe":          _g("forwardPE"),
            "pbv":                 _g("priceToBook"),
            "ps_ttm":              _g("priceToSalesTrailing12Months"),
            "ev_ebitda":           _g("enterpriseToEbitda"),
            "peg_ratio":           _g("trailingPegRatio"),
            "book_value":          _g("bookValue"),
            # Profitability
            "roe":                 _g("returnOnEquity"),
            "roa":                 _g("returnOnAssets"),
            "gross_margin":        _g("grossMargins"),
            "operating_margin":    _g("operatingMargins"),
            "net_margin":          _g("profitMargins"),
            "ebitda":              _g("ebitda"),
            # Growth
            "revenue_growth":      _g("revenueGrowth"),
            "earnings_growth":     _g("earningsGrowth"),
            "quarterly_revenue":   _g("totalRevenue"),
            "quarterly_earnings":  _g("netIncomeToCommon"),
            # Balance Sheet
            "total_assets":        _g("totalAssets"),
            "total_liabilities":   _g("totalDebt"),          # proxy
            "cash":                _g("totalCash"),
            "debt":                _g("totalDebt"),
            "debt_to_equity":      _g("debtToEquity"),
            "current_ratio":       _g("currentRatio"),
            # Cash Flow
            "operating_cf":        _g("operatingCashflow"),
            "free_cf":             _g("freeCashflow"),
            "capex":               _g("capitalExpenditures"),
            # Dividend
            "dividend_yield":      _g("dividendYield"),
            "payout_ratio":        _g("payoutRatio"),
            "dividend_rate":       _g("dividendRate"),
            "trailing_annual_div": _g("trailingAnnualDividendRate"),
            "ex_dividend_date":    _gs("exDividendDate", "N/A"),
            "last_dividend_value": _g("lastDividendValue"),
            "last_dividend_date":  _gs("lastDividendDate", "N/A"),
            # Business Summary
            "business_summary":    _gs("longBusinessSummary", _gs("shortBusinessSummary", "N/A")),
        }
    except Exception:
        return {"ticker": ticker, "company_name": ticker}


def _build_dividend_entries(ticker: str) -> dict:
    """
    Build two dividend entries for the 8-column Dividend group.
    Entry 1: upcoming dividend if exists, else latest historical.
    Entry 2: latest historical if entry 1 is upcoming, else previous year's total.
    Returns keys: div1_year/idr/exdate/paydate  div2_year/idr/exdate/paydate
    """
    empty = {k: "-" for k in
             ["div1_year","div1_idr","div1_exdate","div1_paydate",
              "div2_year","div2_idr","div2_exdate","div2_paydate"]}
    try:
        sym = ticker.upper().strip()
        if not sym.endswith(".JK"):
            sym += ".JK"
        tk  = yf.Ticker(sym)
        info = tk.info or {}
        divs = tk.dividends

        # Normalize timezone
        if divs is not None and not divs.empty:
            if hasattr(divs.index, "tz") and divs.index.tz is not None:
                divs.index = divs.index.tz_localize(None)

        def _fmt_date(raw):
            if raw is None:
                return "-"
            try:
                return pd.Timestamp(raw, unit="s").strftime("%Y-%m-%d")
            except Exception:
                try:
                    return str(raw)[:10]
                except Exception:
                    return "-"

        # Upcoming dividend
        upcoming_amount = float(info.get("dividendRate") or 0)
        upcoming_ex_raw = info.get("exDividendDate")
        upcoming_ex     = _fmt_date(upcoming_ex_raw)
        has_upcoming    = upcoming_amount > 0 and upcoming_ex != "-"

        # Latest historical dividend (most recent entry from yfinance dividends)
        latest_hist = {}
        if divs is not None and not divs.empty:
            last_date = divs.index[-1]
            last_amt  = float(divs.iloc[-1])
            latest_hist = {
                "year": str(last_date.year),
                "idr":  last_amt,
                "exdate": last_date.strftime("%Y-%m-%d"),
                "paydate": "-",
            }

        # Second-latest or previous year
        prev_hist = {}
        if divs is not None and len(divs) >= 2:
            prev_date = divs.index[-2]
            prev_amt  = float(divs.iloc[-2])
            prev_hist = {
                "year": str(prev_date.year),
                "idr":  prev_amt,
                "exdate": prev_date.strftime("%Y-%m-%d"),
                "paydate": "-",
            }

        out = dict(empty)
        if has_upcoming:
            # Entry 1 = upcoming
            out["div1_year"]    = "Upcoming"
            out["div1_idr"]     = upcoming_amount
            out["div1_exdate"]  = upcoming_ex
            out["div1_paydate"] = "-"
            # Entry 2 = latest historical
            if latest_hist:
                out["div2_year"]    = latest_hist["year"]
                out["div2_idr"]     = latest_hist["idr"]
                out["div2_exdate"]  = latest_hist["exdate"]
                out["div2_paydate"] = latest_hist["paydate"]
        else:
            # Entry 1 = latest historical
            if latest_hist:
                out["div1_year"]    = latest_hist["year"]
                out["div1_idr"]     = latest_hist["idr"]
                out["div1_exdate"]  = latest_hist["exdate"]
                out["div1_paydate"] = latest_hist["paydate"]
            # Entry 2 = previous historical
            if prev_hist:
                out["div2_year"]    = prev_hist["year"]
                out["div2_idr"]     = prev_hist["idr"]
                out["div2_exdate"]  = prev_hist["exdate"]
                out["div2_paydate"] = prev_hist["paydate"]
        return out
    except Exception:
        return {k: "-" for k in
                ["div1_year","div1_idr","div1_exdate","div1_paydate",
                 "div2_year","div2_idr","div2_exdate","div2_paydate"]}


def _fetch_dividend_history(ticker: str) -> dict:
    """
    Fetch dividend history from yfinance. Returns per-year dividends (2023-2025),
    CAGR, consecutive years, upcoming ex-date and yield estimates.
    Falls back to N/A gracefully.
    """
    empty = {
        "div_2025": np.nan, "div_2024": np.nan, "div_2023": np.nan,
        "div_cagr": np.nan, "div_consecutive_years": 0,
        "upcoming_dividend": np.nan, "upcoming_ex_date": "N/A",
        "upcoming_pay_date": "N/A", "upcoming_div_yield": np.nan,
    }
    try:
        sym = ticker.upper().strip()
        if not sym.endswith(".JK"):
            sym += ".JK"
        tk = yf.Ticker(sym)
        divs = tk.dividends
        info = tk.info or {}

        if divs is None or divs.empty:
            return empty

        # Normalize timezone
        if hasattr(divs.index, "tz") and divs.index.tz is not None:
            divs.index = divs.index.tz_localize(None)

        def _year_total(yr):
            mask = divs.index.year == yr
            return float(divs[mask].sum()) if mask.any() else np.nan

        div_2025 = _year_total(2025)
        div_2024 = _year_total(2024)
        div_2023 = _year_total(2023)

        # Consecutive years with dividend
        annual_years = sorted({dt.year for dt in divs.index}, reverse=True)
        consecutive = 0
        prev = None
        for yr in annual_years:
            if prev is None or yr == prev - 1:
                consecutive += 1
                prev = yr
            else:
                break

        # CAGR: use last 3 full years
        cagr = np.nan
        if pd.notna(div_2023) and pd.notna(div_2025) and div_2023 > 0:
            cagr = ((div_2025 / div_2023) ** (1 / 2) - 1) * 100

        # Upcoming dividend from yfinance info
        upcoming_dividend = float(info.get("dividendRate") or np.nan) if info.get("dividendRate") else np.nan
        ex_div_raw = info.get("exDividendDate")
        upcoming_ex_date = "N/A"
        if ex_div_raw:
            try:
                upcoming_ex_date = pd.Timestamp(ex_div_raw, unit="s").strftime("%Y-%m-%d")
            except Exception:
                try:
                    upcoming_ex_date = str(ex_div_raw)[:10]
                except Exception:
                    pass

        curr_price = float(info.get("regularMarketPrice") or info.get("currentPrice") or np.nan)
        upcoming_div_yield = (upcoming_dividend / curr_price * 100) if pd.notna(upcoming_dividend) and pd.notna(curr_price) and curr_price > 0 else np.nan

        return {
            "div_2025": div_2025,
            "div_2024": div_2024,
            "div_2023": div_2023,
            "div_cagr": cagr,
            "div_consecutive_years": consecutive,
            "upcoming_dividend": upcoming_dividend,
            "upcoming_ex_date": upcoming_ex_date,
            "upcoming_pay_date": "N/A",  # not available from yfinance
            "upcoming_div_yield": upcoming_div_yield,
        }
    except Exception:
        return empty


def _fetch_google_rss_disclosure(ticker: str, company_name: str = "") -> dict:
    """
    Fetch the latest corporate-event headline from Google News RSS.
    The result is market-date aware: prefer the newest item on/before MARKET_DATE,
    then fall back to the newest available item when no as-of item exists.
    Returns: date, title, url, category, sentiment, days_since, event_risk.
    """
    import html
    import re
    import urllib.parse
    import urllib.request
    import xml.etree.ElementTree as ET
    from datetime import datetime
    from email.utils import parsedate_to_datetime

    empty = {
        "disclosure_date": "N/A", "disclosure_title": "N/A",
        "disclosure_url": "N/A", "disclosure_category": "N/A",
        "disclosure_sentiment": "N/A", "days_since_disclosure": np.nan,
        "event_risk": "N/A",
    }

    keyword_category = {
        "rights issue": ("Dilution", "Bearish", "High"),
        "penawaran umum terbatas": ("Dilution", "Bearish", "High"),
        "private placement": ("Dilution", "Bearish", "High"),
        "hmetd": ("Dilution", "Bearish", "High"),
        "saham baru": ("Dilution", "Bearish", "High"),
        "suspensi": ("Distress", "Bearish", "High"),
        "pailit": ("Distress", "Bearish", "High"),
        "bankruptcy": ("Distress", "Bearish", "High"),
        "gagal bayar": ("Distress", "Bearish", "High"),
        "utang": ("Leverage", "Neutral", "Medium"),
        "obligasi": ("Debt Financing", "Neutral", "Medium"),
        "restrukturisasi": ("Transitional", "Neutral", "Medium"),
        "restructuring": ("Transitional", "Neutral", "Medium"),
        "akuisisi": ("Expansion", "Neutral", "Medium"),
        "acquisition": ("Expansion", "Neutral", "Medium"),
        "merger": ("Corporate Action", "Neutral", "Medium"),
        "penawaran tender": ("Corporate Action", "Neutral", "Medium"),
        "tender offer": ("Corporate Action", "Neutral", "Medium"),
        "rups": ("Shareholder Meeting", "Neutral", "Medium"),
        "rapat umum pemegang saham": ("Shareholder Meeting", "Neutral", "Medium"),
        "laporan keuangan": ("Earnings", "Neutral", "Medium"),
        "laba": ("Earnings", "Bullish", "Medium"),
        "rugi": ("Earnings", "Bearish", "Medium"),
        "dividen": ("Shareholder Return", "Bullish", "Low"),
        "dividend": ("Shareholder Return", "Bullish", "Low"),
        "buyback": ("Shareholder Return", "Bullish", "Low"),
        "pembelian kembali saham": ("Shareholder Return", "Bullish", "Low"),
        "pemecahan saham": ("Liquidity Positive", "Bullish", "Low"),
        "stock split": ("Liquidity Positive", "Bullish", "Low"),
    }

    def _market_asof_or_none():
        raw = globals().get("MARKET_DATE", None)
        if raw in (None, "", "None", "none", "N/A", "-"):
            return None
        try:
            ts = pd.Timestamp(raw)
            if pd.isna(ts):
                return None
            return ts.normalize()
        except Exception:
            return None

    def _clean_google_title(title):
        title = html.unescape(str(title or "")).strip()
        return re.sub(r"\s+-\s+[^-]{2,80}$", "", title).strip() or title

    def _parse_pubdate(raw):
        try:
            dt = parsedate_to_datetime(str(raw or ""))
            if dt.tzinfo is not None:
                dt = dt.astimezone().replace(tzinfo=None)
            return pd.Timestamp(dt).normalize()
        except Exception:
            return pd.NaT

    def _classify(title):
        title_lower = str(title or "").lower()
        for kw, (cat, sent, ev_risk) in keyword_category.items():
            if kw in title_lower:
                return cat, sent, ev_risk
        return "Google News", "Neutral", "Low"

    try:
        clean = ticker.upper().replace(".JK", "").strip()
        company = re.sub(r"\s+", " ", str(company_name or "")).strip()
        company_query = f' OR "{company}"' if company and company.upper() != clean else ""
        query = (
            f'("{clean}" OR "{clean}.JK"{company_query}) '
            f'(saham OR emiten OR IDX OR BEI OR "Bursa Efek Indonesia")'
        )
        params = urllib.parse.urlencode({
            "q": query,
            "hl": "id",
            "gl": "ID",
            "ceid": "ID:id",
        })
        feed_url = f"https://news.google.com/rss/search?{params}"
        req = urllib.request.Request(feed_url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=12) as resp:
            xml_bytes = resp.read()

        root = ET.fromstring(xml_bytes)
        items = []
        for item in root.findall(".//item"):
            title_raw = item.findtext("title") or ""
            link = html.unescape(item.findtext("link") or feed_url).strip()
            pub_ts = _parse_pubdate(item.findtext("pubDate"))
            source = item.findtext("source") or "Google News"
            title = _clean_google_title(title_raw)
            if pd.isna(pub_ts) or not title:
                continue
            searchable = f"{title} {source}".upper()
            if clean not in searchable and company:
                company_tokens = [
                    w.upper()
                    for w in re.findall(r"[A-Za-z0-9]+", company)
                    if len(w) >= 4
                ]
                if not any(tok in searchable for tok in company_tokens[:4]):
                    continue
            items.append({
                "title": title,
                "url": link,
                "date": pub_ts,
                "source": source,
            })

        if not items:
            return empty

        items.sort(key=lambda x: x["date"], reverse=True)
        asof = _market_asof_or_none()
        if asof is not None:
            asof_items = [x for x in items if x["date"] <= asof]
            selected = asof_items[0] if asof_items else items[0]
            days_base = asof.date()
        else:
            selected = items[0]
            days_base = datetime.now().date()

        category, sentiment, risk = _classify(selected["title"])
        pub_date = selected["date"].date()

        return {
            "disclosure_date": pub_date.strftime("%Y-%m-%d"),
            "disclosure_title": selected["title"][:160],
            "disclosure_url": selected["url"],
            "disclosure_category": category,
            "disclosure_sentiment": sentiment,
            "days_since_disclosure": (days_base - pub_date).days,
            "event_risk": risk,
        }
    except Exception:
        return empty


def _fetch_idx_disclosure(ticker: str) -> dict:
    """
    Backward-compatible wrapper. Disclosure data now comes from Google News RSS.
    """
    return _fetch_google_rss_disclosure(ticker)


def _compute_pbv_bands(ticker: str, years: int = 3) -> dict:
    """
    Stockbit-style PBV Band Analysis.
    - Uses adjusted closing price (auto_adjust=True) to handle splits correctly.
    - BVPS = Total Stockholder Equity / Shares Outstanding (quarterly).
    - BVPS is forward-filled to daily frequency; NO look-ahead bias
      (BVPS for quarter Q is only applied from the earnings release date onward).
    - Excludes PBV <= 0, inf, NaN.
    - Returns z-score, percentile, SD bands, and regime.
    """
    empty = {
        "pbv_curr": np.nan, "pbv_mean": np.nan,
        "pbv_plus1": np.nan, "pbv_plus2": np.nan,
        "pbv_minus1": np.nan, "pbv_minus2": np.nan,
        "pbv_zscore": np.nan, "pbv_pct": np.nan,
        "pbv_regime": "N/A",
    }
    try:
        sym = ticker.upper().strip()
        if not sym.endswith(".JK"):
            sym += ".JK"
        tk = yf.Ticker(sym)

        # ── 1. Adjusted price history ──────────────────────────────────────────
        period_str = f"{years + 1}y"
        price_hist = tk.history(period=period_str, auto_adjust=True)
        if price_hist is None or price_hist.empty:
            return empty
        if hasattr(price_hist.index, "tz") and price_hist.index.tz is not None:
            price_hist.index = price_hist.index.tz_localize(None)
        price_hist = price_hist["Close"].squeeze()
        if hasattr(price_hist, "columns"):
            price_hist = price_hist.iloc[:, 0]
        price_hist = price_hist.dropna()
        price_hist.index = safe_parse_datetime_index(price_hist.index, errors="coerce").normalize()

        # ── 2. Quarterly balance sheet — equity + shares ───────────────────────
        bs = tk.quarterly_balance_sheet
        if bs is None or bs.empty:
            return empty

        # Normalize column orientation (tickers sometimes have rows/cols swapped)
        # yfinance quarterly_balance_sheet: columns = report dates, rows = fields
        if isinstance(bs.columns[0], str):
            bs = bs.T   # flip if dates are in rows
        bs.index = safe_parse_datetime_index(bs.index, errors="coerce").normalize()
        bs = bs.sort_index()

        # Equity field — try multiple yfinance key names
        equity_keys = [
            "Stockholders Equity", "Total Stockholder Equity",
            "Common Stock Equity", "Total Equity Gross Minority Interest",
        ]
        equity_series = None
        for ek in equity_keys:
            if ek in bs.columns:
                equity_series = bs[ek].dropna()
                break
        if equity_series is None or equity_series.empty:
            return empty

        # Shares outstanding — try info first, then balance sheet
        shares_keys = ["Ordinary Shares Number", "Share Issued", "Common Stock"]
        shares_series = None
        for sk in shares_keys:
            if sk in bs.columns:
                shares_series = bs[sk].dropna()
                break
        if shares_series is None or shares_series.empty:
            # fall back to static info shares
            static_shares = float(tk.info.get("sharesOutstanding") or 0)
            if static_shares <= 0:
                return empty
            shares_series = pd.Series(static_shares, index=equity_series.index)

        # ── 3. BVPS per quarter (report date → BVPS) ──────────────────────────
        # Align equity and shares on the same dates
        bvps_q = (equity_series / shares_series).dropna()
        bvps_q = bvps_q[bvps_q > 0]   # exclude negative equity
        if bvps_q.empty:
            return empty

        # ── 4. Forward-fill BVPS to daily price index (NO look-ahead bias) ────
        # Create a daily series spanning the price history
        daily_idx = price_hist.index
        bvps_daily = pd.Series(np.nan, index=daily_idx)
        for report_date, bvps_val in bvps_q.items():
            # Apply BVPS from this report date onward (forward-fill)
            mask = daily_idx >= report_date
            bvps_daily[mask] = bvps_val

        # ── 5. Compute daily PBV ───────────────────────────────────────────────
        pbv_daily = (price_hist / bvps_daily).replace([np.inf, -np.inf], np.nan).dropna()
        pbv_daily = pbv_daily[pbv_daily > 0]

        # Restrict to 5-year window
        cutoff = pbv_daily.index[-1] - pd.DateOffset(years=years)
        pbv_window = pbv_daily[pbv_daily.index >= cutoff]
        if len(pbv_window) < 20:
            return empty

        curr_pbv = float(pbv_daily.iloc[-1])
        mean_pbv = float(pbv_window.mean())
        std_pbv  = float(pbv_window.std(ddof=1))
        if std_pbv == 0:
            return empty

        plus1  = mean_pbv + std_pbv
        plus2  = mean_pbv + 2 * std_pbv
        minus1 = mean_pbv - std_pbv
        minus2 = mean_pbv - 2 * std_pbv

        z_score  = (curr_pbv - mean_pbv) / std_pbv
        # Percentile rank — pure numpy, no scipy dependency
        arr      = pbv_window.values
        pct_rank = float(np.sum(arr <= curr_pbv) / len(arr) * 100)

        # Regime (spec labels)
        if curr_pbv < minus2:
            regime = "Deep Undervalued"
        elif curr_pbv < minus1:
            regime = "Undervalued"
        elif curr_pbv <= plus1:
            regime = "Fair Value"
        elif curr_pbv <= plus2:
            regime = "Overvalued"
        else:
            regime = "Bubble"

        return {
            "pbv_curr":   round(curr_pbv, 4),
            "pbv_mean":   round(mean_pbv, 4),
            "pbv_plus1":  round(plus1,    4),
            "pbv_plus2":  round(plus2,    4),
            "pbv_minus1": round(minus1,   4),
            "pbv_minus2": round(minus2,   4),
            "pbv_zscore": round(z_score,  4),
            "pbv_pct":    round(pct_rank, 2),
            "pbv_regime": regime,
        }
    except Exception:
        return empty


# ── FUNDAMENTAL DETAIL SCHEMA ─────────────────────────────────────────────────
# Groups and columns for the IDX Fundamental Detail sheet.
# Format per column: (key, header, col_width, number_format, alignment)
# number_format: "compact" → compact_fmt(), "0.00%" → percent, "#,##0" → integer, None → raw text
FUNDAMENTAL_DETAIL_SCHEMA = [
    ("Stock Info", FILL_GROUP_OWNER, [
        ("ticker",          "Ticker",         10, None,    "left"),
        ("close",           "Price",          10, "#,##0", "center"),
        ("pct_change",      "Price Change %", 10, "0.00%", "center"),
        ("idx_sector",      "IDX Sector",     20, None,    "left"),
    ]),
    ("Company Profile", FILL_GROUP_OWNER, [
        ("yf_business_summary",   "Business Summary",   60, None,      "left"),
        ("yf_shares_outstanding", "Shares Outstanding", 18, "compact", "center"),
        ("yf_float_shares",       "Free Float",         16, "compact", "center"),
        ("yf_market_cap",         "Market Cap",         14, "compact", "center"),
        ("yf_enterprise_value",   "Enterprise Value",   16, "compact", "center"),
        ("yf_ipo_date",           "IPO Date",           14, None,      "center"),
    ]),
    ("Valuation", FILL_GROUP_VAL, [
        ("yf_pe_ratio",        "P/E Ratio",         12, "0.00",   "center"),
        ("yf_forward_pe",      "Forward P/E",       12, "0.00",   "center"),
        ("yf_pb_ratio",        "P/B Ratio",         12, "0.00",   "center"),
        ("yf_ps_ratio",        "P/S Ratio",         12, "0.00",   "center"),
        ("yf_ev_ebitda",       "EV/EBITDA",         12, "0.00",   "center"),
        ("yf_peg_ratio",       "PEG Ratio",         12, "0.00",   "center"),
        ("yf_book_value",      "Book Value",        12, "#,##0",  "center"),
        ("yf_dividend_yield",  "Dividend Yield",    14, "0.00%",  "center"),
    ]),
    ("Profitability", FILL_GROUP_MOM, [
        ("yf_profit_margin",   "Profit Margin",     14, "0.00%",  "center"),
        ("yf_operating_margin","Operating Margin",  14, "0.00%",  "center"),
        ("yf_roe",             "ROE",               12, "0.00%",  "center"),
        ("yf_roa",             "ROA",               12, "0.00%",  "center"),
        ("yf_gross_margin",    "Gross Margin",      14, "0.00%",  "center"),
        ("yf_ebitda",          "EBITDA",            14, "compact","center"),
    ]),
    ("Growth", FILL_GROUP_Q, [
        ("yf_revenue_growth",    "Revenue Growth",    14, "0.00%",  "center"),
        ("yf_earnings_growth",   "Earnings Growth",   14, "0.00%",  "center"),
        ("yf_quarterly_revenue", "Quarterly Revenue", 16, "compact","center"),
        ("yf_quarterly_earnings","Quarterly Earnings",16, "compact","center"),
    ]),
    ("Balance Sheet", FILL_GROUP_MSF, [
        ("yf_total_assets",      "Total Assets",      16, "compact","center"),
        ("yf_total_liabilities", "Total Liabilities", 16, "compact","center"),
        ("yf_total_debt",        "Total Debt",        14, "compact","center"),
        ("yf_cash",              "Cash & Equiv",      14, "compact","center"),
        ("yf_current_ratio",     "Current Ratio",     12, "0.00",   "center"),
        ("yf_debt_equity",       "Debt/Equity",       12, "0.00",   "center"),
    ]),
    ("Cash Flow", FILL_GROUP_PQ, [
        ("yf_operating_cf",  "Operating CF",   14, "compact","center"),
        ("yf_free_cf",       "Free CF",        14, "compact","center"),
        ("yf_capex",         "CapEx",          14, "compact","center"),
    ]),
    ("Upcoming Dividend", FILL_GROUP_VR, [
        ("div1_year",     "Year",              10, None,    "center"),
        ("div1_idr",      "IDR",               12, "#,##0", "center"),
        ("div1_exdate",   "Ex Date",           14, None,    "center"),
        ("div1_paydate",  "Pay Date",          14, None,    "center"),
        ("yf_dividend_yield", "Dividend Yield TTM", 16, "0.00%", "center"),
        ("div_trap_flag", "Div Trap Risk",     28, None,    "center"),
    ]),
    ("Latest Dividend", FILL_GROUP_VR, [
        ("div2_year",     "Year",              10, None,    "center"),
        ("div2_idr",      "IDR",               12, "#,##0", "center"),
        ("div2_exdate",   "Ex Date",           14, None,    "center"),
        ("div2_paydate",  "Pay Date",          14, None,    "center"),
        ("yf_payout_ratio","Payout Ratio",     14, "0.00%", "center"),
    ]),
    ("PBV Band Analysis  ·  3-Year Rolling", FILL_GROUP_Q, [
        ("yf_pbv_curr",   "Current PBV",    12, "#,##0.00","center"),
        ("yf_pbv_mean",   "Mean PBV (3Y)",  14, "#,##0.00","center"),
        ("yf_pbv_plus2",  "+2 SD",          12, "#,##0.00","center"),
        ("yf_pbv_plus1",  "+1 SD",          12, "#,##0.00","center"),
        ("yf_pbv_minus1", "-1 SD",          12, "#,##0.00","center"),
        ("yf_pbv_minus2", "-2 SD",          12, "#,##0.00","center"),
        ("yf_pbv_zscore", "PBV Z-Score",    12, "0.00",   "center"),
        ("yf_pbv_pct",    "PBV Percentile", 14, "0.00",   "center"),
        ("yf_pbv_regime", "PBV Regime",     14, None,     "center"),
    ]),
    ("Disclosure & Corporate Events", FILL_GROUP_MSF, [
        ("yf_disclosure_date",       "Latest Disclosure Date",  16, None,    "center"),
        ("yf_disclosure_title",      "Latest Disclosure Title",  50, None,   "left"),
        ("yf_disclosure_url",        "Latest Disclosure URL",    40, None,   "left"),
        ("yf_disclosure_category",   "Disclosure Category",      18, None,   "center"),
        ("yf_disclosure_sentiment",  "Disclosure Sentiment",     18, None,   "center"),
        ("yf_days_since_disclosure", "Days Since Disclosure",    18, "#,##0","center"),
        ("yf_event_risk",            "Event Risk",               14, None,   "center"),
    ]),
]

# Colour map for PBV Regime labels
PBV_REGIME_COLORS = {
    "Deep Undervalued": "008000",
    "Undervalued":      "375623",
    "Fair Value":       "595959",
    "Overvalued":       "E36C09",
    "Bubble":           "C00000",
    # legacy labels
    "Deep Value":       "008000",
    "Value":            "375623",
    "Premium":          "E36C09",
    "Euphoric":         "C00000",
}

# Compact-format keys in the fundamental sheet
_COMPACT_KEYS = {
    "yf_shares_outstanding", "yf_float_shares", "yf_market_cap", "yf_enterprise_value",
    "yf_ebitda", "yf_quarterly_revenue", "yf_quarterly_earnings",
    "yf_total_assets", "yf_total_liabilities", "yf_total_debt", "yf_cash",
    "yf_operating_cf", "yf_free_cf", "yf_capex", "mcap",
}


def _build_fundamental_row(screener_row: dict) -> dict:
    """
    Merge screener row fields + yfinance fundamentals into a single flat dict
    keyed to FUNDAMENTAL_DETAIL_SCHEMA. All yfinance fields prefixed "yf_".
    """
    ticker = str(screener_row.get("ticker", "")).upper().strip()

    # Fetch yfinance fundamentals
    raw_fund = _fetch_fundamental_data(ticker)
    pbv_band  = _compute_pbv_bands(ticker)

    # Build combined row: screener fields + yf_-prefixed fundamental fields
    combined = {
        "ticker":     ticker,
        "close":      screener_row.get("close", np.nan),
        "pct_change": screener_row.get("pct_change", np.nan),
        "emiten":     screener_row.get("emiten", ""),
        "idx_sector": screener_row.get("idx_sector", ""),
    }

    # yfinance fields with yf_ prefix
    _YF_MAP = {
        "company_name":       "yf_company_name",
        "sector":             "yf_sector",
        "industry":           "yf_industry",
        "shares_outstanding": "yf_shares_outstanding",
        "float_shares":       "yf_float_shares",
        "market_cap":         "yf_market_cap",
        "enterprise_value":   "yf_enterprise_value",
        "ipo_date":           "yf_ipo_date",
        # Valuation
        "pe_ttm":             "yf_pe_ratio",
        "forward_pe":         "yf_forward_pe",
        "pbv":                "yf_pb_ratio",
        "ps_ttm":             "yf_ps_ratio",
        "ev_ebitda":          "yf_ev_ebitda",
        "peg_ratio":          "yf_peg_ratio",
        "book_value":         "yf_book_value",
        # Profitability
        "roe":                "yf_roe",
        "roa":                "yf_roa",
        "gross_margin":       "yf_gross_margin",
        "operating_margin":   "yf_operating_margin",
        "net_margin":         "yf_profit_margin",
        "ebitda":             "yf_ebitda",
        # Growth
        "revenue_growth":     "yf_revenue_growth",
        "earnings_growth":    "yf_earnings_growth",
        "quarterly_revenue":  "yf_quarterly_revenue",
        "quarterly_earnings": "yf_quarterly_earnings",
        # Balance Sheet
        "total_assets":       "yf_total_assets",
        "total_liabilities":  "yf_total_liabilities",
        "cash":               "yf_cash",
        "debt":               "yf_total_debt",
        "debt_to_equity":     "yf_debt_equity",
        "current_ratio":      "yf_current_ratio",
        # Cash Flow
        "operating_cf":       "yf_operating_cf",
        "free_cf":            "yf_free_cf",
        "capex":              "yf_capex",
        # Dividend
        "dividend_yield":     "yf_dividend_yield",
        "payout_ratio":       "yf_payout_ratio",
        "dividend_rate":      "yf_dividend_rate",
        "last_dividend_date": "yf_last_dividend_date",
        # Profile
        "business_summary":   "yf_business_summary",
    }
    _PBV_MAP = {
        "pbv_curr":   "yf_pbv_curr",
        "pbv_mean":   "yf_pbv_mean",
        "pbv_plus1":  "yf_pbv_plus1",
        "pbv_plus2":  "yf_pbv_plus2",
        "pbv_minus1": "yf_pbv_minus1",
        "pbv_minus2": "yf_pbv_minus2",
        "pbv_zscore": "yf_pbv_zscore",
        "pbv_pct":    "yf_pbv_pct",
        "pbv_regime": "yf_pbv_regime",
    }

    for src_key, dest_key in _YF_MAP.items():
        combined[dest_key] = raw_fund.get(src_key, np.nan)

    for src_key, dest_key in _PBV_MAP.items():
        combined[dest_key] = pbv_band.get(src_key, np.nan)

    # Dividend: 8-column 2-entry layout
    div_entries = _build_dividend_entries(ticker)
    for k, v in div_entries.items():
        combined[k] = v

    # Upgrade 8 — Dividend trap detection (uses already-fetched fundamental data)
    combined["div_trap_flag"] = detect_dividend_trap(combined)

    # Legacy dividend history (kept for Guide reference, not in schema)
    div_hist = _fetch_dividend_history(ticker)
    combined["yf_div_2025"]           = div_hist.get("div_2025", np.nan)
    combined["yf_div_2024"]           = div_hist.get("div_2024", np.nan)
    combined["yf_div_2023"]           = div_hist.get("div_2023", np.nan)
    combined["yf_div_cagr"]           = div_hist.get("div_cagr", np.nan)
    combined["yf_div_consecutive"]    = div_hist.get("div_consecutive_years", 0)
    combined["yf_upcoming_dividend"]  = div_hist.get("upcoming_dividend", np.nan)
    combined["yf_upcoming_ex_date"]   = div_hist.get("upcoming_ex_date", "N/A")
    combined["yf_upcoming_pay_date"]  = div_hist.get("upcoming_pay_date", "N/A")
    combined["yf_upcoming_div_yield"] = div_hist.get("upcoming_div_yield", np.nan)

    # Google News RSS corporate-event headline, selected relative to MARKET_DATE.
    disc = _fetch_google_rss_disclosure(
        ticker,
        raw_fund.get("company_name", screener_row.get("emiten", "")),
    )
    combined["yf_disclosure_date"]       = disc.get("disclosure_date", "N/A")
    combined["yf_disclosure_title"]      = disc.get("disclosure_title", "N/A")
    combined["yf_disclosure_url"]        = disc.get("disclosure_url", "N/A")
    combined["yf_disclosure_category"]   = disc.get("disclosure_category", "N/A")
    combined["yf_disclosure_sentiment"]  = disc.get("disclosure_sentiment", "N/A")
    combined["yf_days_since_disclosure"] = disc.get("days_since_disclosure", np.nan)
    combined["yf_event_risk"]            = disc.get("event_risk", "N/A")

    return combined


def _institutional_trade_plan(r: dict, hist: "pd.DataFrame | None" = None) -> dict:
    """
    Institutional Trade Plan Engine (v2 — spec compliant).
    - Anchor = nearest POI to current close (above OR below).
    - Entry = anchor level.
    - Target = next POI above anchor that gives R/R ≥ 1:2; else ≥ 1:1.
    - Invalidation = anchor − ATR proxy.
    - R/R displayed as "1:X" format.
    """
    close = safe_num(r.get("close"), np.nan)
    if pd.isna(close) or close <= 0:
        return {"poi": "-", "anchor": "-", "entry": "-",
                "target": np.nan, "upside_pct": np.nan,
                "invalidation": np.nan, "risk_reward": "-"}

    pois = {}
    for prefix, label in [("pq","PrevQ"), ("py","PrevY")]:
        for key, suf in [("_vwap","VWAP"),("_m1","-1SD"),("_m2","-2SD"),
                          ("_p1","+1SD"),("_p2","+2SD")]:
            v = safe_num(r.get(f"{prefix}{key}"), np.nan)
            if pd.notna(v) and v > 0:
                pois[f"{label} {suf}"] = v

    ob_eq = safe_num(r.get("smc_ob_equilibrium"), np.nan)
    def _mid(s):
        try:
            pts = str(s).replace("–","-").split("-")
            if len(pts)==2:
                return (float(pts[0].replace(",","")) + float(pts[1].replace(",","")))/2
        except Exception:
            pass
        return np.nan

    ob_bull_mid = _mid(str(r.get("smc_closest_ob","") or ""))
    ob_bear_mid = _mid(str(r.get("smc_closest_ob_bear","") or ""))
    eq_mid = (ob_bull_mid + ob_bear_mid)/2 if pd.notna(ob_bull_mid) and pd.notna(ob_bear_mid) else np.nan

    for k, v in [("Bull OB", ob_bull_mid), ("Equilibrium", eq_mid),
                  ("Bear OB", ob_bear_mid), ("EMA25", safe_num(r.get("ema25"), np.nan)),
                  ("EMA50",  safe_num(r.get("ema50"),  np.nan))]:
        if pd.notna(v) and v > 0:
            pois[k] = v

    if not pois:
        return {"poi": "-", "anchor": "-", "entry": "-",
                "target": np.nan, "upside_pct": np.nan,
                "invalidation": np.nan, "risk_reward": "-"}

    # Anchor = nearest POI to close (above OR below)
    anchor_name = min(pois, key=lambda k: abs(pois[k] - close))
    anchor      = pois[anchor_name]

    # ATR proxy for invalidation
    atr_pct   = safe_num(r.get("atr14_pct"), 3.0) / 100.0
    atr_proxy = close * atr_pct if atr_pct > 0 else close * 0.02
    invalidation = anchor - atr_proxy
    risk = anchor - invalidation   # risk from anchor to invalidation

    # Find target with R/R ≥ 1:2, then ≥ 1:1
    higher_pois = sorted([v for v in pois.values() if v > anchor + (atr_proxy * 0.5)])
    target = np.nan
    for t_candidate in higher_pois:
        reward = t_candidate - anchor
        if risk > 0 and reward / risk >= 2.0:
            target = t_candidate
            break
    if pd.isna(target):
        for t_candidate in higher_pois:
            reward = t_candidate - anchor
            if risk > 0 and reward / risk >= 1.0:
                target = t_candidate
                break
    if pd.isna(target) and higher_pois:
        target = higher_pois[-1]
    if pd.isna(target):
        target = anchor * 1.08

    upside_pct = ((target - close) / close * 100) if close > 0 else np.nan
    reward     = target - anchor
    rr_ratio   = (reward / risk) if risk > 0 else np.nan
    rr_str     = f"1:{round(rr_ratio, 1)}" if pd.notna(rr_ratio) else "-"

    return {
        "poi":          anchor_name,
        "anchor":       f"{anchor_name} @ {int(round(anchor))}",
        "entry":        f"{int(round(anchor))}",
        "target":       round(target, 0),
        "upside_pct":   round(upside_pct, 2) if pd.notna(upside_pct) else np.nan,
        "invalidation": round(invalidation, 0),
        "risk_reward":  rr_str,
    }

def _is_swing_bos_today(r: dict) -> bool:
    """
    Filter C: True if the latest swing structural event for today is a BOS
    (Break of Structure) — either Bullish or Bearish.
    Uses smc_latest_swing_struct field.
    """
    swing_struct = str(r.get("smc_latest_swing_struct","") or "").strip()
    if not swing_struct or swing_struct in ("-","N/A",""):
        return False
    # BOS pattern — not CHoCH
    is_bos   = "BOS" in swing_struct.upper()
    is_choch = "CHOCH" in swing_struct.upper() or "CHoCH" in swing_struct

    from datetime import date, timedelta
    # Check if the event date is today or yesterday (within 1 trading day)
    import re
    date_match = re.search(r'\d{4}-\d{2}-\d{2}', swing_struct)
    if date_match:
        try:
            event_date = date.fromisoformat(date_match.group())
            age_days   = (date.today() - event_date).days
            recent     = age_days <= 2   # allow T+1 lag
        except Exception:
            recent = True  # if can't parse, don't exclude
    else:
        recent = True  # no date encoded → field freshly set this run

    return is_bos and not is_choch and recent


def build_idx_screener_sheet(wb, latest_market_day: str, rows: list):
    """
    IDX Screener sheet — three filter categories + Institutional Trade Plan Engine.
    Filter A: Good ADTV (≥1B IDR) + Price > EMA 25 AND EMA 50
    Filter B: Golden Cross event (RSI or MACD)
    Filter C: Swing BOS detected today (SMC)
    Trade plan columns: POI | Anchor | Entry | Target | Upside % | Invalidation | R/R
    Expected output columns (spec): Ticker | Sector | Price | Chg % | RVOL |
                                    ADR & ATR (14) | Zone | MP Profile | MA Position
    """
    sheet_name = "IDX Screener"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 0.5

    FILL_HDR_A  = PatternFill("solid", fgColor="1F4E79")
    FILL_HDR_B  = PatternFill("solid", fgColor="375623")
    FILL_HDR_C  = PatternFill("solid", fgColor="7B3F00")
    FILL_HDR_D  = PatternFill("solid", fgColor="4A0080")   # deep purple — RRG
    FILL_ROW_A  = PatternFill("solid", fgColor="EBF3FB")
    FILL_ROW_B  = PatternFill("solid", fgColor="EBF5EB")
    FILL_ROW_C  = PatternFill("solid", fgColor="FEF3E2")
    FILL_ROW_D  = PatternFill("solid", fgColor="F3E8FF")
    FONT_HDR    = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    FONT_BODY   = Font(name="Calibri", size=10, color="1F1F1F")
    FONT_TICKER = Font(name="Calibri", size=10, bold=True, color="1F4E79")
    FONT_SEC    = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    AL_C        = Alignment(horizontal="center", vertical="center", wrap_text=True)
    AL_L        = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # Spec-defined columns + trade plan
    COLUMNS = [
        # key,                 header,            width, fmt
        ("filter_tag",         "Filter",           12,   None),
        ("ticker",             "Ticker",           10,   None),
        ("idx_sector",         "Sector",           20,   None),
        ("close",              "Price",            10,   "#,##0"),
        ("pct_change",         "Chg %",             9,   "0.00%"),
        ("rvol20",             "RVOL",              9,   "0.00"),
        ("adr_atr_zone",       "ADR & ATR (14)",   16,   None),
        ("q_remarks",          "Zone",             28,   None),
        ("mp_profile",         "MP Profile",       22,   None),
        ("summary_ma",         "MA Position",      24,   None),
        # Trade plan
        ("tp_poi",             "POI",              20,   None),
        ("tp_anchor",          "Anchor",           24,   None),
        ("tp_entry",           "Entry",            18,   None),
        ("tp_target",          "Target",           12,   "#,##0"),
        ("tp_upside_pct",      "Upside %",         10,   "0.00"),
        ("tp_invalidation",    "Invalidation",     14,   "#,##0"),
        ("tp_rr",              "R/R",               8,   None),
    ]
    KEYS   = [k for k,_,_,_ in COLUMNS]
    HDRS   = [h for _,h,_,_ in COLUMNS]
    WIDTHS = [w for _,_,w,_ in COLUMNS]
    FMTS   = [f for _,_,_,f in COLUMNS]

    from openpyxl.utils import get_column_letter
    for i, w in enumerate(WIDTHS, start=2):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Header row 2
    ws.row_dimensions[1].height = 6
    ws.row_dimensions[2].height = 28
    for i, hdr in enumerate(HDRS, start=2):
        c = ws.cell(2, i, hdr)
        c.fill = FILL_HDR_A; c.font = FONT_HDR; c.alignment = AL_C
    ws.freeze_panes = "B3"

    # ── Filters ───────────────────────────────────────────────────────────────
    def _not_suspended(r):
        """Exclude suspended tickers only. ARA/ARB tickers are OK."""
        return not r.get("suspended", False)

    def _good_adtv(r):
        """ADTV ≥ 5B IDR or Avg Volume 20D ≥ 5M shares (pre-filter gate)."""
        return (safe_num(r.get("adtr20"), 0) >= 5_000_000_000 or
                safe_num(r.get("adtv20"), 0) >= 5_000_000)

    def _above_ema25_50(r):
        cl  = safe_num(r.get("close"),  np.nan)
        e25 = safe_num(r.get("ema25"),  np.nan)
        e50 = safe_num(r.get("ema50"),  np.nan)
        return pd.notna(cl) and pd.notna(e25) and pd.notna(e50) and cl > e25 and cl > e50

    def _filter_d_poi_touch(r):
        """Filter D: candle low touches nearest POI and closes above it."""
        close = safe_num(r.get("close"), np.nan)
        low   = safe_num(r.get("last_low"), np.nan)   # today's candle low
        if pd.isna(close) or pd.isna(low):
            return False
        # Collect all POIs (same hierarchy as trade plan engine)
        pois = {}
        for prefix, label in [("pq","PrevQ"), ("py","PrevY")]:
            for key, suf in [("_vwap","VWAP"),("_m1","-1SD"),("_m2","-2SD")]:
                v = safe_num(r.get(f"{prefix}{key}"), np.nan)
                if pd.notna(v) and v > 0:
                    pois[label+suf] = v
        ob_eq = safe_num(r.get("smc_ob_equilibrium"), np.nan)
        if pd.notna(ob_eq): pois["Equilibrium"] = ob_eq
        e25 = safe_num(r.get("ema25"), np.nan)
        e50 = safe_num(r.get("ema50"), np.nan)
        if pd.notna(e25): pois["EMA25"] = e25
        if pd.notna(e50): pois["EMA50"] = e50

        # For each POI: candle low ≤ POI (touch) and close > POI (bounce)
        for _, poi_val in pois.items():
            if low <= poi_val * 1.005 and close > poi_val:
                return True
        return False

    def _filter_e_eq_breakout(r):
        """Filter E: price was ranging on Equilibrium ≥2 days then broke on MARKET_DATE."""
        return str(r.get("today_event","")).startswith("Price Break") and \
               "Equilibrium" in str(r.get("today_event",""))

    # Apply base gate: exclude suspended for all filters
    live_rows = [r for r in rows if _not_suspended(r)]

    # Pre-filter by ADTV gate (Spec §3a point 2)
    adtv_rows = [r for r in live_rows if _good_adtv(r)]

    filter_a = [(r,"A") for r in adtv_rows if _above_ema25_50(r)]
    filter_b = [(r,"B") for r in live_rows
                if str(r.get("cross_status","")).strip()=="Golden"
                or str(r.get("macd_cross","")).strip()=="Golden Cross"]
    filter_c = [(r,"C") for r in live_rows if _is_swing_bos_today(r)]
    filter_d = [(r,"D") for r in live_rows if _filter_d_poi_touch(r)]
    filter_e = [(r,"E") for r in live_rows if _filter_e_eq_breakout(r)]

    # ── Write helpers ─────────────────────────────────────────────────────────
    row_idx = 3

    def _section(label, fill, count):
        nonlocal row_idx
        ws.row_dimensions[row_idx].height = 22
        ws.merge_cells(start_row=row_idx, start_column=2,
                       end_row=row_idx, end_column=len(COLUMNS)+1)
        c = ws.cell(row_idx, 2, f"  {label}   ({count} tickers)")
        c.fill = fill; c.font = FONT_SEC; c.alignment = AL_L
        row_idx += 1

    def _data_row(r, tag, row_fill):
        nonlocal row_idx
        tp = _institutional_trade_plan(r)
        aug = dict(r)
        aug["filter_tag"]      = tag
        aug["tp_poi"]          = tp.get("poi","-")
        aug["tp_anchor"]       = tp.get("anchor","-")
        aug["tp_entry"]        = tp.get("entry","-")
        aug["tp_target"]       = tp.get("target", np.nan)
        aug["tp_upside_pct"]   = tp.get("upside_pct", np.nan)
        aug["tp_invalidation"] = tp.get("invalidation", np.nan)
        aug["tp_rr"]           = tp.get("risk_reward", np.nan)

        ws.row_dimensions[row_idx].height = 20
        for col_i, (key, fmt) in enumerate(zip(KEYS, FMTS), start=2):
            raw = aug.get(key, "-")
            val = raw
            if isinstance(val, float) and pd.isna(val):
                val = "-"
            if val in ("N/A","None",None):
                val = "-"
            c = ws.cell(row_idx, col_i, val)
            c.fill = row_fill
            c.font = FONT_TICKER if key == "ticker" else FONT_BODY
            c.alignment = AL_C
            if fmt and val != "-" and isinstance(raw, (int,float)) and pd.notna(raw):
                c.number_format = fmt
            # Upside % colour
            if key == "tp_upside_pct" and isinstance(raw,(int,float)) and pd.notna(raw):
                c.font = Font(name="Calibri", size=10,
                              color="008000" if raw >= 8 else ("E36C09" if raw >= 3 else "1F1F1F"))
            # Chg % colour
            if key == "pct_change" and isinstance(raw,(int,float)) and pd.notna(raw):
                c.font = Font(name="Calibri", size=10,
                              color="008000" if raw > 0 else ("C00000" if raw < 0 else "1F1F1F"))
        row_idx += 1

    # ── Sections ──────────────────────────────────────────────────────────────
    _section("Filter A  —  Price > EMA 25 & EMA 50  (ADTV ≥5B or AvgVol ≥5M gate applied)", FILL_HDR_A, len(filter_a))
    for r, tag in filter_a:
        _data_row(r, tag, FILL_ROW_A)
    if not filter_a:
        ws.cell(row_idx, 2, "No tickers matched Filter A.").font = FONT_BODY
        row_idx += 1
    row_idx += 1

    _section("Filter B  —  Golden Cross Event  (RSI or MACD)", FILL_HDR_B, len(filter_b))
    for r, tag in filter_b:
        _data_row(r, tag, FILL_ROW_B)
    if not filter_b:
        ws.cell(row_idx, 2, "No tickers matched Filter B.").font = FONT_BODY
        row_idx += 1
    row_idx += 1

    _section("Filter C  —  Swing BOS Detected on MARKET_DATE  (based on candle data)", FILL_HDR_C, len(filter_c))
    for r, tag in filter_c:
        _data_row(r, tag, FILL_ROW_C)
    if not filter_c:
        ws.cell(row_idx, 2, "No tickers matched Filter C.").font = FONT_BODY
        row_idx += 1
    row_idx += 1

    _section("Filter D  —  Candle Low Touched POI → Closed Above POI", FILL_HDR_D, len(filter_d))
    for r, tag in filter_d:
        _data_row(r, tag, FILL_ROW_D)
    if not filter_d:
        ws.cell(row_idx, 2, "No tickers matched Filter D.").font = FONT_BODY
        row_idx += 1
    row_idx += 1

    FILL_HDR_E = PatternFill("solid", fgColor="005050")
    FILL_ROW_E = PatternFill("solid", fgColor="E0F5F5")
    _section("Filter E  —  Price Ranging Equilibrium ≥2D then Broke on MARKET_DATE", FILL_HDR_E, len(filter_e))
    for r, tag in filter_e:
        _data_row(r, tag, FILL_ROW_E)
    if not filter_e:
        ws.cell(row_idx, 2, "No tickers matched Filter E.").font = FONT_BODY
        row_idx += 1

    # Title row 1
    ws.row_dimensions[1].height = 5
    ws.row_dimensions[2].height = 24
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 5
    ws.row_dimensions[6].height = 44

    _bt_label = " [BACKTEST REPLAY]" if BACKTEST_MODE else ""
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=len(COLUMNS)+1)
    t2 = ws.cell(2, 2, f"IDX Screener{_bt_label}")
    t2.font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
    t2.alignment = AL_L

    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=len(COLUMNS)+1)
    try:
        _pretty = pd.Timestamp(latest_market_day).strftime("%B %d, %Y")
    except Exception:
        _pretty = str(latest_market_day)
    t3 = ws.cell(3, 2, f"{'BACKTEST REPLAY as of ' if BACKTEST_MODE else 'Market Data as of '}{_pretty}")
    t3.font = Font(name="Calibri", size=10, italic=True, color="595959")
    t3.alignment = AL_L

    # Row 6: pre-filter gate info
    ws.merge_cells(start_row=6, start_column=2, end_row=6, end_column=len(COLUMNS)+1)
    gate_text = (
        "PRE-FILTER GATE (all filters):  ① Suspended tickers excluded  "
        "② ADTV ≥ 5B IDR or Avg Volume 20D ≥ 5M shares  "
        f"  |  A={len(filter_a)}  B={len(filter_b)}  C={len(filter_c)}  D={len(filter_d)}  E={len(filter_e)}"
    )
    t6 = ws.cell(6, 2, gate_text)
    t6.font = Font(name="Calibri", size=9, italic=True, color="374151")
    t6.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    print(f"[INFO] IDX Screener: A={len(filter_a)}, B={len(filter_b)}, C={len(filter_c)}, D={len(filter_d)}, E={len(filter_e)}")


def build_fundamental_key_stats_sheet(wb, latest_market_day: str, rows: list):
    """
    IDX Fundamental Detail — columnar table matching IDX Technical Detail architecture.
    Uses FUNDAMENTAL_DETAIL_SCHEMA with identical grouped-header engine.
    One row per ticker; Stock Info group leads, then fundamental sections.
    """
    sheet_name = "IDX Fundamental Detail"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 0.5

    # ── Row heights (mirror IDX Technical Detail) ─────────────────────────────
    for r_h, h in {1: 5, 2: 24, 3: 18, 4: 5, 5: 24, 6: 44, 7: 20, 8: 20, 9: 20, 10: 5, 11: 24, 12: 48}.items():
        ws.row_dimensions[r_h].height = h

    # ── Title rows ────────────────────────────────────────────────────────────
    _asof_ts   = get_effective_asof_date()
    _pretty    = _asof_ts.strftime("%B %d, %Y")
    _mode_lbl  = f"BACKTEST REPLAY as of {_pretty}" if BACKTEST_MODE else f"Data as of {_pretty}"
    _bt_note   = " — Note: yfinance fundamentals reflect latest available, not replay-adjusted" if BACKTEST_MODE else ""

    ws["B2"] = "IDX Fundamental Detail" + (" [BACKTEST REPLAY]" if BACKTEST_MODE else "")
    style_plain(ws["B2"], font=FONT_TITLE, align="left")
    ws["B3"] = (
        f"{_mode_lbl}  ·  yfinance  ·  IDX Universe{_bt_note}  ·  "
        ""
    )
    style_plain(ws["B3"], font=FONT_SUBTITLE, align="left")

    # Clear spacer row 4
    for _c in range(2, 250):
        ws.cell(4, _c).value = None

    # ── Flatten schema → flat_cols (identical engine to build_detail_sheet) ──
    flat_cols    = []
    col_idx      = 2
    group_ranges = []

    for group_name, group_fill, cols in FUNDAMENTAL_DETAIL_SCHEMA:
        start_idx = col_idx
        for key, header, width, fmt, align in cols:
            letter = get_column_letter(col_idx)
            ws.column_dimensions[letter].width = width
            flat_cols.append((col_idx, letter, group_name, group_fill, key, header, width, fmt, align))
            col_idx += 1
        end_idx = col_idx - 1
        group_ranges.append((group_name, group_fill, start_idx, end_idx))

    # ── Data vintage labels per group ────────────────────────────────────────
    # Maps group name → short vintage descriptor shown in the group header bar.
    # Sourced from yfinance conventions for IDX (BEI) companies.
    _VINTAGE = {
        "Stock Info":                    "Live Price  ·  IDX",
        "Company Profile":               "Latest Available  ·  yfinance",
        "Valuation":                     "TTM  ·  yfinance",
        "Profitability":                 "TTM  ·  yfinance",
        "Growth":                        "TTM vs Prior Year  ·  yfinance",
        "Balance Sheet":                 "Latest Quarterly Report  ·  yfinance",
        "Cash Flow":                     "TTM  ·  yfinance",
        "Dividend":                      "Latest & Upcoming  ·  yfinance",
        "PBV Band Analysis":             "5-Year Rolling  ·  Daily BVPS Forward-Filled  ·  yfinance",
        "Disclosure & Corporate Events": "Google RSS News  ·  news.google.com",
    }

    # ── Group header row (row 5) ──────────────────────────────────────────────
    for group_name, group_fill, start_idx, end_idx in group_ranges:
        vintage   = _VINTAGE.get(group_name, "yfinance")
        label     = f"{group_name}  ·  {vintage}"
        start_cell = f"{get_column_letter(start_idx)}5"
        end_cell   = f"{get_column_letter(end_idx)}5"
        ws.merge_cells(f"{start_cell}:{end_cell}")
        for c in range(start_idx, end_idx + 1):
            cell = ws.cell(5, c)
            cell.fill = group_fill
            cell.font = FONT_GROUP
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDER
        ws[start_cell] = label

    # ── Subheader row (row 6) ─────────────────────────────────────────────────
    for col_idx, letter, group_name, group_fill, key, header, width, fmt, align in flat_cols:
        c = ws.cell(6, col_idx)
        c.value = header
        style_cell(c, fill=group_fill, font=FONT_SUBHEADER, align="center", wrap=True)

    # ── Data rows ─────────────────────────────────────────────────────────────
    ok_rows   = [r for r in rows if r.get("data_status") in ("OK", "PARTIAL DATA")]
    _COMPACT_KEYS = {
        "yf_shares_outstanding", "yf_float_shares", "yf_market_cap", "yf_enterprise_value",
        "yf_total_assets", "yf_total_liabilities", "yf_cash", "yf_debt",
        "yf_operating_cf", "yf_free_cf", "yf_capex",
    }
    PBV_REGIME_COLORS = {
        "Deep Value":   "1D4ED8",
        "Undervalued":  "16A34A",
        "Fair Value":   "737373",
        "Premium":      "CA8A04",
        "Euphoric":     "DC2626",
    }

    for r_idx, screener_row in enumerate(ok_rows, start=7):
        combined = _build_fundamental_row(screener_row)

        for col_idx, letter, group_name, group_fill, key, header, width, fmt, align in flat_cols:
            c   = ws.cell(r_idx, col_idx)
            val = combined.get(key, np.nan)

            # Compact large-number formatting
            if key in _COMPACT_KEYS and val != "N/A" and pd.notna(val):
                val = compact_fmt(val)

            # NaN → "-"
            if isinstance(val, float) and pd.isna(val):
                val = "-"

            c.value = val
            style_cell(c, fill=None, font=FONT_BODY, align=align,
                       wrap=(key in ("yf_company_name", "yf_industry", "yf_business_summary", "yf_disclosure_title")))

            # Disclosure URL → clickable hyperlink
            if key == "yf_disclosure_url" and isinstance(val, str) and val.startswith("http"):
                c.hyperlink = val
                c.value = "View →"
                c.font = Font(name="Calibri", size=10, color="0563C1", underline="single")

            # % Change colour (green / red)
            if key == "pct_change":
                try:
                    _v = safe_num(combined.get("pct_change"), np.nan)
                    if pd.notna(_v):
                        c.number_format = "0.00%"
                        c.font = Font(name=FONT_BODY.name, size=FONT_BODY.size, bold=False,
                                      color="008000" if _v > 0 else ("C00000" if _v < 0 else "1F1F1F"))
                except Exception:
                    pass

            # PBV Regime colour
            if key == "yf_pbv_regime" and val not in ("-", "", None):
                rc = PBV_REGIME_COLORS.get(str(val), "1F1F1F")
                c.font = Font(name="Calibri", size=10, bold=True, color=rc)

            # Disclosure Sentiment colour
            if key == "yf_disclosure_sentiment":
                sent_colors = {"Bullish": "008000", "Bearish": "C00000", "Neutral": "595959"}
                rc = sent_colors.get(str(val), "1F1F1F")
                c.font = Font(name="Calibri", size=10, bold=True, color=rc)

            # Event Risk colour
            if key == "yf_event_risk":
                risk_colors = {"High": "C00000", "Medium": "E36C09", "Low": "008000"}
                rc = risk_colors.get(str(val), "1F1F1F")
                c.font = Font(name="Calibri", size=10, bold=True, color=rc)

            # Number format (only for true numeric cells)
            if fmt and val != "-" and fmt not in ("compact",):
                if isinstance(combined.get(key), (int, float)) and pd.notna(combined.get(key)):
                    c.number_format = fmt

        ws.row_dimensions[r_idx].height = 26

    # ── Freeze panes + autofilter (mirror IDX Technical Detail) ──────────────
    last_col_letter = get_column_letter(flat_cols[-1][0]) if flat_cols else "Z"
    ws.freeze_panes = "E7"
    ws.auto_filter.ref = f"B6:{last_col_letter}{max(6, ws.max_row)}"



# =========================
# BACKTEST ENGINE  (Phase D — bar-by-bar historical replay)
# =========================
def _run_backtest_replay(rows_cache: dict) -> dict:
    """
    Bar-by-bar replay over BACKTEST_TICKERS.
    rows_cache: {ticker: pd.DataFrame} with full OHLCV history.
    Returns a dict of result DataFrames for each backtest output sheet.
    """
    import warnings
    warnings.filterwarnings("ignore")

    results_by_date = {}   # date -> list of scored rows
    regime_log     = []    # (date, ticker, regime)
    signal_log     = []    # (date, ticker, signal, score, macd_state)

    if not rows_cache:
        return {}

    # Build unified date range from available data
    all_dates = set()
    for df in rows_cache.values():
        if df is not None and not df.empty:
            all_dates.update(df.index.normalize().unique())
    if not all_dates:
        return {}

    try:
        start_dt = pd.Timestamp("2020-01-01")   # fixed historical start for bar-by-bar replay
        end_dt   = get_market_date()             # MARKET_DATE is the replay boundary
    except Exception:
        start_dt = pd.Timestamp("2020-01-01")
        end_dt   = pd.Timestamp.now().normalize()

    date_range = sorted(d for d in all_dates if start_dt <= d <= end_dt)
    if not date_range:
        return {}

    # Sample every 5th trading day for speed (configurable)
    STEP = 5
    sampled_dates = date_range[::STEP]

    for snap_date in sampled_dates:
        day_rows = []
        for ticker, full_hist in rows_cache.items():
            if full_hist is None or full_hist.empty:
                continue
            # Strict as-of: only use bars <= snap_date
            hist = full_hist[full_hist.index.normalize() <= snap_date].copy()
            if len(hist) < MIN_BARS_PARTIAL:
                continue
            try:
                close = safe_num(hist["Close"].iloc[-1])
                if pd.isna(close) or close <= 0:
                    continue

                # MACD
                macd_data = compute_macd_momentum(hist) if len(hist) >= 35 else {}
                # SMC
                smc_data  = compute_smc_engine(hist) if len(hist) >= 60 else {}
                # VWAP
                cq_start = quarter_start(snap_date)
                df_q = hist.loc[hist.index >= cq_start]
                q = anchored_vwap_block(df_q) if not df_q.empty else {}

                # Score (simple composite for backtest)
                score = 0.0
                swing = smc_data.get("smc_swing_trend", "Neutral")
                macd_entry = macd_data.get("macd_entry", "")
                if swing == "Bullish":          score += 3.0
                if "ENTRY SIGNAL" in macd_entry: score += 3.0
                if "Continuation" in macd_entry: score += 2.0
                pd_zone = smc_data.get("smc_premium_discount", "")
                if pd_zone == "Discount":       score += 2.0
                if pd_zone == "Equilibrium":    score += 1.0
                if pd_zone == "Premium":        score -= 1.0

                composite = smc_data.get("smc_composite_state", "Neutral Rotation")
                regime_log.append({
                    "date": snap_date.strftime("%Y-%m-%d"),
                    "ticker": ticker,
                    "regime": composite,
                    "swing_trend": swing,
                    "macd_regime": macd_data.get("macd_regime", ""),
                    "pd_zone": pd_zone,
                    "score": round(score, 2),
                })

                if "ENTRY SIGNAL" in macd_entry or score >= 5.0:
                    signal_log.append({
                        "date": snap_date.strftime("%Y-%m-%d"),
                        "ticker": ticker,
                        "signal": macd_entry,
                        "score": round(score, 2),
                        "smc_state": composite,
                        "pd_zone": pd_zone,
                        "close": close,
                        "qvwap": q.get("vwap", np.nan) if q else np.nan,
                    })

                day_rows.append({
                    "ticker": ticker,
                    "close": close,
                    "score": round(score, 2),
                    "composite": composite,
                    "swing": swing,
                    "macd": macd_entry,
                })
            except Exception:
                continue

        if day_rows:
            day_rows.sort(key=lambda x: x["score"], reverse=True)
            results_by_date[snap_date.strftime("%Y-%m-%d")] = day_rows

    # ── Build result DataFrames ──────────────────────────────────────────────
    df_regime  = pd.DataFrame(regime_log)  if regime_log  else pd.DataFrame()
    df_signal  = pd.DataFrame(signal_log)  if signal_log  else pd.DataFrame()

    # Historical Rankings: top 5 per snapshot date
    ranking_rows = []
    for snap_date_str, rows_day in results_by_date.items():
        for rank_pos, r in enumerate(rows_day[:5], 1):
            ranking_rows.append({"date": snap_date_str, "rank": rank_pos, **r})
    df_rankings = pd.DataFrame(ranking_rows) if ranking_rows else pd.DataFrame()

    # Hit Rate Analytics: signals that were followed by 5%+ return within 10 days
    hit_rows = []
    for sig in signal_log:
        ticker = sig["ticker"]
        sig_date = pd.Timestamp(sig["date"])
        full_hist = rows_cache.get(ticker)
        if full_hist is None or full_hist.empty:
            continue
        entry_close = sig["close"]
        future = full_hist[full_hist.index.normalize() > sig_date].head(10)
        if future.empty or pd.isna(entry_close) or entry_close <= 0:
            continue
        max_ret = float((future["Close"].max() - entry_close) / entry_close * 100)
        hit = max_ret >= 5.0
        hit_rows.append({
            "date": sig["date"], "ticker": ticker,
            "signal": sig["signal"], "score": sig["score"],
            "entry_close": entry_close, "max_return_10d_pct": round(max_ret, 2),
            "hit": "YES" if hit else "NO",
        })
    df_hits = pd.DataFrame(hit_rows) if hit_rows else pd.DataFrame()

    # Equity Curve: daily portfolio value (equal-weight top-5 per snapshot)
    equity_rows = []
    portfolio_val = 100.0   # start at index 100
    prev_closes = {}
    for snap_date_str in sorted(results_by_date.keys()):
        top5 = results_by_date[snap_date_str][:5]
        if top5 and prev_closes:
            total_ret = 0.0
            counted = 0
            for r in top5:
                prev = prev_closes.get(r["ticker"])
                if prev and prev > 0:
                    total_ret += (r["close"] - prev) / prev
                    counted += 1
            if counted > 0:
                portfolio_val *= (1 + total_ret / counted)
        equity_rows.append({"date": snap_date_str, "portfolio_value": round(portfolio_val, 4)})
        for r in top5:
            prev_closes[r["ticker"]] = r["close"]
    df_equity = pd.DataFrame(equity_rows) if equity_rows else pd.DataFrame()

    # Summary stats
    total_signals = len(signal_log)
    total_hits    = df_hits["hit"].eq("YES").sum() if not df_hits.empty and "hit" in df_hits.columns else 0
    win_rate      = total_hits / total_signals * 100 if total_signals > 0 else 0.0
    avg_return    = float(df_hits["max_return_10d_pct"].mean()) if not df_hits.empty and "max_return_10d_pct" in df_hits.columns else 0.0

    eq_vals = df_equity["portfolio_value"].values if not df_equity.empty and "portfolio_value" in df_equity.columns else [100.0]
    peak = 100.0
    max_dd = 0.0
    for v in eq_vals:
        if v > peak:
            peak = v
        dd = (peak - v) / peak * 100 if peak > 0 else 0
        if dd > max_dd:
            max_dd = dd

    expectancy = win_rate/100 * avg_return - (1 - win_rate/100) * abs(avg_return * 0.5) if total_signals > 0 else 0.0

    # Sharpe proxy (annualised returns / std)
    if not df_equity.empty and len(df_equity) > 2 and "portfolio_value" in df_equity.columns:
        returns_ser = df_equity["portfolio_value"].pct_change().dropna()
        ann_ret = float(returns_ser.mean() * 252)
        ann_std = float(returns_ser.std() * (252**0.5))
        sharpe_proxy = ann_ret / ann_std if ann_std > 0 else 0.0
    else:
        ann_ret = 0.0; ann_std = 0.0; sharpe_proxy = 0.0

    # Signal frequency
    if sampled_dates:
        year_span = (sampled_dates[-1] - sampled_dates[0]).days / 365.25
        sig_freq = total_signals / year_span if year_span > 0 else 0.0
    else:
        sig_freq = 0.0

    df_summary = pd.DataFrame([{
        "Metric": "Total Signals",          "Value": total_signals},
        {"Metric": "Total Hits (≥5% / 10D)","Value": total_hits},
        {"Metric": "Win Rate %",             "Value": round(win_rate, 2)},
        {"Metric": "Avg Return (10D max) %", "Value": round(avg_return, 2)},
        {"Metric": "Max Drawdown %",         "Value": round(max_dd, 2)},
        {"Metric": "Expectancy %",           "Value": round(expectancy, 2)},
        {"Metric": "Sharpe Proxy",           "Value": round(sharpe_proxy, 2)},
        {"Metric": "Signal Frequency / Yr",  "Value": round(sig_freq, 1)},
        {"Metric": "Annualised Return %",    "Value": round(ann_ret * 100, 2)},
        {"Metric": "Snapshots Evaluated",    "Value": len(sampled_dates)},
    ])

    return {
        "summary":  df_summary,
        "rankings": df_rankings,
        "signals":  df_signal,
        "regime":   df_regime,
        "equity":   df_equity,
        "hits":     df_hits,
    }


def _write_df_to_ws(ws, df: pd.DataFrame, title: str, col_start: int = 2):
    """Generic helper: write a DataFrame to a worksheet with header row."""
    if df is None or df.empty:
        ws.cell(2, col_start).value = f"{title} — No data available"
        style_plain(ws.cell(2, col_start), font=FONT_SUBTITLE, align="left")
        return
    # Title row
    ws.merge_cells(start_row=2, start_column=col_start,
                   end_row=2,  end_column=col_start + len(df.columns) - 1)
    tc = ws.cell(2, col_start)
    tc.value = title
    style_plain(tc, font=FONT_TITLE, align="left")
    # Header
    for ci, col in enumerate(df.columns, start=col_start):
        hc = ws.cell(3, ci)
        hc.value = str(col).replace("_", " ").title()
        style_cell(hc, fill=FILL_HEADER, font=FONT_HEADER, align="center")
        ws.column_dimensions[get_column_letter(ci)].width = max(14, len(str(col)) + 4)
    # Data rows
    for ri, (_, row) in enumerate(df.iterrows(), start=4):
        row_fill = PatternFill("solid", fgColor="F8FAFC" if ri % 2 == 0 else "FFFFFF")
        for ci, val in enumerate(row.values, start=col_start):
            dc = ws.cell(ri, ci)
            raw = val
            if isinstance(val, float) and np.isnan(val):
                raw = "N/A"
            dc.value = raw
            style_cell(dc, fill=row_fill, font=FONT_BODY, align="center")


def build_backtest_sheets(wb, bt_results: dict):
    """
    Build 6 backtest output sheets from bt_results dict.
    Skips gracefully if any result df is missing.
    """
    SHEET_DEFS = [
        ("BT Summary",          "summary",  "Backtest Summary — Performance Metrics"),
        ("BT Historical Ranks", "rankings", "Historical Rankings — Top 5 by Snapshot Date"),
        ("BT Signal Replay",    "signals",  "Signal Replay — All Entry Signals Generated"),
        ("BT Regime Log",       "regime",   "Regime Transition Log — SMC State by Date"),
        ("BT Equity Curve",     "equity",   "Equity Curve — Portfolio Index (Base 100)"),
        ("BT Hit Rate",         "hits",     "Hit Rate Analytics — Signal Outcomes"),
    ]
    for sheet_name, key, title in SHEET_DEFS:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name)
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 0.5
        ws.row_dimensions[1].height = 4
        df = bt_results.get(key, pd.DataFrame())
        _write_df_to_ws(ws, df, title, col_start=2)


# =========================
# GUIDE & LOGIC REFERENCE  (Full rewrite — reflects all v2 engines)
# =========================
def build_guide_sheet(ws):
    """
    Guide & Logic Reference — comprehensive column-by-column documentation.
    Structure: Group header → Column | Formula | Interpretation | Bullish Signal | Bearish Signal
    """
    ws.title = "Guide & Logic Reference"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 0.5
    ws.column_dimensions["B"].width = 28   # Column
    ws.column_dimensions["C"].width = 42   # Formula / Logic
    ws.column_dimensions["D"].width = 52   # Interpretation
    ws.column_dimensions["E"].width = 36   # Bullish Signal
    ws.column_dimensions["F"].width = 36   # Bearish Signal

    FILL_GROUP  = PatternFill("solid", fgColor="1F4E79")
    FILL_COL_HDR= PatternFill("solid", fgColor="2E75B6")
    FILL_ALT    = PatternFill("solid", fgColor="EBF3FB")
    FONT_GRP    = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    FONT_COL    = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    FONT_BODY   = Font(name="Calibri", size=9,  color="1F1F1F")
    AL_C        = Alignment(horizontal="center", vertical="center", wrap_text=True)
    AL_L        = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    r = 2

    def _group(title):
        nonlocal r
        ws.row_dimensions[r].height = 26
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        c = ws.cell(r, 2, f"  {title}")
        c.fill = FILL_GROUP; c.font = FONT_GRP; c.alignment = AL_L
        r += 1
        # Column headers
        ws.row_dimensions[r].height = 20
        for ci, hdr in enumerate(["Column Name","Formula / Logic","Definition","How to Use",""], 2):
            ch = ws.cell(r, ci, hdr)
            ch.fill = FILL_COL_HDR; ch.font = FONT_COL; ch.alignment = AL_C
        r += 1

    def _row(col, formula, interp, bull, bear, alt=False):
        nonlocal r
        ws.row_dimensions[r].height = 48
        fill = FILL_ALT if alt else None
        vals = [col, formula, interp, bull, bear]
        for ci, v in enumerate(vals, 2):
            c = ws.cell(r, ci, v)
            c.font = FONT_BODY; c.alignment = AL_L
            if fill: c.fill = fill
        r += 1

    # ── STOCK INFO ─────────────────────────────────────────────────────────────
    _group("1 · STOCK INFO")
    _row("Ticker","IDX stock code","Unique exchange identifier for the security","","")
    _row("Closing Price","Last traded close price (IDR)","Current market price; basis for all % diff calculations","Rising session close","Gap-down close")
    _row("Price Change %","(Close − Prev Close) / Prev Close","Daily price momentum; positive = buying pressure","≥ +1%","≤ −1%")
    _row("Beta (vs IHSG)","COVARIANCE.P(stock_ret, IHSG_ret) / VAR.P(IHSG_ret) — 252-day rolling","Market sensitivity. B>1 = amplified moves vs IHSG. B<0 = counter-cyclical","0.8-1.5 (healthy co-movement)","B>2 (excessive volatility) or B<0 (decoupled)")
    _row("Emiten","Company full name","Issuer name for cross-referencing","","")
    _row("IDX Sector","IDX official sector classification","Sector used for RRG rotation analysis","","")

    # ── OWNERSHIP ─────────────────────────────────────────────────────────────
    _group("2 · OWNERSHIP (KSEI)")
    _row("Investors","Top investor names from KSEI","Identifies institutional/strategic holders","Named fund/institution","Anonymous nominee")
    _row("Free Float","% shares freely tradeable","High FF = institutional accessibility","FF > 40%","FF < 10% (illiquid float)")
    _row("Classic HHI","Σ(share_i²) × 10,000","Herfindahl-Hirschman concentration index. >2500 = concentrated","500–2500 (moderate spread)","HHI > 2500 (dominated)")
    _row("CR1","Top-1 holder concentration %","Single largest holder; CR1>50% signals controlling shareholder","","CR1 > 70% (overhang risk)")
    _row("CR3","Top-3 holder concentration %","Combined top-3 weight; proxy for block-holder overhang","","CR3 > 85%")
    _row("CCS","Closely Correlated Shareholders ratio","Identifies aligned block holders acting as a unit","Low CCS","CCS > 0.6 (coordinated)")
    _row("Ownership Type","Conglomerate / SOE / Family / Public","Governance context; affects dividend policy and dilution risk","Public / Institutional","Family-controlled + high CR1")

    # ── STOCK REGIME ──────────────────────────────────────────────────────────
    _group("3 · STOCK REGIME")
    _row("Market Cap","Shares × Close (IDR)","Size tier: LargeCap ≥ 10T, MidCap 1–10T, SmallCap < 1T","LargeCap (institutional eligible)","SmallCap < 500B (thin liquidity)")
    _row("Market Cap Category","LargeCap / MidCap / SmallCap","Determines applicable screening thresholds","LargeCap / MidCap","MicroCap (< 200B)")
    _row("Liquidity Category","High / Medium / Low (ADTV-based)","Governs position-sizing deployability","High (ADTV > 10B)","Low (ADTV < 1B)")
    _row("Verdict Weight Profile","Composite regime label","Aggregates size + liquidity + ownership into a single tradeable profile","Institutional Grade","Speculative / Thin")

    # ── MARKET STRUCTURE & SMC ────────────────────────────────────────────────
    _group("4 · MARKET STRUCTURE & SMC")
    _row("Internal Trend","Direction of internal (short-swing) market structure","Bullish = HH/HL sequence on internal pivots; Bearish = LH/LL","Bullish","Bearish")
    _row("Swing Trend","Direction of major swing market structure","Primary trend bias. Overrides internal on conflict.","Bullish","Bearish")
    _row("Latest Internal Struct","Most recent internal BOS or CHoCH label + date","Identifies the latest internal structural event","BOS Bullish","CHoCH Bearish")
    _row("Latest Swing Struct","Most recent swing BOS or CHoCH label + date","Identifies the latest swing structural shift","BOS Bullish","CHoCH Bearish")
    _row("Strong High","Most recent confirmed swing high (unmitigated)","Resistance level; break above = swing BOS Bullish","Price approaching from below","Price rejected at level")
    _row("Weak High","Highest high of declining sequence","Liquidity magnet above price; targeted by MM sweeps","","Price approaching Weak High")
    _row("Strong Low","Most recent confirmed swing low (unmitigated)","Support level; break below = swing BOS Bearish","Price holding above","Price breaks below")
    _row("Weak Low","Lowest low of rising sequence","Liquidity pool below; targeted in stop-hunts","","Price approaching Weak Low")
    _row("Closest OB Bull","Nearest Bullish Order Block range [Low–High]","Key institutional demand zone; price tends to react here","Price reacting at OB Bull","Price slices through OB Bull")
    _row("Equilibrium","Midpoint of Closest OB Bull range","50% retracement of OB; institutional re-entry reference","Price bounces at Equilibrium","Price closes below Equilibrium")
    _row("Closest OB Bear","Nearest Bearish Order Block range [Low–High]","Institutional supply zone; expect distribution/rejection","","Price entering OB Bear zone")

    # ── LIQUIDITY ─────────────────────────────────────────────────────────────
    _group("5 · LIQUIDITY")
    _row("Lot","Volume ÷ 100","IDX board lot count; 1 lot = 100 shares","","")
    _row("Value (Approx)","(Open+High+Low+Close)/4 × Volume","Rupiah-denominated traded value; OHLC4 Typical Price × Volume","Value > 5B IDR","Value < 500M IDR")
    _row("Average Value 20D (Approx)","Rolling 20D mean of OHLC4 × Volume","Average daily Rupiah turnover; institutional deployability gate","Avg Value > 10B IDR","Avg Value < 1B IDR")
    _row("Volume","Raw shares traded","Raw participation; context only without RVOL","","")
    _row("Average Volume 20D","SMA(Volume, 20)","Baseline participation level","","")
    _row("RVOL 20D","Volume / Avg Volume 20D","Relative Volume = current vol vs 20D baseline. >1.5 = elevated activity","RVOL > 1.5","RVOL < 0.5")
    _row("RVOL 20D Zone","Above 1.5 / Below 1.5","Binary label for institutional scan filters","Above 1.5","Below 1.5")
    _row("ADR %","SMA(High−Low, 14) / Close × 100","Average Daily Range %; volatility proxy","ADR > 3% (tradeable swing)","ADR < 1% (dead range)")
    _row("ATR (14) %","ATR(14) / Close × 100 (Wilder EMA)","True Range volatility; accounts for gaps","ATR > 3%","ATR < 1%")
    _row("ADR & ATR (14) Zone","Above 3% if both ADR and ATR > 3%","Combined volatility gate for swing eligibility","Above 3%","Below 3%")

    # ── VWAP SECTIONS ─────────────────────────────────────────────────────────
    _group("6 · VWAP ZONES  (Current QVWAP / Previous QVWAP / Previous Year VWAP)")
    _row("VWAP","Anchored VWAP from period start = Σ(OHLC4×Vol) / Σ(Vol)","Volume-weighted average price; institutional cost basis for the period","Price above VWAP","Price below VWAP")
    _row("-1 SD / -2 SD / -3 SD","VWAP − n × StdDev(price, volume-weighted)","Standard deviation bands below VWAP; discount zones","Price at -1SD to -2SD (value)","Price at -3SD (distressed)")
    _row("+1 SD / +2 SD (computed)","VWAP + n × StdDev(price, volume-weighted)","SD bands above VWAP; premium zones","","Price at +2SD or higher (overextended)")
    _row("SD Score","(Close − VWAP) / StdDev","Normalised position relative to VWAP. 0 = at VWAP, −1 = at −1SD, +2 = at +2SD","−0.5 to +0.5 (balanced)","< −2 or > +2 (extreme)")
    _row("SD Δ 1D","Today SD Score − Yesterday SD Score","Momentum of SD movement; positive = moving toward premium","Rising toward 0 from negative","Falling from positive")
    _row("VWAP Zone","Price Near [ZONE] • X.XX% if SD Score within ±0.1 of integer level; else '-'","Proximity label to nearest VWAP SD band. Used for entry timing.","Price Near VWAP • <2%","Price Near +2 SD (overextended)")
    _row("VWAP Zone Days","Consecutive days price has closed in current zone","Zone persistence metric; high = institutional conviction","≥ 3 days in discount zone","Repeated rejection at premium zone")

    # ── MOVING AVERAGE ────────────────────────────────────────────────────────
    _group("7 · MOVING AVERAGE")
    _row("EMA 25","EMA(Close, 25) (Exponential weighted, span=25)","Short-to-medium term trend filter. Price > EMA 25 = short-term momentum intact","Price > EMA 25","Price < EMA 25")
    _row("EMA 25 %diff","(Close − EMA25) / EMA25 × 100","% distance from EMA 25; overextension warning","0–5% (healthy)","< 0% (below MA) or > 15% (stretched)")
    _row("EMA 25 Pos","Above / Below","Binary directional label","Above","Below")
    _row("EMA 50","EMA(Close, 50)","Medium-term trend; institutional momentum reference","Price > EMA 50","Price < EMA 50")
    _row("EMA 50 %diff","(Close − EMA50) / EMA50 × 100","% distance from EMA 50","0–8%","< 0%")
    _row("EMA 50 Pos","Above / Below","Binary directional label","Above","Below")
    _row("SMA 200","SMA(Close, 200)","Long-term secular trend; institutional positioning anchor. Golden/Death Cross reference MA","Price > SMA 200","Price < SMA 200 (secular bear)")
    _row("SMA 200 %diff","(Close − SMA200) / SMA200 × 100","% distance from SMA 200","0–20% (healthy uptrend)","< 0% (secular bear)")
    _row("SMA 200 Pos","Above / Below","Long-term regime label","Above","Below")
    _row("MA Zone","Summary label: Above All MA / Above EMA20 ▲ EMA25 ▲ EMA50 ▲ SMA200 / etc.","4-tier MA alignment score. Tier 4 = full bull stack","Above All MA (Tier 4)","Below All MA (Tier 0)")

    # ── RSI MOMENTUM ──────────────────────────────────────────────────────────
    _group("8 · RSI MOMENTUM")
    _row("RSI 14","RSI(Close, 14) — Wilder smoothing","Momentum oscillator 0–100. >70 = Overbought, <30 = Oversold","40–60 (rising) or >50 recovery","<40 declining or <30 extreme")
    _row("RSI 14 Δ 1D","RSI today − RSI yesterday","Day-over-day RSI momentum","Positive delta rising above 50","Negative delta falling below 50")
    _row("RSI Status","Overbought / Oversold / Strong / Neutral / Weak","Regime label based on RSI thresholds","Strong (55–70)","Oversold (<30) or Weak (<45)")
    _row("RSI MA 14","SMA(RSI14, 14)","Smoothed RSI signal line; used for cross detection","","")
    _row("RSI Pos","Above / Below RSI MA 14","Price RSI position vs its own MA. Above = momentum confirmed","Above","Below")
    _row("RSI Cross","Golden / Dead / -","Golden = RSI crosses above RSI MA (bullish momentum trigger) Dead = RSI crosses below RSI MA","Golden","Dead")
    _row("Divergence Signal","Bull Div / Bear Div / Hidden Bull / Hidden Bear / -","Price-RSI divergence; early reversal signal","Bull Div (price lower low, RSI higher low)","Bear Div (price higher high, RSI lower high)")

    # ── MACD MOMENTUM ─────────────────────────────────────────────────────────
    _group("9 · MACD MOMENTUM  (MACD 4C Smooth — 1:1 Pine Script port)")
    _row("MACD Line","EMA(12) − EMA(26) Rounded to integer","Fast momentum line; positive = bullish bias","Positive and rising","Negative and falling")
    _row("Signal Line","EMA(MACD, 9) Rounded to integer","Smoothed MACD trigger line","MACD > Signal and rising","MACD < Signal and falling")
    _row("Histogram (EMA3)","EMA(MACD − Signal, 3) Rounded to integer (smoothed histogram)","Momentum acceleration; colour-coded 4C: Blue=NegRise, Aqua=PosRise, Red=PosFall, Maroon=NegFall","Blue bar (NegRise) = early recovery entry","Maroon bar (NegFall) = deepening sell-off")
    _row("Lines Position","Both Above Zero / Both Below Zero / Mixed","MACD and Signal line zero-line context","Both Above Zero","Both Below Zero")
    _row("Wave Pattern","Mountain Building / Mountain Declining / Valley Deepening / Valley Recovering","Histogram wave shape; describes momentum phase","Mountain Building (positive expanding)","Valley Deepening (negative expanding)")
    _row("MACD Cross","Golden Cross / Dead Cross / -","Golden = MACD crosses above Signal line Dead = MACD crosses below Signal line","Golden Cross","Dead Cross")

    # ── DISCLOSURE & CORPORATE EVENTS ─────────────────────────────────────────
    _group("10 · DISCLOSURE & CORPORATE EVENTS  (Google News RSS)")
    _row("Latest Disclosure Date","Date of most recent Google News headline selected by MARKET_DATE","Recency indicator; fresh corporate headline = potential volatility","≤ 7 days (fresh catalyst)","Stale > 90 days")
    _row("Latest Disclosure Title","Headline text from Google News RSS","Keywords reveal corporate action type","Dividend / Acquisition announcement","Rights Issue / Bankruptcy / Restructuring")
    _row("Latest Disclosure URL","Hyperlink to Google News RSS item","Direct access to the source headline","","")
    _row("Disclosure Category","Dilution / Expansion / Shareholder Return / Liquidity Positive / Transitional / Distress / Corporate Action / General","Keyword-classified corporate action type","Shareholder Return / Liquidity Positive","Dilution / Distress")
    _row("Disclosure Sentiment","Bullish / Neutral / Bearish","Sentiment impact from disclosure keyword mapping","Bullish","Bearish")
    _row("Days Since Disclosure","MARKET_DATE − Latest Disclosure Date; latest item used if MARKET_DATE has no match","Age of most recent disclosure","< 14 days","")
    _row("Event Risk","High / Medium / Low","Rights Issue / Bankruptcy = High Acquisition / Restructuring = Medium Dividend / Stock Split = Low","Low","High")

    # ── DIVIDEND FRAMEWORK ────────────────────────────────────────────────────
    _group("11 · DIVIDEND  (Current · Upcoming · History)")
    _row("Entry 1: Year","Upcoming / Calendar year of dividend","Upcoming = not yet paid; otherwise most recent completed year","Upcoming label","No upcoming")
    _row("Entry 1: IDR","Dividend amount in IDR per share","Full cash return per share for this period","Growing vs prior year","Declining or zero")
    _row("Entry 1: Ex Date","Ex-dividend date for this entry","Own shares before this date to receive dividend","Known date current/next quarter","Missing")
    _row("Entry 1: Pay Date","Payment date for this entry","Cash distribution date to registered holders","","Unknown")
    _row("Entry 2: Year/IDR/Ex Date/Pay Date","Second most recent dividend period","Entry 1 upcoming → Entry 2 = latest paid. Entry 1 latest → Entry 2 = prior year.","Consistent pattern","Irregular / skipped")

    # ── PBV BAND FRAMEWORK ────────────────────────────────────────────────────
    _group("12 · PBV BAND ANALYSIS  (Stockbit-Style)")
    _row("Current PBV","Market Cap / Book Value of Equity (latest)","Current price-to-book multiple; valuation benchmark","< PBV Mean (5Y)","> PBV +2 SD (euphoric)")
    _row("PBV Mean (5Y)","5-year rolling mean of quarterly PBV","Long-run fair value anchor","","")
    _row("PBV +1 SD / +2 SD","PBV Mean + n × StdDev(PBV, 5Y)","Premium valuation bands","","Above +2 SD (richly valued)")
    _row("PBV -1 SD / -2 SD","PBV Mean − n × StdDev(PBV, 5Y)","Discount valuation bands","Below -1 SD (value zone)","Below -2 SD (distressed valuation)")
    _row("PBV Z-Score","(Current PBV − PBV Mean) / PBV StdDev","Standardised deviation from historical PBV mean. 0 = Fair Value, −2 = Deep Value, +2 = Euphoric","Z-Score < 0 (below mean)","Z-Score > 2 (extreme premium)")
    _row("PBV Percentile","Percentile rank of Current PBV in 5Y distribution","Relative valuation ranking vs own history","< 25th pct (historical discount)","≥ 75th pct (historical premium)")
    _row("PBV Regime","Deep Value / Value / Fair Value / Premium / Euphoric","Categorical valuation label from SD bands","Deep Value or Value","Euphoric")

    # ── IDX SCREENER FILTERS ──────────────────────────────────────────────────
    _group("13 · IDX SCREENER SHEET  (Filter Logic)")
    _row("Filter A","adtr20 >= 1B IDR AND Close > EMA 25 AND Close > EMA 50","Institutional liquidity gate + dual MA trend alignment","Both met","Either fails")
    _row("Filter B","RSI Cross = Golden OR MACD Cross = Golden Cross","Event-driven trigger; fresh cross = early momentum ignition","Either Golden Cross","Dead Cross")
    _row("Filter C","Latest Swing BOS within past 2 trading days (SMC engine)","Structure break confirming directional intent; most timely signal","BOS Bullish","BOS Bearish / none")
    _row("POI","Nearest support Point of Interest below current price","Structural level anchoring the trade plan","Named POI near price","No POI")
    _row("Anchor","POI name + level","Best support acting as entry reference","Closest POI below price","Anchor far from price")
    _row("Entry","anchor to anchor x 1.01 (1% band)","Actionable entry zone; buy within this band","Band near anchor","Band far from price")
    _row("Target","Next higher POI above current price","Primary take-profit from VWAP SD or Bear OB","Upside >= 8%","Upside < 3%")
    _row("Upside %","(Target / Close - 1) x 100","Expected return to target",">=8%","<3%")
    _row("Invalidation","Anchor - ATR(14) proxy","Structural stop; close below = setup invalid","Tight stop = high R/R","Wide stop = poor R/R")
    _row("R/R","(Target - Close) / (Close - Invalidation)","Risk-Reward ratio for the trade setup","R/R >= 2.0","R/R < 1.0")

    ws.row_dimensions[1].height = 32
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=6)
    title = ws.cell(1, 2, "  IDX Screener — Guide & Logic Reference")
    title.font      = Font(name="Calibri", size=14, bold=True, color="1F4E79")
    title.alignment = Alignment(horizontal="left", vertical="center")



if __name__ == "__main__":
    try:
        main()
    except Exception:
        import traceback
        print("\n[FATAL ERROR]")
        traceback.print_exc()
        input("\nPress Enter to exit...")
