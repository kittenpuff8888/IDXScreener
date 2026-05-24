from __future__ import annotations

import argparse
import json
import math
import re
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "Output"
DOCS_DIR = ROOT / "docs"
DATA_DIR = DOCS_DIR / "data"
DOWNLOADS_DIR = DOCS_DIR / "downloads"

SCREENER_COLUMNS = [
    "Filter",
    "Ticker",
    "Sector",
    "Price",
    "Chg %",
    "RVOL",
    "ADR & ATR (14)",
    "Zone",
    "MP Profile",
    "MA Position",
    "POI",
    "Anchor",
    "Entry",
    "Target",
    "Upside %",
    "Invalidation",
    "R/R",
]


def clean_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return None
        return value
    return value


def slug_date(value: str) -> str:
    parsed = datetime.strptime(value, "%Y-%m-%d")
    return parsed.strftime("%Y-%m-%d")


def latest_workbook() -> Path:
    files = [
        p
        for p in OUTPUT_DIR.glob("IDX Screener as of *.xlsx")
        if not p.name.startswith("~$")
    ]
    if not files:
        raise FileNotFoundError(f"No IDX Screener workbook found in {OUTPUT_DIR}")
    return max(files, key=lambda p: p.stat().st_mtime)


def find_stat(ws, label: str) -> Any:
    for row in ws.iter_rows(values_only=True):
        cells = list(row)
        for index, value in enumerate(cells):
            if str(value or "").strip().lower() == label.lower():
                return cells[index + 1] if index + 1 < len(cells) else None
    return None


def parse_table(ws, header_row: int, data_start: int) -> list[dict[str, Any]]:
    headers = [clean_value(ws.cell(header_row, col).value) for col in range(2, ws.max_column + 1)]
    rows: list[dict[str, Any]] = []
    for row_idx in range(data_start, ws.max_row + 1):
        item: dict[str, Any] = {}
        has_value = False
        for offset, header in enumerate(headers, start=2):
            if not header:
                continue
            value = clean_value(ws.cell(row_idx, offset).value)
            item[str(header)] = value
            if value not in (None, ""):
                has_value = True
        if has_value and item.get("Ticker"):
            rows.append(item)
    return rows


def parse_idx_screener(ws) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    current_section = ""
    filter_tags = {"A", "B", "C", "D", "E"}

    for row_idx in range(4, ws.max_row + 1):
        first = ws.cell(row_idx, 2).value
        text = str(first or "").strip()
        if text.startswith("Filter "):
            current_section = re.sub(r"\s+", " ", text)
            continue
        if text not in filter_tags:
            continue

        values = [clean_value(ws.cell(row_idx, col).value) for col in range(2, 19)]
        item = dict(zip(SCREENER_COLUMNS, values))
        item["Section"] = current_section or f"Filter {text}"
        rows.append(item)

    return rows


def update_manifest(entry: dict[str, Any]) -> None:
    manifest_path = DATA_DIR / "manifest.json"
    if manifest_path.exists():
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    else:
        manifest = {"latest": None, "dates": []}

    dates = [d for d in manifest.get("dates", []) if d.get("date") != entry["date"]]
    dates.append(entry)
    dates.sort(key=lambda d: d["date"], reverse=True)
    manifest["dates"] = dates
    manifest["latest"] = dates[0]["date"] if dates else None
    manifest_path.write_text(json.dumps(manifest, indent=2, ensure_ascii=False), encoding="utf-8")


def export_workbook(workbook_path: Path) -> Path:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    DOWNLOADS_DIR.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(workbook_path, read_only=False, data_only=True)
    processing = wb["Data Processing Results"]
    market_date = str(find_stat(processing, "As-of Date") or "").strip()
    if not market_date:
        raise ValueError("Could not find As-of Date in Data Processing Results")
    market_date = slug_date(market_date)

    run_time = clean_value(find_stat(processing, "Run Time"))
    total_scanned = clean_value(find_stat(processing, "Total Tickers Scanned"))
    ok_count = clean_value(find_stat(processing, "OK  (full data)"))
    partial_count = clean_value(find_stat(processing, "Partial Data"))
    no_data_count = clean_value(find_stat(processing, "No Data"))

    payload = {
        "date": market_date,
        "runTime": run_time,
        "workbook": f"downloads/{market_date}.xlsx",
        "summary": {
            "totalScanned": total_scanned,
            "ok": ok_count,
            "partial": partial_count,
            "noData": no_data_count,
        },
        "screener": parse_idx_screener(wb["IDX Screener"]),
        "technical": parse_table(wb["IDX Technical Detail"], header_row=6, data_start=7),
        "fundamental": parse_table(wb["IDX Fundamental Detail"], header_row=6, data_start=7),
        "processing": parse_table(wb["Data Processing Results"], header_row=14, data_start=15),
    }

    data_path = DATA_DIR / f"{market_date}.json"
    data_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    shutil.copy2(workbook_path, DOWNLOADS_DIR / f"{market_date}.xlsx")

    update_manifest(
        {
            "date": market_date,
            "runTime": run_time,
            "rows": len(payload["screener"]),
            "file": f"data/{market_date}.json",
            "workbook": f"downloads/{market_date}.xlsx",
        }
    )
    return data_path


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, default=None)
    args = parser.parse_args()

    workbook = args.workbook or latest_workbook()
    data_path = export_workbook(workbook)
    print(f"Exported {workbook.name} -> {data_path}")


if __name__ == "__main__":
    main()
