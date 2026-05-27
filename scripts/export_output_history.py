from __future__ import annotations

import argparse
import json
import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from export_latest import DATA_DIR, DOWNLOADS_DIR, export_workbook, find_stat


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT_DIR = ROOT / "Output"


def filename_market_date(path: Path) -> str | None:
    match = re.search(r"as of (.+?) Running at", path.name)
    if not match:
        return None
    label = re.sub(r"(\d+)(st|nd|rd|th)", r"\1", match.group(1))
    try:
        return datetime.strptime(label, "%B %d %Y").strftime("%Y-%m-%d")
    except ValueError:
        return None


def workbook_market_date(path: Path) -> str | None:
    try:
        wb = load_workbook(path, read_only=False, data_only=True)
        if "Data Processing Results" not in wb.sheetnames:
            return None
        value = find_stat(wb["Data Processing Results"], "As-of Date")
        return str(value).strip() if value else None
    except Exception as exc:
        print(f"[WARN] Skipped {path.name}: {exc}")
        return None


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--days", type=int, default=7)
    args = parser.parse_args()

    if args.days < 1:
        raise ValueError("--days must be at least 1")
    if not args.output_dir.exists():
        raise FileNotFoundError(f"Output folder not found: {args.output_dir}")

    latest_by_date: dict[str, Path] = {}
    workbooks = sorted(
        [p for p in args.output_dir.glob("IDX Screener as of *.xlsx") if not p.name.startswith("~$")],
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    scan_limit = max(args.days * 8, 40)
    for workbook in workbooks[:scan_limit]:
        market_date = filename_market_date(workbook)
        if not market_date:
            market_date = workbook_market_date(workbook)
            if not market_date:
                continue
        current = latest_by_date.get(market_date)
        if current is None or workbook.stat().st_mtime > current.stat().st_mtime:
            latest_by_date[market_date] = workbook
        if len(latest_by_date) >= args.days:
            break

    selected = sorted(latest_by_date.items(), key=lambda item: item[0], reverse=True)[: args.days]
    for market_date, workbook in reversed(selected):
        print(f"Exporting {market_date}: {workbook.name}")
        data_path = DATA_DIR / f"{market_date}.json"
        manifest_path = DATA_DIR / "manifest.json"
        workbook_path = DOWNLOADS_DIR / f"{market_date}.xlsx"
        previous_manifest = manifest_path.read_bytes() if manifest_path.exists() else None
        previous_json = data_path.read_bytes() if data_path.exists() else None
        previous_workbook = workbook_path.read_bytes() if workbook_path.exists() else None
        previous_rows = -1
        if previous_json:
            try:
                previous_rows = len(json.loads(previous_json.decode("utf-8")).get("screener", []))
            except Exception:
                previous_rows = -1
        try:
            export_workbook(workbook)
            if previous_rows >= 0 and data_path.exists():
                new_rows = len(json.loads(data_path.read_text(encoding="utf-8")).get("screener", []))
                if new_rows < previous_rows:
                    data_path.write_bytes(previous_json or b"")
                    if previous_manifest:
                        manifest_path.write_bytes(previous_manifest)
                    if previous_workbook:
                        workbook_path.write_bytes(previous_workbook)
                    print(f"[WARN] Kept existing {market_date} export ({previous_rows} rows > {new_rows} rows).")
        except Exception as exc:
            if previous_manifest:
                manifest_path.write_bytes(previous_manifest)
            if previous_json:
                data_path.write_bytes(previous_json)
            if previous_workbook:
                workbook_path.write_bytes(previous_workbook)
            print(f"[WARN] Could not export {workbook.name}: {exc}")


if __name__ == "__main__":
    main()
